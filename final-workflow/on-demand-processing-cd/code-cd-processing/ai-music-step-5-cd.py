import os
import glob
import datetime
import openpyxl
import requests
import time
from openpyxl.styles import Alignment
from openpyxl import load_workbook

def find_latest_results_folder(prefix):
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None
    return max(matching_folders)

def find_latest_step4_file(results_folder):
    # Find files starting with "ai-music-step-4-" and ending with ".xlsx"
    files = [f for f in os.listdir(results_folder) 
             if f.startswith("ai-music-step-4-") and f.endswith(".xlsx")]
    if not files:
        return None
    latest_file = max(files)
    return os.path.join(results_folder, latest_file)

def create_review_spreadsheet():
    # Set the folder prefix 
    base_dir_prefix = "final-workflow/on-demand-processing-cd/cd-output-folders/results-"
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Please run the previous scripts first.")
        return None
    
    step4_file = find_latest_step4_file(results_folder)
    if not step4_file:
        print("No step 4 file found in the results folder!")
        return None

    print(f"Using source file: {step4_file}")

    # Open the latest step 4 workbook
    wb_src = openpyxl.load_workbook(step4_file)
    sheet_src = wb_src.active

    # Create a new workbook for review
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active

    # Copy columns D, H, I, and K.
    # In the new file these will be columns A, B, C, and D.
    for row in sheet_src.iter_rows(min_row=1, values_only=True):
        new_row = (
            row[3] if len(row) > 3 else None,   # Column D becomes A
            row[7] if len(row) > 7 else None,   # Column H becomes B (OCLC Number)
            row[8] if len(row) > 8 else None,   # Column I becomes C
            row[10] if len(row) > 10 else None  # Column K becomes D
        )
        sheet_new.append(new_row)

    # Set column widths:
    sheet_new.column_dimensions['A'].width = 23
    sheet_new.column_dimensions['B'].width = 23
    sheet_new.column_dimensions['C'].width = 23
    sheet_new.column_dimensions['D'].width = 50

    for cell in sheet_new['D']:
        cell.alignment = Alignment(wrap_text=True)

    # Save the review file using today's date
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    new_file_name = f"results-review-cd-{current_date}.xlsx"
    new_file_path = os.path.join(results_folder, new_file_name)
    
    wb_new.save(new_file_path)
    print(f"Review file created and saved to: {new_file_path}")
    return new_file_path

def get_access_token(client_id, client_secret):
    token_url = "https://oauth.oclc.org/token"
    data = {
        "grant_type": "client_credentials",
        "scope": "wcapi"
    }
    response = requests.post(token_url, data=data, auth=(client_id, client_secret))
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        raise Exception(f"Failed to get access token: {response.text}")

def get_bib_info(oclc_number, access_token):
    """
    Query the OCLC API for bibliographic information for a specific OCLC number.
    """
    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs/{oclc_number}"
    
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    
    try:
        print(f"Making API request to: {endpoint}")
        response = requests.get(endpoint, headers=headers)
        response.raise_for_status()
        
        print(f"API response status: {response.status_code}")
        print(f"API response headers: {response.headers}")
        
        data = response.json()
        print(f"Response keys: {list(data.keys()) if isinstance(data, dict) else 'Not a dictionary'}")
        if isinstance(data, dict) and 'bibRecords' in data:
            print(f"Found {len(data['bibRecords'])} bibliographic records")
        else:
            print("No 'bibRecords' key found in response")
            
        return data
    except requests.RequestException as e:
        print(f"Error getting information for OCLC number {oclc_number}: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"Error response status: {e.response.status_code}")
            print(f"Error response content: {e.response.text}")
        return {"error": str(e)}

def format_bib_info(data):
    """Format bibliographic information for display."""
    output = []
    
    if not isinstance(data, dict) or "identifier" not in data:
        return "No bibliographic information available."
    
    record = data
    
    title = "N/A"
    if "title" in record and "mainTitles" in record["title"]:
        title = record["title"]["mainTitles"][0].get("text", "N/A")
    
    series_title = "N/A"
    if "series" in record and isinstance(record["series"], list) and len(record["series"]) > 0:
        series_title = record["series"][0].get("title", "N/A")
    
    author = "N/A"
    contributors = []
    if "contributor" in record:
        if "creators" in record["contributor"] and record["contributor"]["creators"]:
            for creator in record["contributor"]["creators"]:
                if "nonPersonName" in creator and "text" in creator["nonPersonName"]:
                    author_name = creator["nonPersonName"]["text"]
                    author = author_name
                    contributors.append(author_name)
                elif "firstName" in creator and "secondName" in creator:
                    first_name = creator.get("firstName", {}).get("text", "")
                    second_name = creator.get("secondName", {}).get("text", "")
                    author_name = f"{first_name} {second_name}".strip()
                    author = author_name
                    contributors.append(author_name)
    
    publisher = "N/A"
    place = "N/A"
    date_pub = "N/A"
    if "publishers" in record and record["publishers"]:
        publisher_info = record["publishers"][0]
        if "publisherName" in publisher_info and "text" in publisher_info["publisherName"]:
            publisher = publisher_info["publisherName"]["text"]
        if "publicationPlace" in publisher_info:
            place = publisher_info["publicationPlace"]
    
    if "date" in record and "publicationDate" in record["date"]:
        date_pub = record["date"]["publicationDate"].replace("\u2117", "c")
    
    content_type = "N/A"
    if "format" in record:
        if "generalFormat" in record["format"]:
            content_type = record["format"]["generalFormat"]
        if "specificFormat" in record["format"]:
            content_type += f" - {record['format']['specificFormat']}"
    
    upc = "N/A"
    if "identifier" in record and "otherStandardIdentifiers" in record["identifier"]:
        for identifier in record["identifier"]["otherStandardIdentifiers"]:
            if identifier.get("type") == "Universal Product Code (UPC)":
                upc = identifier.get("id", "N/A")
                break
    
    contents = []
    if "description" in record and "contents" in record["description"]:
        for content_item in record["description"]["contents"]:
            if "titles" in content_item:
                contents = content_item["titles"]
                break
    
    oclc_number = "N/A"
    if "identifier" in record and "oclcNumber" in record["identifier"]:
        oclc_number = record["identifier"]["oclcNumber"]
    
    output.append(f"Title: {title}")
    output.append(f"Series Title: {series_title}")
    output.append(f"Author: {author}")
    output.append(f"Contributors: {', '.join(contributors) if contributors else 'N/A'}")
    output.append(f"Publisher: {publisher}")
    output.append(f"Place of Publication: {place}")
    output.append(f"Date of Publication: {date_pub}")
    output.append(f"Content Type: {content_type}")
    output.append(f"UPC: {upc}")
    
    if contents:
        output.append("Contents:")
        for i, track in enumerate(contents, 1):
            output.append(f"  {i}. {track}")
    
    output.append(f"OCLC Number: {oclc_number}")
    
    return "\n".join(output)

def process_spreadsheet(file_path):
    client_id = os.environ.get("OCLC_CLIENT_ID")
    client_secret = os.environ.get("OCLC_SECRET")

    if not client_id or not client_secret:
        print("Error: OCLC_CLIENT_ID and OCLC_SECRET must be set in environment variables.")
        return
    
    try:
        # Get the access token
        access_token = get_access_token(client_id, client_secret)
        print("Successfully obtained access token.")
        
        print(f"Loading workbook: {file_path}")
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        max_row = sheet.max_row
        print(f"Spreadsheet has {max_row} rows")
        
        # Add header for Column E if it doesn't exist
        if sheet.cell(row=1, column=5).value != "OCLC Record":
            sheet.cell(row=1, column=5).value = "OCLC Record"
        
        # Set column width for column E
        sheet.column_dimensions['E'].width = 60
        
        rows_processed = 0
        for row in range(2, max_row + 1):  # Skip header row
            oclc_number = sheet.cell(row=row, column=2).value  # Column B
            if not oclc_number:
                print(f"Row {row}: No OCLC number, skipping.")
                continue
            
            print(f"Processing row {row}, OCLC number: {oclc_number}")
            oclc_number = str(oclc_number).strip()
            
            # Query OCLC API for bibliographic information
            result = get_bib_info(oclc_number, access_token)
            formatted_info = format_bib_info(result)
            
            # Get holdings information
            base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
            holdings_endpoint = f"{base_url}/bibs-holdings"
            headers = {
                "Authorization": f"Bearer {access_token}",
                "Accept": "application/json"
            }
            params = {
                "oclcNumber": oclc_number,
                "limit": 50
            }
            try:
                print(f"Getting holdings info for OCLC number: {oclc_number}")
                holdings_response = requests.get(holdings_endpoint, params=params, headers=headers)
                holdings_response.raise_for_status()
                holdings_data = holdings_response.json()
                
                is_held_by_IXA = False
                total_holding_count = 0
                
                if "briefRecords" in holdings_data and len(holdings_data["briefRecords"]) > 0:
                    record = holdings_data["briefRecords"][0]
                    if "institutionHolding" in record:
                        holdings = record["institutionHolding"]
                        total_holding_count = holdings.get("totalHoldingCount", 0)
                        if "briefHoldings" in holdings:
                            for holding in holdings["briefHoldings"]:
                                if holding.get("oclcSymbol", "") == "IXA":
                                    is_held_by_IXA = True
                                    break
                
                holdings_info = f"\nTotal Institutions Holding: {total_holding_count}\nHeld by IXA: {'Yes' if is_held_by_IXA else 'No'}"
                formatted_info += holdings_info
                
            except requests.RequestException as e:
                print(f"Error getting holdings for OCLC number {oclc_number}: {str(e)}")
                formatted_info += "\nError retrieving holdings information."
            
            cell = sheet.cell(row=row, column=5)
            cell.value = formatted_info
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            print("Formatted information:")
            print(formatted_info)
            
            rows_processed += 1
            print(f"Row {row} processed successfully.")
            
            # Small delay to avoid API rate limits
            time.sleep(0.5)
        
        workbook.save(file_path)
        print(f"Processed {rows_processed} rows. Results saved to {file_path}")
        
    except Exception as e:
        print(f"Error processing spreadsheet: {str(e)}")

def main():
    # Part 1: Create the review spreadsheet
    new_file_path = create_review_spreadsheet()
    if not new_file_path:
        return
    
    # Part 2: Process the new spreadsheet to add bibliographic info in Column E
    process_spreadsheet(new_file_path)

if __name__ == "__main__":
    main()
