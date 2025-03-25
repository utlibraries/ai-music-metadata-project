import os
import glob
import time
from openai import OpenAI
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import re

def find_latest_results_folder(prefix):
    # Get the parent directory of the prefix
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None

    latest_folder = max(matching_folders)
    
    return latest_folder

# Load the API key from environment variable
client = OpenAI(api_key=os.getenv('OPENAI_HMRC_API_KEY'))

def main():
    # Start timing the entire script execution
    script_start_time = time.time()
    
    # Specify the folder prefix (adjust if needed)
    base_dir_prefix = "final-workflow/on-demand-processing-cd/cd-output-folders/results-"

    # Find the latest results folder using the prefix.
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Run the first script first.")
        exit()
        
    print(f"Using results folder: {results_folder}")

    # Look for step 2 files in the results folder
    step2_files = [f for f in os.listdir(results_folder)
                if f.startswith('ai-music-step-2-') and f.endswith('.xlsx')]

    if not step2_files:
        print("No step 2 files found in the results folder!")
        exit()
        
    # Get the latest step 2 file
    latest_file = max(step2_files)
    workbook_path = os.path.join(results_folder, latest_file)

    print(f"Processing file: {workbook_path}")

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb.active
    
    # Create a temporary workbook for frequent saving (no images)
    temp_wb = openpyxl.Workbook()
    temp_sheet = temp_wb.active
    
    # Copy headers and column settings
    for col_idx, cell in enumerate(sheet[1], 1):
        temp_sheet.cell(row=1, column=col_idx, value=cell.value)
        # Copy column widths where available
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        if column_letter in sheet.column_dimensions:
            temp_sheet.column_dimensions[column_letter].width = sheet.column_dimensions[column_letter].width

    # Create a summary sheet for token and timing metrics
    if "TokenSummary" not in wb.sheetnames:
        summary_sheet = wb.create_sheet("TokenSummary")
    else:
        summary_sheet = wb["TokenSummary"]
    
    # Also create summary in temp workbook
    temp_summary_sheet = temp_wb.create_sheet("TokenSummary")
    
    # Set up the summary sheet headers
    summary_headers = [
        "Total Rows Processed", "Total API Calls", "Failed API Calls",
        "Total Processing Time (s)", "API Time (s)", "Average Time per Call (s)",
        "Total Prompt Tokens", "Total Completion Tokens", "Total Tokens",
        "Average Tokens per Call", "Estimated Cost ($0.003/1K)", "Date Completed"
    ]
    summary_sheet.append(summary_headers)
    temp_summary_sheet.append(summary_headers)
    
    for col, header in enumerate(summary_headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col)
        summary_sheet.column_dimensions[col_letter].width = 20
        temp_summary_sheet.column_dimensions[col_letter].width = 20

    # Temporary file path for saving progress
    temp_output_file = "temp_step3_progress.xlsx"
    temp_output_path = os.path.join(results_folder, temp_output_file)

    # Define the columns
    METADATA_COLUMN = 'E'
    OCLC_RESULTS_COLUMN = 'G'
    RESULT_COLUMN = 'H'
    CONFIDENCE_SCORE_COLUMN = 'I'
    EXPLANATION_COLUMN = 'J'
    OTHER_MATCHES_COLUMN = 'K'  # Column for other potential matches
    PROCESSING_TIME_COLUMN = 'L'  # New column for processing time
    PROMPT_TOKENS_COLUMN = 'M'   # New column for prompt tokens
    COMPLETION_TOKENS_COLUMN = 'N'  # New column for completion tokens
    TOTAL_TOKENS_COLUMN = 'O'    # New column for total tokens


    # Add headers for all columns including new ones
    sheet[f'{RESULT_COLUMN}1'] = 'LLM-Assessed Correct OCLC #'
    sheet[f'{CONFIDENCE_SCORE_COLUMN}1'] = 'LLM Confidence Score'
    sheet[f'{EXPLANATION_COLUMN}1'] = 'LLM Explanation'
    sheet[f'{OTHER_MATCHES_COLUMN}1'] = 'Other Potential Matches'
    sheet[f'{PROCESSING_TIME_COLUMN}1'] = 'Processing Time (s)'
    sheet[f'{PROMPT_TOKENS_COLUMN}1'] = 'Prompt Tokens'
    sheet[f'{COMPLETION_TOKENS_COLUMN}1'] = 'Completion Tokens'
    sheet[f'{TOTAL_TOKENS_COLUMN}1'] = 'Total Tokens'
    
    # Add the same headers to temp sheet
    temp_sheet[f'{RESULT_COLUMN}1'] = 'LLM-Assessed Correct OCLC #'
    temp_sheet[f'{CONFIDENCE_SCORE_COLUMN}1'] = 'LLM Confidence Score'
    temp_sheet[f'{EXPLANATION_COLUMN}1'] = 'LLM Explanation'
    temp_sheet[f'{OTHER_MATCHES_COLUMN}1'] = 'Other Potential Matches'
    temp_sheet[f'{PROCESSING_TIME_COLUMN}1'] = 'Processing Time (s)'
    temp_sheet[f'{PROMPT_TOKENS_COLUMN}1'] = 'Prompt Tokens'
    temp_sheet[f'{COMPLETION_TOKENS_COLUMN}1'] = 'Completion Tokens'
    temp_sheet[f'{TOTAL_TOKENS_COLUMN}1'] = 'Total Tokens'

    # Set column widths
    sheet.column_dimensions[RESULT_COLUMN].width = 30
    sheet.column_dimensions[CONFIDENCE_SCORE_COLUMN].width = 20
    sheet.column_dimensions[EXPLANATION_COLUMN].width = 40
    sheet.column_dimensions[OTHER_MATCHES_COLUMN].width = 70
    sheet.column_dimensions[PROCESSING_TIME_COLUMN].width = 18
    sheet.column_dimensions[PROMPT_TOKENS_COLUMN].width = 15
    sheet.column_dimensions[COMPLETION_TOKENS_COLUMN].width = 18
    sheet.column_dimensions[TOTAL_TOKENS_COLUMN].width = 15
    
    # Mirror column widths in temp sheet
    temp_sheet.column_dimensions[RESULT_COLUMN].width = 30
    temp_sheet.column_dimensions[CONFIDENCE_SCORE_COLUMN].width = 20
    temp_sheet.column_dimensions[EXPLANATION_COLUMN].width = 40
    temp_sheet.column_dimensions[OTHER_MATCHES_COLUMN].width = 70
    temp_sheet.column_dimensions[PROCESSING_TIME_COLUMN].width = 18
    temp_sheet.column_dimensions[PROMPT_TOKENS_COLUMN].width = 15
    temp_sheet.column_dimensions[COMPLETION_TOKENS_COLUMN].width = 18
    temp_sheet.column_dimensions[TOTAL_TOKENS_COLUMN].width = 15

    # Initialize counters for summary
    total_rows = 0
    successful_calls = 0
    failed_calls = 0
    total_api_time = 0
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    processed_rows = 0

    for row in range(2, sheet.max_row + 1):  # Row 1 is the header
        row_start_time = time.time()
        metadata = sheet[f'{METADATA_COLUMN}{row}'].value
        oclc_results = sheet[f'{OCLC_RESULTS_COLUMN}{row}'].value

        # Skip rows with missing data or "No matching records" message
        if not metadata or not oclc_results or oclc_results == "No matching records found" or oclc_results.strip() == "":
            print(f"Skipping row {row}: Missing data or no OCLC results")
            # Explicitly mark these rows as skipped in the results
            sheet[f'{RESULT_COLUMN}{row}'].value = "No OCLC data to process"
            sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = 0
            sheet[f'{EXPLANATION_COLUMN}{row}'].value = "Skipped: No valid OCLC results to analyze"
            processed_rows += 1
            
            # Mirror in temp sheet
            temp_sheet[f'{RESULT_COLUMN}{row}'].value = "No OCLC data to process"
            temp_sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = 0
            temp_sheet[f'{EXPLANATION_COLUMN}{row}'].value = "Skipped: No valid OCLC results to analyze"
            
            # Copy data from other columns to temp sheet
            for col in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(col)
                temp_sheet[f'{col_letter}{row}'].value = sheet[f'{col_letter}{row}'].value
                
            # Increment the skipped counter for summary metrics
            total_rows += 1  # Still count this as a processed row
            failed_calls += 1  # Count as a "failed" call for metrics
            continue
        
        # Only rows with valid data will reach this point
        total_rows += 1
            
        prompt = (
f'''Analyze the following OCLC results based on the given metadata and determine which result is the best match. Methodically go through each record, choose the top 3, then consider them again and choose the record that matches the most elements in the metadata. If two or more records tie for best match, prioritize records that have more holdings and that are held by IXA. If there is no likely match, write "No matching records found".

**Important Instructions**:
1. Confidence Score: 0% indicates no confidence, and 100% indicates high confidence that we have found the correct OCLC number. At or below 80%, the record will be checked by a cataloger. 
2. ***Key Fields in order of importance***:
   - UPC/Product Code (a match is HIGHEST priority if available in both metadata and OCLC record - if not available in one or the other, skip this field.  Occasionally, the UPC is partially obscured in the metadata - if some of the numbers in a UPC are incorrect but other fields are matching, it is still a match)
   - Title 
   - Artist/Performer
   - Contributors (some matching - not all need to match)
   - Publisher Name (exact match preferred; corporate ownership relationships like "Columbia is part of Sony" should NOT be considered a match)
   - Physical Description (should make sense for a CD)
   - Content (track listings - these should be mostly similar, with small differences in spelling or punctuation)
   - Year (this is a key field only if present in the metadata, AND not marked as being a sticker date)
3. ***Notes on Matching Special Cases***:
   - Titles in non-Latin scripts that match in meaning or transliteration should also be considered equivalent.
   - If a field is marked as partially obscured, lessen the importance of that field in the matching process.
   - Different releases in different regions (e.g., Japanese release vs. US release) should be treated as different records even if title and content match.
4. When information is not visible in the metadata, do not use that field in your consideration of a match. 
5. If there is no likely match, write "No matching records found" and set the confidence score as 0.

Format for Response:
- Your response must follow this format exactly:
  1. OCLC number: [number or 'No matching records found']
  2. Confidence score: [%]
  3. Explanation: [List of things that match as key value pairs. If there are multiple records that could be a match, explain why you chose the one you did. If there are no matches, explain why.]
  4. Other potential good matches: [List of other OCLC numbers that could be good matches and a one sentence explanation for each match as key value pairs. If there are no other potential matches, write 'No other potential good matches.']
  
Once you have responded, go back through the response that you wrote and carefully verify each piece of information. If you find a mistake, look for a better record. If there isn't one, reduce the confidence score to 80% or lower. If there is one, once again carefully verify all the facts that support your choice. If you still can't find a match, write "No matching records found" and set the confidence score as 0.

Metadata: {metadata}

OCLC Results: {oclc_results}
''')
        try:
            # Time the API call
            api_call_start = time.time()
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a music cataloger.  You are very knowledgeable about music cataloging best practices, and also have incredible attention to detail.  Read through the metadata and OCLC results carefully, and determine which of the OCLC results looks like the best match. If there is no likely match, write 'No matching records found'.  If you make a mistake, you would feel very bad about it, so you always double check your work."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1000,
                temperature=0.5
            )
            api_call_duration = time.time() - api_call_start
            total_api_time += api_call_duration
            
            # Extract token usage
            prompt_tokens = response.usage.prompt_tokens
            completion_tokens = response.usage.completion_tokens
            tokens_used = prompt_tokens + completion_tokens
            
            # Add to totals
            total_prompt_tokens += prompt_tokens
            total_completion_tokens += completion_tokens
            total_tokens += tokens_used
            
            successful_calls += 1

            analysis_result = response.choices[0].message.content.strip()

            oclc_number = "Not found"
            confidence_score = "0"
            explanation = "Could not parse response"
            other_matches = ""

            try:
                if "OCLC number:" in analysis_result:
                    oclc_part = analysis_result.split("OCLC number:")[1].split("\n")[0].strip()
                    oclc_number = ''.join(char for char in oclc_part if char.isdigit())

                if "Confidence score:" in analysis_result:
                    confidence_part = analysis_result.split("Confidence score:")[1].split("%")[0].strip()
                    try:
                        confidence_score = int(float(confidence_part))
                        confidence_score = min(100, max(0, confidence_score))
                    except ValueError:
                        confidence_score = 0

                try:
                    confidence_score = min(100, int(confidence_score))
                except ValueError:
                    confidence_score = 0

                if "Explanation:" in analysis_result:
                    explanation_parts = analysis_result.split("Explanation:")[1].split("Other potential good matches:")
                    explanation = explanation_parts[0].strip()
                    if explanation.endswith("4."):
                        explanation = explanation[:-2].strip()
                    explanation = re.sub(r'\s+\d+\.\s*$', '', explanation)

                if "Other potential good matches:" in analysis_result:
                    other_matches_part = analysis_result.split("Other potential good matches:")[1].strip()
                    if other_matches_part and oclc_results:
                        # Use improved regex to extract OCLC numbers from the LLM response
                        # Look for patterns like "OCLC Number: 1234567890" or "- OCLC Number: 1234567890"
                        oclc_patterns = re.findall(r'OCLC Number:?\s*(\d{8,10})', other_matches_part, re.IGNORECASE)
                        
                        # If the above pattern doesn't find matches, try a more generic pattern
                        if not oclc_patterns:
                            # Look for patterns that might include a dash before OCLC
                            oclc_patterns = re.findall(r'[- ]*OCLC(?:\s+Number)?:?\s*(\d{8,10})', other_matches_part, re.IGNORECASE)
                        
                        # If still no matches, try to find any 8-10 digit number that might be an OCLC number
                        if not oclc_patterns:
                            oclc_patterns = re.findall(r'\b(\d{8,10})\b', other_matches_part)
                        
                        formatted_matches = []
                        
                        # Process each potential match
                        for match_num in oclc_patterns:
                            match_info = ""
                            # Find the section in OCLC results that contains this number
                            match_section = None
                            
                            # Look for the OCLC section in the results
                            if f"OCLC Number: {match_num}" in oclc_results:
                                # Split by the OCLC number and take the part after it
                                split_results = oclc_results.split(f"OCLC Number: {match_num}")
                                if len(split_results) > 1:
                                    # Take the part after the OCLC number and split by the next dashed line
                                    match_section = split_results[1].split("----------------------------------------")[0]
                            
                            if match_section:
                                held_by_ixa = "No"
                                if "Held by IXA: Yes" in match_section:
                                    held_by_ixa = "Yes"
                                
                                total_holdings = "0"
                                if "Total Institutions Holding:" in match_section:
                                    holdings_part = match_section.split("Total Institutions Holding:")[1].split("\n")[0].strip()
                                    total_holdings = holdings_part
                                
                                title = ""
                                if "- Main Title:" in match_section:
                                    title_part = match_section.split("- Main Title:")[1].split("\n")[0].strip()
                                    title = title_part
                                
                                specific_format = ""
                                if "- specificFormat:" in match_section:
                                    format_part = match_section.split("- specificFormat:")[1].split("\n")[0].strip()
                                    specific_format = format_part
                                
                                # Always include the match info regardless of format - this ensures we get the matches
                                match_info = f"OCLC: {match_num} | IXA: {held_by_ixa} | Holdings: {total_holdings}"
                                if specific_format:
                                    match_info += f" | Format: {specific_format}"
                                if title:
                                    match_info += f" | Title: {title}"
                                
                                if match_info:
                                    formatted_matches.append(match_info)
                        
                        # Add the structured information followed by the original text
                        if formatted_matches:
                            other_matches = "\n".join(formatted_matches) + "\n\nOriginal LLM response:\n" + other_matches_part
                        else:
                            other_matches = "No structured matches found.\n\nOriginal LLM response:\n" + other_matches_part
                    else:
                        other_matches = other_matches_part
            
            except Exception as parsing_error:
                print(f"Error parsing response in row {row}: {parsing_error}")

            # Calculate total processing time for this row
            row_duration = time.time() - row_start_time
            processed_rows += 1
            
            # Update cells with results and metrics
            result_cell = sheet[f'{RESULT_COLUMN}{row}']
            confidence_cell = sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}']
            explanation_cell = sheet[f'{EXPLANATION_COLUMN}{row}']
            other_matches_cell = sheet[f'{OTHER_MATCHES_COLUMN}{row}']
            time_cell = sheet[f'{PROCESSING_TIME_COLUMN}{row}']
            prompt_tokens_cell = sheet[f'{PROMPT_TOKENS_COLUMN}{row}']
            completion_tokens_cell = sheet[f'{COMPLETION_TOKENS_COLUMN}{row}']
            total_tokens_cell = sheet[f'{TOTAL_TOKENS_COLUMN}{row}']

            result_cell.value = oclc_number if isinstance(oclc_number, str) else int(oclc_number)
            confidence_cell.value = confidence_score
            explanation_cell.value = explanation
            other_matches_cell.value = other_matches
            time_cell.value = round(row_duration, 2)
            prompt_tokens_cell.value = prompt_tokens
            completion_tokens_cell.value = completion_tokens
            total_tokens_cell.value = tokens_used

            # Update temp workbook with the same data
            temp_result_cell = temp_sheet[f'{RESULT_COLUMN}{row}']
            temp_confidence_cell = temp_sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}']
            temp_explanation_cell = temp_sheet[f'{EXPLANATION_COLUMN}{row}']
            temp_other_matches_cell = temp_sheet[f'{OTHER_MATCHES_COLUMN}{row}']
            temp_time_cell = temp_sheet[f'{PROCESSING_TIME_COLUMN}{row}']
            temp_prompt_tokens_cell = temp_sheet[f'{PROMPT_TOKENS_COLUMN}{row}']
            temp_completion_tokens_cell = temp_sheet[f'{COMPLETION_TOKENS_COLUMN}{row}']
            temp_total_tokens_cell = temp_sheet[f'{TOTAL_TOKENS_COLUMN}{row}']
            
            temp_result_cell.value = oclc_number if isinstance(oclc_number, str) else int(oclc_number)
            temp_confidence_cell.value = confidence_score
            temp_explanation_cell.value = explanation
            temp_other_matches_cell.value = other_matches
            temp_time_cell.value = round(row_duration, 2)
            temp_prompt_tokens_cell.value = prompt_tokens
            temp_completion_tokens_cell.value = completion_tokens
            temp_total_tokens_cell.value = tokens_used
            
            # Copy data from other columns to temp sheet
            for col in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(col)
                temp_sheet[f'{col_letter}{row}'].value = sheet[f'{col_letter}{row}'].value

            # Set cell alignment in both workbooks
            for cell in [result_cell, confidence_cell, explanation_cell, other_matches_cell, 
                        time_cell, prompt_tokens_cell, completion_tokens_cell, total_tokens_cell]:
                cell.alignment = Alignment(wrap_text=True)
                
            for cell in [temp_result_cell, temp_confidence_cell, temp_explanation_cell, temp_other_matches_cell, 
                        temp_time_cell, temp_prompt_tokens_cell, temp_completion_tokens_cell, temp_total_tokens_cell]:
                cell.alignment = Alignment(wrap_text=True)

            print(f"\n--- Processed row {row}/{sheet.max_row} ---")
            print(f"OCLC Number: {result_cell.value}")
            print(f"Confidence Score: {confidence_cell.value}%")
            print(f"Processing Time: {round(row_duration, 2)}s (API: {round(api_call_duration, 2)}s)")
            print(f"Tokens: {tokens_used} (Prompt: {prompt_tokens}, Completion: {completion_tokens})")
            print(f"Explanation: {explanation_cell.value}")
            if other_matches:
                print(f"Other Potential Matches: \n{other_matches}")
            print("-----------------------------------")
            
            # Save temporary workbook every 10 processed rows
            if processed_rows % 10 == 0:
                # Update summary data in temp workbook
                avg_call_time = total_api_time / successful_calls if successful_calls > 0 else 0
                avg_tokens_per_call = total_tokens / successful_calls if successful_calls > 0 else 0
                estimated_cost = (total_tokens / 1000) * 0.003  # Estimate based on $0.003 per 1K tokens
                
                # Clear old summary data and add new
                for sum_row in temp_summary_sheet.iter_rows(min_row=2, max_row=2):
                    for cell in sum_row:
                        cell.value = None
                
                temp_summary_sheet.append([
                    total_rows,
                    successful_calls,
                    failed_calls,
                    round(time.time() - script_start_time, 2),
                    round(total_api_time, 2),
                    round(avg_call_time, 2),
                    total_prompt_tokens,
                    total_completion_tokens,
                    total_tokens,
                    round(avg_tokens_per_call, 2),
                    round(estimated_cost, 4),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
                
                try:
                    temp_wb.save(temp_output_path)
                    print(f"Progress saved ({row}/{sheet.max_row} rows)")
                except Exception as save_error:
                    print(f"Warning: Could not save temporary progress: {save_error}")

        except Exception as e:
            failed_calls += 1
            print(f"Error processing row {row}: {e}")
            # Update both workbooks to show the error
            sheet[f'{RESULT_COLUMN}{row}'].value = "Error processing"
            sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = 0
            sheet[f'{EXPLANATION_COLUMN}{row}'].value = f"Error: {str(e)}"
            
            temp_sheet[f'{RESULT_COLUMN}{row}'].value = "Error processing"
            temp_sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = 0
            temp_sheet[f'{EXPLANATION_COLUMN}{row}'].value = f"Error: {str(e)}"
            
            # Copy data from other columns to temp sheet
            for col in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(col)
                temp_sheet[f'{col_letter}{row}'].value = sheet[f'{col_letter}{row}'].value

    # Calculate script metrics
    script_duration = time.time() - script_start_time
    avg_call_time = total_api_time / successful_calls if successful_calls > 0 else 0
    avg_tokens_per_call = total_tokens / successful_calls if successful_calls > 0 else 0
    
    # Add summary data
    summary_sheet.append([
        total_rows,
        successful_calls,
        failed_calls,
        round(script_duration, 2),
        round(total_api_time, 2),
        round(avg_call_time, 2),
        total_prompt_tokens,
        total_completion_tokens,
        total_tokens,
        round(avg_tokens_per_call, 2),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    # Create a token usage log file
    log_file_path = os.path.join(results_folder, "oclc_token_usage_log.txt")
    with open(log_file_path, "w") as log_file:
        log_file.write(f"Processing completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        log_file.write(f"Total rows processed: {total_rows}\n")
        log_file.write(f"Successful API calls: {successful_calls}\n")
        log_file.write(f"Failed API calls: {failed_calls}\n")
        log_file.write(f"Total script execution time: {round(script_duration, 2)} seconds\n")
        log_file.write(f"Total API time: {round(total_api_time, 2)} seconds\n")
        log_file.write(f"Average API call time: {round(avg_call_time, 2)} seconds\n")
        log_file.write(f"Total prompt tokens: {total_prompt_tokens}\n")
        log_file.write(f"Total completion tokens: {total_completion_tokens}\n")
        log_file.write(f"Total tokens: {total_tokens}\n")
        log_file.write(f"Average tokens per call: {round(avg_tokens_per_call, 2)}\n")

    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"ai-music-step-3-{current_date}.xlsx"
    full_output_path = os.path.join(results_folder, output_file)
        
    wb.save(full_output_path)
    
    # Clean up temporary file
    try:
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
            print(f"Temporary progress file removed: {temp_output_path}")
    except Exception as remove_error:
        print(f"Warning: Could not remove temporary progress file: {remove_error}")
    
    print(f"\nResults saved to {full_output_path}")
    print(f"Token usage log saved to {log_file_path}")
    print("\nSummary:")
    print(f"- Total rows processed: {total_rows}")
    print(f"- Successful API calls: {successful_calls}")
    print(f"- Failed API calls: {failed_calls}")
    print(f"- Total script execution time: {round(script_duration, 2)} seconds")
    print(f"- Total API time: {round(total_api_time, 2)} seconds")
    print(f"- Total tokens used: {total_tokens} (Prompt: {total_prompt_tokens}, Completion: {total_completion_tokens})")

if __name__ == "__main__":
    main()