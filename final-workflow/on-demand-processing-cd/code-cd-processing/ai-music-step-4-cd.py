import os
import glob
import re
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from difflib import SequenceMatcher

def find_latest_results_folder(prefix):
    base_dir = os.path.dirname(prefix)
    folder_prefix = os.path.basename(prefix)
    pattern = os.path.join(base_dir, folder_prefix + "*")
    
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None

    latest_folder = max(matching_folders, key=os.path.getmtime)
    return latest_folder

def extract_tracks_from_metadata(metadata_str):
    """Extract track listings from metadata string."""
    tracks = []
    
    # First try to find tracks in the structured JSON-like format
    content_section = re.search(r'Contents:\s*-\s*tracks:\s*\[(.*?)\]', metadata_str, re.DOTALL)
    if content_section:
        tracks_content = content_section.group(1)
        
        # Try to extract track objects using regex for JSON-like structures
        track_objects = re.finditer(r'\{\s*"number":\s*\d+,\s*"title":\s*"([^"]+)"', tracks_content)
        for match in track_objects:
            title = match.group(1)
            if title and title.strip() and title.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                tracks.append(title.strip())
        
        # If the above didn't work, try a simpler approach for title extraction
        if not tracks:
            title_matches = re.finditer(r'"title":\s*([^,\n]+)', tracks_content)
            for match in title_matches:
                title_text = match.group(1).strip()
                # Remove quotes if present
                if title_text.startswith('"') and title_text.endswith('"'):
                    title_text = title_text[1:-1]
                # Remove trailing comma if present
                if title_text.endswith(','):
                    title_text = title_text[:-1]
                
                if title_text and title_text.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                    tracks.append(title_text)
    
    # If no tracks found or tracks list is suspiciously short, try alternative methods
    if len(tracks) < 3:
        # Try to find individual track objects with more flexible patterns
        track_patterns = [
            r'"number":\s*\d+,\s*"title":\s*"([^"]+)"',
            r'"number":\s*\d+,\s*"title":\s*([^,\n]+),',
            r'"title":\s*"([^"]+)"[^}]*?"duration":\s*(\d+:\d+)',
            r'"title":\s*"([^"]+)"'
        ]
        
        for pattern in track_patterns:
            if len(tracks) < 3:
                found_tracks = re.findall(pattern, metadata_str)
                for found in found_tracks:
                    title = found[0] if isinstance(found, tuple) else found
                    cleaned = title.strip().rstrip(',')
                    if cleaned and cleaned.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                        if cleaned not in tracks:
                            tracks.append(cleaned)
    
    # Last resort: look for tracks in an unstructured format
    if len(tracks) < 3:
        track_sections = re.findall(r'(?:Track\s+list(?:ing)?|Contents|Tracks):\s*(.*?)(?:(?:\n\s*\w+:)|$)', 
                                   metadata_str, re.DOTALL | re.IGNORECASE)
        
        for section in track_sections:
            potential_tracks = re.findall(r'(?:\d+[\.\)]\s*|"\s*)([^"\n\(]+)(?:"|\n|\(|$)', section)
            potential_tracks += re.findall(r'([^,;]+)\s*\(\d+:\d+\)', section)
            
            for track in potential_tracks:
                cleaned = track.strip()
                if cleaned and cleaned.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                    if cleaned not in tracks:
                        tracks.append(cleaned)
    
    # Filter out field names rather than track titles
    tracks = [t for t in tracks if t.lower() not in [
        "number", "title", "titletransliteration", "composer", "lyricist", 
        "duration", "isrc", "not applicable", "not visible"
    ]]
    
    # Filter out any track that looks like a note or is too long
    tracks = [t for t in tracks if not (
        "note" in t.lower() or 
        t.lower().startswith("contains") or 
        len(t.split()) > 8
    )]
    
    return tracks

def extract_tracks_from_oclc(oclc_results, oclc_number):
    """Extract track listings from OCLC results for a specific OCLC number."""
    tracks = []
    
    # First, find the section for this OCLC number
    oclc_section_pattern = rf"OCLC Number: {oclc_number}.*?(?:(?:Record \d+:|----------------------------------------)|$)"
    oclc_section = re.search(oclc_section_pattern, oclc_results, re.DOTALL)
    
    if oclc_section:
        section_text = oclc_section.group(0)
        
        # Look for a Content section containing track listings
        content_patterns = [
            r'Content:\s*(.*?)(?:(?:\n\s*[A-Z][a-z]+:)|$)',
            r'Description:.*?Content:\s*(.*?)(?:(?:\n\s*[A-Z][a-z]+:)|$)'
        ]
        
        content_text = None
        for pattern in content_patterns:
            content_match = re.search(pattern, section_text, re.DOTALL)
            if content_match:
                content_text = content_match.group(1).strip()
                break
        
        if content_text:
            if " -- " in content_text:
                track_parts = content_text.split(" -- ")
                for part in track_parts:
                    track_name = part.strip()
                    if track_name.endswith('.'):
                        track_name = track_name[:-1].strip()
                    track_name = re.sub(r'\s*/\s*[^(]+', '', track_name)
                    track_name = re.sub(r'\s*\(\d+[:\.]\d+\)\.?$', '', track_name)
                    track_name = re.sub(r'\s*\([^)]*\)$', '', track_name)
                    
                    if track_name and track_name.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                        tracks.append(track_name.strip())
            else:
                for delimiter in ['\n', ';', ',']:
                    if delimiter in content_text and not tracks:
                        parts = content_text.split(delimiter)
                        for part in parts:
                            clean_part = part.strip()
                            if clean_part.endswith('.'):
                                clean_part = clean_part[:-1].strip()
                            clean_part = re.sub(r'\s*/\s*[^(]+', '', clean_part)
                            clean_part = re.sub(r'\s*\(\d+[:\.]\d+\)\.?$', '', clean_part)
                            
                            if clean_part and clean_part.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                                tracks.append(clean_part)
        
        if not tracks:
            track_pattern = r'([^-\(\)]+?)\s*\(\d+[:\.]\d+\)'
            track_matches = re.findall(track_pattern, section_text)
            for match in track_matches:
                clean_track = match.strip()
                if clean_track and clean_track.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                    if clean_track not in tracks:
                        tracks.append(clean_track)
    
    return tracks

def normalize_track(track):
    """Normalize track titles for better comparison."""
    norm = track.lower()
    if norm.startswith("the "):
        norm = norm[4:] + ", the"
    norm = norm.replace(" is a ", " is ").replace(" is the ", " is ")
    norm = norm.replace("(stripped)", "").replace("(edit)", "").replace("stripped", "").replace("edit", "")
    norm = re.sub(r'\s*\(with [^)]+\)', '', norm)
    norm = re.sub(r'\s*\([^)]+\)', '', norm)
    norm = re.sub(r'[^\w\s]', '', norm)
    norm = re.sub(r'\s+', ' ', norm).strip()
    return norm

def calculate_track_similarity(metadata_tracks, oclc_tracks):
    """Calculate the similarity between two track listings."""
    if not metadata_tracks or not oclc_tracks:
        return 0.0
    
    processed_metadata_tracks = []
    processed_oclc_tracks = oclc_tracks.copy()
    
    multi_part_groups = {}
    for i, track in enumerate(metadata_tracks):
        part_match = re.match(r'^(?:Part|Movement)\s*(\d+|[IVX]+)$', track, re.IGNORECASE)
        if part_match:
            if i > 0 and not re.match(r'^(?:Part|Movement)', metadata_tracks[i-1], re.IGNORECASE):
                main_title = metadata_tracks[i-1]
                if main_title not in multi_part_groups:
                    multi_part_groups[main_title] = []
                multi_part_groups[main_title].append(track)
    
    for track in metadata_tracks:
        if track not in multi_part_groups:
            is_part = False
            for parts in multi_part_groups.values():
                if track in parts:
                    is_part = True
                    break
            if not is_part:
                processed_metadata_tracks.append(track)
    
    if multi_part_groups:
        for main_title, parts in multi_part_groups.items():
            processed_metadata_tracks.append(f"{main_title} (with {len(parts)} parts)")
    
    if not processed_metadata_tracks:
        processed_metadata_tracks = metadata_tracks
    
    norm_metadata_tracks = [normalize_track(t) for t in processed_metadata_tracks]
    norm_oclc_tracks = [normalize_track(t) for t in processed_oclc_tracks]
    
    print(f"\nNormalized metadata tracks: {norm_metadata_tracks}")
    print(f"Normalized OCLC tracks: {norm_oclc_tracks}")
    
    matches = 0
    matched_tracks = []
    
    for i, meta_track in enumerate(norm_metadata_tracks):
        best_match = 0
        best_match_index = -1
        is_substring_match = False
        is_part_match = False
        
        if "with" in meta_track and "parts" in meta_track:
            main_title = re.sub(r'\s+with \d+ parts', '', meta_track)
            for j, oclc_track in enumerate(norm_oclc_tracks):
                if (main_title in oclc_track) or (oclc_track in main_title):
                    similarity = 0.95
                    is_part_match = True
                else:
                    similarity = SequenceMatcher(None, main_title, oclc_track).ratio()
                
                if similarity > best_match:
                    best_match = similarity
                    best_match_index = j
        else:
            meta_words = set(meta_track.split())
            for j, oclc_track in enumerate(norm_oclc_tracks):
                oclc_words = set(oclc_track.split())
                common_words = meta_words.intersection(oclc_words)
                
                shorter_length = min(len(meta_words), len(oclc_words))
                if shorter_length > 0 and len(common_words) >= max(1, int(shorter_length * 0.6)):
                    word_similarity = len(common_words) / shorter_length
                    similarity = max(0.8, word_similarity)
                    is_substring_match = True
                elif (meta_track in oclc_track) or (oclc_track in meta_track):
                    similarity = max(0.85, SequenceMatcher(None, meta_track, oclc_track).ratio())
                    is_substring_match = True
                else:
                    similarity = SequenceMatcher(None, meta_track, oclc_track).ratio()
                
                if similarity > best_match:
                    best_match = similarity
                    best_match_index = j
        
        orig_track = processed_metadata_tracks[i]
        match_info = f"{i+1}. {orig_track} => "
        if best_match >= 0.8:
            match_symbol = "✓"
            if is_part_match:
                match_symbol += "(multi-part)"
            elif is_substring_match:
                match_symbol += "(substring)"
            match_info += f"{match_symbol} {processed_oclc_tracks[best_match_index]} ({best_match:.2f})"
            matches += best_match
        else:
            if best_match_index >= 0:
                match_info += f"✗ {processed_oclc_tracks[best_match_index]} ({best_match:.2f})"
            else:
                match_info += "✗ No match"
        
        matched_tracks.append(match_info)
    
    if len(norm_metadata_tracks) == 0:
        return 0.0
    
    print("\nTrack matching details:")
    for match in matched_tracks:
        print(f"  {match}")
    
    similarity = matches / len(norm_metadata_tracks)
    print(f"Total matches: {matches:.2f} out of {len(norm_metadata_tracks)} tracks")
    
    if multi_part_groups and similarity * 100 < 80:
        adjusted_similarity = min(80.0, similarity * 100 + 10.0)
        print(f"Applying multi-part track bonus: final similarity {adjusted_similarity:.2f}%")
        return adjusted_similarity
    
    return similarity * 100

def extract_and_normalize_year(text, is_oclc=False):
    """Extract and normalize publication year to YYYY format."""
    if not text:
        return None
    
    # Look for publication date in structured format
    if is_oclc:
        date_patterns = [
            r'publicationDate:\s*[©℗]?(\d{4})',
            r'machineReadableDate:\s*(\d{4})',
            r'publicationDate:\s*[©℗]?(\d{4})[^\d]',
            r'Dates:[^p]*publicationDate:\s*[©℗]?(\d{4})',
            r'Date:[^p]*publicationDate:\s*[©℗]?(\d{4})',
            r'publicationDate:\s*[©℗]?c?(\d{4})',
            r'publication(?:Date)?:\s*[©℗]?c?(\d{4})'
        ]
    else:
        date_patterns = [
            r'publicationDate:\s*(\d{4})',
            r'Dates:[^p]*publicationDate:\s*(\d{4})',
            r'Date:[^p]*publicationDate:\s*(\d{4})',
            r'publication(?:Date)?:\s*(\d{4})'
        ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            year = match.group(1)
            if year.isdigit() and 1900 <= int(year) <= datetime.now().year:
                return year
    
    # Look for copyright or phonogram year markers
    year_markers = [r'[©℗](\d{4})', r'[©℗](?:\s*)(\d{4})', r'copyright\s+(\d{4})', r'phonogram\s+(\d{4})']
    for marker in year_markers:
        matches = re.findall(marker, text, re.IGNORECASE)
        if matches:
            for year in matches:
                if year.isdigit() and 1900 <= int(year) <= datetime.now().year:
                    return year
    
    # Look for standalone 4-digit years
    year_pattern = r'(?<!\d)(\d{4})(?!\d)'
    matches = re.findall(year_pattern, text)
    valid_years = [y for y in matches if 1900 <= int(y) <= datetime.now().year]
    if valid_years:
        # Return the most frequently occurring year
        from collections import Counter
        return Counter(valid_years).most_common(1)[0][0]
    
    return None

def compare_publication_years(metadata_year, oclc_year):
    """
    Compare publication years and return a tuple (match_status, details)
    match_status is True if years match or if either year is missing
    match_status is False only if both years are present and don't match
    """
    # If either year is missing, don't count it against the match
    if metadata_year is None or oclc_year is None:
        return (True, f"Incomplete year data: metadata_year={metadata_year}, oclc_year={oclc_year}")
    
    # Normalize years to strings and compare
    if metadata_year == oclc_year:
        return (True, f"Years match: {metadata_year} == {oclc_year}")
    else:
        return (False, f"Years do not match: {metadata_year} != {oclc_year}")

def main():
    # Specify the folder prefix based on your output location
    base_dir_prefix = "final-workflow/on-demand-processing-cd/cd-output-folders/results-"
    
    # Find the latest results folder using the prefix
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Run the previous scripts first.")
        exit()
        
    print(f"Using results folder: {results_folder}")
    
    # Look for step 3 files in the results folder
    step3_files = [f for f in os.listdir(results_folder) 
                   if f.startswith('ai-music-step-3-') and f.endswith('.xlsx')]
    
    if not step3_files:
        print("No step 3 files found in the results folder!")
        exit()
        
    latest_file = max(step3_files)
    workbook_path = os.path.join(results_folder, latest_file)
    
    print(f"Processing file: {workbook_path}")
    
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb.active
    
    # Define the columns
    METADATA_COLUMN = 'E'
    OCLC_RESULTS_COLUMN = 'G'
    OCLC_NUMBER_COLUMN = 'H'
    CONFIDENCE_SCORE_COLUMN = 'I'
    EXPLANATION_COLUMN = 'J'
    VERIFICATION_COLUMN = 'L'  # Column for track verification results
    YEAR_VERIFICATION_COLUMN = 'M'  # New column for year verification results
    
    sheet[f'{VERIFICATION_COLUMN}1'] = 'Track Verification Results'
    sheet[f'{YEAR_VERIFICATION_COLUMN}1'] = 'Year Verification Results'
    
    sheet.column_dimensions[VERIFICATION_COLUMN].width = 40
    sheet.column_dimensions[YEAR_VERIFICATION_COLUMN].width = 40
    
    records_processed = 0
    records_adjusted_tracks = 0
    records_adjusted_years = 0
    records_skipped = 0
    
    print(f"Starting verification for records with confidence ≥ 85% that mention tracks...")
    print(f"Total rows in spreadsheet: {sheet.max_row - 1}")
    
    for row in range(2, sheet.max_row + 1):
        try:
            metadata = sheet[f'{METADATA_COLUMN}{row}'].value
            oclc_results = sheet[f'{OCLC_RESULTS_COLUMN}{row}'].value
            oclc_number = sheet[f'{OCLC_NUMBER_COLUMN}{row}'].value
            confidence_score = sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value
            explanation = sheet[f'{EXPLANATION_COLUMN}{row}'].value
            
            if not all([metadata, oclc_results, oclc_number, confidence_score, explanation]):
                continue
                
            try:
                conf_score = float(confidence_score)
                if conf_score < 85:
                    continue
            except (ValueError, TypeError):
                continue
                
            track_related_terms = ["track", "content", "song", "listing"]
            if not explanation or not any(term in explanation.lower() for term in track_related_terms):
                print(f"Skipping row {row}: No track-related terms in explanation")
                continue
                
            records_processed += 1
            print(f"\nProcessing row {row} with OCLC number {oclc_number} (confidence: {confidence_score}%)")
            
            metadata_tracks = extract_tracks_from_metadata(metadata)
            oclc_tracks = extract_tracks_from_oclc(oclc_results, oclc_number)
            
            print(f"Metadata tracks ({len(metadata_tracks)}): {metadata_tracks}")
            print(f"OCLC tracks ({len(oclc_tracks)}): {oclc_tracks}")
            
            # Extract and compare publication years
            metadata_year = extract_and_normalize_year(metadata, is_oclc=False)
            oclc_year = extract_and_normalize_year(oclc_results, is_oclc=True)
            
            print(f"Extracted years - Metadata: {metadata_year}, OCLC: {oclc_year}")
            
            year_match, year_details = compare_publication_years(metadata_year, oclc_year)
            
            if not metadata_year and not oclc_year:
                match_status = "N/A - No years to compare"
            elif not metadata_year or not oclc_year:
                match_status = "Considered match - Incomplete data"
            else:
                match_status = "Yes" if year_match else "No"
                
            year_verification_result = f"Metadata year: {metadata_year if metadata_year else 'Not found'}\nOCLC year: {oclc_year if oclc_year else 'Not found'}\nMatch: {match_status}"
            
            # Skip track similarity check if no tracks found
            track_similarity = 0
            if len(metadata_tracks) == 0 or len(oclc_tracks) == 0:
                print(f"Skipping similarity check: {'No metadata tracks' if len(metadata_tracks) == 0 else 'No OCLC tracks'}")
                verification_result = f"Metadata tracks: {len(metadata_tracks)}\nOCLC tracks: {len(oclc_tracks)}\nSkipped: insufficient track data"
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = verification_result
                sheet[f'{VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            else:
                track_similarity = calculate_track_similarity(metadata_tracks, oclc_tracks)
                print(f"Track similarity: {track_similarity:.2f}%")
                
                matching_tracks = 0
                for i, meta_track in enumerate(metadata_tracks):
                    best_match = 0
                    for oclc_track in oclc_tracks:
                        norm_meta = normalize_track(meta_track)
                        norm_oclc = normalize_track(oclc_track)
                        
                        meta_words = set(norm_meta.split())
                        oclc_words = set(norm_oclc.split())
                        common_words = meta_words.intersection(oclc_words)
                        
                        shorter_length = min(len(meta_words), len(oclc_words))
                        if shorter_length > 0 and len(common_words) >= max(1, int(shorter_length * 0.6)):
                            word_similarity = len(common_words) / shorter_length
                            similarity = max(0.8, word_similarity)
                        elif (norm_meta in norm_oclc) or (norm_oclc in norm_meta):
                            similarity = max(0.85, SequenceMatcher(None, norm_meta, norm_oclc).ratio())
                        else:
                            similarity = SequenceMatcher(None, norm_meta, norm_oclc).ratio()
                        
                        if similarity > best_match:
                            best_match = similarity
                    
                    if best_match >= 0.8:
                        matching_tracks += 1
                
                verification_result = f"Metadata tracks: {len(metadata_tracks)}\nOCLC tracks: {len(oclc_tracks)}\nMatching tracks: {matching_tracks}/{len(metadata_tracks)}\nSimilarity: {track_similarity:.2f}%"
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = verification_result
                sheet[f'{VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = year_verification_result
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            
            # Determine if confidence needs to be adjusted based on both track and year verification
            adjust_confidence = False
            adjustment_reasons = []
            
            # Check track similarity
            if len(metadata_tracks) > 0 and len(oclc_tracks) > 0 and track_similarity < 80:
                adjust_confidence = True
                adjustment_reasons.append(f"track listing mismatch (similarity {track_similarity:.2f}%, below 80% threshold)")
            
            # Check year match - only adjust if both years are present and don't match
            if metadata_year and oclc_year and not year_match:
                adjust_confidence = True
                adjustment_reasons.append(f"publication year mismatch (metadata: {metadata_year}, OCLC: {oclc_year})")
            elif not metadata_year or not oclc_year:
                print(f"Not penalizing for missing year data: metadata_year={metadata_year}, oclc_year={oclc_year}")
            
            # Apply confidence adjustment if needed
            if adjust_confidence:
                old_confidence = confidence_score
                new_confidence = 80
                sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = new_confidence
                
                note = f"\n\n[AUTOMATIC REVIEW: Confidence reduced due to: {'; '.join(adjustment_reasons)}. Please verify manually.]"
                
                # Add track comparison details if needed
                if len(metadata_tracks) > 0 and len(oclc_tracks) > 0 and track_similarity < 80:
                    note += "\n\nTrack comparison:"
                    for i, meta_track in enumerate(metadata_tracks):
                        best_match = 0
                        best_match_track = "No match"
                        
                        for oclc_track in oclc_tracks:
                            norm_meta = normalize_track(meta_track)
                            norm_oclc = normalize_track(oclc_track)
                            
                            meta_words = set(norm_meta.split())
                            oclc_words = set(norm_oclc.split())
                            common_words = meta_words.intersection(oclc_words)
                            
                            shorter_length = min(len(meta_words), len(oclc_words))
                            if shorter_length > 0 and len(common_words) >= max(1, int(shorter_length * 0.6)):
                                word_similarity = len(common_words) / shorter_length
                                similarity = max(0.8, word_similarity)
                            elif (norm_meta in norm_oclc) or (norm_oclc in norm_meta):
                                similarity = max(0.85, SequenceMatcher(None, norm_meta, norm_oclc).ratio())
                            else:
                                similarity = SequenceMatcher(None, norm_meta, norm_oclc).ratio()
                            
                            if similarity > best_match:
                                best_match = similarity
                                best_match_track = oclc_track
                        
                        match_status = "✓" if best_match >= 0.8 else "✗"
                        note += f"\n{i+1}. {meta_track} {match_status} {best_match_track} ({best_match:.2f})"
                
                # Add year comparison details - only for actual mismatches
                if metadata_year and oclc_year and not year_match:
                    note += f"\n\nYear comparison: Metadata year {metadata_year} ≠ OCLC year {oclc_year}"
                
                sheet[f'{EXPLANATION_COLUMN}{row}'].value = explanation + note
                
                if len(metadata_tracks) > 0 and len(oclc_tracks) > 0 and track_similarity < 80:
                    records_adjusted_tracks += 1
                
                if metadata_year and oclc_year and not year_match:
                    records_adjusted_years += 1
                
                # Update verification result with action taken
                actions = []
                verification_result = sheet[f'{VERIFICATION_COLUMN}{row}'].value
                year_verification_result = sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value
                
                if track_similarity < 80 and len(metadata_tracks) > 0 and len(oclc_tracks) > 0:
                    actions.append("track mismatch")
                
                # Only count year mismatch when both years exist but don't match
                if metadata_year and oclc_year and not year_match:
                    actions.append("year mismatch")
                
                if actions:
                    action_text = f"\nAction: Reduced confidence from {old_confidence}% to {new_confidence}% due to {' and '.join(actions)}"
                    if verification_result:
                        sheet[f'{VERIFICATION_COLUMN}{row}'].value = verification_result + action_text
                    else:
                        sheet[f'{VERIFICATION_COLUMN}{row}'].value = action_text
                    
                    sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = year_verification_result + action_text
            else:
                if sheet[f'{VERIFICATION_COLUMN}{row}'].value:
                    sheet[f'{VERIFICATION_COLUMN}{row}'].value += "\nAction: None (similarity is acceptable)"
                
                # For year verification, provide appropriate message based on year data
                year_action = "\nAction: "
                if not metadata_year and not oclc_year:
                    year_action += "None (no year data to compare)"
                elif not metadata_year or not oclc_year:
                    year_action += "None (incomplete year data, not penalized)"
                else:
                    year_action += "None (years match)"
                
                if sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value:
                    sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value += year_action
            
        except Exception as e:
            print(f"Error processing row {row}: {e}")
            sheet[f'{VERIFICATION_COLUMN}{row}'].value = f"Error: {str(e)}"
            sheet[f'{VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = f"Error: {str(e)}"
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"ai-music-step-4-{current_date}.xlsx"
    full_output_path = os.path.join(results_folder, output_file)
    
    wb.save(full_output_path)
    print(f"\nResults saved to {full_output_path}")
    print(f"Summary:")
    print(f"  - Processed: {records_processed} records with confidence ≥ 85% and track listings mentioned")
    print(f"  - Adjusted for tracks: {records_adjusted_tracks} records due to low track similarity (< 80% match)")
    print(f"  - Adjusted for years: {records_adjusted_years} records due to publication year mismatch (only when both years present and don't match)")
    print(f"  - Skipped: {sheet.max_row - 1 - records_processed} records (low confidence or no track listings)")

if __name__ == "__main__":
    main()