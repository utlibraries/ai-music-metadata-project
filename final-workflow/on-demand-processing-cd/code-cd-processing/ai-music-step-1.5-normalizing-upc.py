import re
import os
import glob
from pathlib import Path
import openpyxl

def clean_number(number_text):
    """Remove spaces, dashes, and other non-digit characters from a number string."""
    return re.sub(r'[^0-9]', '', number_text)

def is_valid_upc(number):
    """Check if a number is a valid UPC (12 digits) or EAN (13 digits)."""
    return len(number) in [12, 13]

def extract_upc_from_metadata(metadata_text):
    """Extract and validate UPC numbers from the metadata text."""
    # Find the Numbers line in the Publishers section
    numbers_pattern = r'Publishers:.*?Numbers:\s*(.*?)(?:\n|$)'
    numbers_match = re.search(numbers_pattern, metadata_text, re.DOTALL | re.MULTILINE)
    
    if not numbers_match:
        return None, "No 'Numbers' field found"
    
    numbers_text = numbers_match.group(1).strip()
    
    # Handle "Not visible" case
    if numbers_text.lower() == "not visible":
        return None, "Numbers field marked as 'Not visible'"
    
    # Clean the number
    clean_num = clean_number(numbers_text)
    
    if not clean_num:
        return None, "No digits found in Numbers field"
    
    # Check if it's a valid UPC/EAN
    if is_valid_upc(clean_num):
        return clean_num, None
    else:
        return None, f"Number '{clean_num}' (from '{numbers_text}') is not a valid UPC/EAN (expected 12-13 digits, got {len(clean_num)})"

def process_excel_file(input_file_path):
    """Process the Excel file containing metadata entries and update it in place."""
    # Load the workbook
    wb = openpyxl.load_workbook(input_file_path)
    ws = wb.active
    
    # Get the metadata column index
    headers = [cell.value for cell in ws[1]]
    metadata_col_idx = headers.index('AI-Generated Metadata') + 1
    
    # Process each row
    for row_idx in range(2, ws.max_row + 1):
        metadata_cell = ws.cell(row=row_idx, column=metadata_col_idx)
        
        if metadata_cell.value:
            # Process UPC
            upc, _ = extract_upc_from_metadata(metadata_cell.value)
            
            # Only modify the metadata if it contains a numbers section
            if 'Numbers:' in metadata_cell.value:
                # Replace number section with cleaned version
                updated_metadata = re.sub(
                    r'(Publishers:.*?Numbers:)\s*(.*?)(\n|$)',
                    lambda m: f"{m.group(1)} {upc if upc else 'Not visible'}{m.group(3)}",
                    metadata_cell.value,
                    flags=re.DOTALL | re.MULTILINE
                )
                metadata_cell.value = updated_metadata
    
    # Save the updated workbook back to the same file
    wb.save(input_file_path)

def find_latest_results_folder(prefix):
    # Get the parent directory of the prefix
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None

    latest_folder = max(matching_folders)
    
    return latest_folder

def main():
    # Base directory prefix
    base_dir_prefix = "final-workflow/on-demand-processing-cd/cd-output-folders/results-"
    
    # Find the latest results folder
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        return
    
    # Look for previous step files in the results folder
    input_files = [f for f in os.listdir(results_folder) 
                   if f.startswith('ai-music-step-1-') and f.endswith('.xlsx')]
    
    if not input_files:
        return
    
    latest_file = max(input_files)
    input_file = os.path.join(results_folder, latest_file)
    
    # Process the file
    process_excel_file(input_file)

if __name__ == "__main__":
    main()