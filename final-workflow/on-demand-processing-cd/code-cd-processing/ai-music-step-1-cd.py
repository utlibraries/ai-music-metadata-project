import os
import base64
import re
import time
from datetime import datetime
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openai import OpenAI
from collections import defaultdict

client = OpenAI(api_key=os.getenv('OPENAI_HMRC_API_KEY'))

def get_llm_prompt():
    return """Analyze these images of a compact disc and extract the following key metadata fields in the specified format. You are a music cataloger, and know that you are responsible for the accuracy of the information you produce.  If ANY information is unclear, partially visible, or not visible: mark it as 'Not visible' in the metadadata. 

Match this format:
Title Information:
  - Main Title: [Main Title in original language if using latin characters.  Transliterated if in non-latin characters.]
  - Subtitle: [Subtitle in original language if using latin characters.  Transliterated if in non-latin characters.]
  - Primary Contributor: [Artist/Performer Name]
  - Additional Contributors: [arrangers, engineers, producers, session musicians]
Publishers:
  - Name: [Publisher Name]
  - Place: [Place of publication if available]
  - Numbers: [UPC/EAN/ISBN]
Dates:
  - publicationDate: [Publication Year - if written on a sticker, mark as 'Sticker Date']
Language:
  - sungLanguage: [Languages of sung text]
  - printedLanguage: [All languages of printed text]
Format:
  - generalFormat: [Sound Recording]
  - specificFormat: [CD, CD-ROM, Enhanced CD, etc.]
  - materialTypes: [List of Material Types]
Physical Description:
  - size: [4.75" (standard CD)]
  - material: [aluminum/polycarbonate]
  - labelDesign: [Description of disc face design and color]
  - physicalCondition: [Condition notes]
  - specialFeatures: [Booklet details, packaging type (jewel case/digipak), inserts, bonus materials]
Contents:
  - tracks: [
      {
        "number": [Track number],
        "title": [Title in original language],
        "titleTransliteration": [Title transliteration if applicable],
      }
    ]
Series:
  - seriesTitle: [Series name if any]
  - seriesNumber: [Number within series]
Notes:
  - generalNotes: [{'text': [Note Text]}]

***Important: These CD's were donated by a university radio station to our library. Handwritten information on white stickers should be ignored.  Publication dates or years on stickers should be marked as 'sticker dates', as they may indicate the date of acquisition.  The back cover is often the best place to look for the publication year.  When in doubt, mark fields in the metadata as 'Not visible'*** 
 
Analyze the provided images and return metadata formatted exactly as above. Pay special attention to capturing only text that is clearly legible."""

def group_images_by_barcode(folder_path):
    """Group image files by their barcode number."""
    image_groups = defaultdict(list)
    
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.jpg', '.jpeg')):  
            # Extract barcode number (everything before the letter)
            match = re.match(r'(\d+)[abc]\.jpe?g', filename)
            if match:
                barcode = match.group(1)
                image_groups[barcode].append(os.path.join(folder_path, filename))
            else:
                print(f"Filename does not match pattern: {filename}")
    
    # Sort files within each group by the letter (a, b, c)
    for barcode in image_groups:
        image_groups[barcode].sort(key=lambda x: os.path.basename(x)[-5])  # Sort by the letter before .jpg/.jpeg
        
    return image_groups


def process_folder(folder_path, wb, results_folder_path):
    ws = wb.active
    headers = ['Input Image 1', 'Input Image 2', 'Input Image 3', 'Barcode', 'AI-Generated Metadata', 
               'Processing Time (s)', 'Prompt Tokens', 'Completion Tokens', 'Total Tokens']
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        if col == 4:  # Barcode column
            ws.column_dimensions[get_column_letter(col)].width = 15
        elif col >= 6:  # Token and timing columns
            ws.column_dimensions[get_column_letter(col)].width = 15
        else:
            ws.column_dimensions[get_column_letter(col)].width = 30 if col <= 3 else 52

    # Add a summary worksheet
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(['Total Items', 'Items with Issues', 'Total Time (s)', 
                       'Total Prompt Tokens', 'Total Completion Tokens', 'Total Tokens',
                       'Average Time per Item (s)', 'Average Tokens per Item'])

    image_groups = group_images_by_barcode(folder_path)
    total_items = len(image_groups)
    items_with_issues = 0
    processed_items = 0
    
    # Token and time tracking
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    total_time = 0

    for barcode, image_paths in sorted(image_groups.items()):
        processed_items += 1
        item_start_time = time.time()

        try:
            # Take up to first 3 images for each barcode
            image_paths = image_paths[:3]
            prompt_text = get_llm_prompt()
            uploaded_files_info = ""

            for i, img_path in enumerate(image_paths):
                # Determine image type based on filename
                filename = os.path.basename(img_path)
                if filename.endswith('a.jpg') or filename.endswith('a.jpeg'):
                    image_type = "FRONT COVER"
                elif filename.endswith('b.jpg') or filename.endswith('b.jpeg'):
                    image_type = "BACK COVER"
                elif filename.endswith('c.jpg') or filename.endswith('c.jpeg'):
                    image_type = "ADDITIONAL IMAGE"
                else:
                    image_type = "IMAGE"
                
                uploaded_files_info += f"[Image {i+1} - {image_type}: {img_path}]\n"

            prompt = prompt_text + "\n" + uploaded_files_info

            try:
                base64_images = []
                for img_path in image_paths:
                    with open(img_path, "rb") as image_file:
                        base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                        base64_images.append(base64_image)

                # Start API call timing
                api_start_time = time.time()
                
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            *[{
                                "type": "image_url",
                                "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                            } for base64_image in base64_images]
                        ]
                    }],
                    max_tokens=2000
                )
                
                # Calculate API call duration
                api_duration = time.time() - api_start_time
                
                # Extract token information
                prompt_tokens = response.usage.prompt_tokens
                completion_tokens = response.usage.completion_tokens
                total_item_tokens = prompt_tokens + completion_tokens
                
                # Update totals
                total_prompt_tokens += prompt_tokens
                total_completion_tokens += completion_tokens
                total_tokens += total_item_tokens

                metadata_output = response.choices[0].message.content.strip()
                print(f"Extracted Metadata for barcode {barcode}: {metadata_output}")

                # Calculate total processing time for this item
                item_duration = time.time() - item_start_time
                total_time += item_duration
                
                row_data = [
                    '', '', '', barcode, metadata_output, 
                    round(item_duration, 2), prompt_tokens, completion_tokens, total_item_tokens
                ]
                ws.append(row_data)

                # Add thumbnail images to Excel
                for i, img_path in enumerate(image_paths, start=1):
                    img = PILImage.open(img_path)
                    img.thumbnail((200, 200))

                    output = BytesIO()
                    img.save(output, format='PNG')
                    output.seek(0)

                    img_openpyxl = Image(output)
                    img_openpyxl.anchor = ws.cell(row=ws.max_row, column=i).coordinate
                    ws.add_image(img_openpyxl)

                ws.row_dimensions[ws.max_row].height = 215

                for cell in ws[ws.max_row]:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            except Exception as e:
                print(f"Error generating content for barcode {barcode}: {str(e)}")
                ws.append(['', '', '', barcode, f"Error: {str(e)}", 0, 0, 0, 0])
                items_with_issues += 1

        except Exception as e:
            print(f"Error processing barcode {barcode}: {str(e)}")
            ws.append(['', '', '', barcode, f"Error: {str(e)}", 0, 0, 0, 0])
            items_with_issues += 1

        # Log progress with token usage
        print(f"Processed {processed_items}/{total_items} items. Barcode: {barcode}. Time: {round(item_duration, 2)}s. Tokens: {total_item_tokens if 'total_item_tokens' in locals() else 0}")

    # Add summary data
    avg_time = total_time / total_items if total_items > 0 else 0
    avg_tokens = total_tokens / total_items if total_items > 0 else 0
    
    summary_ws.append([
        total_items, 
        items_with_issues, 
        round(total_time, 2),
        total_prompt_tokens,
        total_completion_tokens,
        total_tokens,
        round(avg_time, 2),
        round(avg_tokens, 2)
    ])
    
    # Create a token usage log file
    log_file_path = os.path.join(results_folder_path, "token_usage_log.txt")
    with open(log_file_path, "w") as log_file:
        log_file.write(f"Processing completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        log_file.write(f"Total items processed: {total_items}\n")
        log_file.write(f"Items with issues: {items_with_issues}\n")
        log_file.write(f"Total processing time: {round(total_time, 2)} seconds\n")
        log_file.write(f"Total prompt tokens: {total_prompt_tokens}\n")
        log_file.write(f"Total completion tokens: {total_completion_tokens}\n")
        log_file.write(f"Total tokens: {total_tokens}\n")
        log_file.write(f"Average time per item: {round(avg_time, 2)} seconds\n")
        log_file.write(f"Average tokens per item: {round(avg_tokens, 2)}\n")

    print(f"Processed {total_items} items. {items_with_issues} items had issues.")
    print(f"Total processing time: {round(total_time, 2)} seconds")
    print(f"Total tokens used: {total_tokens} (Prompt: {total_prompt_tokens}, Completion: {total_completion_tokens})")
    
    return total_items, items_with_issues, total_time, total_tokens


def main():
    start_time = time.time()
    
    base_dir = "final-workflow/on-demand-processing-cd"
    images_folder = os.path.join(base_dir, "cd-input-folders/cd-scans-testing-dates")
    base_dir_outputs = os.path.join(base_dir, "cd-output-folders")
    
    # Create results folder with today's date
    current_date = datetime.now().strftime("%Y-%m-%d")
    results_folder_name = f"results-{current_date}"
    results_folder_path = os.path.join(base_dir_outputs, results_folder_name)

    # Create the folder if it doesn't exist
    if not os.path.exists(results_folder_path):
        os.makedirs(results_folder_path)
    
    wb = Workbook()
    total_items, items_with_issues, total_time, total_tokens = process_folder(images_folder, wb, results_folder_path)

    for row in wb.active.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.active.freeze_panes = 'A2'

    # Save output to the results folder
    output_file = f"ai-music-step-1-{current_date}.xlsx"
    full_output_path = os.path.join(results_folder_path, output_file)

    wb.save(full_output_path)
    
    total_execution_time = time.time() - start_time
    
    print(f"Results saved to {full_output_path}")
    print(f"Summary: Processed {total_items} items, {items_with_issues} with issues.")
    print(f"Total execution time: {round(total_execution_time, 2)} seconds")
    print(f"Total OpenAI API time: {round(total_time, 2)} seconds")
    print(f"Total tokens used: {total_tokens}")
    
if __name__ == "__main__":
    main()