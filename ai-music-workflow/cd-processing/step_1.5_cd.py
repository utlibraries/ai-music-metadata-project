# Clean up publication numbers and dates
import re
import os
import openpyxl

#custom modules
from json_workflow import log_error, log_processing_metrics
from shared_utilities import find_latest_results_folder, get_workflow_json_path
from cd_workflow_config import get_file_path_config

def clean_number(number_text):
    """Clean numeric codes by removing spaces/dashes, but preserve spaces in alphanumeric catalog numbers."""
    # If it contains letters, it's likely a catalog number - remove dashes but keep structure
    if re.search(r'[A-Za-z]', number_text):
        return number_text.strip().replace('-', '')
    # If it's all digits/spaces/dashes, remove spaces and dashes
    else:
        return re.sub(r'[\s\-]+', '', number_text)

def is_valid_upc(number):
    """Check if a number is valid (UPC/EAN/ISBN/catalog) but exclude UT Libraries barcodes and other irrelevant bits of information."""
    # Remove spaces only for digit-only validation
    digits_only = re.sub(r'\s', '', number)
    
    # Exclude UT Libraries barcode stickers (10 digits or 15 digits starting with 05917)
    if digits_only.isdigit() and (len(digits_only) == 10 or (len(digits_only) == 15 and digits_only.startswith('05917'))):
        return False
    
    # Accept catalog numbers (contain letters) with reasonable constraints
    if re.search(r'[A-Za-z]', number):
        # Exclude obvious addresses (contain "Road", "Street", "Ave", etc.)
        address_indicators = ['road', 'street', 'ave', 'avenue', 'blvd', 'boulevard', 'drive', 'lane', 'way', 'place', 'court']
        if any(indicator in number.lower() for indicator in address_indicators):
            return False
        # Exclude if it contains state abbreviations with ZIP codes (like "IL 60618")
        if re.search(r'\b[A-Z]{2}\s+\d{5}', number):
            return False
        # Reasonable length for catalog numbers (3-20 characters)
        return 3 <= len(number.strip()) <= 20
    
    # Accept UPC/EAN (12-13 digits only)
    if digits_only.isdigit():
        return len(digits_only) in [12, 13]
    
    return False

def extract_upc_from_metadata(metadata_text):
    """Extract and validate UPC numbers from the metadata text."""
    # Find the Numbers line in the Publishers section
    numbers_pattern = r'Publishers:.*?Numbers:\s*(.*?)(?:\n|$)'
    numbers_match = re.search(numbers_pattern, metadata_text, re.DOTALL | re.MULTILINE)
    
    if not numbers_match:
        return None, "No 'Numbers' field found"
    
    numbers_text = numbers_match.group(1).strip()
    
    # Handle "Not visible" case
    if numbers_text.lower() == "not visible":
        return None, "Numbers field marked as 'Not visible'"
    
    # Remove square brackets that might wrap the numbers
    numbers_text = re.sub(r'[\[\]]', '', numbers_text)
    
    # Split multiple numbers by comma and process each
    number_parts = [part.strip() for part in numbers_text.split(',')]
    valid_numbers = []
    
    for part in number_parts:
        if part:  # Skip empty parts
            clean_num = clean_number(part)
            if clean_num and is_valid_upc(clean_num):
                valid_numbers.append(clean_num)
    
    if valid_numbers:
        return ', '.join(valid_numbers), None
    else:
        return None, f"No valid numbers found in '{numbers_text}'"

def clean_dates_in_metadata(metadata_text):
    """
    Replace dates that include month or day components with 'Not visible'.
    Keep standalone years (e.g., '2002') after stripping copyright symbols.
    """
    # First, find the Dates section
    dates_section_pattern = r'(Dates:.*?publicationDate:)\s*(.*?)(?:\n|$)'
    dates_match = re.search(dates_section_pattern, metadata_text, re.DOTALL | re.MULTILINE)
    
    if not dates_match:
        return metadata_text  # No dates section found
    
    date_value = dates_match.group(2).strip()
    
    # Check if the date is already "Not visible"
    if date_value.lower() == "not visible":
        return metadata_text
    
    # Strip copyright and phonogram symbols
    cleaned_date = re.sub(r'[©℗]', '', date_value).strip()
    
    # Check if the cleaned date is just a 4-digit year (which we want to keep)
    if re.match(r'^[0-9]{4}$', cleaned_date):
        # Replace with cleaned year if symbols were removed
        if cleaned_date != date_value:
            cleaned_text = re.sub(
                dates_section_pattern,
                r'\1 ' + cleaned_date + r'\n',
                metadata_text,
                flags=re.DOTALL | re.MULTILINE
            )
            return cleaned_text
        else:
            return metadata_text  # Keep as-is if no symbols to remove
    
    # Otherwise, it's a complex date with month/day components - replace with "Not visible"
    cleaned_text = re.sub(
        dates_section_pattern,
        r'\1 Not visible\n',
        metadata_text,
        flags=re.DOTALL | re.MULTILINE
    )
    
    return cleaned_text

def process_excel_file(input_file_path, results_folder_path, workflow_json_path):
    """Process the Excel file containing metadata entries and update it in place."""
    # Load the workbook
    wb = openpyxl.load_workbook(input_file_path)
    ws = wb.active
    
    # Get the metadata column index
    headers = [cell.value for cell in ws[1]]
    metadata_col_idx = headers.index('AI-Generated Metadata') + 1
    
    # Initialize tracking variables
    records_processed = 0
    records_with_upc_changes = 0
    records_with_date_changes = 0
    records_with_errors = 0

    print(f"STEP 1.5: METADATA CLEANING")
    print(f"Processing file: {input_file_path}")
    print("-" * 50)
    
    # Process each row
    for row_idx in range(2, ws.max_row + 1):
        metadata_cell = ws.cell(row=row_idx, column=metadata_col_idx)
        barcode = ws.cell(row=row_idx, column=4).value  # Assuming barcode is in column D
        
        if metadata_cell.value:
            try:
                records_processed += 1
                original_metadata = metadata_cell.value
                
                # Clean dates in the metadata
                updated_metadata = clean_dates_in_metadata(metadata_cell.value)
                date_changed = updated_metadata != original_metadata
                
                # Process UPC
                upc, upc_error = extract_upc_from_metadata(updated_metadata)
                upc_changed = False
                
                # Only modify the metadata if it contains a numbers section
                if 'Numbers:' in updated_metadata:
                    old_numbers_section = updated_metadata
                    # Replace number section with cleaned version
                    updated_metadata = re.sub(
                        r'(Publishers:.*?Numbers:)\s*(.*?)(\n|$)',
                        lambda m: f"{m.group(1)} {upc if upc else 'Not visible'}{m.group(3)}",
                        updated_metadata,
                        flags=re.DOTALL | re.MULTILINE
                    )
                    upc_changed = old_numbers_section != updated_metadata
                
                # Update the cell with cleaned metadata
                metadata_cell.value = updated_metadata
                
                # Track changes
                if date_changed:
                    records_with_date_changes += 1
                if upc_changed:
                    records_with_upc_changes += 1
                
                # Update JSON workflow with cleaning results
                try:
                    from json_workflow import update_record_step15_cleaning
                    update_record_step15_cleaning(
                        json_path=workflow_json_path,
                        barcode=str(barcode) if barcode else f"row_{row_idx}",
                        changes_made={
                            "numbers_edited": upc_changed,
                            "date_edited": date_changed
                        },
                        upc_extracted=upc
                    )
                except Exception as json_error:
                    print(f"   JSON logging error for {barcode}: {json_error}")
                    log_error(
                        results_folder_path=results_folder_path,
                        step="step1.5",
                        barcode=str(barcode) if barcode else f"row_{row_idx}",
                        error_type="json_update_error",
                        error_message=str(json_error)
                    )
                    
                # Log significant changes
                if date_changed or upc_changed:
                    print(f"  Row {row_idx} (Barcode: {barcode}): " + 
                          f"{'Date cleaned' if date_changed else ''}" +
                          f"{', ' if date_changed and upc_changed else ''}" +
                          f"{'UPC standardized' if upc_changed else ''}")
            
            except Exception as e:
                records_with_errors += 1
                print(f"  Error processing row {row_idx} (Barcode: {barcode}): {str(e)}")
                
                # Log error to JSON
                try:
                    log_error(
                        results_folder_path=results_folder_path,
                        step="step1.5",
                        barcode=str(barcode) if barcode else f"row_{row_idx}",
                        error_type="metadata_cleaning_error",
                        error_message=str(e),
                        additional_context={
                            "row_number": row_idx,
                            "operation": "metadata_cleaning"
                        }
                    )
                except:
                    pass  # Don't let JSON logging errors break the cleaning process
    
    # Print summary AFTER processing all rows
    print(f"\nCLEANING SUMMARY:")
    print(f"  Total records processed: {records_processed}")
    print(f"  Records with date changes: {records_with_date_changes}")
    print(f"  Records with pub number changes: {records_with_upc_changes}")
    print(f"  Records with errors: {records_with_errors}")

    # Log processing metrics AFTER processing all rows
    try:
        cleaning_metrics = {
            "total_records": records_processed,
            "date_changes": records_with_date_changes,
            "upc_changes": records_with_upc_changes,
            "errors": records_with_errors,
            "success_rate": ((records_processed - records_with_errors) / records_processed * 100) if records_processed > 0 else 0,
            "processing_type": "metadata_cleaning",
            "file_processed": input_file_path
        }
        
        log_processing_metrics(
            results_folder_path=results_folder_path,
            step="step1.5_metadata_cleaning",
            batch_metrics=cleaning_metrics
        )
    except Exception as metrics_error:
        print(f"Warning: Could not log cleaning metrics: {metrics_error}")
    
    # Save the updated workbook back to the same file
    wb.save(input_file_path)
    print(f"Updated spreadsheet to leave only UPC numbers, EAN numbers, and YYYY dates: {input_file_path}")

def main():
    file_paths = get_file_path_config()
    results_folder = find_latest_results_folder(file_paths["results_prefix"])
    workflow_json_path = get_workflow_json_path(results_folder)  
    
    if not results_folder:
        print("No results folder found! Run the previous scripts first.")
        return
    
    # Look for previously created spreadsheet in the results folder
    input_files = [f for f in os.listdir(results_folder) 
                   if f.startswith('full-workflow-data-cd-') and f.endswith('.xlsx')]
    
    if not input_files:
        error_msg = "No full-workflow-data-cd- files found for cleaning"
        print(f"{error_msg}")
        try:
            log_error(
                results_folder_path=results_folder,
                step="step1.5",
                barcode="system",
                error_type="no_files_found",
                error_message=error_msg
            )
        except:
            pass
        return

    latest_file = max(input_files)
    input_file = os.path.join(results_folder, latest_file)

    print(f"Found file to clean: {latest_file}")
    
    # Process the file - now passes workflow_json_path
    process_excel_file(input_file, results_folder, workflow_json_path)

if __name__ == "__main__":
    main()