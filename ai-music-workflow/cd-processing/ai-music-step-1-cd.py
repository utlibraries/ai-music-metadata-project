# Extract metadata from CD images using GPT-4o with batch processing support
import os
import base64
import time
from datetime import datetime
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openai import OpenAI

# Import custom modules
from token_logging import create_token_usage_log, log_individual_response
from batch_processor import BatchProcessor  
from model_pricing import calculate_cost, get_model_info
from json_workflow import initialize_workflow_json, update_record_step1, log_error, log_processing_metrics
from shared_utilities import get_workflow_json_path, extract_metadata_fields, group_images_by_barcode, create_batch_summary
from cd_workflow_config import get_current_timestamp, get_file_path_config, get_model_config

client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

def get_llm_prompt():
    return """Analyze these images of a compact disc and extract the following key metadata fields in the specified format. You are a music cataloger, and know that you are responsible for the accuracy of the information you produce.  If ANY information is unclear, partially visible, or not visible: mark it as 'Not visible' in the metadadata. If you have reason to believe that a sticker may be covering part of a key field, like the title or primary contributor, either mark it as 'Not visible' or make an educated guess based on the visible information and note that it may be partially obscured in parentheses.

Match this format:
Title Information:
  - Main Title: [Main Title in original language if using latin characters.  Transliterated if in non-latin characters.]
  - Subtitle: [Subtitle in original language if using latin characters.  Transliterated if in non-latin characters.]
  - Primary Contributor: [Artist/Performer Name]
  - Additional Contributors: [arrangers, engineers, producers, session musicians]
Publishers:
  - Name: [Publisher Name - please list all publisher, label, series, and distributor names visible on the disc]
  - Place: [Place of publication if available]
  - Numbers: [UPC/EAN/ISBN/Catalog Numbers. Bear in mind that numbers may at times be vertically oriented]
Dates:
  - publicationDate: [Record all dates as printed on the disc - separate multiple dates with commas]
Language:
  - sungLanguage: [Languages of sung text]
  - printedLanguage: [All languages of printed text]
Format:
  - generalFormat: [Sound Recording]
  - specificFormat: [CD, CD-ROM, Enhanced CD, etc.]
  - materialTypes: [List of Material Types]
Physical Description:
  - size: [4.75" (standard CD)]
  - material: [aluminum/polycarbonate]
  - labelDesign: [Description of disc face design and color]
  - physicalCondition: [Condition notes]
  - specialFeatures: [Booklet details, packaging type (jewel case/digipak), inserts, bonus materials]
Contents:
  - tracks: [
      {
        "number": [Track number],
        "title": [Title in original language],
        "titleTransliteration": [Title transliteration if applicable],
      }
    ]
Notes:
  - generalNotes: [{'text': [Note Text]}]

***Important: These are images of CD's that were donated by a university radio station to our library. Handwritten information on white stickers should be ignored.  When in doubt, mark fields in the metadata as 'Not visible'*** 
 
Analyze the provided images and return metadata formatted exactly as above. Pay special attention to capturing only text that is clearly legible."""

def prepare_batch_requests(image_groups, model_name):
    """Prepare all requests for batch processing."""
    batch_requests = []
    custom_id_mapping = {}
    
    for i, (barcode, image_paths) in enumerate(sorted(image_groups.items())):
        # Take up to first 3 images for each barcode
        image_paths = image_paths[:3]
        prompt_text = get_llm_prompt()
        uploaded_files_info = ""

        for j, img_path in enumerate(image_paths):
            # Determine image type based on filename
            filename = os.path.basename(img_path).lower()
            if filename.endswith('a.png') or filename.endswith('a.jpg') or filename.endswith('a.jpeg'):
                image_type = "FRONT COVER"
            elif filename.endswith('b.png') or filename.endswith('b.jpg') or filename.endswith('b.jpeg'):
                image_type = "BACK COVER"
            elif filename.endswith('c.png') or filename.endswith('c.jpg') or filename.endswith('c.jpeg'):
                image_type = "ADDITIONAL IMAGE"
            else:
                image_type = "IMAGE"
            
            uploaded_files_info += f"[Image {j+1} - {image_type}: {img_path}]\n"

        prompt = prompt_text + "\n" + uploaded_files_info

        # Prepare base64 images
        base64_images = []
        content_types = []
        
        for img_path in image_paths:
            with open(img_path, "rb") as image_file:
                base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                base64_images.append(base64_image)
            
            ext = os.path.splitext(img_path)[1].lower()
            if ext == '.png':
                content_types.append("image/png")
            else:
                content_types.append("image/jpeg")

        # Create messages with appropriate content types
        image_contents = []
        for j, base64_image in enumerate(base64_images):
            image_contents.append({
                "type": "image_url",
                "image_url": {"url": f"data:{content_types[j]};base64,{base64_image}"}
            })

        # Create request data
        request_data = {
            "model": model_name,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    *image_contents
                ]
            }],
            "max_tokens": 2000
        }
        
        batch_requests.append(request_data)
        custom_id_mapping[f"req_{i}"] = {
            "barcode": barcode,
            "image_paths": image_paths,
            "row_number": i + 2  # +2 for header row
        }
    
    return batch_requests, custom_id_mapping

def process_folder_with_batch(folder_path, wb, results_folder_path, workflow_json_path):
    """Process folder using batch processing when appropriate."""
    model_config = get_model_config("step1")
    model_name = model_config.get("model", "gpt-4o")
    ws = wb.active
    headers = ['Input Image 1', 'Input Image 2', 'Input Image 3', 'Barcode', 'AI-Generated Metadata']
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        if col == 4:  # Barcode column
            ws.column_dimensions[get_column_letter(col)].width = 17
        else:
            ws.column_dimensions[get_column_letter(col)].width = 30 if col <= 3 else 52

    # Create logs folder
    logs_folder_path = os.path.join(results_folder_path, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)

    # Group images by barcode
    image_groups = group_images_by_barcode(folder_path)
    total_items = len(image_groups)
    
    print(f"\nSTEP 1: METADATA EXTRACTION")
    print(f"Found {total_items} CD image groups to process")
    print(f"Starting metadata extraction using {model_name}...")
    print("-" * 50)

    # Initialize batch processor and check if we should use batch processing
    processor = BatchProcessor()
    use_batch = processor.should_use_batch(total_items)
    
    print(f"Processing mode: {'BATCH' if use_batch else 'INDIVIDUAL'}")
    
    if use_batch:
        print(f"Preparing {total_items} requests for batch processing...")
        
        # Estimate costs
        batch_requests, custom_id_mapping = prepare_batch_requests(image_groups, model_name)
        cost_estimate = processor.estimate_batch_cost(batch_requests, model_name)
        
        print(f"Cost estimate:")
        print(f"   Regular API: ${cost_estimate['regular_cost']:.4f}")
        print(f"   Batch API: ${cost_estimate['batch_cost']:.4f}")
        print(f"   Savings: ${cost_estimate['savings']:.4f} ({cost_estimate['savings_percentage']:.1f}%)")
        
        # Convert to batch format
        formatted_requests = processor.create_batch_requests(batch_requests, "cd_metadata")

        # Use adaptive batch processing that automatically splits based on file size
        results = processor.submit_adaptive_batch(
            batch_requests=formatted_requests,
            custom_id_mapping=custom_id_mapping,
            description=f"CD Metadata Extraction - {total_items} items - {datetime.now().strftime('%Y-%m-%d')}",
            max_file_size_mb=180  # Conservative limit under 200 MB
        )
        
        if results:
            # Process batch results
            processed_results = processor.process_batch_results(results, custom_id_mapping)
            
            print(f"Processing batch results...")
            items_with_issues = 0
            
            for custom_id, result_data in processed_results["results"].items():
                    # Handle both single batch and chunked batch custom IDs
                    if custom_id.startswith("cd_metadata_"):
                        # Single batch format: cd_metadata_0_hash
                        index = int(custom_id.split("_")[2])
                        mapping_key = f"req_{index}"
                    elif custom_id.startswith("chunk_"):
                        # Chunked batch format: chunk_0_1_hash
                        parts = custom_id.split("_")
                        if len(parts) >= 3:
                            index = int(parts[2])  # Third part is the original request index
                            mapping_key = f"req_{index}"
                        else:
                            continue  # Skip malformed custom_ids
                    else:
                        continue  # Skip unknown custom_id formats
                    
                    if mapping_key in custom_id_mapping:
                        barcode = custom_id_mapping[mapping_key]["barcode"]
                        image_paths = custom_id_mapping[mapping_key]["image_paths"]
                        row_number = custom_id_mapping[mapping_key]["row_number"]
                        
                        if result_data["success"]:
                            metadata_output = result_data["content"]
                            usage = result_data["usage"]
                            
                            # Log individual response
                            log_individual_response(
                                logs_folder_path=logs_folder_path,
                                script_name="step1",
                                row_number=row_number,
                                barcode=barcode,
                                response_text=metadata_output,
                                model_name=model_name,
                                prompt_tokens=usage.get("prompt_tokens", 0),
                                completion_tokens=usage.get("completion_tokens", 0),
                                processing_time=0  # Batch processing doesn't track individual timing
                            )
                            
                            try:
                                extracted_fields = extract_metadata_fields(metadata_output)
                                update_record_step1(
                                    json_path=workflow_json_path,
                                    barcode=barcode,
                                    raw_metadata=metadata_output,
                                    extracted_fields=extracted_fields,
                                    model=model_name,
                                    prompt_tokens=usage.get("prompt_tokens", 0),
                                    completion_tokens=usage.get("completion_tokens", 0),
                                    processing_time=0
                                )
                            except Exception as json_error:
                                log_error(
                                    results_folder_path=results_folder_path,
                                    step="step1",
                                    barcode=barcode,
                                    error_type="json_update_error",
                                    error_message=str(json_error)
                                )
                        else:
                            metadata_output = f"Error: {result_data['error']}"
                            items_with_issues += 1
                            
                            # Log error
                            log_individual_response(
                                logs_folder_path=logs_folder_path,
                                script_name="step1",
                                row_number=row_number,
                                barcode=barcode,
                                response_text=metadata_output,
                                model_name=model_name,
                                prompt_tokens=0,
                                completion_tokens=0,
                                processing_time=0
                            )
                        
                        # Add to spreadsheet
                        row_data = ['', '', '', barcode, metadata_output]
                        ws.append(row_data)
                        
                        # Add thumbnail images
                        for i, img_path in enumerate(image_paths, start=1):
                            img = PILImage.open(img_path)
                            img.thumbnail((200, 200))
                            output = BytesIO()
                            img.save(output, format='PNG')
                            output.seek(0)
                            img_openpyxl = Image(output)
                            img_openpyxl.anchor = ws.cell(row=ws.max_row, column=i).coordinate
                            ws.add_image(img_openpyxl)
                        
                        ws.row_dimensions[ws.max_row].height = 215
                        for cell in ws[ws.max_row]:
                            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Return batch processing metrics
            summary = processed_results["summary"]
            return (total_items, items_with_issues, 0,  # 0 for total_time since batch doesn't track individual timing
                   summary["total_prompt_tokens"], summary["total_completion_tokens"], 
                   summary["total_prompt_tokens"] + summary["total_completion_tokens"])
        
        else:
            print("Batch processing failed, falling back to individual processing...")
            use_batch = False
    
    # Fall back to individual processing if batch fails or isn't used
    if not use_batch:
        return process_folder_individual(image_groups, ws, logs_folder_path, model_name, total_items, workflow_json_path, results_folder_path)

def process_folder_individual(image_groups, ws, logs_folder_path, model_name, total_items, workflow_json_path, results_folder_path):
    """Process using individual API calls (original logic)."""
    items_with_issues = 0
    processed_items = 0
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    total_time = 0

    for barcode, image_paths in sorted(image_groups.items()):
        processed_items += 1
        item_start_time = time.time()
        row_number = processed_items + 1

        print(f"\nProcessing CD {processed_items}/{total_items}")
        print(f"   Barcode: {barcode}")
        print(f"   Progress: {(processed_items/total_items)*100:.1f}%")

        try:
            # Take up to first 3 images for each barcode
            image_paths = image_paths[:3]
            prompt_text = get_llm_prompt()
            uploaded_files_info = ""

            for i, img_path in enumerate(image_paths):
                filename = os.path.basename(img_path).lower()
                if filename.endswith('a.png') or filename.endswith('a.jpg') or filename.endswith('a.jpeg'):
                    image_type = "FRONT COVER"
                elif filename.endswith('b.png') or filename.endswith('b.jpg') or filename.endswith('b.jpeg'):
                    image_type = "BACK COVER"
                elif filename.endswith('c.png') or filename.endswith('c.jpg') or filename.endswith('c.jpeg'):
                    image_type = "ADDITIONAL IMAGE"
                else:
                    image_type = "IMAGE"
                
                uploaded_files_info += f"[Image {i+1} - {image_type}: {img_path}]\n"

            prompt = prompt_text + "\n" + uploaded_files_info

            try:
                print(f"Calling OpenAI API...")
                
                base64_images = []
                for img_path in image_paths:
                    with open(img_path, "rb") as image_file:
                        base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                        base64_images.append(base64_image)

                api_start_time = time.time()
                
                content_types = []
                for img_path in image_paths:
                    ext = os.path.splitext(img_path)[1].lower()
                    if ext == '.png':
                        content_types.append("image/png")
                    else:
                        content_types.append("image/jpeg")
                
                image_contents = []
                for i, base64_image in enumerate(base64_images):
                    image_contents.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:{content_types[i]};base64,{base64_image}"}
                    })
                
                response = client.chat.completions.create(
                    model=model_name,
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            *image_contents
                        ]
                    }],
                    max_tokens=2000
                )
                
                api_duration = time.time() - api_start_time
                
                prompt_tokens = response.usage.prompt_tokens
                completion_tokens = response.usage.completion_tokens
                total_item_tokens = prompt_tokens + completion_tokens
                
                total_prompt_tokens += prompt_tokens
                total_completion_tokens += completion_tokens
                total_tokens += total_item_tokens

                metadata_output = response.choices[0].message.content.strip()
                
                print(f"API call successful! Tokens: {total_item_tokens:,}")
                
                log_individual_response(
                    logs_folder_path=logs_folder_path,
                    script_name="step1",
                    row_number=row_number,
                    barcode=barcode,
                    response_text=metadata_output,
                    model_name=model_name,
                    prompt_tokens=prompt_tokens,
                    completion_tokens=completion_tokens,
                    processing_time=api_duration
                )
                try:
                    # Extract structured metadata fields
                    extracted_fields = extract_metadata_fields(metadata_output)
                    
                    # Update workflow JSON with Step 1 results
                    update_record_step1(
                        json_path=workflow_json_path,
                        barcode=barcode,
                        raw_metadata=metadata_output,
                        extracted_fields=extracted_fields,
                        model=model_name,
                        prompt_tokens=prompt_tokens,
                        completion_tokens=completion_tokens,
                        processing_time=api_duration
                    )
                except Exception as json_error:
                    log_error(
                        results_folder_path=results_folder_path,
                        step="step1",
                        barcode=barcode,
                        error_type="json_update_error",
                        error_message=str(json_error)
                    )
                row_data = ['', '', '', barcode, metadata_output]
                ws.append(row_data)

                # Add thumbnail images
                for i, img_path in enumerate(image_paths, start=1):
                    img = PILImage.open(img_path)
                    img.thumbnail((200, 200))
                    output = BytesIO()
                    img.save(output, format='PNG')
                    output.seek(0)
                    img_openpyxl = Image(output)
                    img_openpyxl.anchor = ws.cell(row=ws.max_row, column=i).coordinate
                    ws.add_image(img_openpyxl)

                ws.row_dimensions[ws.max_row].height = 215
                for cell in ws[ws.max_row]:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            except Exception as e:
                print(f"API call failed: {str(e)}")
                error_message = f"Error: {str(e)}"
                ws.append(['', '', '', barcode, error_message])
                items_with_issues += 1

        except Exception as e:
            print(f"Processing failed: {str(e)}")
            error_message = f"Error: {str(e)}"
            ws.append(['', '', '', barcode, error_message])
            items_with_issues += 1

        item_duration = time.time() - item_start_time
        total_time += item_duration

    return total_items, items_with_issues, total_time, total_prompt_tokens, total_completion_tokens, total_tokens

def main():
    # Get configuration
    file_paths = get_file_path_config()
    model_config = get_model_config("step1")
    
    model_name = model_config.get("model", "gpt-4o")
    
    # Start timing the entire script execution
    script_start_time = time.time()
    
    images_folder = file_paths["images_folder"]
    
    current_timestamp = get_current_timestamp()
    results_folder_name = f"results-{current_timestamp}"
    results_folder_path = os.path.join(file_paths["output_base"], results_folder_name)

    if not os.path.exists(results_folder_path):
        os.makedirs(results_folder_path)
    workflow_json_path = get_workflow_json_path(results_folder_path)
    if not os.path.exists(workflow_json_path):
        initialize_workflow_json(results_folder_path)
        print(f"Initialized workflow JSON: {workflow_json_path}")
    
    logs_folder_path = os.path.join(results_folder_path, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)
    
    # Get image groups to show count in model info
    image_groups = group_images_by_barcode(images_folder)
    total_items = len(image_groups)
    
    # Show model pricing info at start
    model_info = get_model_info(model_name)
    if model_info:
        print(f"STEP 1: METADATA EXTRACTION")
        print(f"Using model: {model_name}")
        print(f"Pricing: ${model_info['input_per_1k']:.5f}/1K input, ${model_info['output_per_1k']:.5f}/1K output")
        print(f"Batch discount: {model_info['batch_discount']*100:.0f}%")
        print(f"Total CDs to process: {total_items}")
        print("-" * 50)
    
    wb = Workbook()
    total_items, items_with_issues, total_time, total_prompt_tokens, total_completion_tokens, total_tokens = process_folder_with_batch(images_folder, wb, results_folder_path, workflow_json_path)

    # Apply formatting to all cells
    for row in wb.active.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.active.freeze_panes = 'A2'

    output_file = f"cd-metadata-ai-{current_timestamp}.xlsx"
    full_output_path = os.path.join(results_folder_path, output_file)
    wb.save(full_output_path)
    
    # Calculate script metrics
    script_duration = time.time() - script_start_time
    
    # Determine if batch processing was used (check if we have many items but zero processing time)
    was_batch_processed = total_items > 10 and total_time == 0
    
    # Calculate actual cost using the model pricing
    estimated_cost = calculate_cost(
        model_name=model_name,
        prompt_tokens=total_prompt_tokens,
        completion_tokens=total_completion_tokens,
        is_batch=was_batch_processed
    )
    
    # Create standardized token usage log with enhanced metrics
    create_token_usage_log(
        logs_folder_path=logs_folder_path,
        script_name="step1",
        model_name=model_name,
        total_items=total_items,
        items_with_issues=items_with_issues,
        total_time=total_time,
        total_prompt_tokens=total_prompt_tokens,
        total_completion_tokens=total_completion_tokens,
        additional_metrics={
            "Total script execution time": f"{script_duration:.2f}s",
            "Processing time percentage": f"{(total_time/script_duration)*100:.1f}%" if script_duration > 0 else "0%",
            "Items successfully processed": total_items - items_with_issues,
            "Processing mode": "BATCH" if was_batch_processed else "INDIVIDUAL",
            "Actual cost": f"${estimated_cost:.4f}",
            "Average tokens per item": f"{total_tokens/total_items:.0f}" if total_items > 0 else "0"
        }
    )
    
    # Log full responses if configured
    print(f"\nSTEP 1 COMPLETED!")
    print(f"Successfully processed: {total_items - items_with_issues}/{total_items} CDs")
    print(f"Items with issues: {items_with_issues}")
    print(f"Total script time: {script_duration:.1f}s ({script_duration/60:.1f} minutes)")
    print(f"Processing time: {total_time:.1f}s")
    print(f"Total tokens: {total_tokens:,} (Input: {total_prompt_tokens:,}, Output: {total_completion_tokens:,})")
    print(f"Processing mode: {'BATCH' if was_batch_processed else 'INDIVIDUAL'}")
    print(f"Actual cost: ${estimated_cost:.4f}")
    
    # Show batch savings if applicable
    if was_batch_processed:
        regular_cost = calculate_cost(model_name, total_prompt_tokens, total_completion_tokens, is_batch=False)
        savings = regular_cost - estimated_cost
        savings_percentage = (savings / regular_cost) * 100 if regular_cost > 0 else 0
        print(f"Regular API cost would have been: ${regular_cost:.4f}")
        print(f"Batch savings: ${savings:.4f} ({savings_percentage:.1f}%)")
    
    print(f"Results saved to: {full_output_path}")
    print(f"Token usage log saved to: {os.path.join(logs_folder_path, 'step1_token_usage_log.txt')}")
    print(f"Full responses log saved to: {os.path.join(logs_folder_path, 'step1_llm_responses_log.txt')}")
    
    try:
        batch_summary = create_batch_summary(
            total_items=total_items,
            successful_items=total_items - items_with_issues,
            failed_items=items_with_issues,
            total_time=total_time,
            total_tokens=total_tokens,
            estimated_cost=estimated_cost,
            processing_mode="BATCH" if was_batch_processed else "INDIVIDUAL"
        )
        
        log_processing_metrics(
            results_folder_path=results_folder_path,
            step="step1_metadata_extraction",
            batch_metrics=batch_summary
        )
    except Exception as metrics_error:
        print(f"Warning: Could not log processing metrics: {metrics_error}")

if __name__ == "__main__":
    main()