import os
import glob
import datetime
import re
import openpyxl
import requests
import time
import csv
from difflib import SequenceMatcher
from openpyxl.styles import Alignment
from openpyxl import load_workbook

def find_latest_results_folder(prefix):
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None
    return max(matching_folders)

def find_latest_step4_file(results_folder):
    # Find files starting with "ai-music-step-4-" and ending with ".xlsx"
    files = [f for f in os.listdir(results_folder) 
             if f.startswith("ai-music-step-4-") and f.endswith(".xlsx")]
    if not files:
        return None
    latest_file = max(files)
    return os.path.join(results_folder, latest_file)

def get_access_token(client_id, client_secret):
    token_url = "https://oauth.oclc.org/token"
    data = {
        "grant_type": "client_credentials",
        "scope": "wcapi"
    }
    response = requests.post(token_url, data=data, auth=(client_id, client_secret))
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        raise Exception(f"Failed to get access token: {response.text}")

def get_bib_info(oclc_number, access_token):
    """
    Query the OCLC API for bibliographic information for a specific OCLC number.
    """
    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs/{oclc_number}"
    
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    
    try:
        response = requests.get(endpoint, headers=headers)
        response.raise_for_status()
        
        data = response.json()
        # The current API returns data directly, not in a bibRecords array
        return data
    except requests.RequestException as e:
        print(f"Error getting information for OCLC number {oclc_number}: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"Error response status: {e.response.status_code}")
            print(f"Error response content: {e.response.text}")
        return {"error": str(e)}

def clean_title(title):
    """Clean up title by removing strange punctuation but keeping slashes."""
    # Replace any double sword or other special characters, but keep slashes
    title = re.sub(r'[^\w\s\/\-\:\;\,\.\(\)\[\]\&\']', ' ', title)
    # Normalize whitespace
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def extract_title_from_bib_info(data):
    """Extract and clean title from bibliographic information."""
    if not isinstance(data, dict):
        return "No title available"
    
    # Check for error in response
    if "error" in data:
        return "No title available"
    
    # Current API response typically has title at the top level
    if "title" in data:
        # For responses with mainTitles structure
        if "mainTitles" in data["title"] and data["title"]["mainTitles"]:
            title = data["title"]["mainTitles"][0].get("text", "No title available")
            return clean_title(title)
        
        # For responses with title as direct property
        if isinstance(data["title"], str):
            return clean_title(data["title"])
    
    # Check for title in different locations that might be in the API response
    if "name" in data and isinstance(data["name"], str):
        return clean_title(data["name"])
    
    # Look for other possible title fields
    for field in ["titleInfo", "titleStatement", "uniformTitle"]:
        if field in data and isinstance(data[field], str):
            return clean_title(data[field])
    
    return "No title available"

def calculate_title_similarity(title1, title2):
    """Calculate similarity between two titles using SequenceMatcher."""
    return SequenceMatcher(None, title1.lower(), title2.lower()).ratio()

def create_high_confidence_spreadsheet():
    # Set the folder prefix 
    base_dir_prefix = "ai-music-workflow/cd-processing/cd-output-folders/results-"
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Please run the previous scripts first.")
        return None
    
    step4_file = find_latest_step4_file(results_folder)
    if not step4_file:
        print("No step 4 file found in the results folder!")
        return None

    print(f"Using source file: {step4_file}")

    # Get OCLC API credentials
    client_id = os.environ.get("OCLC_CLIENT_ID")
    client_secret = os.environ.get("OCLC_SECRET")

    if not client_id or not client_secret:
        print("Error: OCLC_CLIENT_ID and OCLC_SECRET must be set in environment variables.")
        return None
    
    try:
        # Get the access token
        access_token = get_access_token(client_id, client_secret)
        print("Successfully obtained access token.")
        
        # Open the latest step 4 workbook
        wb_src = load_workbook(step4_file)
        sheet_src = wb_src.active

        # Create a new workbook for high confidence matches
        wb_new = openpyxl.Workbook()
        sheet_new = wb_new.active

        # Set header row
        header_row = ["Barcode", "Confidence Score", "OCLC Number", "Title from OCLC"]
        sheet_new.append(header_row)

        # Set column widths
        sheet_new.column_dimensions['A'].width = 23  # Barcode
        sheet_new.column_dimensions['B'].width = 15  # Confidence Score
        sheet_new.column_dimensions['C'].width = 23  # OCLC Number
        sheet_new.column_dimensions['D'].width = 50  # Title

        # Get the column indices from the source workbook
        # In source: Column D=Barcode, Column H=OCLC Number, Column I=Confidence Score
        BARCODE_COL_IDX = 4  # Column D
        OCLC_NUM_COL_IDX = 8  # Column H
        CONF_SCORE_COL_IDX = 9  # Column I

        # Collect high confidence matches (80% or higher)
        high_confidence_matches = []
        
        # Process rows and collect high confidence matches
        for row_idx in range(2, sheet_src.max_row + 1):  # Skip header row
            barcode = sheet_src.cell(row=row_idx, column=BARCODE_COL_IDX).value
            oclc_number = sheet_src.cell(row=row_idx, column=OCLC_NUM_COL_IDX).value
            confidence_score = sheet_src.cell(row=row_idx, column=CONF_SCORE_COL_IDX).value
            
            # Skip rows with missing data or "No matching records found"
            if not barcode or not oclc_number or not confidence_score:
                continue
                
            if isinstance(oclc_number, str) and oclc_number.strip() in ["Not found", "Error processing"]:
                continue
                
            # Convert confidence score to float and check if it's high confidence
            try:
                conf_score = float(confidence_score)
                if conf_score >= 80:  # High confidence threshold
                    # Query OCLC API for title information
                    oclc_data = get_bib_info(str(oclc_number).strip(), access_token)
                    oclc_title = extract_title_from_bib_info(oclc_data)
                    
                    if oclc_title != "No title available":
                        # Add to high confidence matches
                        high_confidence_matches.append({
                            "barcode": barcode,
                            "confidence_score": conf_score,
                            "oclc_number": oclc_number,
                            "title": oclc_title
                        })
                        print(f"Found high confidence match: Barcode={barcode}, OCLC={oclc_number}, Score={conf_score}, Title={oclc_title}")
                    else:
                        print(f"Could not extract title for: Barcode={barcode}, OCLC={oclc_number}, Score={conf_score}")
                    
            except ValueError:
                # Skip if confidence score is not a valid number
                continue
                
            # Small delay to avoid API rate limits
            time.sleep(0.5)
        
        # Sort high confidence matches by title for deduplication
        sorted_matches = sorted(high_confidence_matches, key=lambda x: x["title"])
        
        # Deduplicate matches based on title similarity
        deduplicated_matches = []
        skip_indices = set()
        
        for i in range(len(sorted_matches)):
            if i in skip_indices:
                continue
                
            current_match = sorted_matches[i]
            
            # Compare with all subsequent matches
            for j in range(i + 1, len(sorted_matches)):
                if j in skip_indices:
                    continue
                    
                compare_match = sorted_matches[j]
                
                # Check title similarity
                similarity = calculate_title_similarity(current_match["title"], compare_match["title"])
                
                # If titles are similar (threshold 0.9), keep only the higher confidence match
                if similarity >= 0.9:
                    print(f"Found similar titles: '{current_match['title']}' and '{compare_match['title']}' (similarity: {similarity:.2f})")
                    
                    if current_match["confidence_score"] >= compare_match["confidence_score"]:
                        skip_indices.add(j)  # Skip the lower confidence match
                        print(f"Keeping match with higher confidence score: {current_match['confidence_score']} vs {compare_match['confidence_score']}")
                    else:
                        skip_indices.add(i)  # Skip current match if the other has higher confidence
                        print(f"Keeping match with higher confidence score: {compare_match['confidence_score']} vs {current_match['confidence_score']}")
                        break  # No need to continue checking if current match is skipped
            
            # Add current match to deduplicated list if not skipped
            if i not in skip_indices:
                deduplicated_matches.append(current_match)
        
        # Add deduplicated matches to spreadsheet
        for match in deduplicated_matches:
            sheet_new.append([
                match["barcode"],
                match["confidence_score"],
                match["oclc_number"],
                match["title"]
            ])
            
        # Save the high confidence spreadsheet
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        high_conf_file = f"high-confidence-matches-{current_date}.xlsx"
        high_conf_path = os.path.join(results_folder, high_conf_file)
        wb_new.save(high_conf_path)
        
        print(f"High confidence matches spreadsheet created with {len(deduplicated_matches)} records: {high_conf_path}")
        
        # Create final output spreadsheet with OCLC #, Barcode, Title format
        wb_final = openpyxl.Workbook()
        sheet_final = wb_final.active
        
        # Rearrange columns for final format
        for row_idx, match in enumerate(deduplicated_matches, 1):
            sheet_final.cell(row=row_idx, column=1).value = match["oclc_number"]  # OCLC Number first
            sheet_final.cell(row=row_idx, column=2).value = match["barcode"]      # Barcode second
            sheet_final.cell(row=row_idx, column=3).value = match["title"]        # Title third
        
        # Set column widths for final spreadsheet
        sheet_final.column_dimensions['A'].width = 23  # OCLC Number
        sheet_final.column_dimensions['B'].width = 23  # Barcode
        sheet_final.column_dimensions['C'].width = 50  # Title
        
        # Save the final spreadsheet
        final_file = f"final-unique-matches-{current_date}.xlsx"
        final_path = os.path.join(results_folder, final_file)
        wb_final.save(final_path)
        
        print(f"Final unique matches spreadsheet created: {final_path}")
        
        # Create pipe-delimited text file
        text_file = f"unique-matches-{current_date}.txt"
        text_path = os.path.join(results_folder, text_file)
        
        with open(text_path, 'w', newline='', encoding='utf-8') as f:
            for match in deduplicated_matches:
                line = f"{match['oclc_number']}|{match['barcode']}|{match['title']}\n"
                f.write(line)
        
        print(f"Pipe-delimited text file created: {text_path}")
        
        return {
            "high_conf_path": high_conf_path,
            "final_path": final_path,
            "text_path": text_path,
            "match_count": len(deduplicated_matches)
        }
        
    except Exception as e:
        print(f"Error creating high confidence spreadsheet: {str(e)}")
        return None

def main():
    result = create_high_confidence_spreadsheet()
    if result:
        print("\nSummary:")
        print(f"- Total high confidence unique matches: {result['match_count']}")
        print(f"- High confidence spreadsheet: {result['high_conf_path']}")
        print(f"- Final unique matches spreadsheet: {result['final_path']}")
        print(f"- Pipe-delimited text file: {result['text_path']}")
    else:
        print("Failed to create high confidence matches spreadsheet.")

if __name__ == "__main__":
    main()