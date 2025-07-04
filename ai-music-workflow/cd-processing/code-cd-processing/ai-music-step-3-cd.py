#!/usr/bin/env python3
import os
import glob
import time
from openai import OpenAI
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import re
from token_logging import create_token_usage_log, log_individual_response
from batch_processor import BatchProcessor 
from model_pricing import calculate_cost, estimate_cost, get_model_info

def find_latest_results_folder(prefix):
    # Get the parent directory of the prefix
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None

    latest_folder = max(matching_folders)
    
    return latest_folder

# Load the API key from environment variable
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

def prepare_batch_requests(sheet, model_name):
    """Prepare all requests for batch processing."""
    batch_requests = []
    custom_id_mapping = {}
    
    # Define the columns
    BARCODE_COLUMN = 'D'  
    METADATA_COLUMN = 'E'
    OCLC_RESULTS_COLUMN = 'G'
    
    valid_rows = []
    
    # First pass: collect all valid rows
    for row in range(2, sheet.max_row + 1):
        metadata = sheet[f'{METADATA_COLUMN}{row}'].value
        oclc_results = sheet[f'{OCLC_RESULTS_COLUMN}{row}'].value
        barcode = sheet[f'{BARCODE_COLUMN}{row}'].value
        
        # Skip rows with missing data or "No matching records" message
        if (metadata and oclc_results and 
            oclc_results != "No matching records found" and 
            oclc_results.strip() != ""):
            valid_rows.append((row, metadata, oclc_results, barcode))
    
    # Second pass: create batch requests for valid rows
    for i, (row, metadata, oclc_results, barcode) in enumerate(valid_rows):
        prompt = (
    f'''Analyze the following OCLC results based on the given metadata and determine which result is the best match. Methodically go through each record, choose the top 3, then consider them again and choose the record that matches the most elements in the metadata. If two or more records tie for best match, prioritize records that have more holdings and that are held by IXA. If there is no likely match, write "No matching records found".

    **Important Instructions**:
    1. Confidence Score: 0% indicates no confidence, and 100% indicates high confidence that we have found the correct OCLC number. If the confidence is below 79%, the record will be checked by a cataloger. 
    2. ***Key Fields in order of importance***:
    - UPC/Product Code (a match is HIGHEST priority if available in both metadata and OCLC record - if not available in one or the other, skip this field.  Occasionally, the UPC is partially obscured in the metadata - if some of the numbers in a UPC are incorrect but other fields are matching, it is still a match)
    - Title 
    - Artist/Performer
    - Contributors (some matching - not all need to match)
    - Publisher Name (the metadata record may include multiple names of publishers and distributor.  These do not all need to match the OCLC record, but there should be at least one exact match unless there are no visible publishers in the metadata; corporate ownership relationships like "Columbia is part of Sony" should NOT be considered a match.)
    - Physical Description (should make sense for a CD)
    - Content (track listings - these should be mostly similar, with small differences in spelling or punctuation)
    - Year (Should be an exact match if present in both the metadata and oclc record. If there are two years written in a record, the latter of the years is the reissue date, which is what we want to match)
    3. ***Notes on Matching Special Cases***:
    - Titles in non-Latin scripts that match in meaning or transliteration should also be considered equivalent.
    - If a field is marked as partially obscured, lessen the importance of that field in the matching process.
    - Different releases in different regions (e.g., Japanese release vs. US release) should be treated as different records even if title and content match.
    4. When information is not visible in the metadata, DO NOT use that field in your consideration of a match. It may be written in the metadata as 'not visible' or 'not available', etc.
    5. If there is a publisher in the OCLC record but it cannot be found anywhere in the metadata, the OCLC record or the CD may be a reissue - mark it as 79 because that way it will be checked by a cataloger. 
    6. The publisher should have at least one match between the metadata and OCLC record.  This may be a partial match, but it needs to be at least a fuzzy match.  No corporate relationships or associations unless explicitly mentioned in both the metadata and the OCLC record.  If the publisher is not visible in the metadata, do not use this field in your consideration of a match.
    7. If there is no likely match, write "No matching records found" and set the confidence score as 0.

    Format for Response:
    - Your response must follow this format exactly:
    1. OCLC number: [number or 'No matching records found']
    2. Confidence score: [%]
    3. Explanation: [List of things that match as key value pairs. If there are multiple records that could be a match, explain why you chose the one you did. If there are no matches, explain why.]
    4. Other potential good matches: [List of other OCLC numbers that could be good matches and a one sentence explanation for each match as key value pairs. If there are no other potential matches, write 'No other potential good matches.']
    
    Once you have responded, go back through the response that you wrote and carefully verify each piece of information. If you find a mistake, look for a better record. If there isn't one, reduce the confidence score to 79% or lower. If there is one, once again carefully verify all the facts that support your choice. If you still can't find a match, write "No matching records found" and set the confidence score as 0.

    Metadata: {metadata}

    OCLC Results: {oclc_results}
    ''')
        
        # Create request data
        request_data = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": "You are a music cataloger.  You are very knowledgeable about music cataloging best practices, and also have incredible attention to detail.  Read through the metadata and OCLC results carefully, and determine which of the OCLC results looks like the best match. If there is no likely match, write 'No matching records found'.  If you make a mistake, you would feel very bad about it, so you always double check your work."},
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 1500,
            "temperature": 0.5
        }
        
        batch_requests.append(request_data)
        custom_id_mapping[f"req_{i}"] = {
            "barcode": barcode,
            "row_number": row,
            "metadata": metadata,
            "oclc_results": oclc_results
        }
    
    return batch_requests, custom_id_mapping, valid_rows

def process_with_batch(sheet, temp_sheet, logs_folder_path, model_name, results_folder):
    """Process using batch processing when appropriate."""
    
    # Define the columns
    RESULT_COLUMN = 'H'
    CONFIDENCE_SCORE_COLUMN = 'I'
    EXPLANATION_COLUMN = 'J'
    OTHER_MATCHES_COLUMN = 'K'  
    PROCESSING_TIME_COLUMN = 'L'  
    PROMPT_TOKENS_COLUMN = 'M'   
    COMPLETION_TOKENS_COLUMN = 'N'  
    TOTAL_TOKENS_COLUMN = 'O'
    
    # Initialize batch processor and check if we should use batch processing
    processor = BatchProcessor()
    
    # Count valid rows for batch decision
    total_valid_rows = 0
    for row in range(2, sheet.max_row + 1):
        metadata = sheet[f'E{row}'].value
        oclc_results = sheet[f'G{row}'].value
        if (metadata and oclc_results and 
            oclc_results != "No matching records found" and 
            oclc_results.strip() != ""):
            total_valid_rows += 1
    
    use_batch = processor.should_use_batch(total_valid_rows)
    
    print(f"🤖 Processing mode: {'BATCH' if use_batch else 'INDIVIDUAL'}")
    print(f"📊 Valid rows to process: {total_valid_rows}")
    
    if use_batch and total_valid_rows > 0:
        print(f"📦 Preparing {total_valid_rows} requests for batch processing...")
        
        # Prepare batch requests
        batch_requests, custom_id_mapping, valid_rows = prepare_batch_requests(sheet, model_name)
        
        # Estimate costs
        cost_estimate = processor.estimate_batch_cost(batch_requests, model_name)
        
        print(f"💰 Cost estimate:")
        print(f"   Regular API: ${cost_estimate['regular_cost']:.4f}")
        print(f"   Batch API: ${cost_estimate['batch_cost']:.4f}")
        print(f"   Savings: ${cost_estimate['savings']:.4f} ({cost_estimate['savings_percentage']:.1f}%)")
        
        # Convert to batch format
        formatted_requests = processor.create_batch_requests(batch_requests, "oclc_analysis")
        
        # Submit batch
        batch_id = processor.submit_batch(
            formatted_requests, 
            f"OCLC Analysis - {total_valid_rows} items - {datetime.now().strftime('%Y-%m-%d')}"
        )
        
        # Wait for completion
        results = processor.wait_for_completion(batch_id, max_wait_hours=24, check_interval_minutes=5)
        
        if results:
            # Process batch results
            processed_results = processor.process_batch_results(results, custom_id_mapping)
            
            print(f"📊 Processing batch results...")
            
            # Initialize counters
            successful_calls = 0
            failed_calls = 0
            total_prompt_tokens = 0
            total_completion_tokens = 0
            total_tokens = 0
            
            # Process successful results
            for custom_id, result_data in processed_results["results"].items():
                if custom_id.startswith("oclc_analysis_"):
                    # Extract the index from custom_id
                    index = int(custom_id.split("_")[2])
                    mapping_key = f"req_{index}"
                    
                    if mapping_key in custom_id_mapping:
                        barcode = custom_id_mapping[mapping_key]["barcode"]
                        row = custom_id_mapping[mapping_key]["row_number"]
                        metadata = custom_id_mapping[mapping_key]["metadata"]
                        oclc_results = custom_id_mapping[mapping_key]["oclc_results"]
                        
                        if result_data["success"]:
                            analysis_result = result_data["content"]
                            usage = result_data["usage"]
                            
                            # Track tokens
                            prompt_tokens = usage.get("prompt_tokens", 0)
                            completion_tokens = usage.get("completion_tokens", 0)
                            tokens_used = prompt_tokens + completion_tokens
                            
                            total_prompt_tokens += prompt_tokens
                            total_completion_tokens += completion_tokens
                            total_tokens += tokens_used
                            successful_calls += 1
                            
                            # Parse the results
                            oclc_number, confidence_score, explanation, other_matches = parse_analysis_result(
                                analysis_result, oclc_results
                            )
                            
                            # Log individual response
                            log_individual_response(
                                logs_folder_path=logs_folder_path,
                                script_name="metadata_analysis",
                                row_number=row,
                                barcode=barcode,
                                response_text=analysis_result,
                                model_name=model_name,
                                prompt_tokens=prompt_tokens,
                                completion_tokens=completion_tokens,
                                processing_time=0  # Batch processing doesn't track individual timing
                            )
                            
                        else:
                            # Handle errors
                            failed_calls += 1
                            oclc_number = "Error processing"
                            confidence_score = 0
                            explanation = f"Error: {result_data['error']}"
                            other_matches = ""
                            prompt_tokens = completion_tokens = tokens_used = 0
                            
                            # Log error
                            log_individual_response(
                                logs_folder_path=logs_folder_path,
                                script_name="metadata_analysis",
                                row_number=row,
                                barcode=barcode,
                                response_text=f"ERROR: {result_data['error']}",
                                model_name=model_name,
                                prompt_tokens=0,
                                completion_tokens=0,
                                processing_time=0
                            )
                        
                        # Update both workbooks
                        update_workbook_row(sheet, temp_sheet, row, oclc_number, confidence_score, 
                                          explanation, other_matches, 0, prompt_tokens, 
                                          completion_tokens, tokens_used)
            
            # Handle skipped rows (no valid data)
            total_rows = process_skipped_rows(sheet, temp_sheet)
            
            # Return batch processing metrics
            summary = processed_results["summary"]
            return (total_rows, successful_calls, failed_calls, 0,  # 0 for total_time
                   total_prompt_tokens, total_completion_tokens, total_tokens)
        
        else:
            print("❌ Batch processing failed, falling back to individual processing...")
            use_batch = False
    
    # Fall back to individual processing if batch fails or isn't used
    if not use_batch:
        return process_individual(sheet, temp_sheet, logs_folder_path, model_name, results_folder)

def process_individual(sheet, temp_sheet, logs_folder_path, model_name, results_folder):
    """Process using individual API calls (original logic)."""
    
    # Define the columns
    BARCODE_COLUMN = 'D'  
    METADATA_COLUMN = 'E'
    OCLC_RESULTS_COLUMN = 'G'
    
    # Initialize counters for summary
    total_rows = 0
    successful_calls = 0
    failed_calls = 0
    total_api_time = 0
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    processed_rows = 0
    
    # Temporary file path for saving progress
    temp_output_file = "temp_cd_metadata_progress.xlsx"
    temp_output_path = os.path.join(results_folder, temp_output_file)
    
    for row in range(2, sheet.max_row + 1):  # Row 1 is the header
        row_start_time = time.time()
        metadata = sheet[f'{METADATA_COLUMN}{row}'].value
        oclc_results = sheet[f'{OCLC_RESULTS_COLUMN}{row}'].value
        barcode = sheet[f'{BARCODE_COLUMN}{row}'].value

        print(f"\n🔍 Analyzing Row {row}/{sheet.max_row}")
        print(f"   Barcode: {barcode}")
        print(f"   Progress: {((row-1)/(sheet.max_row-1))*100:.1f}%")

        # Skip rows with missing data or "No matching records" message
        if not metadata or not oclc_results or oclc_results == "No matching records found" or oclc_results.strip() == "":
            print(f"   ⏭️  Skipping: Missing data or no OCLC results")
            # Mark these rows as skipped in the results
            update_workbook_row(sheet, temp_sheet, row, "No OCLC data to process", 0, 
                              "Skipped: No valid OCLC results to analyze", "", 0, 0, 0, 0)
            
            # Copy data from other columns to temp sheet
            for col in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(col)
                temp_sheet[f'{col_letter}{row}'].value = sheet[f'{col_letter}{row}'].value
                
            total_rows += 1
            failed_calls += 1
            continue
        
        # Only rows with valid data will reach this point
        total_rows += 1
        print(f"   📝 Valid data found - proceeding with analysis")
        
        # Show OCLC results summary
        if oclc_results:
            oclc_record_count = oclc_results.count("OCLC Number:")
            print(f"   📚 Found {oclc_record_count} OCLC records to analyze")
            
        prompt = (
    f'''Analyze the following OCLC results based on the given metadata and determine which result is the best match. Methodically go through each record, choose the top 3, then consider them again and choose the record that matches the most elements in the metadata. If two or more records tie for best match, prioritize records that have more holdings and that are held by IXA. If there is no likely match, write "No matching records found".

    **Important Instructions**:
    1. Confidence Score: 0% indicates no confidence, and 100% indicates high confidence that we have found the correct OCLC number. If the confidence is below 79%, the record will be checked by a cataloger. 
    2. ***Key Fields in order of importance***:
    - UPC/Product Code (a match is HIGHEST priority if available in both metadata and OCLC record - if not available in one or the other, skip this field.  Occasionally, the UPC is partially obscured in the metadata - if some of the numbers in a UPC are incorrect but other fields are matching, it is still a match)
    - Title 
    - Artist/Performer
    - Contributors (some matching - not all need to match)
    - Publisher Name (the metadata record may include multiple names of publishers and distributor.  These do not all need to match the OCLC record, but there should be at least one exact match unless there are no visible publishers in the metadata; corporate ownership relationships like "Columbia is part of Sony" should NOT be considered a match.)
    - Physical Description (should make sense for a CD)
    - Content (track listings - these should be mostly similar, with small differences in spelling or punctuation)
    - Year (Should be an exact match if present in both the metadata and oclc record. If there are two years written in a record, the latter of the years is the reissue date, which is what we want to match)
    3. ***Notes on Matching Special Cases***:
    - Titles in non-Latin scripts that match in meaning or transliteration should also be considered equivalent.
    - If a field is marked as partially obscured, lessen the importance of that field in the matching process.
    - Different releases in different regions (e.g., Japanese release vs. US release) should be treated as different records even if title and content match.
    4. When information is not visible in the metadata, DO NOT use that field in your consideration of a match. It may be written in the metadata as 'not visible' or 'not available', etc.
    5. If there is a publisher in the OCLC record but it cannot be found anywhere in the metadata, the OCLC record or the CD may be a reissue - mark it as 79 because that way it will be checked by a cataloger. 
    6. The publisher should have at least one match between the metadata and OCLC record.  This may be a partial match, but it needs to be at least a fuzzy match.  No corporate relationships or associations unless explicitly mentioned in both the metadata and the OCLC record.  If the publisher is not visible in the metadata, do not use this field in your consideration of a match.
    7. If there is no likely match, write "No matching records found" and set the confidence score as 0.

    Format for Response:
    - Your response must follow this format exactly:
    1. OCLC number: [number or 'No matching records found']
    2. Confidence score: [%]
    3. Explanation: [List of things that match as key value pairs. If there are multiple records that could be a match, explain why you chose the one you did. If there are no matches, explain why.]
    4. Other potential good matches: [List of other OCLC numbers that could be good matches and a one sentence explanation for each match as key value pairs. If there are no other potential matches, write 'No other potential good matches.']
    
    Once you have responded, go back through the response that you wrote and carefully verify each piece of information. If you find a mistake, look for a better record. If there isn't one, reduce the confidence score to 79% or lower. If there is one, once again carefully verify all the facts that support your choice. If you still can't find a match, write "No matching records found" and set the confidence score as 0.

    Metadata: {metadata}

    OCLC Results: {oclc_results}
    ''')
        try:
            print(f"   🤖 Calling OpenAI API for analysis...")
            
            # Time the API call
            api_call_start = time.time()
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are a music cataloger.  You are very knowledgeable about music cataloging best practices, and also have incredible attention to detail.  Read through the metadata and OCLC results carefully, and determine which of the OCLC results looks like the best match. If there is no likely match, write 'No matching records found'.  If you make a mistake, you would feel very bad about it, so you always double check your work."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1500,
                temperature=0.5
            )
            api_call_duration = time.time() - api_call_start
            total_api_time += api_call_duration
        
            # Extract token information
            prompt_tokens = response.usage.prompt_tokens
            completion_tokens = response.usage.completion_tokens
            tokens_used = prompt_tokens + completion_tokens
            
            # Add to totals
            total_prompt_tokens += prompt_tokens
            total_completion_tokens += completion_tokens
            total_tokens += tokens_used
            
            successful_calls += 1

            analysis_result = response.choices[0].message.content.strip()
            
            print(f"   ✅ API call successful!")
            print(f"   ⏱️  API time: {api_call_duration:.2f}s")
            print(f"   🎯 Tokens: {tokens_used:,} (P:{prompt_tokens:,}, C:{completion_tokens:,})")
            
            # Log individual response
            log_individual_response(
                logs_folder_path=logs_folder_path,
                script_name="metadata_analysis",
                row_number=row,
                barcode=barcode,
                response_text=analysis_result,
                model_name=model_name,
                prompt_tokens=prompt_tokens,
                completion_tokens=completion_tokens,
                processing_time=api_call_duration
            )
            
            # Parse the results
            oclc_number, confidence_score, explanation, other_matches = parse_analysis_result(
                analysis_result, oclc_results
            )
            
            # Calculate total processing time for this row
            row_duration = time.time() - row_start_time
            processed_rows += 1
            
            print(f"   🎯 Selected OCLC: {oclc_number}")
            print(f"   📊 Confidence: {confidence_score}%")
            
            # Update workbooks
            update_workbook_row(sheet, temp_sheet, row, oclc_number, confidence_score, 
                              explanation, other_matches, row_duration, prompt_tokens, 
                              completion_tokens, tokens_used)
            
            # Copy data from other columns to temp sheet
            for col in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(col)
                temp_sheet[f'{col_letter}{row}'].value = sheet[f'{col_letter}{row}'].value

            # Progress tracking
            progress = (row - 1) / (sheet.max_row - 1)
            bar_length = 30
            filled_length = int(bar_length * progress)
            bar = '█' * filled_length + '-' * (bar_length - filled_length)
            print(f"   Progress: |{bar}| {progress*100:.1f}%")

        except Exception as e:
            failed_calls += 1
            print(f"   ❌ API call failed: {str(e)}")
            
            # Log errors
            log_individual_response(
                logs_folder_path=logs_folder_path,
                script_name="metadata_analysis",
                row_number=row,
                barcode=barcode,
                response_text=f"ERROR: {str(e)}",
                model_name=model_name,
                prompt_tokens=0,
                completion_tokens=0,
                processing_time=0
            )
                
            # Update both workbooks to show the error
            update_workbook_row(sheet, temp_sheet, row, "Error processing", 0, 
                              f"Error: {str(e)}", "", 0, 0, 0, 0)
            
            # Copy data from other columns to temp sheet
            for col in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(col)
                temp_sheet[f'{col_letter}{row}'].value = sheet[f'{col_letter}{row}'].value

    return total_rows, successful_calls, failed_calls, total_api_time, total_prompt_tokens, total_completion_tokens, total_tokens

def parse_analysis_result(analysis_result, oclc_results):
    """Parse the AI analysis result into structured components."""
    oclc_number = "Not found"
    confidence_score = 0
    explanation = "Could not parse response"
    other_matches = ""

    try:
        # Parse the results and show key findings
        if "OCLC number:" in analysis_result:
            oclc_part = analysis_result.split("OCLC number:")[1].split("\n")[0].strip()
            oclc_number = ''.join(char for char in oclc_part if char.isdigit())

        if "Confidence score:" in analysis_result:
            confidence_part = analysis_result.split("Confidence score:")[1].split("%")[0].strip()
            try:
                confidence_score = int(float(confidence_part))
                confidence_score = min(100, max(0, confidence_score))
            except ValueError:
                confidence_score = 0

        if "Explanation:" in analysis_result:
            explanation_parts = analysis_result.split("Explanation:")[1].split("Other potential good matches:")
            explanation = explanation_parts[0].strip()
            if explanation.endswith("4."):
                explanation = explanation[:-2].strip()
            explanation = re.sub(r'\s+\d+\.\s*$', '', explanation)

        if "Other potential good matches:" in analysis_result:
            other_matches_part = analysis_result.split("Other potential good matches:")[1].strip()
            if other_matches_part and oclc_results:
                # Extract OCLC numbers from the LLM response
                oclc_patterns = re.findall(r'OCLC Number:?\s*(\d{8,10})', other_matches_part, re.IGNORECASE)
                
                if not oclc_patterns:
                    oclc_patterns = re.findall(r'[- ]*OCLC(?:\s+Number)?:?\s*(\d{8,10})', other_matches_part, re.IGNORECASE)
                
                if not oclc_patterns:
                    oclc_patterns = re.findall(r'\b(\d{8,10})\b', other_matches_part)
                
                formatted_matches = []
                
                # Process each potential match
                for match_num in oclc_patterns:
                    if f"OCLC Number: {match_num}" in oclc_results:
                        split_results = oclc_results.split(f"OCLC Number: {match_num}")
                        if len(split_results) > 1:
                            match_section = split_results[1].split("----------------------------------------")[0]
                            
                            held_by_ixa = "Yes" if "Held by IXA: Yes" in match_section else "No"
                            
                            total_holdings = "0"
                            if "Total Institutions Holding:" in match_section:
                                holdings_part = match_section.split("Total Institutions Holding:")[1].split("\n")[0].strip()
                                total_holdings = holdings_part
                            
                            title = ""
                            if "- Main Title:" in match_section:
                                title_part = match_section.split("- Main Title:")[1].split("\n")[0].strip()
                                title = title_part
                            
                            specific_format = ""
                            if "- specificFormat:" in match_section:
                                format_part = match_section.split("- specificFormat:")[1].split("\n")[0].strip()
                                specific_format = format_part
                            
                            match_info = f"OCLC: {match_num} | IXA: {held_by_ixa} | Holdings: {total_holdings}"
                            if specific_format:
                                match_info += f" | Format: {specific_format}"
                            if title:
                                match_info += f" | Title: {title}"
                            
                            formatted_matches.append(match_info)
                
                if formatted_matches:
                    other_matches = "\n".join(formatted_matches) + "\n\nOriginal LLM response:\n" + other_matches_part
                else:
                    other_matches = "No structured matches found.\n\nOriginal LLM response:\n" + other_matches_part
            else:
                other_matches = other_matches_part
    
    except Exception as parsing_error:
        print(f"   ⚠️  Error parsing response: {parsing_error}")
        
    return oclc_number, confidence_score, explanation, other_matches

def update_workbook_row(sheet, temp_sheet, row, oclc_number, confidence_score, explanation, 
                       other_matches, row_duration, prompt_tokens, completion_tokens, tokens_used):
    """Update both main and temp workbooks with results."""
    
    # Define the columns
    RESULT_COLUMN = 'H'
    CONFIDENCE_SCORE_COLUMN = 'I'
    EXPLANATION_COLUMN = 'J'
    OTHER_MATCHES_COLUMN = 'K'  
    PROCESSING_TIME_COLUMN = 'L'  
    PROMPT_TOKENS_COLUMN = 'M'   
    COMPLETION_TOKENS_COLUMN = 'N'  
    TOTAL_TOKENS_COLUMN = 'O'
    
    # Update main workbook
    sheet[f'{RESULT_COLUMN}{row}'].value = oclc_number if isinstance(oclc_number, str) else int(oclc_number)
    sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = confidence_score
    sheet[f'{EXPLANATION_COLUMN}{row}'].value = explanation
    sheet[f'{OTHER_MATCHES_COLUMN}{row}'].value = other_matches
    sheet[f'{PROCESSING_TIME_COLUMN}{row}'].value = round(row_duration, 2)
    sheet[f'{PROMPT_TOKENS_COLUMN}{row}'].value = prompt_tokens
    sheet[f'{COMPLETION_TOKENS_COLUMN}{row}'].value = completion_tokens
    sheet[f'{TOTAL_TOKENS_COLUMN}{row}'].value = tokens_used

    # Update temp workbook
    temp_sheet[f'{RESULT_COLUMN}{row}'].value = oclc_number if isinstance(oclc_number, str) else int(oclc_number)
    temp_sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = confidence_score
    temp_sheet[f'{EXPLANATION_COLUMN}{row}'].value = explanation
    temp_sheet[f'{OTHER_MATCHES_COLUMN}{row}'].value = other_matches
    temp_sheet[f'{PROCESSING_TIME_COLUMN}{row}'].value = round(row_duration, 2)
    temp_sheet[f'{PROMPT_TOKENS_COLUMN}{row}'].value = prompt_tokens
    temp_sheet[f'{COMPLETION_TOKENS_COLUMN}{row}'].value = completion_tokens
    temp_sheet[f'{TOTAL_TOKENS_COLUMN}{row}'].value = tokens_used
    
    # Set cell alignment in both workbooks
    for col in [RESULT_COLUMN, CONFIDENCE_SCORE_COLUMN, EXPLANATION_COLUMN, OTHER_MATCHES_COLUMN, 
                PROCESSING_TIME_COLUMN, PROMPT_TOKENS_COLUMN, COMPLETION_TOKENS_COLUMN, TOTAL_TOKENS_COLUMN]:
        sheet[f'{col}{row}'].alignment = Alignment(wrap_text=True)
        temp_sheet[f'{col}{row}'].alignment = Alignment(wrap_text=True)

def process_skipped_rows(sheet, temp_sheet):
    """Process rows that were skipped due to missing data."""
    total_rows = 0
    
    for row in range(2, sheet.max_row + 1):
        metadata = sheet[f'E{row}'].value
        oclc_results = sheet[f'G{row}'].value
        
        # Count all rows
        total_rows += 1
        
        # Handle skipped rows that weren't processed in batch
        if (not metadata or not oclc_results or 
            oclc_results == "No matching records found" or 
            oclc_results.strip() == ""):
            
            # Only update if not already processed
            if not sheet[f'H{row}'].value:
                update_workbook_row(sheet, temp_sheet, row, "No OCLC data to process", 0, 
                                  "Skipped: No valid OCLC results to analyze", "", 0, 0, 0, 0)
            
            # Copy data from other columns to temp sheet
            for col in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(col)
                temp_sheet[f'{col_letter}{row}'].value = sheet[f'{col_letter}{row}'].value
    
    return total_rows

def main():
    model_name = "01-mini"  
    
    # Start timing the entire script execution
    script_start_time = time.time()
    
    # Specify the folder prefix (adjust if needed)
    base_dir_prefix = "ai-music-workflow/cd-processing/cd-output-folders/results-"

    # Find the latest results folder using the prefix.
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Run the first script first.")
        exit()
        
    print(f"Using results folder: {results_folder}")

    # Create logs folder within the results folder
    logs_folder_path = os.path.join(results_folder, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)

    # Look for step 2 files in the results folder
    step2_files = [f for f in os.listdir(results_folder)
            if f.startswith('cd-metadata-ai-') and f.endswith('.xlsx')]

    if not step2_files:
        print("No step 2 files found in the results folder!")
        exit()
        
    # Get the latest step 2 file
    latest_file = max(step2_files)
    workbook_path = os.path.join(results_folder, latest_file)

    print(f"Processing file: {workbook_path}")

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb.active
    
    # Show model pricing info at start
    model_info = get_model_info(model_name)
    if model_info:
        print(f"🧠 STEP 3: AI ANALYSIS OF OCLC MATCHES")
        print(f"Using model: {model_name}")
        print(f"Pricing: ${model_info['input_per_1k']:.5f}/1K input, ${model_info['output_per_1k']:.5f}/1K output")
        print(f"Batch discount: {model_info['batch_discount']*100:.0f}%")
        print(f"Total rows to analyze: {sheet.max_row - 1}")
        print("-" * 50)
    
    # Create a temporary workbook for frequent saving (no images)
    temp_wb = openpyxl.Workbook()
    temp_sheet = temp_wb.active
    
    # Copy headers and column settings
    for col_idx, cell in enumerate(sheet[1], 1):
        temp_sheet.cell(row=1, column=col_idx, value=cell.value)
        # Copy column widths where available
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        if column_letter in sheet.column_dimensions:
            temp_sheet.column_dimensions[column_letter].width = sheet.column_dimensions[column_letter].width

    # Create a summary sheet for token and timing metrics
    if "TokenSummary" not in wb.sheetnames:
        summary_sheet = wb.create_sheet("TokenSummary")
    else:
        summary_sheet = wb["TokenSummary"]
    
    # Also create summary in temp workbook
    temp_summary_sheet = temp_wb.create_sheet("TokenSummary")
    
    # Set up the summary sheet headers
    summary_headers = [
        "Total Rows Processed", "Total API Calls", "Failed API Calls",
        "Total Processing Time (s)", "API Time (s)", "Average Time per Call (s)",
        "Total Prompt Tokens", "Total Completion Tokens", "Total Tokens",
        "Average Tokens per Call", "Estimated Cost", "Date Completed"
    ]
    summary_sheet.append(summary_headers)
    temp_summary_sheet.append(summary_headers)
    
    for col, header in enumerate(summary_headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col)
        summary_sheet.column_dimensions[col_letter].width = 20
        temp_summary_sheet.column_dimensions[col_letter].width = 20

    # Define the columns
    RESULT_COLUMN = 'H'
    CONFIDENCE_SCORE_COLUMN = 'I'
    EXPLANATION_COLUMN = 'J'
    OTHER_MATCHES_COLUMN = 'K'  
    PROCESSING_TIME_COLUMN = 'L'  
    PROMPT_TOKENS_COLUMN = 'M'   
    COMPLETION_TOKENS_COLUMN = 'N'  
    TOTAL_TOKENS_COLUMN = 'O'    

    # Add headers for all columns including new ones
    sheet[f'{RESULT_COLUMN}1'] = 'LLM-Assessed Correct OCLC #'
    sheet[f'{CONFIDENCE_SCORE_COLUMN}1'] = 'LLM Confidence Score'
    sheet[f'{EXPLANATION_COLUMN}1'] = 'LLM Explanation'
    sheet[f'{OTHER_MATCHES_COLUMN}1'] = 'Other Potential Matches'
    sheet[f'{PROCESSING_TIME_COLUMN}1'] = 'Processing Time (s)'
    sheet[f'{PROMPT_TOKENS_COLUMN}1'] = 'Prompt Tokens'
    sheet[f'{COMPLETION_TOKENS_COLUMN}1'] = 'Completion Tokens'
    sheet[f'{TOTAL_TOKENS_COLUMN}1'] = 'Total Tokens'
    
    # Add the same headers to temp sheet
    temp_sheet[f'{RESULT_COLUMN}1'] = 'LLM-Assessed Correct OCLC #'
    temp_sheet[f'{CONFIDENCE_SCORE_COLUMN}1'] = 'LLM Confidence Score'
    temp_sheet[f'{EXPLANATION_COLUMN}1'] = 'LLM Explanation'
    temp_sheet[f'{OTHER_MATCHES_COLUMN}1'] = 'Other Potential Matches'
    temp_sheet[f'{PROCESSING_TIME_COLUMN}1'] = 'Processing Time (s)'
    temp_sheet[f'{PROMPT_TOKENS_COLUMN}1'] = 'Prompt Tokens'
    temp_sheet[f'{COMPLETION_TOKENS_COLUMN}1'] = 'Completion Tokens'
    temp_sheet[f'{TOTAL_TOKENS_COLUMN}1'] = 'Total Tokens'

    # Set column widths
    sheet.column_dimensions[RESULT_COLUMN].width = 30
    sheet.column_dimensions[CONFIDENCE_SCORE_COLUMN].width = 20
    sheet.column_dimensions[EXPLANATION_COLUMN].width = 40
    sheet.column_dimensions[OTHER_MATCHES_COLUMN].width = 70
    sheet.column_dimensions[PROCESSING_TIME_COLUMN].width = 18
    sheet.column_dimensions[PROMPT_TOKENS_COLUMN].width = 15
    sheet.column_dimensions[COMPLETION_TOKENS_COLUMN].width = 18
    sheet.column_dimensions[TOTAL_TOKENS_COLUMN].width = 15
    
    # Mirror column widths in temp sheet
    temp_sheet.column_dimensions[RESULT_COLUMN].width = 30
    temp_sheet.column_dimensions[CONFIDENCE_SCORE_COLUMN].width = 20
    temp_sheet.column_dimensions[EXPLANATION_COLUMN].width = 40
    temp_sheet.column_dimensions[OTHER_MATCHES_COLUMN].width = 70
    temp_sheet.column_dimensions[PROCESSING_TIME_COLUMN].width = 18
    temp_sheet.column_dimensions[PROMPT_TOKENS_COLUMN].width = 15
    temp_sheet.column_dimensions[COMPLETION_TOKENS_COLUMN].width = 18
    temp_sheet.column_dimensions[TOTAL_TOKENS_COLUMN].width = 15
    
    # Process with batch or individual logic
    (total_rows, successful_calls, failed_calls, total_api_time, 
     total_prompt_tokens, total_completion_tokens, total_tokens) = process_with_batch(
        sheet, temp_sheet, logs_folder_path, model_name, results_folder)

    # Calculate script metrics
    script_duration = time.time() - script_start_time
    avg_call_time = total_api_time / successful_calls if successful_calls > 0 else 0
    avg_tokens_per_call = total_tokens / successful_calls if successful_calls > 0 else 0
    
    # Determine if batch processing was used (check if we have many successful calls but zero API time)
    was_batch_processed = successful_calls > 10 and total_api_time == 0
    
    # Calculate actual cost using the model pricing
    estimated_cost = calculate_cost(
        model_name=model_name,
        prompt_tokens=total_prompt_tokens,
        completion_tokens=total_completion_tokens,
        is_batch=was_batch_processed
    )

    # Add summary data
    summary_sheet.append([
        total_rows,
        successful_calls,
        failed_calls,
        round(script_duration, 2),
        round(total_api_time, 2),
        round(avg_call_time, 2),
        total_prompt_tokens,
        total_completion_tokens,
        total_tokens,
        round(avg_tokens_per_call, 2),
        round(estimated_cost, 4),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    # Create standardized token usage log
    create_token_usage_log(
        logs_folder_path=logs_folder_path,
        script_name="metadata_analysis",
        model_name=model_name,
        total_items=total_rows,
        items_with_issues=failed_calls,
        total_time=script_duration,
        total_prompt_tokens=total_prompt_tokens,
        total_completion_tokens=total_completion_tokens,
        additional_metrics={
            "Successful API calls": successful_calls,
            "Average API call time": f"{avg_call_time:.2f}s",
            "Total script execution time": f"{script_duration:.2f}s",
            "API time percentage": f"{(total_api_time/script_duration)*100:.1f}%" if script_duration > 0 else "0%",
            "Rows with valid data": successful_calls,
            "Rows skipped (no data)": total_rows - successful_calls - failed_calls,
            "Processing mode": "BATCH" if was_batch_processed else "INDIVIDUAL",
            "Actual cost": f"${estimated_cost:.4f}"
        }
    )

    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"cd-metadata-ai-{current_date}.xlsx"
    full_output_path = os.path.join(results_folder, output_file)
        
    wb.save(full_output_path)

    # Clean up temporary file
    temp_output_path = os.path.join(results_folder, "temp_cd_metadata_progress.xlsx")
    try:
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
            print(f"Temporary progress file removed: {temp_output_path}")
    except Exception as remove_error:
        print(f"Warning: Could not remove temporary progress file: {remove_error}")

    print(f"\nResults saved to {full_output_path}")
    print(f"Token usage log saved to {os.path.join(logs_folder_path, 'metadata_analysis_token_usage_log.txt')}")
    print(f"Full responses log saved to {os.path.join(logs_folder_path, 'metadata_analysis_full_responses_log.txt')}")

    # Enhanced final summary with correct cost calculation
    print(f"\n🎉 STEP 3 COMPLETED!")
    print(f"✅ Successfully analyzed: {successful_calls} records")
    print(f"❌ Failed calls: {failed_calls}")
    print(f"⏱️  Total script time: {script_duration:.1f}s ({script_duration/60:.1f} minutes)")
    print(f"⏱️  Total API time: {total_api_time:.1f}s")
    print(f"🎯 Total tokens: {total_tokens:,} (Input: {total_prompt_tokens:,}, Output: {total_completion_tokens:,})")
    print(f"🤖 Processing mode: {'BATCH' if was_batch_processed else 'INDIVIDUAL'}")
    print(f"💰 Actual cost: ${estimated_cost:.4f}")
    
    # Show batch savings if applicable
    if was_batch_processed:
        regular_cost = calculate_cost(model_name, total_prompt_tokens, total_completion_tokens, is_batch=False)
        savings = regular_cost - estimated_cost
        savings_percentage = (savings / regular_cost) * 100 if regular_cost > 0 else 0
        print(f"💰 Regular API cost would have been: ${regular_cost:.4f}")
        print(f"💰 Batch savings: ${savings:.4f} ({savings_percentage:.1f}%)")

if __name__ == "__main__":
    main()