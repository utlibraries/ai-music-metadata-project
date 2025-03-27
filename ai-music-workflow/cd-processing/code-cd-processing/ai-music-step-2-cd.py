import os
import glob
import json
import requests
import time
import shutil  # Added for file operations
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import re

api_calls = {'count': 0, 'reset_time': time.time()}

def find_latest_results_folder(prefix):
    # Get the parent directory of the prefix
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None

    latest_folder = max(matching_folders)
    
    return latest_folder

def get_access_token(client_id, client_secret):
    token_url = "https://oauth.oclc.org/token"
    data = {
        "grant_type": "client_credentials",
        "scope": "wcapi"
    }
    response = requests.post(token_url, data=data, auth=(client_id, client_secret))
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        raise Exception(f"Failed to get access token: {response.text}")
    
def extract_metadata_fields(metadata_str):
    # Initialize the same structure you already expect
    fields = {
        "Main Title": None,
        "English Title": None,
        "Subtitle": None,
        "Primary Contributor": {"Artist/Performer": None},
        "Publishers": [{"Name": None, "Numbers": None}],
        "Contents": {"tracks": []}
    }
    
    def clean_value(value):
        """
        Strip out placeholders like "not visible",
        remove any leading field-label strings, etc.
        """
        if not value or any(x in str(value).lower() for x in ["not visible", "not available", "n/a", "unavailable", "unknown", "[none]", "none", "not present", "not listed", "not applicable"]):
            return None
        value = re.sub(r'^-\s*', '', value)
        value = re.sub(r'^(Main Title:|English Title:|Subtitle:|Primary Contributor:|Artist/Performer:|Name:|Numbers:)\s*', '', value, flags=re.IGNORECASE)
        return value.strip()

    # Split metadata into individual lines
    lines = metadata_str.splitlines()
    
    # Regex that detects the known fields.
    pattern = re.compile(r'^\-\s*(Main Title|English Title|Subtitle|Primary Contributor|Artist/Performer|Name|Numbers):\s*(.*)$')
    newer_pattern = re.compile(r'^\s*\-\s*(Main Title|English Title|Subtitle|Primary Contributor|Artist/Performer|Name|Numbers):\s*(.*)$')

    for line in lines:
        line = line.strip()
        if not line:
            continue  # skip blank lines

        match = pattern.match(line) or newer_pattern.match(line)
        if match:
            field_name = match.group(1)  # e.g. "Main Title"
            field_value = clean_value(match.group(2))

            if not field_value:
                continue

            field_name_lower = field_name.lower()

            if field_name_lower in ["main title", "english title", "subtitle"]:
                fields[field_name.title()] = field_value  # "Main Title", "English Title", "Subtitle"
            elif field_name_lower in ["primary contributor", "artist/performer"]:
                fields["Primary Contributor"]["Artist/Performer"] = field_value
            elif field_name_lower == "name":
                fields["Publishers"][0]["Name"] = field_value
            elif field_name_lower == "numbers":
                fields["Publishers"][0]["Numbers"] = field_value

    # Handle track titles - try both formats
    track_matches_quoted = re.finditer(r'"title":\s*"([^"]+)"', metadata_str)
    for match in track_matches_quoted:
        track_title = clean_value(match.group(1))
        if track_title:
            fields["Contents"]["tracks"].append({"title": track_title})
    
    if not fields["Contents"]["tracks"]:
        track_matches_unquoted = re.finditer(r'"title":\s*([^,\n]+)', metadata_str)
        for match in track_matches_unquoted:
            track_title = clean_value(match.group(1))
            if track_title:
                fields["Contents"]["tracks"].append({"title": track_title})

    return fields


def construct_queries_from_metadata(metadata):
    """Generate all possible query combinations without limiting to just 5."""
    def safe_get(value):
        if not value or not isinstance(value, str):
            return None
        if any(x in value.lower() for x in ["not visible", "not available", "n/a", "unavailable", "unknown", " [none]", "none", "not present", "not listed", "not applicable"]):
            return None
        cleaned = re.sub(r'^-\s', '', value.strip())
        cleaned = re.sub(r'^(Primary Contributor:|Artist/Performer:|Name:)\s', '', cleaned)
        return cleaned if cleaned else None

    title = safe_get(metadata.get('Main Title'))
    subtitle = safe_get(metadata.get('Subtitle'))
    artist = safe_get(metadata.get('Primary Contributor', {}).get('Artist/Performer'))
    publisher = safe_get(metadata.get('Publishers', [{}])[0].get('Name'))
    pub_numbers = safe_get(metadata.get('Publishers', [{}])[0].get('Numbers'))
    pub_year = metadata.get('Dates', {}).get('publicationDate')

    product_code = None
    if isinstance(pub_numbers, str):
        code_match = re.search(r'(UPC|EAN):\s*([^,\]]+)', pub_numbers, re.IGNORECASE)
        if not code_match:
            code_candidates = re.findall(r'\d[\d\s-]{10,}\d', pub_numbers)
            for candidate in code_candidates:
                digits_only = re.sub(r'\D', '', candidate)
                if len(digits_only) in [12, 13]:
                    code_match = re.match(r'(.*)', digits_only)
                    break
        
        if code_match:
            if isinstance(code_match.group(1), str) and code_match.group(1).upper() in ['UPC', 'EAN']:
                potential_code = code_match.group(2).strip()
            else:
                potential_code = code_match.group(1).strip()
                
            digits_only = re.sub(r'\D', '', potential_code)
            if len(digits_only) in [12, 13]:
                product_code = digits_only

    tracks = metadata.get('Contents', {}).get('tracks', [])
    first_track = next((safe_get(track.get('title')) for track in tracks if track.get('title')), None)
    
    second_track = None
    if len(tracks) > 1:
        second_track = safe_get(tracks[1].get('title')) if tracks[1].get('title') else None
    
    third_track = None
    if len(tracks) > 2:
        third_track = safe_get(tracks[2].get('title')) if tracks[2].get('title') else None

    queries = []

    if artist and first_track and second_track:
        queries.append(f'"{artist}" "{first_track}" "{second_track}"')
        
    if first_track and second_track:
        queries.append(f'"{first_track}" "{second_track}"')
        
    if title:
        if all([title, subtitle, artist]):
            queries.append(f'"{title}" "{subtitle}" "{artist}"')
            
        if all([title, first_track, second_track]):
            queries.append(f'{title} "{first_track}" "{second_track}"')
            
        if all([title, artist, publisher]):
            queries.append(f'"{title}" "{artist}" {publisher}')

        if title and artist:
            queries.append(f'"{title}" "{artist}"')

        if all([title, first_track]):
            queries.append(f'"{title}" "{first_track}"')
            
        if all([title, subtitle]):
            queries.append(f'"{title}" "{subtitle}"')

        if title and publisher:
            queries.append(f'"{title}" {publisher}')

        if title and product_code:
            queries.append(f'"{title}" {product_code}')
        
    if artist and publisher and product_code:
        queries.append(f'"{artist}" {publisher} "{product_code}"')

    if artist and publisher and pub_year:
        queries.append(f'{artist} {publisher} {pub_year}')

    if artist and publisher and first_track:
        queries.append(f'{artist} {publisher} {first_track}')
        
    if artist and third_track:
        queries.append(f'{artist} {third_track}')
    
    if first_track:
        queries.append(f'"{first_track}"')
        
    if third_track:
        queries.append(f'"{third_track}"')

    if artist:
        queries.append(f'{artist}')

    if product_code:
        queries.append(f'{product_code}')

    seen = set()
    unique_queries = []
    for q in queries:
        if q not in seen:
            seen.add(q)
            unique_queries.append(q)

    # Return ALL unique queries instead of just the first 5
    return unique_queries

def format_oclc_results(json_response, access_token):
    try:
        data = json.loads(json_response)
        if not isinstance(data, dict):
            return "Error: Invalid JSON response"
            
        total_records = data.get('numberOfRecords', 0)
        if total_records == 0:
            return "No records found"
            
        formatted_results = []
        valid_records = []
        
        for record in data.get('bibRecords', []):
            include_record = False
            if 'format' in record and 'specificFormat' in record['format']:
                specific_format = record['format']['specificFormat']
                if isinstance(specific_format, str) and any(cd_term in specific_format for cd_term in ["CD", "compact disc", "Compact Disc"]):
                    include_record = True
            
            if include_record:
                valid_records.append(record)
        
        filtered_total = len(valid_records)
        if filtered_total == 0:
            return "No matching records with CD format found"
            
        formatted_results.append(f"Total Records Found (CD format only): {filtered_total}\n")
        
        for idx, record in enumerate(valid_records[:5], 1):
            formatted_results.append(f"\nRecord {idx}:")
            formatted_results.append("-" * 40)
            
            oclc_number = None
            if 'identifier' in record and 'oclcNumber' in record['identifier']:
                oclc_number = record['identifier']['oclcNumber']
                formatted_results.append(f"OCLC Number: {oclc_number}")
            
            if oclc_number:
                is_held_by_IXA, total_holding_count, holding_institutions = get_holdings_info(oclc_number, access_token)
                formatted_results.append(f"\nHeld by IXA: {'Yes' if is_held_by_IXA else 'No'}")
                formatted_results.append(f"Total Institutions Holding: {total_holding_count}")
            
            if 'identifier' in record:
                formatted_results.append("\nIdentifier:")
                # Add OCLC number
                if 'oclcNumber' in record['identifier']:
                    formatted_results.append(f"  - oclcNumber: {record['identifier']['oclcNumber']}")
                
                # Add UPC if it exists
                if 'otherStandardIdentifiers' in record['identifier']:
                    for id_item in record['identifier']['otherStandardIdentifiers']:
                        if isinstance(id_item, dict) and id_item.get('type') == 'Universal Product Code (UPC)':
                            formatted_results.append(f"  - UPC: {id_item.get('id', 'N/A')}")
            
            if 'title' in record:
                formatted_results.append("Title Information:")
                if 'mainTitles' in record['title']:
                    for title in record['title']['mainTitles']:
                        formatted_results.append(f"  - Main Title: {title.get('text', 'N/A')}")
                if 'subtitles' in record['title']:
                    for subtitle in record['title']['subtitles']:
                        formatted_results.append(f"  - Subtitle: {subtitle.get('text', 'N/A')}")
            
            if 'contributor' in record:
                formatted_results.append("Contributors:")
                for creator_type in ['creators', 'contributors']:
                    if creator_type in record['contributor']:
                        for person in record['contributor'][creator_type]:
                            if 'firstName' in person and 'secondName' in person:
                                name = f"{person.get('firstName', {}).get('text', '')} {person.get('secondName', {}).get('text', '')}"
                            elif 'nonPersonName' in person:
                                name = person['nonPersonName'].get('text', '')
                            else:
                                name = 'N/A'
                            role = person.get('type', 'N/A')
                            formatted_results.append(f"  - {name.strip()} ({role})")
                        
            if 'publishers' in record:
                formatted_results.append("Publishers:")
                for pub in record['publishers']:
                    pub_name = pub.get('publisherName', {}).get('text', 'N/A')
                    pub_place = pub.get('publicationPlace', 'N/A')
                    formatted_results.append(f"  - Name: {pub_name}")
                    formatted_results.append(f"    Place: {pub_place}")
            
            if 'date' in record:
                formatted_results.append("Dates:")
                # Only include publicationDate
                if 'publicationDate' in record['date']:
                    formatted_results.append(f"  - publicationDate: {record['date']['publicationDate']}")
            
            if 'language' in record:
                formatted_results.append("Language:")
                for key, value in record['language'].items():
                    formatted_results.append(f"  - {key}: {value}")
                        
            if 'description' in record:
                formatted_results.append("Description:")
                if 'physicalDescription' in record['description']:
                    formatted_results.append(f"  - Physical: {record['description']['physicalDescription']}")
                if 'contents' in record['description']:
                    for content in record['description']['contents']:
                        if 'contentNote' in content:
                            formatted_results.append(f"  - Content: {content['contentNote'].get('text', '')}")
                        
            formatted_results.append("-" * 40)
            
        return "\n".join(formatted_results)
        
    except json.JSONDecodeError:
        return "Error: Invalid JSON response"
    except Exception as e:
        return f"Error formatting results: {str(e)}"
    
def format_oclc_api_response_for_accumulation(data, access_token, seen_oclc_numbers=None):
    if seen_oclc_numbers is None:
        seen_oclc_numbers = set()
        
    try:
        if not isinstance(data, dict):
            return "Error: Invalid JSON response"
            
        total_records = data.get('numberOfRecords', 0)
        if total_records == 0:
            return "No records found"
            
        formatted_results = []
        valid_records = []
        
        for record in data.get('bibRecords', []):
            # Skip records we've already seen
            oclc_number = None
            if 'identifier' in record and 'oclcNumber' in record['identifier']:
                oclc_number = record['identifier']['oclcNumber']
                if oclc_number in seen_oclc_numbers:
                    continue
            
            include_record = False
            if 'format' in record and 'specificFormat' in record['format']:
                specific_format = record['format']['specificFormat']
                if isinstance(specific_format, str) and any(cd_term in specific_format for cd_term in ["CD", "compact disc", "Compact Disc"]):
                    include_record = True
            
            if include_record:
                valid_records.append(record)
        
        filtered_total = len(valid_records)
        if filtered_total == 0:
            return "No matching records with CD format found"
        
        for record in valid_records[:5]:
            # Add a divider line between records
            formatted_results.append("\n" + "-" * 40)
            
            oclc_number = None
            if 'identifier' in record and 'oclcNumber' in record['identifier']:
                oclc_number = record['identifier']['oclcNumber']
                formatted_results.append(f"OCLC Number: {oclc_number}")
            
            if oclc_number:
                is_held_by_IXA, total_holding_count, holding_institutions = get_holdings_info(oclc_number, access_token)
                formatted_results.append(f"\nHeld by IXA: {'Yes' if is_held_by_IXA else 'No'}")
                formatted_results.append(f"Total Institutions Holding: {total_holding_count}")
            
            if 'identifier' in record:
                formatted_results.append("\nIdentifier:")
                # Add OCLC number
                if 'oclcNumber' in record['identifier']:
                    formatted_results.append(f"  - oclcNumber: {record['identifier']['oclcNumber']}")
                
                # Add UPC if it exists
                if 'otherStandardIdentifiers' in record['identifier']:
                    for id_item in record['identifier']['otherStandardIdentifiers']:
                        if isinstance(id_item, dict) and id_item.get('type') == 'Universal Product Code (UPC)':
                            formatted_results.append(f"  - UPC: {id_item.get('id', 'N/A')}")
            
            if 'title' in record:
                formatted_results.append("Title Information:")
                if 'mainTitles' in record['title']:
                    for title in record['title']['mainTitles']:
                        formatted_results.append(f"  - Main Title: {title.get('text', 'N/A')}")
                if 'subtitles' in record['title']:
                    for subtitle in record['title']['subtitles']:
                        formatted_results.append(f"  - Subtitle: {subtitle.get('text', 'N/A')}")
                if 'seriesTitles' in record['title']:
                    for series in record['title']['seriesTitles']:
                        formatted_results.append(f"  - Series Title: {series.get('seriesTitle', 'N/A')}")
            
            if 'contributor' in record:
                formatted_results.append("Contributors:")
                for creator_type in ['creators', 'contributors']:
                    if creator_type in record['contributor']:
                        for person in record['contributor'][creator_type]:
                            if 'firstName' in person and 'secondName' in person:
                                name = f"{person.get('firstName', {}).get('text', '')} {person.get('secondName', {}).get('text', '')}"
                            elif 'nonPersonName' in person:
                                name = person['nonPersonName'].get('text', '')
                            else:
                                name = 'N/A'
                            role = person.get('type', 'N/A')
                            formatted_results.append(f"  - {name.strip()} ({role})")
            
            if 'publishers' in record:
                formatted_results.append("Publishers:")
                for pub in record['publishers']:
                    pub_name = pub.get('publisherName', {}).get('text', 'N/A')
                    pub_place = pub.get('publicationPlace', 'N/A')
                    formatted_results.append(f"  - Name: {pub_name}")
                    formatted_results.append(f"    Place: {pub_place}")
            
            if 'date' in record:
                formatted_results.append("Dates:")
                # Only include publicationDate
                if 'publicationDate' in record['date']:
                    formatted_results.append(f"  - publicationDate: {record['date']['publicationDate']}")
            
            if 'language' in record:
                formatted_results.append("Language:")
                for key, value in record['language'].items():
                    formatted_results.append(f"  - {key}: {value}")
                        
            if 'musicInfo' in record:
                formatted_results.append("Music Information:")
                for key, value in record['musicInfo'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            if 'description' in record:
                formatted_results.append("Description:")
                if 'physicalDescription' in record['description']:
                    formatted_results.append(f"  - Physical: {record['description']['physicalDescription']}")
                if 'contents' in record['description']:
                    for content in record['description']['contents']:
                        if 'contentNote' in content:
                            formatted_results.append(f"  - Content: {content['contentNote'].get('text', '')}")
            if 'note' in record:
                formatted_results.append("Notes:")
                if isinstance(record['note'], dict):
                    for key, value in record['note'].items():
                        formatted_results.append(f"  - {key}: {value}")
                elif isinstance(record['note'], list):
                    for note in record['note']:
                        formatted_results.append(f"  - {note}")
                                    
            formatted_results.append("-" * 40)
            
        return "\n".join(formatted_results), filtered_total
        
    except Exception as e:
        return f"Error formatting results: {str(e)}", 0
    
def get_holdings_info(oclc_number, access_token):
    global api_calls
    api_calls['count'] += 1
    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs-holdings"
    
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    
    params = {
        "oclcNumber": oclc_number
    }
    
    try:
        response = requests.get(endpoint, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        is_held_by_IXA = False
        total_holding_count = 0
        holding_institutions = []
        
        if "briefRecords" in data and len(data["briefRecords"]) > 0:
            record = data["briefRecords"][0]
            
            if "institutionHolding" in record:
                holdings = record["institutionHolding"]
                total_holding_count = holdings.get("totalHoldingCount", 0)
                
                if "briefHoldings" in holdings:
                    for holding in holdings["briefHoldings"]:
                        symbol = holding.get("oclcSymbol", "")
                        inst_name = holding.get("institutionName", "")
                        if inst_name:
                            import html
                            inst_name = html.unescape(inst_name)
                        formatted_holding = f"{symbol} ({inst_name})" if inst_name else symbol
                        holding_institutions.append(formatted_holding)
                        if symbol == "IXA":
                            is_held_by_IXA = True
        
        return is_held_by_IXA, total_holding_count, holding_institutions
    
    except requests.RequestException as e:
        print(f"Error getting holdings for OCLC number {oclc_number}: {str(e)}")
        return False, 0, []
    
def truncate_contributors(performers, max_performers=3):
    return performers[:max_performers]

def remove_non_latin(text):
    cleaned = re.sub(r'[^\w\s\-\/\(\)áéíóúãõñâêîôûÁÉÍÓÚÃÕÑÂÊÎÔÛ]', '', text)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = re.sub(r'\(\s*\)', '', cleaned)
    return cleaned.strip()

def query_oclc_api(metadata, barcode, limit=10):
    global api_calls
    current_time = time.time()
    if current_time - api_calls['reset_time'] >= 86400:
        api_calls = {'count': 0, 'reset_time': current_time}

    if api_calls['count'] >= 50000:
        return "Rate limit exceeded. Please try again later.", {}

    client_id = os.environ.get("OCLC_CLIENT_ID")
    client_secret = os.environ.get("OCLC_SECRET")
    
    if not client_id or not client_secret:
        return "Error: OCLC_CLIENT_ID and OCLC_SECRET must be set in environment variables", {}

    try:
        access_token = get_access_token(client_id, client_secret)
    except Exception as e:
        return f"Error getting access token: {str(e)}", {}

    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs"

    queries = metadata.get("Queries", [])
    if not isinstance(queries, list):
        return "Error: Invalid query format", "Queries must be a list of strings"

    cleaned_queries = []
    for q in queries:
        if isinstance(q, str) and q.strip():
            cleaned = q.replace(str(barcode), "x").strip()
            if len(cleaned) >= 3:
                cleaned_queries.append(cleaned)

    if not cleaned_queries:
        return "No valid queries could be constructed", "Please check the metadata format"

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }

    query_log = ["Attempted Queries:"]
    attempted_queries = []
    
    # Track unique OCLC numbers to avoid duplicates
    seen_oclc_numbers = set()
    accumulated_results = []
    total_records_found = 0
    max_results_to_show = 10  # Our target - 10 CD results

    # Try ALL queries - don't stop until we've tried everything
    for idx, query in enumerate(cleaned_queries, 1):
        query_log.append(f"\nQuery {idx}: {query}")
        attempted_queries.append(idx)
        
        # Note if we already have enough results, but still try the query
        if len(seen_oclc_numbers) >= max_results_to_show:
            query_log.append(f"Already have {max_results_to_show} unique records, but trying this query anyway.")
        
        params = {
            "q": query,
            "limit": limit,
            "offset": 1,
            "itemType": "music",
            "inCatalogLanguage": "eng",
            "itemSubType": "music-cd"
        }

        try:
            response = requests.get(endpoint, params=params, headers=headers)
            api_calls['count'] += 1
            response.raise_for_status()
            data = response.json()
            
            total_records = data.get("numberOfRecords", 0)
            if total_records > 1000:
                query_log.append(f"Too many results ({total_records}), skipping")
                continue
                
            if total_records > 0:
                # Extract OCLC numbers from the current response to check for new records
                current_oclc_numbers = set()
                for record in data.get('bibRecords', []):
                    # Only consider CD format records
                    include_record = False
                    if 'format' in record and 'specificFormat' in record['format']:
                        specific_format = record['format']['specificFormat']
                        if isinstance(specific_format, str) and any(cd_term in specific_format for cd_term in ["CD", "compact disc", "Compact Disc"]):
                            include_record = True
                    
                    if not include_record:
                        continue
                        
                    # Check OCLC number
                    if 'identifier' in record and 'oclcNumber' in record['identifier']:
                        oclc_number = record['identifier']['oclcNumber']
                        if oclc_number and oclc_number not in seen_oclc_numbers:
                            current_oclc_numbers.add(oclc_number)
                
                if current_oclc_numbers:
                    # We found new unique CD format records
                    results, record_count = format_oclc_api_response_for_accumulation(data, access_token, seen_oclc_numbers)
                    
                    if results and "No matching records with CD format found" not in results and "No records found" not in results:
                        accumulated_results.append(results)
                        total_records_found += record_count
                        # Add the new OCLC numbers to our seen set
                        seen_oclc_numbers.update(current_oclc_numbers)
                        query_log.append(f"Added new CD format matches (now have {len(seen_oclc_numbers)} unique records)")
                    else:
                        query_log.append(f"No new CD format matches found")
                else:
                    query_log.append(f"No new unique CD format matches found")
            else:
                query_log.append(f"No matches found")
            
        except requests.RequestException as e:
            api_calls['count'] += 1
            query_log.append(f"Query failed: {str(e)}")

    # Combine all accumulated results with a single count at the top
    if accumulated_results:
        # Add the total count at the top, but limit to max_results_to_show
        displayed_record_count = min(total_records_found, max_results_to_show)
        total_header = f"Total CD Format Records Found: {total_records_found} (Displaying: {displayed_record_count})"
        
        # Limit the actual content to show only max_results_to_show records
        limited_results = []
        record_count = 0
        
        # Parse through accumulated results to extract and limit individual records
        for result_set in accumulated_results:
            if record_count >= max_results_to_show:
                break
                
            # Split by record divider
            record_sections = result_set.split("-" * 40)
            
            for section in record_sections:
                if record_count >= max_results_to_show:
                    break
                    
                if section.strip() and "OCLC Number:" in section:
                    limited_results.append("-" * 40)
                    limited_results.append(section.strip())
                    record_count += 1
        
        combined_results = total_header + "\n\n" + "\n".join(limited_results)
        return combined_results, "\n".join(query_log)
    else:
        return "No matching records with CD format found after trying all queries", "\n".join(query_log)
    
def process_metadata_file(input_file, results_folder_path):
    wb = load_workbook(input_file)
    ws = wb.active

    # Create a temporary workbook for periodic saving
    temp_wb = Workbook()
    temp_ws = temp_wb.active
    
    # Copy headers from main workbook
    for col_idx, cell in enumerate(ws[1], 1):
        temp_ws.cell(row=1, column=col_idx, value=cell.value)
        # Copy column widths
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(col_idx)
        if column_letter in ws.column_dimensions:
            temp_ws.column_dimensions[column_letter].width = ws.column_dimensions[column_letter].width

    if ws['F1'].value != 'OCLC Query':
        ws.insert_cols(6)
        ws['F1'] = 'OCLC Query'
        temp_ws['F1'] = 'OCLC Query'
    if ws['G1'].value != 'OCLC API Results':
        ws.insert_cols(7)
        ws['G1'] = 'OCLC API Results'
        temp_ws['G1'] = 'OCLC API Results'
        
    ws.column_dimensions['F'].width = 52
    ws.column_dimensions['G'].width = 52
    temp_ws.column_dimensions['F'].width = 52
    temp_ws.column_dimensions['G'].width = 52
    
    # Temporary file path
    temp_output_file = "temp_step2_progress.xlsx"
    temp_output_path = os.path.join(results_folder_path, temp_output_file)
    
    total_rows = ws.max_row
    processed_rows = 0
    
    for row in range(2, total_rows + 1):
        metadata_str = ws.cell(row=row, column=5).value  # Column E
        barcode = ws.cell(row=row, column=4).value       # Column D
        if not metadata_str or metadata_str.startswith('Error'):
            # Still copy this row to temp workbook
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                temp_cell = temp_ws.cell(row=row, column=col_idx, value=cell_value)
                if ws.cell(row=row, column=col_idx).alignment:
                    temp_cell.alignment = Alignment(vertical='top', wrap_text=True)
            processed_rows += 1
            continue

        try:
            metadata_fields = extract_metadata_fields(metadata_str)
            if not isinstance(metadata_fields, dict):
                raise ValueError("Invalid metadata format")
            
            queries = construct_queries_from_metadata(metadata_fields)
            results, query_log = query_oclc_api({"Queries": queries}, barcode)
            
            # Update main workbook
            ws.cell(row=row, column=6, value=query_log)
            ws.cell(row=row, column=7, value=results)
            ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            # Update temp workbook
            temp_ws.cell(row=row, column=6, value=query_log)
            temp_ws.cell(row=row, column=7, value=results)
            temp_ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            temp_ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            # Copy image cells and other data from main to temp
            for col_idx in range(1, 6):  # Columns A-E
                cell_value = ws.cell(row=row, column=col_idx).value
                temp_cell = temp_ws.cell(row=row, column=col_idx, value=cell_value)
                if ws.cell(row=row, column=col_idx).alignment:
                    temp_cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            processed_rows += 1
            print(f"Processed row {row}/{total_rows}")
            
            # Save temporary workbook every 10 rows
            if processed_rows % 10 == 0:
                try:
                    temp_wb.save(temp_output_path)
                    print(f"Progress saved ({processed_rows}/{total_rows} rows)")
                except Exception as save_error:
                    print(f"Warning: Could not save temporary progress: {save_error}")
                    
            time.sleep(0.1)

        except Exception as e:
            error_message = f"Error processing row {row}: {str(e)}"
            print(error_message)
            
            # Update both workbooks with error
            ws.cell(row=row, column=6, value="Error processing")
            ws.cell(row=row, column=7, value=error_message)
            ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            temp_ws.cell(row=row, column=6, value="Error processing")
            temp_ws.cell(row=row, column=7, value=error_message)
            temp_ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            temp_ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            # Copy other columns from main to temp
            for col_idx in range(1, 6):  # Columns A-E
                cell_value = ws.cell(row=row, column=col_idx).value
                temp_cell = temp_ws.cell(row=row, column=col_idx, value=cell_value)
                if ws.cell(row=row, column=col_idx).alignment:
                    temp_cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            processed_rows += 1

    # Clean up temporary file
    try:
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
            print(f"Temporary progress file removed: {temp_output_path}")
    except Exception as remove_error:
        print(f"Warning: Could not remove temporary progress file: {remove_error}")

    return wb

def main():
    # Instead of specifying the full folder name, just provide the prefix.
    base_dir_prefix = "ai-music-workflow/cd-processing/cd-output-folders/results-"
    
    # Find the latest results folder using the prefix.
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Run the first script first.")
        return
        
    print(f"Using results folder: {results_folder}")
    
    # Look for previous step files in the results folder.
    input_files = [f for f in os.listdir(results_folder) 
                   if f.startswith('ai-music-step-1-') and f.endswith('.xlsx')]
    
    if not input_files:
        print("No step 1 files found in the results folder!")
        return
        
    latest_file = max(input_files)
    input_file = os.path.join(results_folder, latest_file)
    
    print(f"Processing file: {input_file}")
    wb = process_metadata_file(input_file, results_folder)
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"ai-music-step-2-{current_date}.xlsx"
    full_output_path = os.path.join(results_folder, output_file)
    
    wb.save(full_output_path)
    print(f"Results saved to {full_output_path}")
    print("Summary: Process completed.")

if __name__ == "__main__":
    main()