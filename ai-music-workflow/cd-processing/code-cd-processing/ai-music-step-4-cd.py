import os
import glob
import re
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from difflib import SequenceMatcher

def find_latest_results_folder(prefix):
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None

    latest_folder = max(matching_folders)
    
    return latest_folder

def extract_tracks_from_metadata(metadata_str):
    """Extract track listings from metadata string."""
    tracks = []
    
    # First try to find tracks in the structured JSON-like format
    content_section = re.search(r'Contents:\s*-\s*tracks:\s*\[(.*?)\]', metadata_str, re.DOTALL)
    if content_section:
        tracks_content = content_section.group(1)
        
        # Try to extract track objects using regex for JSON-like structures
        track_objects = re.finditer(r'\{\s*"number":\s*\d+,\s*"title":\s*"([^"]+)"', tracks_content)
        for match in track_objects:
            title = match.group(1)
            if title and title.strip() and title.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                tracks.append(title.strip())
        
        # If the above didn't work, try a simpler approach for title extraction
        if not tracks:
            title_matches = re.finditer(r'"title":\s*([^,\n]+)', tracks_content)
            for match in title_matches:
                title_text = match.group(1).strip()
                # Remove quotes if present
                if title_text.startswith('"') and title_text.endswith('"'):
                    title_text = title_text[1:-1]
                # Remove trailing comma if present
                if title_text.endswith(','):
                    title_text = title_text[:-1]
                
                if title_text and title_text.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                    tracks.append(title_text)
    
    # If no tracks found or tracks list is suspiciously short, try alternative methods
    if len(tracks) < 3:
        # Try to find individual track objects with more flexible patterns
        track_patterns = [
            r'"number":\s*\d+,\s*"title":\s*"([^"]+)"',
            r'"number":\s*\d+,\s*"title":\s*([^,\n]+),',
            r'"title":\s*"([^"]+)"[^}]*?"duration":\s*(\d+:\d+)',
            r'"title":\s*"([^"]+)"'
        ]
        
        for pattern in track_patterns:
            if len(tracks) < 3:
                found_tracks = re.findall(pattern, metadata_str)
                for found in found_tracks:
                    title = found[0] if isinstance(found, tuple) else found
                    cleaned = title.strip().rstrip(',')
                    if cleaned and cleaned.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                        if cleaned not in tracks:
                            tracks.append(cleaned)
    
    # Last resort: look for tracks in an unstructured format
    if len(tracks) < 3:
        track_sections = re.findall(r'(?:Track\s+list(?:ing)?|Contents|Tracks):\s*(.*?)(?:(?:\n\s*\w+:)|$)', 
                                   metadata_str, re.DOTALL | re.IGNORECASE)
        
        for section in track_sections:
            potential_tracks = re.findall(r'(?:\d+[\.\)]\s*|"\s*)([^"\n\(]+)(?:"|\n|\(|$)', section)
            potential_tracks += re.findall(r'([^,;]+)\s*\(\d+:\d+\)', section)
            
            for track in potential_tracks:
                cleaned = track.strip()
                if cleaned and cleaned.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                    if cleaned not in tracks:
                        tracks.append(cleaned)
    
    # Filter out field names rather than track titles
    tracks = [t for t in tracks if t.lower() not in [
        "number", "title", "titletransliteration", "composer", "lyricist", 
        "duration", "isrc", "not applicable", "not visible"
    ]]
    
    # Filter out any track that looks like a note or is too long
    tracks = [t for t in tracks if not (
        "note" in t.lower() or 
        t.lower().startswith("contains") or 
        len(t.split()) > 8
    )]
    
    return tracks

def extract_tracks_from_oclc(oclc_results, oclc_number):
    """Extract track listings from OCLC results for a specific OCLC number."""
    tracks = []
    
    # First, find the section for this OCLC number
    oclc_section_pattern = rf"OCLC Number: {oclc_number}.*?(?:(?:----------------------------------------)|$)"
    oclc_section = re.search(oclc_section_pattern, oclc_results, re.DOTALL)
    
    if oclc_section:
        section_text = oclc_section.group(0)
        
        # Look for a Content section containing track listings
        content_patterns = [
            r'Content:\s*(.*?)(?:(?:\n\s*[A-Z][a-z]+:)|$)',
            r'Description:.*?Content:\s*(.*?)(?:(?:\n\s*[A-Z][a-z]+:)|$)'
        ]
        
        content_text = None
        for pattern in content_patterns:
            content_match = re.search(pattern, section_text, re.DOTALL)
            if content_match:
                content_text = content_match.group(1).strip()
                break
        
        if content_text:
            if " -- " in content_text:
                track_parts = content_text.split(" -- ")
                for part in track_parts:
                    track_name = part.strip()
                    if track_name.endswith('.'):
                        track_name = track_name[:-1].strip()
                    track_name = re.sub(r'\s*/\s*[^(]+', '', track_name)
                    track_name = re.sub(r'\s*\(\d+[:\.]\d+\)\.?$', '', track_name)
                    track_name = re.sub(r'\s*\([^)]*\)$', '', track_name)
                    
                    if track_name and track_name.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                        tracks.append(track_name.strip())
            else:
                for delimiter in ['\n', ';', ',']:
                    if delimiter in content_text and not tracks:
                        parts = content_text.split(delimiter)
                        for part in parts:
                            clean_part = part.strip()
                            if clean_part.endswith('.'):
                                clean_part = clean_part[:-1].strip()
                            clean_part = re.sub(r'\s*/\s*[^(]+', '', clean_part)
                            clean_part = re.sub(r'\s*\(\d+[:\.]\d+\)\.?$', '', clean_part)
                            
                            if clean_part and clean_part.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                                tracks.append(clean_part)
        
        if not tracks:
            track_pattern = r'([^-\(\)]+?)\s*\(\d+[:\.]\d+\)'
            track_matches = re.findall(track_pattern, section_text)
            for match in track_matches:
                clean_track = match.strip()
                if clean_track and clean_track.lower() not in ["not visible", "n/a", "unavailable", "none"]:
                    if clean_track not in tracks:
                        tracks.append(clean_track)
    
    return tracks

def normalize_track(track):
    """Normalize track titles for better comparison."""
    norm = track.lower()
    if norm.startswith("the "):
        norm = norm[4:] + ", the"
    norm = norm.replace(" is a ", " is ").replace(" is the ", " is ")
    norm = norm.replace("(stripped)", "").replace("(edit)", "").replace("stripped", "").replace("edit", "")
    norm = re.sub(r'\s*\(with [^)]+\)', '', norm)
    norm = re.sub(r'\s*\([^)]+\)', '', norm)
    norm = re.sub(r'[^\w\s]', '', norm)
    norm = re.sub(r'\s+', ' ', norm).strip()
    return norm

def calculate_track_similarity(metadata_tracks, oclc_tracks):
    """Calculate the similarity between two track listings."""
    if not metadata_tracks or not oclc_tracks:
        return 0.0
    
    processed_metadata_tracks = []
    processed_oclc_tracks = oclc_tracks.copy()
    
    multi_part_groups = {}
    for i, track in enumerate(metadata_tracks):
        part_match = re.match(r'^(?:Part|Movement)\s*(\d+|[IVX]+)$', track, re.IGNORECASE)
        if part_match:
            if i > 0 and not re.match(r'^(?:Part|Movement)', metadata_tracks[i-1], re.IGNORECASE):
                main_title = metadata_tracks[i-1]
                if main_title not in multi_part_groups:
                    multi_part_groups[main_title] = []
                multi_part_groups[main_title].append(track)
    
    for track in metadata_tracks:
        if track not in multi_part_groups:
            is_part = False
            for parts in multi_part_groups.values():
                if track in parts:
                    is_part = True
                    break
            if not is_part:
                processed_metadata_tracks.append(track)
    
    if multi_part_groups:
        for main_title, parts in multi_part_groups.items():
            processed_metadata_tracks.append(f"{main_title} (with {len(parts)} parts)")
    
    if not processed_metadata_tracks:
        processed_metadata_tracks = metadata_tracks
    
    norm_metadata_tracks = [normalize_track(t) for t in processed_metadata_tracks]
    norm_oclc_tracks = [normalize_track(t) for t in processed_oclc_tracks]
    
    print(f"\nNormalized metadata tracks: {norm_metadata_tracks}")
    print(f"Normalized OCLC tracks: {norm_oclc_tracks}")
    
    matches = 0
    matched_tracks = []
    
    for i, meta_track in enumerate(norm_metadata_tracks):
        best_match = 0
        best_match_index = -1
        is_substring_match = False
        is_part_match = False
        
        if "with" in meta_track and "parts" in meta_track:
            main_title = re.sub(r'\s+with \d+ parts', '', meta_track)
            for j, oclc_track in enumerate(norm_oclc_tracks):
                if (main_title in oclc_track) or (oclc_track in main_title):
                    similarity = 0.95
                    is_part_match = True
                else:
                    similarity = SequenceMatcher(None, main_title, oclc_track).ratio()
                
                if similarity > best_match:
                    best_match = similarity
                    best_match_index = j
        else:
            meta_words = set(meta_track.split())
            for j, oclc_track in enumerate(norm_oclc_tracks):
                oclc_words = set(oclc_track.split())
                common_words = meta_words.intersection(oclc_words)
                
                shorter_length = min(len(meta_words), len(oclc_words))
                if shorter_length > 0 and len(common_words) >= max(1, int(shorter_length * 0.6)):
                    word_similarity = len(common_words) / shorter_length
                    similarity = max(0.8, word_similarity)
                    is_substring_match = True
                elif (meta_track in oclc_track) or (oclc_track in meta_track):
                    similarity = max(0.85, SequenceMatcher(None, meta_track, oclc_track).ratio())
                    is_substring_match = True
                else:
                    similarity = SequenceMatcher(None, meta_track, oclc_track).ratio()
                
                if similarity > best_match:
                    best_match = similarity
                    best_match_index = j
        
        orig_track = processed_metadata_tracks[i]
        match_info = f"{i+1}. {orig_track} => "
        if best_match >= 0.8:
            match_symbol = "✓"
            if is_part_match:
                match_symbol += "(multi-part)"
            elif is_substring_match:
                match_symbol += "(substring)"
            match_info += f"{match_symbol} {processed_oclc_tracks[best_match_index]} ({best_match:.2f})"
            matches += best_match
        else:
            if best_match_index >= 0:
                match_info += f"✗ {processed_oclc_tracks[best_match_index]} ({best_match:.2f})"
            else:
                match_info += "✗ No match"
        
        matched_tracks.append(match_info)
    
    if len(norm_metadata_tracks) == 0:
        return 0.0
    
    print("\nTrack matching details:")
    for match in matched_tracks:
        print(f"  {match}")
    
    similarity = matches / len(norm_metadata_tracks)
    print(f"Total matches: {matches:.2f} out of {len(norm_metadata_tracks)} tracks")
    
    if multi_part_groups and similarity * 100 < 80:
        adjusted_similarity = min(80.0, similarity * 100 + 10.0)
        print(f"Applying multi-part track bonus: final similarity {adjusted_similarity:.2f}%")
        return adjusted_similarity
    
    return similarity * 100

def extract_and_normalize_year(text, is_oclc=False):
    """Extract and normalize publication year to YYYY format."""
    if not text:
        return None
    
    # First check if publication date is explicitly marked as not visible/available
    not_visible_patterns = [
        r'publicationDate:\s*(?:Not\s+visible|N/A|None|Unavailable)',
        r'Dates:[^p]*publicationDate:\s*(?:Not\s+visible|N/A|None|Unavailable)',
        r'Date:[^p]*publicationDate:\s*(?:Not\s+visible|N/A|None|Unavailable)'
    ]
    
    for pattern in not_visible_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return None  # Publication date is explicitly marked as not available
    
    # Look for publication date in structured format
    if is_oclc:
        date_patterns = [
            r'publicationDate:\s*[©℗]?(\d{4})',
            r'machineReadableDate:\s*(\d{4})',
            r'publicationDate:\s*[©℗]?(\d{4})[^\d]',
            r'Dates:[^p]*publicationDate:\s*[©℗]?(\d{4})',
            r'Date:[^p]*publicationDate:\s*[©℗]?(\d{4})',
            r'publicationDate:\s*[©℗]?c?(\d{4})',
            r'publication(?:Date)?:\s*[©℗]?c?(\d{4})'
        ]
    else:
        date_patterns = [
            r'publicationDate:\s*(\d{4})',
            r'Dates:[^p]*publicationDate:\s*(\d{4})',
            r'Date:[^p]*publicationDate:\s*(\d{4})',
            r'publication(?:Date)?:\s*(\d{4})'
        ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            year = match.group(1)
            if year.isdigit() and 1900 <= int(year) <= datetime.now().year:
                return year
    
    # Look for copyright or phonogram year markers
    year_markers = [r'[©℗](\d{4})', r'[©℗](?:\s*)(\d{4})', r'copyright\s+(\d{4})', r'phonogram\s+(\d{4})']
    for marker in year_markers:
        matches = re.findall(marker, text, re.IGNORECASE)
        if matches:
            for year in matches:
                if year.isdigit() and 1900 <= int(year) <= datetime.now().year:
                    return year
    
    # Look for standalone 4-digit years - BUT ONLY IF WE'RE LOOKING AT OCLC DATA
    # This prevents picking up random years from notes sections in metadata
    if is_oclc:
        year_pattern = r'(?<!\d)(\d{4})(?!\d)'
        matches = re.findall(year_pattern, text)
        valid_years = [y for y in matches if 1900 <= int(y) <= datetime.now().year]
        if valid_years:
            # Return the most frequently occurring year
            from collections import Counter
            return Counter(valid_years).most_common(1)[0][0]
    
    return None

def extract_year_from_oclc_by_number(oclc_results, oclc_number):
    """Extract publication year from OCLC results for a specific OCLC number."""
    
    # First, find the section for this OCLC number
    oclc_section_pattern = rf"OCLC Number: {oclc_number}.*?(?:(?:Record \d+:|----------------------------------------)|$)"
    oclc_section = re.search(oclc_section_pattern, oclc_results, re.DOTALL)
    
    if oclc_section:
        section_text = oclc_section.group(0)
        
        # Look for publication date in structured format with multiple years
        multi_year_patterns = [
            r'publicationDate:\s*\[(\d{4})\],\s*[©℗](\d{4})',
            r'publicationDate:\s*[©℗]?(\d{4})[^\d]*[©℗](\d{4})',
            r'Dates:[^p]*publicationDate:\s*\[(\d{4})\],\s*[©℗](\d{4})',
            r'Date:[^p]*publicationDate:\s*\[(\d{4})\],\s*[©℗](\d{4})'
        ]
        
        for pattern in multi_year_patterns:
            match = re.search(pattern, section_text, re.IGNORECASE)
            if match:
                year1 = match.group(1)
                year2 = match.group(2)
                # Return the later year when there are multiple years
                if year1.isdigit() and year2.isdigit():
                    return str(max(int(year1), int(year2)))
        
        # Handle "pc" or "p c" notation (e.g., pc2004)
        pc_patterns = [
            r'publicationDate:\s*p\s*c\s*(\d{4})',
            r'publicationDate:\s*pc\s*(\d{4})',
            r'Dates:[^p]*publicationDate:\s*p\s*c\s*(\d{4})',
            r'Date:[^p]*publicationDate:\s*pc\s*(\d{4})'
        ]
        
        for pattern in pc_patterns:
            match = re.search(pattern, section_text, re.IGNORECASE)
            if match:
                year = match.group(1)
                if year.isdigit() and 1900 <= int(year) <= datetime.now().year:
                    return year
        
        # Look for specific date patterns in this section (original patterns)
        date_patterns = [
            r'publicationDate:\s*[©℗]?(\d{4})',
            r'machineReadableDate:\s*(\d{4})',
            r'publicationDate:\s*[©℗]?(\d{4})[^\d]',
            r'Dates:[^p]*publicationDate:\s*[©℗]?(\d{4})',
            r'Date:[^p]*publicationDate:\s*[©℗]?(\d{4})',
            r'publicationDate:\s*[©℗]?c?(\d{4})',
            r'publication(?:Date)?:\s*[©℗]?c?(\d{4})'
        ]
        
        # Add specific pattern for bracketed years
        bracketed_patterns = [
            r'publicationDate:\s*\[(\d{4})\]',
            r'Dates:[^p]*publicationDate:\s*\[(\d{4})\]',
            r'Date:[^p]*publicationDate:\s*\[(\d{4})\]'
        ]
        
        date_patterns.extend(bracketed_patterns)
        
        for pattern in date_patterns:
            match = re.search(pattern, section_text, re.IGNORECASE)
            if match:
                year = match.group(1)
                if year.isdigit() and 1900 <= int(year) <= datetime.now().year:
                    return year
        
        # Look for copyright or phonogram year markers
        year_markers = [
            r'[©℗](\d{4})', 
            r'[©℗](?:\s*)(\d{4})', 
            r'copyright\s+(\d{4})', 
            r'phonogram\s+(\d{4})'
        ]
        
        years_found = []
        for marker in year_markers:
            matches = re.findall(marker, section_text, re.IGNORECASE)
            for year in matches:
                if year.isdigit() and 1900 <= int(year) <= datetime.now().year:
                    years_found.append(int(year))
        
        if years_found:
            # Return the most recent year when we find multiple copyright/phonogram years
            return str(max(years_found))
        
        # Look for standalone 4-digit years
        year_pattern = r'(?<!\d)(\d{4})(?!\d)'
        matches = re.findall(year_pattern, section_text)
        valid_years = [int(y) for y in matches if 1900 <= int(y) <= datetime.now().year]
        if valid_years:
            # Return the most frequently occurring year
            from collections import Counter
            return str(Counter(valid_years).most_common(1)[0][0])
    
    return None

def compare_publication_years(metadata_year, oclc_year):
    """
    Compare publication years and return a match status.
    """
    # If either year is missing, don't count it against the match
    if metadata_year is None or oclc_year is None:
        return (True, f"Incomplete year data: metadata_year={metadata_year}, oclc_year={oclc_year}")
    
    # Convert to integers for numerical comparison
    metadata_year_int = int(metadata_year)
    oclc_year_int = int(oclc_year)
    
    # Check if the absolute difference is less than or equal to 1
    year_diff = abs(metadata_year_int - oclc_year_int)
    
    if year_diff <= 1:
        if year_diff == 0:
            return (True, f"Years match exactly: {metadata_year} == {oclc_year}")
        else:
            return (True, f"Years are within 1 year: {metadata_year} vs {oclc_year}")
    else:
        return (False, f"Years differ by more than 1 year: {metadata_year} vs {oclc_year} (difference: {year_diff} years)")
def check_oclc_held_by_ixa(oclc_results, oclc_number):
    """
    Check if a specific OCLC number is held by IXA.
    Returns 'Y' if held by IXA, 'N' if not, or 'N/A' if there are no other potential matches.
    """
    if not oclc_results or not oclc_number:
        return 'N/A'
    
    # Find the section for this OCLC number
    oclc_section_pattern = rf"OCLC Number: {oclc_number}.*?(?:(?:----------------------------------------)|$)"
    oclc_section = re.search(oclc_section_pattern, oclc_results, re.DOTALL)
    
    if oclc_section:
        section_text = oclc_section.group(0)
        
        # Direct check for "Held by IXA: Yes" string
        if re.search(r'Held by IXA:\s*Yes', section_text, re.IGNORECASE):
            return 'Y'
        elif re.search(r'Held by IXA:\s*No', section_text, re.IGNORECASE):
            return 'N'
    
    return 'N/A'

def check_other_matches_held_by_ixa(other_matches_text):
    """
    Check if any of the other potential matches are held by IXA.
    Returns 'Y' if at least one is held by IXA, 'N' if none are, 
    or 'N/A' if can't determine or no other matches exist.
    """
    if not other_matches_text or other_matches_text.lower() in ['none', 'none.', 'no other potential good matches.']:
        return 'N'
    
    # Look for instances of "IXA: Yes" in the structured match information
    if re.search(r'IXA: Yes', other_matches_text, re.IGNORECASE):
        return 'Y'
    
    # If we have structured data but no "IXA: Yes", then none are held by IXA
    if re.search(r'IXA: No', other_matches_text, re.IGNORECASE):
        return 'N'
    
    return 'N/A'

def main():
    # Specify the folder prefix based on your output location
    base_dir_prefix = "ai-music-workflow/cd-processing/cd-output-folders/results-"
    
    # Find the latest results folder using the prefix
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Run the previous scripts first.")
        exit()
        
    print(f"Using results folder: {results_folder}")
    
    # Look for step 3 files in the results folder
    step3_files = [f for f in os.listdir(results_folder) 
                   if f.startswith('ai-music-step-3-') and f.endswith('.xlsx')]
    
    if not step3_files:
        print("No step 3 files found in the results folder!")
        exit()
        
    latest_file = max(step3_files)
    workbook_path = os.path.join(results_folder, latest_file)
    
    print(f"Processing file: {workbook_path}")
    
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb.active
    
    # Define the columns
    METADATA_COLUMN = 'E'
    OCLC_RESULTS_COLUMN = 'G'
    OCLC_NUMBER_COLUMN = 'H'
    CONFIDENCE_SCORE_COLUMN = 'I'
    EXPLANATION_COLUMN = 'J'
    OTHER_POTENTIAL_MATCHES_COLUMN = 'K'  # Column for other potential matches
    VERIFICATION_COLUMN = 'L'  # Column for track verification results
    YEAR_VERIFICATION_COLUMN = 'M'  # New column for year verification results
    IXA_HOLDING_COLUMN = 'N'   # New column for IXA holding status of chosen OCLC
    OTHER_IXA_HOLDING_COLUMN = 'O'  # New column for IXA holding status of other matches
    
    sheet[f'{VERIFICATION_COLUMN}1'] = 'Track Verification Results'
    sheet[f'{YEAR_VERIFICATION_COLUMN}1'] = 'Year Verification Results'
    sheet[f'{IXA_HOLDING_COLUMN}1'] = 'Match Held at IXA?'
    sheet[f'{OTHER_IXA_HOLDING_COLUMN}1'] = 'Potential Matches at IXA?'
    
    sheet.column_dimensions[VERIFICATION_COLUMN].width = 40
    sheet.column_dimensions[YEAR_VERIFICATION_COLUMN].width = 40
    sheet.column_dimensions[IXA_HOLDING_COLUMN].width = 20
    sheet.column_dimensions[OTHER_IXA_HOLDING_COLUMN].width = 25
    
    records_processed = 0
    records_adjusted_tracks = 0
    records_adjusted_years = 0
    records_skipped = 0
    records_skipped_none_matches = 0
    records_main_match_at_ixa = 0
    records_other_matches_at_ixa = 0
    
    print(f"Starting verification for records with confidence ≥ 85% that mention tracks...")
    print(f"Total rows in spreadsheet: {sheet.max_row - 1}")
    
    for row in range(2, sheet.max_row + 1):
        try:
            metadata = sheet[f'{METADATA_COLUMN}{row}'].value
            oclc_results = sheet[f'{OCLC_RESULTS_COLUMN}{row}'].value
            oclc_number = sheet[f'{OCLC_NUMBER_COLUMN}{row}'].value
            confidence_score = sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value
            explanation = sheet[f'{EXPLANATION_COLUMN}{row}'].value
            other_potential_matches = sheet[f'{OTHER_POTENTIAL_MATCHES_COLUMN}{row}'].value
            
            # Check and populate IXA holdings status regardless of other processing
            if oclc_number and str(oclc_number).strip() != "" and oclc_number != "Not found" and oclc_results:
                # Check if the chosen OCLC match is held by IXA
                ixa_holding_status = check_oclc_held_by_ixa(oclc_results, str(oclc_number).strip())
                sheet[f'{IXA_HOLDING_COLUMN}{row}'].value = ixa_holding_status
                
                # Update counter for IXA holdings
                if ixa_holding_status == 'Y':
                    records_main_match_at_ixa += 1
            else:
                sheet[f'{IXA_HOLDING_COLUMN}{row}'].value = 'N/A'
            
            # Check if any other potential matches are held by IXA
            if other_potential_matches:
                other_ixa_status = check_other_matches_held_by_ixa(str(other_potential_matches))
                sheet[f'{OTHER_IXA_HOLDING_COLUMN}{row}'].value = other_ixa_status
                
                # Update counter for other matches at IXA
                if other_ixa_status == 'Y':
                    records_other_matches_at_ixa += 1
            else:
                sheet[f'{OTHER_IXA_HOLDING_COLUMN}{row}'].value = 'N/A'
            
            if not oclc_number or str(oclc_number).strip() == "":
                # Clear the verification columns when no OCLC number is present
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = None
                sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = None
                records_skipped += 1
                print(f"Skipping row {row}: No OCLC number provided")
                continue
            
            if not oclc_number or str(oclc_number).strip() == "":
            # Clear the verification columns when no OCLC number is present
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = None
                sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = None
                records_skipped += 1
                print(f"Skipping row {row}: No OCLC number provided")
                continue
        
            # Skip processing if "other potential matches" column is set to "None" or "None."
            if other_potential_matches is not None:
                # Convert to string if it's not already a string
                other_potential_matches_str = str(other_potential_matches) if not isinstance(other_potential_matches, str) else other_potential_matches
                if other_potential_matches_str.strip().lower() in ["none", "none."]:
                    # Clear the verification columns for rows with "None" in other potential matches
                    sheet[f'{VERIFICATION_COLUMN}{row}'].value = "Not Applicable"
                    sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = "Not Applicable"
                    records_skipped_none_matches += 1
                    print(f"Skipping row {row}")
                    continue
            
            if not all([metadata, oclc_results, oclc_number, confidence_score, explanation]):
                continue
                
            try:
                conf_score = float(confidence_score)
                if conf_score < 85:
                    # Clear the verification columns for skipped rows
                    sheet[f'{VERIFICATION_COLUMN}{row}'].value = None
                    sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = None
                    records_skipped += 1
                    print(f"Skipping row {row}: Confidence score {conf_score}% is below threshold")
                    continue
            except (ValueError, TypeError):
                # Clear the verification columns for rows with invalid confidence scores
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = None
                sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = None
                records_skipped += 1
                continue
                
            track_related_terms = ["track", "content", "song", "listing"]
            if not explanation or not any(term in explanation.lower() for term in track_related_terms):
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = None
                sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = None
                records_skipped += 1
                print(f"Skipping row {row}: No track-related terms in explanation")
                continue
                            
            records_processed += 1
            print(f"\nProcessing row {row} with OCLC number {oclc_number} (confidence: {confidence_score}%)")
            
            metadata_tracks = extract_tracks_from_metadata(metadata)
            oclc_tracks = extract_tracks_from_oclc(oclc_results, oclc_number)
            
            print(f"Metadata tracks ({len(metadata_tracks)}): {metadata_tracks}")
            print(f"OCLC tracks ({len(oclc_tracks)}): {oclc_tracks}")
            
            # Extract and compare publication years
            metadata_year = extract_and_normalize_year(metadata, is_oclc=False)
            
            # Use the new function to extract year specifically for this OCLC number
            oclc_year = extract_year_from_oclc_by_number(oclc_results, oclc_number)
            
            print(f"Extracted years - Metadata: {metadata_year}, OCLC: {oclc_year}")
            
            year_match, _ = compare_publication_years(metadata_year, oclc_year)
            
            if not metadata_year and not oclc_year:
                match_status = "N/A - No years to compare"
            elif not metadata_year or not oclc_year:
                match_status = "Considered match - Incomplete data"
            else:
                year_diff = abs(int(metadata_year) - int(oclc_year))
                if year_diff <= 1:
                    if year_diff == 0:
                        match_status = "Yes - Exact match"
                    else:
                        match_status = "Yes - Within 1 year"
                else:
                    match_status = "No - More than 1 year difference"
                
            year_verification_result = f"Metadata year: {metadata_year if metadata_year else 'Not found'}\nOCLC year: {oclc_year if oclc_year else 'Not found'}\nMatch: {match_status}"
            
            # Skip track similarity check if no tracks found
            track_similarity = 0
            if len(metadata_tracks) == 0 or len(oclc_tracks) == 0:
                print(f"Skipping similarity check: {'No metadata tracks' if len(metadata_tracks) == 0 else 'No OCLC tracks'}")
                verification_result = f"Metadata tracks: {len(metadata_tracks)}\nOCLC tracks: {len(oclc_tracks)}\nSkipped: insufficient track data"
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = verification_result
                sheet[f'{VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            else:
                track_similarity = calculate_track_similarity(metadata_tracks, oclc_tracks)
                print(f"Track similarity: {track_similarity:.2f}%")
                
                matching_tracks = 0
                for i, meta_track in enumerate(metadata_tracks):
                    best_match = 0
                    for oclc_track in oclc_tracks:
                        norm_meta = normalize_track(meta_track)
                        norm_oclc = normalize_track(oclc_track)
                        
                        meta_words = set(norm_meta.split())
                        oclc_words = set(norm_oclc.split())
                        common_words = meta_words.intersection(oclc_words)
                        
                        shorter_length = min(len(meta_words), len(oclc_words))
                        if shorter_length > 0 and len(common_words) >= max(1, int(shorter_length * 0.6)):
                            word_similarity = len(common_words) / shorter_length
                            similarity = max(0.8, word_similarity)
                        elif (norm_meta in norm_oclc) or (norm_oclc in norm_meta):
                            similarity = max(0.85, SequenceMatcher(None, norm_meta, norm_oclc).ratio())
                        else:
                            similarity = SequenceMatcher(None, norm_meta, norm_oclc).ratio()
                        
                        if similarity > best_match:
                            best_match = similarity
                    
                    if best_match >= 0.8:
                        matching_tracks += 1
                
                verification_result = f"Metadata tracks: {len(metadata_tracks)}\nOCLC tracks: {len(oclc_tracks)}\nMatching tracks: {matching_tracks}/{len(metadata_tracks)}\nSimilarity: {track_similarity:.2f}%"
                sheet[f'{VERIFICATION_COLUMN}{row}'].value = verification_result
                sheet[f'{VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = year_verification_result
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            
            # Determine if confidence needs to be adjusted based on both track and year verification
            adjust_confidence = False
            adjustment_reasons = []
            
            # Check track similarity
            if len(metadata_tracks) > 0 and len(oclc_tracks) > 0 and track_similarity < 80:
                adjust_confidence = True
                adjustment_reasons.append(f"track listing mismatch (similarity {track_similarity:.2f}%, below 80% threshold)")
            
            # Check year match - only adjust if both years are present and differ by more than 1 year
            if metadata_year and oclc_year and not year_match:
                adjust_confidence = True
                year_diff = abs(int(metadata_year) - int(oclc_year))
                adjustment_reasons.append(f"publication year mismatch (metadata: {metadata_year}, OCLC: {oclc_year}, difference: {year_diff} years)")
            elif not metadata_year or not oclc_year:
                print(f"Not penalizing for missing year data: metadata_year={metadata_year}, oclc_year={oclc_year}")
            
            # Apply confidence adjustment if needed
            if adjust_confidence:
                old_confidence = confidence_score
                new_confidence = 80
                sheet[f'{CONFIDENCE_SCORE_COLUMN}{row}'].value = new_confidence
                
                note = f"\n\n[AUTOMATIC REVIEW: Confidence reduced due to: {'; '.join(adjustment_reasons)}. Please verify manually.]"
                
                # Add track comparison details if needed
                if len(metadata_tracks) > 0 and len(oclc_tracks) > 0 and track_similarity < 80:
                    note += "\n\nTrack comparison:"
                    for i, meta_track in enumerate(metadata_tracks):
                        best_match = 0
                        best_match_track = "No match"
                        
                        for oclc_track in oclc_tracks:
                            norm_meta = normalize_track(meta_track)
                            norm_oclc = normalize_track(oclc_track)
                            
                            meta_words = set(norm_meta.split())
                            oclc_words = set(norm_oclc.split())
                            common_words = meta_words.intersection(oclc_words)
                            
                            shorter_length = min(len(meta_words), len(oclc_words))
                            if shorter_length > 0 and len(common_words) >= max(1, int(shorter_length * 0.6)):
                                word_similarity = len(common_words) / shorter_length
                                similarity = max(0.8, word_similarity)
                            elif (norm_meta in norm_oclc) or (norm_oclc in norm_meta):
                                similarity = max(0.85, SequenceMatcher(None, norm_meta, norm_oclc).ratio())
                            else:
                                similarity = SequenceMatcher(None, norm_meta, norm_oclc).ratio()
                            
                            if similarity > best_match:
                                best_match = similarity
                                best_match_track = oclc_track
                        
                        match_status = "✓" if best_match >= 0.8 else "✗"
                        note += f"\n{i+1}. {meta_track} {match_status} {best_match_track} ({best_match:.2f})"
                
                # Add year comparison details - only for actual mismatches that are more than 1 year apart
                if metadata_year and oclc_year and not year_match:
                    year_diff = abs(int(metadata_year) - int(oclc_year))
                    note += f"\n\nYear comparison: Metadata year {metadata_year} differs from OCLC year {oclc_year} by {year_diff} years"
                
                sheet[f'{EXPLANATION_COLUMN}{row}'].value = explanation + note
                
                if len(metadata_tracks) > 0 and len(oclc_tracks) > 0 and track_similarity < 80:
                    records_adjusted_tracks += 1
                
                if metadata_year and oclc_year and not year_match:
                    records_adjusted_years += 1
                
                # Update verification result with action taken
                actions = []
                verification_result = sheet[f'{VERIFICATION_COLUMN}{row}'].value
                year_verification_result = sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value
                
                if track_similarity < 80 and len(metadata_tracks) > 0 and len(oclc_tracks) > 0:
                    actions.append("track mismatch")
                
                # Only count year mismatch when both years exist but differ by more than 1 year
                if metadata_year and oclc_year and not year_match:
                    actions.append("year mismatch of more than 1 year")
                
                if actions:
                    action_text = f"\nAction: Reduced confidence from {old_confidence}% to {new_confidence}% due to {' and '.join(actions)}"
                    if verification_result:
                        sheet[f'{VERIFICATION_COLUMN}{row}'].value = verification_result + action_text
                    else:
                        sheet[f'{VERIFICATION_COLUMN}{row}'].value = action_text
                    
                    sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = year_verification_result + action_text
            else:
                if sheet[f'{VERIFICATION_COLUMN}{row}'].value:
                    sheet[f'{VERIFICATION_COLUMN}{row}'].value += "\nAction: None (similarity is acceptable)"
                
                # For year verification, provide appropriate message based on year data
                year_action = "\nAction: "
                if not metadata_year and not oclc_year:
                    year_action += "None (no year data to compare)"
                elif not metadata_year or not oclc_year:
                    year_action += "None (incomplete year data, not penalized)"
                else:
                    year_diff = abs(int(metadata_year) - int(oclc_year))
                    if year_diff <= 1:
                        if year_diff == 0:
                            year_action += "None (years match exactly)"
                        else:
                            year_action += f"None (years are within 1 year: {metadata_year} vs {oclc_year})"
                    else:
                        year_action += f"Reduced confidence (years differ by {year_diff} years)"
                
                if sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value:
                    sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value += year_action
            
        except Exception as e:
            print(f"Error processing row {row}: {e}")
            sheet[f'{VERIFICATION_COLUMN}{row}'].value = f"Error: {str(e)}"
            sheet[f'{VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].value = f"Error: {str(e)}"
            sheet[f'{YEAR_VERIFICATION_COLUMN}{row}'].alignment = Alignment(wrap_text=True)
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"ai-music-step-4-{current_date}.xlsx"
    full_output_path = os.path.join(results_folder, output_file)
    
    wb.save(full_output_path)
    print(f"\nResults saved to {full_output_path}")
    # Updated summary statistics
    print(f"Summary:")
    print(f"  - Processed: {records_processed} records with confidence ≥ 85% and track listings mentioned")
    print(f"  - Adjusted for tracks: {records_adjusted_tracks} records due to low track similarity (< 80% match)")
    print(f"  - Adjusted for years: {records_adjusted_years} records due to publication year mismatch (only when both years present and differ by more than 1 year)")
    print(f"  - Skipped: {records_skipped} records (low confidence or no track listings)")
    print(f"IXA Holdings:")
    print(f"  - Records where LLM's chosen OCLC match is held by IXA: {records_main_match_at_ixa}")
    print(f"  - Records where at least one other potential match is held by IXA: {records_other_matches_at_ixa}")
    
if __name__ == "__main__":
    main()