import os
import glob
import datetime
import re
import openpyxl
import requests
import time
from difflib import SequenceMatcher
from openpyxl.styles import Alignment
from openpyxl import load_workbook

def find_latest_results_folder(prefix):
    base_dir = os.path.dirname(prefix)
    pattern = os.path.join(base_dir, "results-*")
    matching_folders = glob.glob(pattern)
    if not matching_folders:
        return None
    return max(matching_folders)

def find_latest_cd_metadata_file(results_folder):
    # Find files starting with "cd-metadata-ai-" and ending with ".xlsx"
    files = [f for f in os.listdir(results_folder) 
             if f.startswith("cd-metadata-ai-") and f.endswith(".xlsx")]
    if not files:
        return None
    latest_file = max(files)
    return os.path.join(results_folder, latest_file)

def get_access_token(client_id, client_secret):
    token_url = "https://oauth.oclc.org/token"
    data = {
        "grant_type": "client_credentials",
        "scope": "wcapi"
    }
    response = requests.post(token_url, data=data, auth=(client_id, client_secret))
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        raise Exception(f"Failed to get access token: {response.text}")

def get_bib_info(oclc_number, access_token):
    """
    Query the OCLC API for bibliographic information for a specific OCLC number.
    """
    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs/{oclc_number}"
    
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    
    try:
        response = requests.get(endpoint, headers=headers)
        response.raise_for_status()
        
        data = response.json()
        # The current API returns data directly, not in a bibRecords array
        return data
    except requests.RequestException as e:
        print(f"Error getting information for OCLC number {oclc_number}: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"Error response status: {e.response.status_code}")
            print(f"Error response content: {e.response.text}")
        return {"error": str(e)}

def get_holdings_info(oclc_number, access_token):
    """
    Query the OCLC API for holdings information for a specific OCLC number.
    """
    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    holdings_endpoint = f"{base_url}/bibs-holdings"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    params = {
        "oclcNumber": oclc_number,
        "limit": 50
    }
    
    try:
        holdings_response = requests.get(holdings_endpoint, params=params, headers=headers)
        holdings_response.raise_for_status()
        holdings_data = holdings_response.json()
        
        is_held_by_IXA = False
        total_holding_count = 0
        
        if "briefRecords" in holdings_data and len(holdings_data["briefRecords"]) > 0:
            record = holdings_data["briefRecords"][0]
            if "institutionHolding" in record:
                holdings = record["institutionHolding"]
                total_holding_count = holdings.get("totalHoldingCount", 0)
                if "briefHoldings" in holdings:
                    for holding in holdings["briefHoldings"]:
                        if holding.get("oclcSymbol", "") == "IXA":
                            is_held_by_IXA = True
                            break
        
        return {
            "held_by_ixa": is_held_by_IXA,
            "total_holdings": total_holding_count
        }
        
    except requests.RequestException as e:
        print(f"Error getting holdings for OCLC number {oclc_number}: {str(e)}")
        return {
            "held_by_ixa": False,
            "total_holdings": 0,
            "error": str(e)
        }

def clean_title(title):
    """Clean up title by removing strange punctuation but keeping slashes."""
    # Replace any double sword or other special characters, but keep slashes
    title = re.sub(r'[^\w\s\/\-\:\;\,\.\(\)\[\]\&\']', ' ', title)
    # Normalize whitespace
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def extract_title_from_bib_info(data):
    """Extract and clean title from bibliographic information."""
    if not isinstance(data, dict):
        return "No title available"
    
    # Check for error in response
    if "error" in data:
        return "No title available"
    
    # Current API response typically has title at the top level
    if "title" in data:
        # For responses with mainTitles structure
        if "mainTitles" in data["title"] and data["title"]["mainTitles"]:
            title = data["title"]["mainTitles"][0].get("text", "No title available")
            return clean_title(title)
        
        # For responses with title as direct property
        if isinstance(data["title"], str):
            return clean_title(data["title"])
    
    # Check for title in different locations that might be in the API response
    if "name" in data and isinstance(data["name"], str):
        return clean_title(data["name"])
    
    # Look for other possible title fields
    for field in ["titleInfo", "titleStatement", "uniformTitle"]:
        if field in data and isinstance(data[field], str):
            return clean_title(data[field])
    
    return "No title available"

def extract_author_from_bib_info(data):
    """Extract author from bibliographic information."""
    if not isinstance(data, dict) or "error" in data:
        return "No author available"
    
    # Look for contributor information
    if "contributor" in data:
        if "creators" in data["contributor"] and data["contributor"]["creators"]:
            for creator in data["contributor"]["creators"]:
                if "nonPersonName" in creator and "text" in creator["nonPersonName"]:
                    return creator["nonPersonName"]["text"]
                elif "firstName" in creator and "secondName" in creator:
                    first_name = creator.get("firstName", {}).get("text", "")
                    second_name = creator.get("secondName", {}).get("text", "")
                    return f"{first_name} {second_name}".strip()
    
    return "No author available"

def extract_publication_date_from_bib_info(data):
    """Extract publication date from bibliographic information."""
    if not isinstance(data, dict) or "error" in data:
        return "No date available"
    
    if "date" in data and "publicationDate" in data["date"]:
        return data["date"]["publicationDate"].replace("\u2117", "c")
    
    return "No date available"

def format_bib_info(data):
    """Format bibliographic information for display."""
    output = []
    
    if not isinstance(data, dict) or "identifier" not in data:
        return "No bibliographic information available."
    
    record = data
    
    title = "N/A"
    if "title" in record and "mainTitles" in record["title"]:
        title = record["title"]["mainTitles"][0].get("text", "N/A")
    
    series_title = "N/A"
    if "series" in record and isinstance(record["series"], list) and len(record["series"]) > 0:
        series_title = record["series"][0].get("title", "N/A")
    
    author = "N/A"
    contributors = []
    if "contributor" in record:
        if "creators" in record["contributor"] and record["contributor"]["creators"]:
            for creator in record["contributor"]["creators"]:
                if "nonPersonName" in creator and "text" in creator["nonPersonName"]:
                    author_name = creator["nonPersonName"]["text"]
                    author = author_name
                    contributors.append(author_name)
                elif "firstName" in creator and "secondName" in creator:
                    first_name = creator.get("firstName", {}).get("text", "")
                    second_name = creator.get("secondName", {}).get("text", "")
                    author_name = f"{first_name} {second_name}".strip()
                    author = author_name
                    contributors.append(author_name)
    
    publisher = "N/A"
    place = "N/A"
    date_pub = "N/A"
    if "publishers" in record and record["publishers"]:
        publisher_info = record["publishers"][0]
        if "publisherName" in publisher_info and "text" in publisher_info["publisherName"]:
            publisher = publisher_info["publisherName"]["text"]
        if "publicationPlace" in publisher_info:
            place = publisher_info["publicationPlace"]
    
    if "date" in record and "publicationDate" in record["date"]:
        date_pub = record["date"]["publicationDate"].replace("\u2117", "c")
    
    content_type = "N/A"
    if "format" in record:
        if "generalFormat" in record["format"]:
            content_type = record["format"]["generalFormat"]
        if "specificFormat" in record["format"]:
            content_type += f" - {record['format']['specificFormat']}"
    
    upc = "N/A"
    if "identifier" in record and "otherStandardIdentifiers" in record["identifier"]:
        for identifier in record["identifier"]["otherStandardIdentifiers"]:
            if identifier.get("type") == "Universal Product Code (UPC)":
                upc = identifier.get("id", "N/A")
                break
    
    # Enhanced content parsing logic
    contents = []
    
    # Method 1: Look for contents in the description field
    if "description" in record:
        # Check for the contentNote field in contents array - this is the format in the example
        if "contents" in record["description"]:
            for content_item in record["description"]["contents"]:
                # Check for contentNote object which contains track listings
                if "contentNote" in content_item and "text" in content_item["contentNote"]:
                    content_text = content_item["contentNote"]["text"]
                    # Common pattern: tracks separated by " -- "
                    if " -- " in content_text:
                        # Split by " -- " and clean up each track
                        tracks = []
                        for track in content_text.split(" -- "):
                            track = track.strip()
                            # Remove trailing period if it exists
                            if track.endswith('.'):
                                track = track[:-1].strip()
                            tracks.append(track)
                        contents.extend(tracks)
                        continue
                
                # Check for titles array format (original method)
                if "titles" in content_item:
                    for title_item in content_item["titles"]:
                        if isinstance(title_item, str):
                            contents.append(title_item)
                        elif isinstance(title_item, dict) and "text" in title_item:
                            contents.append(title_item["text"])
                
                # Check for different content formats
                if "items" in content_item:
                    for item in content_item["items"]:
                        if isinstance(item, str):
                            contents.append(item)
                        elif isinstance(item, dict) and "text" in item:
                            contents.append(item["text"])

        # Method 2: Check for TOC as a note
        if "notes" in record["description"]:
            for note in record["description"]["notes"]:
                # Look for various types of content notes
                is_content_note = False
                if "text" in note and any(marker in note["text"].lower() for marker in ["contents:", "tracks:", "track listing"]):
                    is_content_note = True
                
                if is_content_note and "text" in note:
                    toc_text = note["text"]
                    
                    # Try multiple approaches to parse the track list
                    # Approach 1: Split by common delimiters
                    for delimiter in ["--", ";", "/"]:
                        if delimiter in toc_text:
                            # Extract the content part (after any header like "Contents:" if present)
                            content_part = toc_text
                            for header in ["Contents:", "Tracks:", "Track listing:"]:
                                if header.lower() in toc_text.lower():
                                    content_part = toc_text.split(header, 1)[-1].strip()
                                    break
                                
                            # Split by delimiter and clean up
                            parts = [part.strip() for part in content_part.split(delimiter) if part.strip()]
                            if parts:
                                contents.extend(parts)
                                break

    # Method 3: Check for special MARC21 fields often used for music contents
    if "varFields" in record:
        for field in record.get("varFields", []):
            if field.get("marcTag") in ["505", "500"] and "subfields" in field:
                for subfield in field["subfields"]:
                    if subfield.get("code") == "a" and subfield.get("content"):
                        content = subfield["content"]
                        # Split content by common delimiters in track listings
                        for delimiter in ["--", ";", "/"]:
                            if delimiter in content:
                                parts = [part.strip() for part in content.split(delimiter) if part.strip()]
                                if parts:
                                    contents.extend(parts)
                                    break
    
    # Method 4: Check for $t prefixed content
    if "description" in record and "notes" in record["description"]:
        for note in record["description"]["notes"]:
            if "text" in note and "**$t**" in note["text"]:
                parts = note["text"].split("**$t**")
                # Skip the first part as it's usually empty or a header
                for part in parts[1:]:
                    # Clean up each part and add to contents
                    cleaned_part = part.strip()
                    if cleaned_part:
                        # Remove trailing -- if present
                        if cleaned_part.endswith("--"):
                            cleaned_part = cleaned_part[:-2].strip()
                        contents.append(cleaned_part)
    
    oclc_number = "N/A"
    if "identifier" in record and "oclcNumber" in record["identifier"]:
        oclc_number = record["identifier"]["oclcNumber"]
    
    output.append(f"Title: {title}")
    output.append(f"Series Title: {series_title}")
    output.append(f"Author: {author}")
    output.append(f"Contributors: {', '.join(contributors) if contributors else 'N/A'}")
    output.append(f"Publisher: {publisher}")
    output.append(f"Place of Publication: {place}")
    output.append(f"Date of Publication: {date_pub}")
    output.append(f"Content Type: {content_type}")
    output.append(f"UPC: {upc}")
    
    if contents:
        output.append("Contents:")
        for i, track in enumerate(contents, 1):
            # Clean up the track listing
            cleaned_track = track
            # Remove any trailing punctuation
            if cleaned_track.endswith(('.', ';')):
                cleaned_track = cleaned_track[:-1]
            output.append(f"  {i}. {cleaned_track}")
    
    output.append(f"OCLC Number: {oclc_number}")
    
    return "\n".join(output)

def calculate_title_similarity(title1, title2):
    """Calculate similarity between two titles using SequenceMatcher."""
    return SequenceMatcher(None, title1.lower(), title2.lower()).ratio()

def create_low_confidence_review_spreadsheet(results_folder, step4_file, all_records, access_token, current_date):
    """
    Create a review spreadsheet for unique low confidence matches with detailed information.
    """
    print("Creating low confidence review spreadsheet...")
    
    # Filter for unique low confidence matches (including those without OCLC numbers)
    low_confidence_records = [record for record in all_records 
                             if record["sort_group"] == "Cataloger Review (Low Confidence)"]
    
    if not low_confidence_records:
        print("No low confidence matches found to review.")
        return None
    
    # Open the original step 4 file to get additional data
    wb_src = load_workbook(step4_file)
    sheet_src = wb_src.active
    
    # Create mapping from barcode to source row data
    barcode_to_source = {}
    for row_idx in range(2, sheet_src.max_row + 1):  # Skip header row
        barcode = sheet_src.cell(row=row_idx, column=4).value  # Column D
        if barcode:
            # Get relevant columns: D (Barcode), E (AI-Generated Metadata), H (OCLC Number), other potential columns
            row_data = {
                "barcode": barcode,
                "metadata": sheet_src.cell(row=row_idx, column=5).value,  # Column E - AI-Generated Metadata
                "other_oclc_numbers": sheet_src.cell(row=row_idx, column=11).value   # Column K - Other Potential Matches
            }
            barcode_to_source[barcode] = row_data
    
    # Create new workbook for review
    wb_review = openpyxl.Workbook()
    sheet_review = wb_review.active
    
    # Set header row
    header_row = ["Barcode", "AI-Generated Metadata", "OCLC Number Chosen", "Other Potential Matches", "Confidence Score", "OCLC Record"]
    sheet_review.append(header_row)
    
    # Set column widths
    sheet_review.column_dimensions['A'].width = 20  # Barcode
    sheet_review.column_dimensions['B'].width = 50  # AI-Generated Metadata
    sheet_review.column_dimensions['C'].width = 18  # OCLC Number Chosen
    sheet_review.column_dimensions['D'].width = 60  # Other Potential Matches 
    sheet_review.column_dimensions['E'].width = 15  # Confidence Score
    sheet_review.column_dimensions['F'].width = 60  # OCLC Record
    
    # Set column A to text format
    from openpyxl.styles import NamedStyle
    text_style = NamedStyle(name="text_style")
    text_style.number_format = '@'  # Text format
    
    # Process each low confidence record
    processed_count = 0
    for record in low_confidence_records:
        barcode = record["barcode"]
        oclc_number = record["oclc_number"]
        
        print(f"Processing low confidence review record {processed_count + 1}/{len(low_confidence_records)} - Barcode: {barcode}")
        
        # Get source data
        source_data = barcode_to_source.get(barcode, {})
        metadata = source_data.get("metadata", "No AI-generated metadata available")
        other_oclc_numbers = source_data.get("other_oclc_numbers", "No other candidates")
        
        # Get detailed OCLC record information
        if oclc_number and record["has_valid_oclc"]:
            oclc_data = get_bib_info(oclc_number, access_token)
            formatted_info = format_bib_info(oclc_data)
            
            # Add holdings information
            holdings_info = get_holdings_info(oclc_number, access_token)
            holdings_text = f"\nTotal Institutions Holding: {holdings_info.get('total_holdings', 0)}\nHeld by IXA: {'Yes' if holdings_info.get('held_by_ixa', False) else 'No'}"
            formatted_info += holdings_text
            
            # Small delay to avoid API rate limits
            time.sleep(0.5)
        else:
            formatted_info = "No OCLC record available - no valid OCLC number found"
        
        # Add row to spreadsheet
        new_row = [
            barcode,
            metadata if metadata else "No AI-generated metadata available",
            oclc_number if oclc_number else "No OCLC number",
            other_oclc_numbers if other_oclc_numbers else "No other candidates",
            record.get("confidence_score", "No confidence score"),
            formatted_info
        ]
        sheet_review.append(new_row)
        
        # Format the newly added row
        current_row = sheet_review.max_row
        
        # Set column A (Barcode) to text format
        barcode_cell = sheet_review.cell(row=current_row, column=1)
        barcode_cell.number_format = '@'  # Text format
        
        # Set text wrapping for column B (metadata)
        metadata_cell = sheet_review.cell(row=current_row, column=2)
        metadata_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set text wrapping for column D (Other Potential Matches)
        other_matches_cell = sheet_review.cell(row=current_row, column=4)
        other_matches_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set text wrapping for column F (OCLC record) 
        oclc_record_cell = sheet_review.cell(row=current_row, column=6)
        oclc_record_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        processed_count += 1
        
        # Only add delay if we made API calls
        if oclc_number and record["has_valid_oclc"]:
            # Small delay to avoid API rate limits
            time.sleep(0.5)
    
    # Also format the header row for column A to text
    header_barcode_cell = sheet_review.cell(row=1, column=1)
    header_barcode_cell.number_format = '@'
    
    # Save the review spreadsheet
    review_file = f"low-confidence-review-{current_date}.xlsx"
    review_path = os.path.join(results_folder, review_file)
    wb_review.save(review_path)
    
    print(f"Low confidence review spreadsheet created with {len(low_confidence_records)} records: {review_path}")
    return review_path

def find_duplicate_groups(all_records, similarity_threshold=0.9, confidence_threshold=80):
    """
    Find groups of duplicate records based on similar OCLC numbers or titles.
    Only includes records with confidence >= confidence_threshold in duplicate detection.
    
    Args:
        all_records: List of all record dictionaries
        similarity_threshold: Threshold for determining similar titles
        confidence_threshold: Minimum confidence to consider for duplicate detection
    
    Returns:
        List of lists, where each inner list contains duplicate records
    """
    # Only consider records with confidence >= threshold for duplicate detection
    high_confidence_records = []
    high_confidence_indices = []
    
    for i, record in enumerate(all_records):
        try:
            confidence = float(record.get("confidence_score", 0))
            if confidence >= confidence_threshold:
                high_confidence_records.append(record)
                high_confidence_indices.append(i)
        except (ValueError, TypeError):
            # If confidence can't be parsed, treat as low confidence
            pass
    
    duplicate_groups = []
    processed_indices = set()
    
    for i, record in enumerate(high_confidence_records):
        original_index = high_confidence_indices[i]
        if original_index in processed_indices:
            continue
            
        oclc_number = str(record.get("oclc_number", "")).strip()
        title = record.get("title", "")
        
        # Find all duplicates for this record (only among high confidence records)
        duplicates = [record]
        duplicate_original_indices = [original_index]
        
        for j, other_record in enumerate(high_confidence_records[i+1:], i+1):
            other_original_index = high_confidence_indices[j]
            if other_original_index in processed_indices:
                continue
                
            other_oclc = str(other_record.get("oclc_number", "")).strip()
            other_title = other_record.get("title", "")
            
            is_duplicate = False
            
            # Check for similar OCLC numbers (might indicate duplicates)
            try:
                if abs(int(oclc_number) - int(other_oclc)) <= 5:
                    is_duplicate = True
            except (ValueError, TypeError):
                pass
            
            # Check for very similar titles
            if not is_duplicate and title and other_title:
                similarity = calculate_title_similarity(title, other_title)
                if similarity >= similarity_threshold:
                    is_duplicate = True
            
            if is_duplicate:
                duplicates.append(other_record)
                duplicate_original_indices.append(other_original_index)
        
        # Mark all indices as processed
        for idx in duplicate_original_indices:
            processed_indices.add(idx)
        
        # Only add to duplicate groups if there are actually duplicates
        if len(duplicates) > 1:
            duplicate_groups.append(duplicates)
    
    return duplicate_groups

def determine_sort_group_for_duplicates(duplicate_group, confidence_threshold=80):
    """
    Determine sort groups for a group of duplicate records.
    Priority: IXA held > highest confidence > first encountered
    
    Args:
        duplicate_group: List of duplicate record dictionaries
        confidence_threshold: Threshold for high confidence matches
    
    Returns:
        None (modifies records in place)
    """
    if len(duplicate_group) <= 1:
        return
    
    # Sort by priority: IXA held first, then by confidence score (highest first), then by original order
    def sort_key(record):
        held_by_ixa = record.get("held_by_ixa", False)
        try:
            confidence = float(record.get("confidence_score", 0))
        except (ValueError, TypeError):
            confidence = 0
        
        # Return tuple for sorting: (not held_by_ixa, -confidence)
        # This puts IXA held items first, then highest confidence
        return (not held_by_ixa, -confidence)
    
    sorted_duplicates = sorted(duplicate_group, key=sort_key)
    
    # The first item (highest priority) gets its normal classification
    primary_record = sorted_duplicates[0]
    
    # Determine classification for primary record
    if primary_record.get("held_by_ixa", False):
        # For IXA held items, check confidence to determine if it's truly the same item
        try:
            conf_score = float(primary_record.get("confidence_score", 0))
            if conf_score >= confidence_threshold:
                primary_record["sort_group"] = "Held by UT Libraries (IXA)"
            else:
                # Low confidence IXA match - treat as uncertain, not as held item
                primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
        except (ValueError, TypeError):
            primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
    else:
        try:
            conf_score = float(primary_record.get("confidence_score", 0))
            if conf_score >= confidence_threshold:
                primary_record["sort_group"] = "Alma Batch Upload (High Confidence)"
            else:
                primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
        except (ValueError, TypeError):
            primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
    
    # All other records in the group are marked as duplicates
    for record in sorted_duplicates[1:]:
        record["sort_group"] = "Duplicate"

def determine_sort_group(record, confidence_threshold=80):
    """
    Determine the sort group for a non-duplicate record.
    
    Args:
        record: Dictionary containing record information
        confidence_threshold: Threshold for high confidence matches
    
    Returns:
        String indicating the sort group
    """
    held_by_ixa = record.get("held_by_ixa", False)
    confidence_score = record.get("confidence_score", 0)
    
    # Check confidence level first - low confidence items are always for review
    try:
        conf_score = float(confidence_score)
        if conf_score < confidence_threshold:
            return "Cataloger Review (Low Confidence)"
    except (ValueError, TypeError):
        return "Cataloger Review (Low Confidence)"
    
    # Only for high confidence items, check if held by IXA
    if held_by_ixa:
        return "Held by UT Libraries (IXA)"
    else:
        return "Alma Batch Upload (High Confidence)"

def create_all_records_spreadsheet():
    # Set the folder prefix 
    base_dir_prefix = "ai-music-workflow/cd-processing/cd-output-folders/results-"
    results_folder = find_latest_results_folder(base_dir_prefix)
    if not results_folder:
        print("No results folder found! Please run the previous scripts first.")
        return None
    
    step4_file = find_latest_cd_metadata_file(results_folder)
    if not step4_file:
        print("No cd-metadata-ai file found in the results folder!")

    print(f"Using source file: {step4_file}")

    # Get OCLC API credentials
    client_id = os.environ.get("OCLC_CLIENT_ID")
    client_secret = os.environ.get("OCLC_SECRET")

    if not client_id or not client_secret:
        print("Error: OCLC_CLIENT_ID and OCLC_SECRET must be set in environment variables.")
        return None
    
    try:
        # Get the access token
        access_token = get_access_token(client_id, client_secret)
        print("Successfully obtained access token.")
        
        # Open the latest step 4 workbook
        wb_src = load_workbook(step4_file)
        sheet_src = wb_src.active

        # Create a new workbook for all records
        wb_new = openpyxl.Workbook()
        sheet_new = wb_new.active

        # Set header row
        header_row = ["Barcode", "Sort Group", "OCLC Number", "OCLC Title", "OCLC Author", "OCLC Date of Publication", "Confidence Score"]
        sheet_new.append(header_row)

        # Set column widths
        sheet_new.column_dimensions['A'].width = 16  # Barcode
        sheet_new.column_dimensions['B'].width = 28  # Sort Group
        sheet_new.column_dimensions['C'].width = 12  # OCLC Number
        sheet_new.column_dimensions['D'].width = 50  # OCLC Title
        sheet_new.column_dimensions['E'].width = 50  # OCLC Author
        sheet_new.column_dimensions['F'].width = 20  # OCLC Date
        sheet_new.column_dimensions['G'].width = 15  # Confidence Score

        # Get the column indices from the source workbook
        # In source: Column D=Barcode, Column H=OCLC Number, Column I=Confidence Score
        BARCODE_COL_IDX = 4  # Column D
        OCLC_NUM_COL_IDX = 8  # Column H
        CONF_SCORE_COL_IDX = 9  # Column I

        # First pass: collect all records with basic info
        all_records = []
        all_records_dict = {}  # For duplicate detection
        
        print("First pass: Collecting all records...")
        for row_idx in range(2, sheet_src.max_row + 1):  # Skip header row
            barcode = sheet_src.cell(row=row_idx, column=BARCODE_COL_IDX).value
            oclc_number = sheet_src.cell(row=row_idx, column=OCLC_NUM_COL_IDX).value
            confidence_score = sheet_src.cell(row=row_idx, column=CONF_SCORE_COL_IDX).value
            
            # Skip rows with missing barcode (essential identifier)
            if not barcode:
                continue
            
            # Handle records with no OCLC number or invalid OCLC numbers
            oclc_number_str = ""
            has_valid_oclc = False
            
            if oclc_number and str(oclc_number).strip() not in ["", "Not found", "Error processing"]:
                oclc_number_str = str(oclc_number).strip()
                has_valid_oclc = True
            
            record = {
                "barcode": barcode,
                "oclc_number": oclc_number_str,
                "confidence_score": confidence_score if confidence_score else 0,
                "title": "",
                "author": "",
                "publication_date": "",
                "held_by_ixa": False,
                "sort_group": "",
                "has_valid_oclc": has_valid_oclc
            }
            
            all_records.append(record)
            
            # Only add to dict for duplicate detection if it has a valid OCLC number
            if has_valid_oclc:
                all_records_dict[oclc_number_str] = record
            
        print(f"Found {len(all_records)} records to process.")

        # Second pass: Get OCLC data and holdings information
        print("Second pass: Getting OCLC data and holdings information...")
        processed_count = 0
        
        for record in all_records:
            oclc_number = record["oclc_number"]
            has_valid_oclc = record["has_valid_oclc"]
            
            print(f"Processing record {processed_count + 1}/{len(all_records)} - Barcode: {record['barcode']}")
            
            if has_valid_oclc:
                print(f"  OCLC Number: {oclc_number}")
                
                # Get bibliographic information
                oclc_data = get_bib_info(oclc_number, access_token)
                record["title"] = extract_title_from_bib_info(oclc_data)
                record["author"] = extract_author_from_bib_info(oclc_data)
                record["publication_date"] = extract_publication_date_from_bib_info(oclc_data)
                
                # Get holdings information
                holdings_info = get_holdings_info(oclc_number, access_token)
                record["held_by_ixa"] = holdings_info["held_by_ixa"]
                
                # Small delay to avoid API rate limits
                time.sleep(0.5)
            else:
                print(f"  No valid OCLC number - skipping API calls")
                record["title"] = "No OCLC match found"
                record["author"] = "No author available"
                record["publication_date"] = "No date available"
                record["held_by_ixa"] = False
            
            processed_count += 1
        
        # Third pass: Handle duplicates and determine sort groups
        print("Third pass: Finding duplicate groups...")
        
        # Only look for duplicates among records with valid OCLC numbers AND high confidence
        records_with_oclc = [record for record in all_records if record["has_valid_oclc"]]
        duplicate_groups = find_duplicate_groups(records_with_oclc, confidence_threshold=80)
        
        print(f"Found {len(duplicate_groups)} duplicate groups (high confidence only)")
        for i, group in enumerate(duplicate_groups):
            print(f"  Group {i+1}: {len(group)} duplicates")
        
        # Process duplicate groups
        print("Processing duplicate groups...")
        for group in duplicate_groups:
            determine_sort_group_for_duplicates(group, confidence_threshold=80)
        
        # Process remaining records (including low confidence items that were excluded from duplicate detection)
        print("Processing non-duplicate records...")
        processed_records = set()
        for group in duplicate_groups:
            for record in group:
                processed_records.add(id(record))
        
        for record in all_records:
            if id(record) not in processed_records:
                if record["has_valid_oclc"]:
                    record["sort_group"] = determine_sort_group(record, confidence_threshold=80)
                else:
                    # Records without valid OCLC numbers are always low confidence
                    record["sort_group"] = "Cataloger Review (Low Confidence)"
        
        # Add all records to spreadsheet
        for record in all_records:
            sheet_new.append([
                record["barcode"],
                record["sort_group"],
                record["oclc_number"],
                record["title"],
                record["author"],
                record["publication_date"],
                record["confidence_score"]
            ])
        
        # Save the all records spreadsheet
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        all_records_file = f"sort-groups-all-records-{current_date}.xlsx"
        all_records_path = os.path.join(results_folder, all_records_file)
        wb_new.save(all_records_path)
        
        print(f"All records spreadsheet created with {len(all_records)} records: {all_records_path}")
        
        # Create summary statistics
        sort_group_counts = {}
        for record in all_records:
            group = record["sort_group"]
            sort_group_counts[group] = sort_group_counts.get(group, 0) + 1
        
        print("\nSort Group Summary:")
        for group, count in sorted(sort_group_counts.items()):
            print(f"  {group}: {count} records")
        
        # Create high confidence matches for Alma batch upload (text file only)
        unique_matches = [record for record in all_records 
                         if record["sort_group"] == "Alma Batch Upload (High Confidence)"]
        
        # Create pipe-delimited text file for Alma processing
        text_file = f"batch-upload-alma-cd-{current_date}.txt"
        text_path = os.path.join(results_folder, text_file)
        
        with open(text_path, 'w', newline='', encoding='utf-8') as f:
            for record in unique_matches:
                line = f"{record['oclc_number']}|{record['barcode']}|{record['title']}\n"
                f.write(line)
        
        print(f"Pipe-delimited text file for Alma processing created: {text_path}")
        
        # Create low confidence review spreadsheet
        review_path = create_low_confidence_review_spreadsheet(
            results_folder, step4_file, all_records, access_token, current_date
        )
        
        return {
            "all_records_path": all_records_path,
            "text_file_path": text_path,
            "review_path": review_path,
            "total_records": len(all_records),
            "sort_group_counts": sort_group_counts,
            "unique_matches_count": len(unique_matches)
        }
        
    except Exception as e:
        print(f"Error creating all records spreadsheet: {str(e)}")
        return None

def main():
    result = create_all_records_spreadsheet()
    if result:
        print(f"\n=== All Records Processing Complete ===")
        print(f"Total records processed: {result['total_records']}")
        print(f"All records file: {result['all_records_path']}")
        print(f"Alma processing file: {result['text_file_path']}")
        if result['review_path']:
            print(f"Low confidence review file: {result['review_path']}")
        print(f"Unique high-confidence matches: {result['unique_matches_count']}")
        print("\nBreakdown by Sort Group:")
        for group, count in sorted(result['sort_group_counts'].items()):
            print(f"  {group}: {count}")
    else:
        print("Failed to create all records spreadsheet.")

if __name__ == "__main__":
    main()