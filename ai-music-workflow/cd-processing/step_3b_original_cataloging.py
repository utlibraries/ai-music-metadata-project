"""
Step 3b: AI Original Cataloging for No-OCLC-Match Records

Reads Step 1 extracted metadata for records with no OCLC match,
generates complete MARC records using GPT-4.1, and produces:
  1. A section in a new HTML review page (same interface as Step 6, distinct orange styling)
  2. A receipt/report saved to deliverables/ and AI_Music_Operations/

Does NOT touch Alma or OCLC. No duplicates are possible here.
Approved records are processed by step_3c and step_3d.

Reuses: find_latest_results_folder, get_workflow_json_path, get_file_path_config,
        load_workflow_json, save_workflow_json, log_error, log_processing_metrics,
        BatchProcessor, get_model_config, get_token_limit_param, get_temperature_param,
        get_current_timestamp
"""

import os
import json
import math
import shutil
import datetime
from openai import OpenAI

# Reuse existing modules exactly - no duplication
from shared_utilities import find_latest_results_folder, get_workflow_json_path
from cd_workflow_config import (
    get_file_path_config, get_model_config, get_token_limit_param,
    get_temperature_param, get_current_timestamp
)
from json_workflow import load_workflow_json, save_workflow_json, log_error, log_processing_metrics
from batch_processor import BatchProcessor
from shared_utilities import create_batch_summary

STEP_NAME = "step3b"
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
bp = BatchProcessor(default_step="step3")  # Reuse step3 config for batch decisions


def get_no_match_barcodes(results_folder):
    """
    Read the sorting spreadsheet to find all barcodes with no OCLC match
    (sort group = Cataloger Review (Low Confidence) AND no OCLC number).
    Returns list of barcodes.
    """
    from openpyxl import load_workbook
    import glob

    deliverables = os.path.join(results_folder, "deliverables")
    files = [f for f in os.listdir(deliverables) if f.startswith("cd-workflow-sorting-") and f.endswith(".xlsx")]
    if not files:
        print("No sorting spreadsheet found in deliverables.")
        return []

    sorting_file = os.path.join(deliverables, max(files))
    wb = load_workbook(sorting_file, read_only=True)
    ws = wb.active

    no_match = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        barcode = row[0]
        sort_group = row[1]
        oclc = row[2]
        if (sort_group == "Cataloger Review (Low Confidence)" and
                (not oclc or str(oclc).strip() in ["", "None", "Not found", "Error processing"])):
            no_match.append(str(barcode))

    wb.close()
    print(f"Found {len(no_match)} records with no OCLC match.")
    return no_match


def title_case_catalog(text):
    """
    Apply cataloging standard title case:
    Capitalize first word only, preserve proper nouns as detected.
    Input may be ALL CAPS or all lowercase from disc scan.
    """
    if not text:
        return text
    # Convert to lowercase first to normalize ALL CAPS input
    words = text.strip().lower().split()
    if not words:
        return text
    # Capitalize first word always
    words[0] = words[0].capitalize()
    # Known proper noun indicators (add as needed)
    # For music cataloging we capitalize first word + after : and after --
    result = []
    capitalize_next = False
    for i, word in enumerate(words):
        if i == 0 or capitalize_next:
            result.append(word.capitalize())
            capitalize_next = False
        elif word in ('--',):
            result.append(word)
            capitalize_next = True
        else:
            result.append(word)
    return ' '.join(result)


def build_marc_prompt(barcode, extracted_fields):
    """
    Build the MARC generation prompt.
    Generates a full enriched MARC record with only Whit Williams approved changes:
    1. 500: 'AI-generated minimum viable record.' (removed 'Cataloger review completed.')
    2. 650 Popular music: no $z United States subdivision
    3. No 500 Property of KUT Radio note
    4. Sentence case on 245 and 505 (first word only)
    All other fields remain as original enriched record.
    """
    today_str = datetime.datetime.now().strftime('%Y-%m-%d')
    return f"""You are an expert music cataloger following RDA and MARC 21 standards for sound recordings.

Generate a complete, enriched MARC record for this CD based on the extracted metadata below.
This CD had no match in OCLC WorldCat and will be submitted as an original catalog record.

EXTRACTED METADATA:
{json.dumps(extracted_fields, indent=2, ensure_ascii=False)}

Return a JSON object with these exact keys. Use null for fields that cannot be determined.

{{
  "leader": "00000njm a2200000 i 4500",
  "field_007": "sd fsngnnmmned",
  "field_024": null,
  "field_028": null,
  "field_100": null,
  "field_110": null,
  "field_245": null,
  "field_246": null,
  "field_264": null,
  "field_300": "1 audio disc : $b digital ; $c 4 3/4 in.",
  "field_336": "performed music $b prm $2 rdacontent",
  "field_337": "audio $b s $2 rdamedia",
  "field_338": "audio disc $b sd $2 rdacarrier",
  "field_500_general": null,
  "field_500_ai": "AI-generated minimum viable record.",
  "field_505": null,
  "field_518": null,
  "field_588": "Description based on AI-assisted metadata extraction from cover image, {today_str}.",
  "field_650": [],
  "field_700": [],
  "confidence_score": 0,
  "completeness_notes": null
}}

RULES — follow exactly:

TITLE CASE — CRITICAL:
All text in field_245 and field_505 MUST use sentence case:
- Capitalize the FIRST word of the title only
- Capitalize the first word after each " -- " separator in track listings
- Do NOT capitalize every word regardless of what appears on the CD
- The CD may show ALL CAPS or all lowercase — normalize both to sentence case
- CORRECT 245: "Written all my letters / $c Karen Abrahams."
- WRONG 245:   "Written All My Letters / $c Karen Abrahams."
- CORRECT 505: "Can't find my way home -- Pain and sorrow lamini -- Little wing."
- WRONG 505:   "CAN'T FIND MY WAY HOME -- PAIN AND SORROW LAMINI -- LITTLE WING."

field_024 — UPC/EAN barcode:
- Look for a 12-digit UPC or 13-digit EAN barcode printed on the disc, back cover, or spine
- Format: "$a 012345678901" — digits only, no spaces, no dashes
- Both 12-digit (UPC-A) and 13-digit (EAN-13) barcodes are valid — include whichever is visible
- CRITICAL: NEVER use barcodes starting with 05917 — those are UT Libraries institutional barcodes
- NEVER guess or construct a barcode — only include if clearly visible on the CD packaging
- If the only barcode visible starts with 05917, set field_024 to null
- null if no UPC/EAN barcode is visible

field_028: always null — do not use this field

field_100 (personal name main entry):
- Use for a single person as primary creator/performer
- Format: "Lastname, Firstname, $e role."
- Example: "Abrahams, Karen, $e performer."
- Use ONE of 100 or 110, never both

field_110 (corporate/group main entry):
- Use for bands, ensembles, orchestras, groups
- Format: "Bandname (Musical group), $e performer." or "Bandname, $e performer."
- Example: "Antibalas (Musical group), $e performer."
- Use ONE of 100 or 110, never both

field_245:
- Format: "Title of album / $c Statement of responsibility."
- ind1=1 if field_100 or field_110 is present, ind1=0 if both are null
- ind2=0 unless title begins with an article: The=3, A=2, An=3
- Examples: "245 10" when 1xx present, "245 00" when no 1xx
- Apply sentence case as described above
- End with period
- Do NOT include the MARC tag in the value itself

field_246:
- Include ONLY if there is a clear variant title
- null if no variant title

field_264:
- Format: "$a Place : $b Publisher, $c ℗year." or "$a Place : $b Publisher, $c ©year."
- Use ℗ for sound recording copyright
- Omit $a if place not visible, omit $b if publisher not visible
- null if year not visible
- Do NOT include the MARC tag in the value itself

field_028:
- Publisher/label catalog number if visible on disc or sleeve
- Format: "LABELNUMBER $b PUBLISHER"
- Include if clearly visible — helps with identification
- null if not visible or unclear

field_500_general:
- Any important general note visible on the CD
- Include: recording info, dedications, liner notes, recommended tracks, website URLs
- Do NOT include: Property of KUT Radio, or any local library notes
- null if no significant note visible

field_500_ai:
- Always exactly: "AI-generated minimum viable record."
- Do not change this text

field_505:
- Format: "Track one -- Track two -- Track three."
- Apply sentence case: capitalize first word of each track title only
- Include ALL tracks visible on disc or back cover
- null if no track listing visible
- Do NOT include the MARC tag in the value itself

field_518:
- Recording date and/or place if explicitly stated on the CD
- Format: "Recorded at STUDIO, CITY, DATE."
- null if not visible

field_588:
- Always exactly: "Description based on AI-assisted metadata extraction from cover image, {today_str}."
- Do not change this text

field_650:
- ALWAYS include: "Popular music." — NO geographic subdivision, no $z United States
- ALWAYS include: "Sound recordings."
- Add specific genre if determinable: "Afrobeat (Music).", "Blues (Music).", "Jazz.", "Rock music.", "Country music.", "Folk music.", "Classical music.", etc.
- Format each as plain string without MARC tag prefix
- Example: ["Popular music.", "Sound recordings.", "Afrobeat (Music)."]

field_700:
- Added entries for additional contributors with roles
- Include producers, engineers, photographers, designers if visible in credits
- Format: "Lastname, Firstname, $e role, $e role." 
- Example: "Fitzgerald, Michael, $e producer, $e engineer."
- Empty array if no additional contributors visible

CONFIDENCE SCORE (0-100):
- 100: title + contributor + year + tracks all clearly visible and readable
- 85-99: title + contributor clearly visible, year or tracks partially visible
- 70-84: title clearly visible, contributor or year uncertain
- 50-69: title visible but other key fields very uncertain
- 0-49: title unclear or not visible
Set confidence_score to this integer.

completeness_notes:
- Brief note on any fields that were uncertain or could not be determined
- null if all fields were clearly determinable

Return ONLY the JSON object. No markdown, no preamble, no explanation."""


def generate_marc_records_batch(barcodes, workflow_json_path, workflow_data, model_name):
    """
    Submit all records to OpenAI batch API for MARC generation.
    Returns dict of barcode -> marc_fields.
    """
    print(f"Preparing {len(barcodes)} MARC generation requests for batch...")

    batch_requests = []
    id_map = {}

    for i, barcode in enumerate(barcodes):
        record = workflow_data.get("records", {}).get(barcode, {})
        extracted = record.get("step1_metadata_extraction", {}).get("extracted_fields", {})

        if not extracted:
            print(f"  Skipping {barcode} — no Step 1 data")
            continue

        prompt = build_marc_prompt(barcode, extracted)
        custom_id = f"marc_{i}"

        request_body = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": "You are an expert music cataloger. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            "response_format": {"type": "json_object"},
            **get_token_limit_param(model_name, 2000),
            **get_temperature_param(model_name, 0.1)
        }

        batch_requests.append({
            "custom_id": custom_id,
            "method": "POST",
            "url": "/v1/chat/completions",
            "body": request_body
        })

        id_map[custom_id] = barcode

    if not batch_requests:
        return {}

    print(f"Submitting batch of {len(batch_requests)} MARC requests...")
    cost_estimate = bp.estimate_batch_cost(batch_requests, model_name)
    print(f"  Estimated cost: ${cost_estimate['batch_cost']:.4f} (regular: ${cost_estimate['regular_cost']:.4f})")

    batch_id = bp.submit_batch(
        batch_requests,
        f"Original Cataloging MARC — {len(batch_requests)} items — {datetime.datetime.now().strftime('%Y-%m-%d')}"
    )

    print(f"  Batch submitted: {batch_id}")
    results_raw = bp.wait_for_completion(batch_id, max_wait_hours=24, check_interval_minutes=5)

    if not results_raw:
        print("Batch failed — falling back to individual processing.")
        return generate_marc_records_individual(barcodes, workflow_data, model_name)

    processed = bp.process_batch_results(results_raw, id_map)
    marc_results = {}

    for custom_id, result in processed.get("results", {}).items():
        barcode = id_map.get(custom_id)
        if not barcode:
            continue
        if result.get("success"):
            try:
                content = result["content"]
                # Strip markdown fences if present
                content = content.strip()
                if content.startswith("```"):
                    content = content.split("```")[1]
                    if content.startswith("json"):
                        content = content[4:]
                marc_fields = json.loads(content.strip())
                marc_results[barcode] = {"marc_fields": marc_fields, "success": True, "tokens": result.get("usage", {})}
            except Exception as e:
                print(f"  Parse error for {barcode}: {e}")
                marc_results[barcode] = {"marc_fields": {}, "success": False, "error": str(e)}
        else:
            marc_results[barcode] = {"marc_fields": {}, "success": False, "error": result.get("error", "Unknown")}

    return marc_results


def generate_marc_records_individual(barcodes, workflow_data, model_name):
    """Fallback: generate MARC records one at a time."""
    marc_results = {}
    total = len(barcodes)

    for i, barcode in enumerate(barcodes, 1):
        print(f"  [{i}/{total}] Generating MARC for {barcode}...")
        record = workflow_data.get("records", {}).get(barcode, {})
        extracted = record.get("step1_metadata_extraction", {}).get("extracted_fields", {})

        if not extracted:
            marc_results[barcode] = {"marc_fields": {}, "success": False, "error": "No Step 1 data"}
            continue

        prompt = build_marc_prompt(barcode, extracted)

        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are an expert music cataloger. Return only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                **get_token_limit_param(model_name, 2000),
                **get_temperature_param(model_name, 0.1)
            )
            content = response.choices[0].message.content.strip()
            marc_fields = json.loads(content)
            marc_results[barcode] = {
                "marc_fields": marc_fields,
                "success": True,
                "tokens": {
                    "prompt_tokens": response.usage.prompt_tokens,
                    "completion_tokens": response.usage.completion_tokens
                }
            }
        except Exception as e:
            print(f"    Error: {e}")
            marc_results[barcode] = {"marc_fields": {}, "success": False, "error": str(e)}

    return marc_results


def save_marc_to_workflow_json(workflow_json_path, barcode, marc_result, extracted_fields):
    """Save generated MARC fields to workflow JSON under step3b key."""
    data = load_workflow_json(workflow_json_path)

    if barcode not in data.get("records", {}):
        return

    data["records"][barcode]["step3b_original_cataloging"] = {
        "marc_fields": marc_result.get("marc_fields", {}),
        "generation_success": marc_result.get("success", False),
        "error": marc_result.get("error"),
        "tokens_used": marc_result.get("tokens", {}),
        "generated_at": datetime.datetime.now().isoformat(),
        "status": "pending_review"  # Set to approved/rejected by step_3c
    }

    data["records"][barcode]["updated_at"] = datetime.datetime.now().isoformat()
    save_workflow_json(workflow_json_path, data)


def format_marc_for_display(marc_fields, barcode):
    """Format MARC fields as readable text for HTML display.
    Strips any accidental MARC tag prefixes the AI included in field values."""
    if not marc_fields:
        return "MARC generation failed — no fields available."

    today_str = datetime.datetime.now().strftime('%Y-%m-%d')

    def strip_tag(val, prefixes):
        """Remove any accidental tag prefix from a field value."""
        if not val:
            return val
        v = str(val).strip()
        for p in prefixes:
            if v.startswith(p):
                v = v[len(p):].strip()
        return v

    lines = []

    # 007
    lines.append(f"007    {strip_tag(marc_fields.get('field_007','sd fsngnnmmned'), ['007'])}")

    # 024 UPC
    f024 = marc_fields.get('field_024')
    if f024:
        val = strip_tag(str(f024), ['024 1_','024 1 ','024'])
        # Never show barcodes starting with 05917
        if val and not val.startswith('05917') and not val.startswith('$a 05917'):
            lines.append(f"024 1_ {val}")

    # 1xx
    f100 = marc_fields.get('field_100')
    if f100:
        lines.append(f"100 1_ {strip_tag(f100, ['100 1_','100 1 ','100'])}")
    f110 = marc_fields.get('field_110')
    if f110 and not f100:
        lines.append(f"110 2_ {strip_tag(f110, ['110 2_','110 2 ','110'])}")

    # 245
    f245 = marc_fields.get('field_245')
    if f245:
        lines.append(f"245 10 {strip_tag(f245, ['245 10','245 00','245 1 ','245 0 ','245'])}")

    # 246
    f246 = marc_fields.get('field_246')
    if f246:
        lines.append(f"246 1_ {strip_tag(f246, ['246 1_','246 1 ','246'])}")

    # 264
    f264 = marc_fields.get('field_264')
    if f264:
        lines.append(f"264 _1 {strip_tag(f264, ['264 _1','264  1','264'])}")

    # 300/336/337/338
    lines.append(f"300    {strip_tag(marc_fields.get('field_300','1 audio disc : $b digital ; $c 4 3/4 in.'), ['300'])}")
    lines.append(f"336    {strip_tag(marc_fields.get('field_336','performed music $b prm $2 rdacontent'), ['336'])}")
    lines.append(f"337    {strip_tag(marc_fields.get('field_337','audio $b s $2 rdamedia'), ['337'])}")
    lines.append(f"338    {strip_tag(marc_fields.get('field_338','audio disc $b sd $2 rdacarrier'), ['338'])}")

    # 500 general note
    f500g = marc_fields.get('field_500_general')
    if f500g:
        lines.append(f"500    {strip_tag(f500g, ['500   ','500  ','500'])}")

    # 500 AI note
    f500ai = marc_fields.get('field_500_ai', 'AI-generated minimum viable record.')
    if f500ai:
        lines.append(f"500    {strip_tag(f500ai, ['500   ','500  ','500'])}")

    # 505
    f505 = marc_fields.get('field_505')
    if f505:
        lines.append(f"505 0_ {strip_tag(f505, ['505 0_','505 0 ','505'])}")

    # 518
    f518 = marc_fields.get('field_518')
    if f518:
        lines.append(f"518    {strip_tag(f518, ['518   ','518  ','518'])}")

    # 588
    f588 = marc_fields.get('field_588',
        f'Description based on AI-assisted metadata extraction from cover image, {today_str}.')
    lines.append(f"588    {strip_tag(f588, ['588   ','588  ','588'])}")

    # 650 subjects
    for subj in (marc_fields.get('field_650') or []):
        if subj:
            lines.append(f"650  _0 {strip_tag(subj, ['650  _0','650 _0','650  0','650'])}")

    # 700 added entries
    for ae in (marc_fields.get('field_700') or []):
        if ae:
            lines.append(f"700 1_ {strip_tag(ae, ['700 1_','700 1 ','700'])}")

    conf = marc_fields.get('confidence_score', 0)
    notes = marc_fields.get('completeness_notes')
    lines.append(f"\n[Confidence: {conf}%]")
    if notes:
        lines.append(f"[Note: {notes}]")

    return "\n".join(lines)


def create_original_cataloging_html(results_folder, barcodes_with_marc, workflow_json_path,
                                     workflow_data, images_folder, current_ts, records_per_page=100):
    """
    Create HTML review pages for original cataloging records.
    Wrapper that orchestrates index and page creation.
    """
    import math
    print(f"Creating original cataloging HTML review ({len(barcodes_with_marc)} records)...")
    images_folder_name = os.path.basename(images_folder)
    total_pages = math.ceil(len(barcodes_with_marc) / records_per_page)
    index_file = f"original-catalog-index-{current_ts}.html"
    index_path = os.path.join(results_folder, index_file)
    _create_orig_index(index_path, len(barcodes_with_marc), total_pages, current_ts, images_folder_name)
    page_files = []
    for page_num in range(1, total_pages + 1):
        start = (page_num - 1) * records_per_page
        end = min(start + records_per_page, len(barcodes_with_marc))
        page_records = barcodes_with_marc[start:end]
        page_file = f"original-catalog-page-{page_num}-{current_ts}.html"
        page_path = os.path.join(results_folder, page_file)
        _create_orig_page(
            page_path, page_records, workflow_data, images_folder, results_folder,
            page_num, total_pages, records_per_page, start, current_ts, images_folder_name
        )
        page_files.append(page_path)
        print(f"  Created page {page_num}/{total_pages} with {len(page_records)} records")
    return {"index_path": index_path, "page_files": page_files, "total_pages": total_pages}


def _create_orig_index(index_path, total_records, total_pages, current_ts, images_folder_name):
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Original Cataloging Review — {current_ts}</title>
<style>
body{{font-family:Arial,sans-serif;margin:20px;background:#fef9f0;}}
.header{{background:#d35400;color:white;padding:20px;border-radius:5px;margin-bottom:30px;}}
.summary{{background:white;padding:20px;border-radius:8px;margin-bottom:20px;box-shadow:0 2px 4px rgba(0,0,0,.1);}}
.page-links{{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:15px;margin-top:20px;}}
.page-link{{background:#e67e22;color:white;padding:15px;text-decoration:none;border-radius:5px;text-align:center;font-weight:bold;}}
.page-link:hover{{background:#d35400;}}
.notice{{background:#fdebd0;border-left:4px solid #e67e22;padding:12px;border-radius:4px;margin-bottom:15px;}}
</style>
</head>
<body>
<div class="header">
<h1>AI Original Cataloging Review</h1>
<p>Generated: {current_ts} | Records with no OCLC match: {total_records} | Pages: {total_pages}</p>
</div>
<div class="summary">
<div class="notice"><strong>What this is:</strong> These {total_records} CDs had no matching record in OCLC WorldCat.
The AI has generated a complete MARC record for each one based on the cover image metadata.
Review each record alongside the CD image. Approve to send to OCLC and Alma, or reject with notes.</div>
<div class="notice"><strong>After review:</strong> Export your decisions to CSV, then run <code>step_3c_oclc_original_record.py</code>
followed by <code>step_3d_original_catalog_alma_import.py</code>.</div>
</div>
<div class="summary">
<h2>Review Pages</h2>
<div class="page-links">"""

    for p in range(1, total_pages + 1):
        html += f'<a href="original-catalog-page-{p}-{current_ts}.html" class="page-link">Page {p}</a>'

    html += f"""
</div></div>
<div class="summary">
<h2>Export All Decisions</h2>
<button onclick="exportAll()" style="background:#c0392b;color:white;border:none;padding:15px 30px;border-radius:5px;cursor:pointer;font-weight:bold;font-size:16px;">
Export All Original Cataloging Decisions to CSV</button>
</div>
<script>
const PFX='orig-catalog-{current_ts}-';
function exportAll(){{
  const name=prompt('Your name:'); if(!name) return;
  const rows=[];
  for(let i=0;i<localStorage.length;i++){{
    const k=localStorage.key(i); if(!k||!k.startsWith(PFX)) continue;
    const m=k.replace(PFX,'').match(/^decision-(.+)$/); if(!m) continue;
    const barcode=m[1];
    const decision=localStorage.getItem(k)||'';
    const notes=localStorage.getItem(PFX+'notes-'+barcode)||'';
    rows.push([barcode,decision,notes,name,new Date().toISOString().split('T')[0],'original_cataloging'].join(','));
  }}
  if(!rows.length){{alert('No decisions found. Open a review page first.');return;}}
  const csv='Barcode,Cataloger Decision,Notes,Cataloger,Review Date,Record Type\\n'+rows.join('\\n');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([csv],{{type:'text/csv'}}));
  a.download='original-catalog-decisions-{images_folder_name}-{current_ts}.csv';
  document.body.appendChild(a);a.click();a.remove();
  alert('Exported '+rows.length+' decisions.');
}}
</script>
</body></html>"""

    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(html)


def _create_orig_page(page_path, page_records, workflow_data, images_folder, results_folder,
                       page_num, total_pages, records_per_page, start_idx, current_ts, images_folder_name):
    """Create a single original cataloging review page."""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Original Cataloging — Page {page_num} — {current_ts}</title>
<style>
body{{font-family:Arial,sans-serif;margin:20px;background:#fef9f0;}}
.header{{background:#d35400;color:white;padding:20px;border-radius:5px;margin-bottom:20px;}}
.nav{{background:white;padding:15px;border-radius:5px;margin-bottom:20px;text-align:center;}}
.nav a,.nav span{{background:#e67e22;color:white;padding:10px 20px;text-decoration:none;border-radius:5px;margin:0 8px;font-weight:bold;display:inline-block;}}
.nav .disabled{{background:#95a5a6;pointer-events:none;}}
.record{{background:white;border:2px solid #e67e22;border-radius:8px;margin-bottom:30px;padding:20px;box-shadow:0 2px 4px rgba(0,0,0,.1);}}
.record-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:15px;padding-bottom:10px;border-bottom:2px solid #f39c12;}}
.barcode{{font-size:22px;font-weight:bold;color:#d35400;}}
.badge{{background:#e67e22;color:white;padding:6px 14px;border-radius:15px;font-weight:bold;font-size:14px;}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px;}}
.images{{display:flex;flex-direction:column;gap:12px;}}
.img-label{{font-weight:bold;color:#555;margin-bottom:4px;}}
.images img{{max-width:100%;max-height:480px;border:2px solid #e67e22;border-radius:5px;cursor:pointer;object-fit:contain;}}
.marc-section{{background:#fef9f0;padding:15px;border-radius:5px;border:1px solid #f39c12;max-height:80vh;overflow-y:auto;}}
.marc-title{{font-weight:bold;color:#d35400;margin-bottom:8px;font-size:16px;}}
.marc-field{{font-family:monospace;font-size:13px;margin:3px 0;white-space:pre-wrap;word-break:break-word;}}
.tag{{color:#2980b9;font-weight:bold;min-width:60px;display:inline-block;}}
.val{{color:#2c3e50;}}
.decision-area{{grid-column:1/-1;background:#fdebd0;border:1px solid #e67e22;border-radius:5px;padding:15px;margin-top:10px;}}
.decision-area h3{{color:#d35400;margin-bottom:10px;}}
.btns{{display:flex;gap:10px;flex-wrap:wrap;margin-top:8px;}}
.btn{{padding:9px 16px;border:none;border-radius:5px;cursor:pointer;font-weight:bold;font-size:14px;}}
.btn-approve{{background:#27ae60;color:white;}}
.btn-reject{{background:#e74c3c;color:white;}}
.btn-hold{{background:#f39c12;color:white;}}
textarea,input[type=text]{{width:100%;margin-top:10px;padding:8px;border:1px solid #ddd;border-radius:5px;font-size:14px;box-sizing:border-box;}}
textarea{{resize:vertical;min-height:60px;}}
.no-img{{color:#999;font-style:italic;text-align:center;padding:20px;border:2px dashed #e67e22;border-radius:5px;}}
.confidence-note{{background:#fef5e7;border:1px solid #f39c12;padding:8px 12px;border-radius:4px;margin-top:8px;font-size:13px;color:#7d6608;}}
.export-bar{{background:white;padding:15px;margin-bottom:20px;border:1px solid #ddd;border-radius:8px;}}
</style>
</head>
<body>
<div class="header">
<h1>AI Original Cataloging Review — Page {page_num} of {total_pages}</h1>
<p>Records {start_idx+1}–{start_idx+len(page_records)} | No OCLC match found — AI-generated MARC records</p>
</div>
<div class="nav">
<a href="original-catalog-index-{current_ts}.html">Index</a>
{'<a href="original-catalog-page-'+str(page_num-1)+'-'+current_ts+'.html">Previous</a>' if page_num > 1 else '<span class="disabled">Previous</span>'}
<strong style="margin:0 15px;">Page {page_num} of {total_pages}</strong>
{'<a href="original-catalog-page-'+str(page_num+1)+'-'+current_ts+'.html">Next</a>' if page_num < total_pages else '<span class="disabled">Next</span>'}
</div>
<div class="export-bar">
<strong>Export this page:</strong>&nbsp;
<button onclick="exportPage()" style="background:#c0392b;color:white;border:none;padding:8px 18px;border-radius:5px;cursor:pointer;font-weight:bold;">Export Page CSV</button>
&nbsp;<span style="color:#666;font-size:13px;">Use the Index page to export all pages at once.</span>
</div>
"""

    images_subfolder = os.path.join(results_folder, "images")
    os.makedirs(images_subfolder, exist_ok=True)

    for i, barcode in enumerate(page_records, 1):
        record_id = start_idx + i
        record = workflow_data.get("records", {}).get(str(barcode), {})
        step3b = record.get("step3b_original_cataloging", {})
        marc_fields = step3b.get("marc_fields", {})
        marc_display = format_marc_for_display(marc_fields, barcode)
        confidence_note = str(marc_fields.get("confidence_notes", "") or "") if marc_fields else ""

        # Get title for display
        title_display = ""
        if marc_fields:
            t = marc_fields.get("field_245") or marc_fields.get("field_100") or ""
            title_display = str(t)[:80] if t else ""

        # Copy images
        image_files = []
        if os.path.exists(images_folder):
            for fname in sorted(os.listdir(images_folder)):
                if fname.startswith(str(barcode)) and fname.lower().endswith(('.jpg', '.jpeg', '.png')):
                    src = os.path.join(images_folder, fname)
                    dst = os.path.join(images_subfolder, fname)
                    try:
                        shutil.copy2(src, dst)
                        image_files.append((os.path.join("images", fname).replace("\\", "/"), fname))
                    except Exception:
                        pass

        html += f"""
<div class="record" id="rec-{record_id}" data-barcode="{barcode}">
  <div class="record-header">
    <div class="barcode">Barcode: {barcode}</div>
    <div class="badge">AI Original Record</div>
  </div>
  {'<div style="color:#7d6608;background:#fef5e7;padding:6px 10px;border-radius:4px;margin-bottom:12px;font-size:13px;">'+title_display+'</div>' if title_display else ''}
  <div class="grid">
    <div class="images">
      <h3 style="color:#d35400;">CD Images</h3>"""

        if image_files:
            for img_path, fname in image_files[:3]:
                lbl = "Front Cover" if 'a.' in fname.lower() else ("Back Cover" if 'b.' in fname.lower() else "Additional")
                html += f"""
      <div><div class="img-label">{lbl}</div>
      <img src="{img_path}" alt="{lbl}" onclick="window.open(this.src,'_blank')"
           onerror="this.style.display='none'"></div>"""
        else:
            html += '<div class="no-img">No images found</div>'

        html += f"""
    </div>
    <div class="marc-section">
      <div class="marc-title">AI-Generated MARC Record</div>
      <div class="marc-field"><pre style="margin:0;white-space:pre-wrap;font-size:13px;">{marc_display}</pre></div>
      {'<div class="confidence-note"><strong>Cataloger note:</strong> '+confidence_note+'</div>' if confidence_note else ''}
    </div>
    <div class="decision-area">
      <h3>Cataloger Decision</h3>
      <p style="color:#555;font-size:14px;">Review the CD images and AI-generated MARC record. Approve to proceed to OCLC and Alma, or reject with notes.</p>
      <div class="btns">
        <button class="btn btn-approve" onclick="decide('{barcode}','Approved',event)">Approve — Send to OCLC &amp; Alma</button>
        <button class="btn btn-reject" onclick="decide('{barcode}','Rejected',event)">Reject — Do Not Import</button>
        <button class="btn btn-hold" onclick="decide('{barcode}','Needs Further Review',event)">Hold for Further Review</button>
      </div>
      <textarea placeholder="Notes for cataloger (corrections, issues, etc.)..." id="notes-{barcode}"></textarea>
    </div>
  </div>
</div>"""

    # JavaScript — same localStorage pattern as Step 6
    html += f"""
<script>
const PFX='orig-catalog-{current_ts}-';
const PAGE={page_num};
const BARCODES={json.dumps(page_records)};

function decide(barcode, decision, e){{
  const rec=document.getElementById('rec-'+(BARCODES.indexOf(barcode)+1+{start_idx}));
  rec.querySelectorAll('.btn').forEach(b=>b.style.opacity='0.5');
  e.target.style.opacity='1'; e.target.style.transform='scale(1.05)';
  localStorage.setItem(PFX+'decision-'+barcode, decision);
  saveNotes();
}}

function saveNotes(){{
  BARCODES.forEach(b=>{{
    const el=document.getElementById('notes-'+b);
    if(el) localStorage.setItem(PFX+'notes-'+b, el.value);
  }});
}}

document.addEventListener('input', e=>{{
  if(e.target.tagName==='TEXTAREA') saveNotes();
}});

document.addEventListener('DOMContentLoaded', ()=>{{
  BARCODES.forEach((b,i)=>{{
    const recId=i+1+{start_idx};
    const decision=localStorage.getItem(PFX+'decision-'+b);
    const notes=localStorage.getItem(PFX+'notes-'+b);
    const rec=document.getElementById('rec-'+recId);
    if(!rec) return;
    if(decision){{
      rec.querySelectorAll('.btn').forEach(btn=>btn.style.opacity='0.5');
      rec.querySelectorAll('.btn').forEach(btn=>{{
        if(btn.textContent.includes(decision.split(' ')[0])) {{btn.style.opacity='1';btn.style.transform='scale(1.05)';}}
      }});
    }}
    if(notes){{const el=document.getElementById('notes-'+b); if(el) el.value=notes;}}
  }});
}});

function exportPage(){{
  const name=prompt('Your name:'); if(!name) return;
  const rows=[];
  BARCODES.forEach(b=>{{
    const d=localStorage.getItem(PFX+'decision-'+b)||'';
    if(!d) return;
    const n=localStorage.getItem(PFX+'notes-'+b)||'';
    rows.push([b,'"'+d+'"','"'+n.replace(/"/g,'""')+'"','"'+name+'"',new Date().toISOString().split('T')[0],'original_cataloging'].join(','));
  }});
  if(!rows.length){{alert('No decisions on this page yet.');return;}}
  const csv='Barcode,Cataloger Decision,Notes,Cataloger,Review Date,Record Type\\n'+rows.join('\\n');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([csv],{{type:'text/csv'}}));
  a.download='original-catalog-decisions-page{page_num}-{current_ts}.csv';
  document.body.appendChild(a);a.click();a.remove();
  alert('Exported '+rows.length+' decisions from page {page_num}.');
}}
</script>
</body></html>"""

    with open(page_path, 'w', encoding='utf-8') as f:
        f.write(html)


def create_report(results_folder, barcodes_processed, marc_results, workflow_data, current_ts):
    """
    Create receipt/report of original cataloging processing.
    Also creates batch-ready import file for records scoring >= confidence threshold.
    Saves to:
      1. deliverables/original-cataloging-report-[timestamp].txt
      2. deliverables/original-cataloging-batch-ready-[timestamp].txt  (high confidence)
      3. AI_Music_Operations/original-cataloging/[date]/
    """
    from cd_workflow_config import get_threshold_config
    threshold = get_threshold_config("confidence").get("high_confidence", 70)
    ops_dir = os.environ.get("AI_MUSIC_OPERATIONS_DIR")
    deliverables = os.path.join(results_folder, "deliverables")

    successful = [b for b in barcodes_processed if marc_results.get(b, {}).get("success")]
    failed = [b for b in barcodes_processed if not marc_results.get(b, {}).get("success")]

    # Split into high confidence (batch ready) and needs review
    batch_ready = []
    needs_review = []
    for barcode in successful:
        marc = marc_results.get(barcode, {}).get("marc_fields", {})
        score = marc.get("confidence_score", 0) if marc else 0
        record = workflow_data.get("records", {}).get(str(barcode), {})
        extracted = record.get("step1_metadata_extraction", {}).get("extracted_fields", {})
        title = (extracted.get("title_information") or {}).get("main_title") or "Unknown title"
        if score >= threshold:
            batch_ready.append((barcode, score, title))
        else:
            needs_review.append((barcode, score, title))

    # Write batch-ready file (same format as copy cataloging: OCLC|barcode|title)
    # For original cataloging OCLC number is not yet assigned — use placeholder
    # Step 3c will fill in the real OCLC number after creation
    batch_filename = f"original-cataloging-batch-ready-{current_ts}.txt"
    batch_path = os.path.join(deliverables, batch_filename)
    with open(batch_path, 'w', encoding='utf-8') as f:
        for barcode, score, title in batch_ready:
            f.write(f"PENDING|{barcode}|{title}\n")
    print(f"Batch-ready file saved: {batch_path} ({len(batch_ready)} records)")

    lines = [
        "=" * 65,
        "AI MUSIC METADATA PROJECT — ORIGINAL CATALOGING REPORT",
        "=" * 65,
        f"Generated:              {current_ts}",
        f"Results folder:         {os.path.basename(results_folder)}",
        f"Confidence threshold:   {threshold}%",
        f"Total processed:        {len(barcodes_processed)}",
        f"MARC generated:         {len(successful)}",
        f"Failed:                 {len(failed)}",
        f"Batch-ready (>={threshold}%):   {len(batch_ready)}",
        f"Needs review (<{threshold}%):   {len(needs_review)}",
        "",
        "NEXT STEPS:",
        f"  HIGH CONFIDENCE ({len(batch_ready)} records):",
        "  1. Run step_3c_oclc_original_record.py to create OCLC records",
        "  2. Run step_3d_original_catalog_alma_import.py to import to Alma",
        f"  LOW CONFIDENCE ({len(needs_review)} records):",
        "  1. Open original-catalog-index-[timestamp].html to review",
        "  2. Export decisions CSV",
        "  3. Run step_3c then step_3d for approved records",
        "",
        f"BATCH-READY RECORDS (>={threshold}% confidence):",
        "-" * 40,
    ]
    for barcode, score, title in batch_ready:
        lines.append(f"  {barcode}  |  {score:3d}%  |  {title[:45]}")

    if needs_review:
        lines += ["", f"NEEDS REVIEW (<{threshold}% confidence):", "-" * 40]
        for barcode, score, title in needs_review:
            lines.append(f"  {barcode}  |  {score:3d}%  |  {title[:45]}")

    if failed:
        lines += ["", "FAILED (MARC generation error):", "-" * 40]
        for barcode in failed:
            lines.append(f"  {barcode}")

    lines += [
        "",
        "=" * 65,
        "NOTE: All records use the approved minimal MARC template.",
        "Fields: 007, 1xx, 245, 264, 300, 336-338, 500, 505, 588, 650x2",
        "Omitted: 024, 028, 518, local notes",
        "500: AI-generated minimum viable record.",
        "650: Popular music. / Sound recordings. (no geographic subdivision)",
        "=" * 65,
    ]

    report_text = "\n".join(lines)
    report_filename = f"original-cataloging-report-{current_ts}.txt"
    report_path = os.path.join(deliverables, report_filename)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_text)
    print(f"Report saved: {report_path}")

    if ops_dir:
        date_str = datetime.datetime.now().strftime('%Y-%m-%d')
        ops_folder = os.path.join(ops_dir, "original-cataloging", date_str)
        os.makedirs(ops_folder, exist_ok=True)
        for fname, fpath in [(report_filename, report_path), (batch_filename, batch_path)]:
            import shutil
            shutil.copy2(fpath, os.path.join(ops_folder, fname))
        print(f"Reports also saved to: {ops_folder}")

    return report_path, batch_path
    """
    Create receipt/report of original cataloging processing.
    Saves to:
      1. deliverables/original-cataloging-report-[timestamp].txt
      2. AI_Music_Operations/original-cataloging/[date]/  (if env var set)
    """
    ops_dir = os.environ.get("AI_MUSIC_OPERATIONS_DIR")

    successful = [b for b in barcodes_processed if marc_results.get(b, {}).get("success")]
    failed = [b for b in barcodes_processed if not marc_results.get(b, {}).get("success")]

    lines = [
        "=" * 65,
        "AI MUSIC METADATA PROJECT — ORIGINAL CATALOGING REPORT",
        "=" * 65,
        f"Generated:          {current_ts}",
        f"Results folder:     {os.path.basename(results_folder)}",
        f"Total processed:    {len(barcodes_processed)}",
        f"MARC generated:     {len(successful)}",
        f"Failed:             {len(failed)}",
        f"Status:             Pending cataloger review",
        "",
        "NEXT STEPS:",
        "  1. Open original-catalog-index-[timestamp].html in a browser",
        "  2. Review each AI-generated MARC record alongside the CD image",
        "  3. Approve, Reject, or Hold each record",
        "  4. Export decisions to CSV",
        "  5. Run step_3c_oclc_original_record.py to create OCLC records",
        "  6. Run step_3d_original_catalog_alma_import.py to import to Alma",
        "",
        "MARC GENERATION SUMMARY:",
        "-" * 40,
    ]

    for barcode in barcodes_processed:
        result = marc_results.get(barcode, {})
        record = workflow_data.get("records", {}).get(str(barcode), {})
        extracted = record.get("step1_metadata_extraction", {}).get("extracted_fields", {})
        title = extracted.get("title_information", {}).get("main_title", "Unknown")
        artist = extracted.get("title_information", {}).get("primary_contributor", "Unknown")
        status = "OK" if result.get("success") else f"FAILED: {result.get('error','')[:50]}"
        lines.append(f"  {barcode}  |  {(title or chr(63))[:35]:<35}  |  {(artist or chr(63))[:25]:<25}  |  {status}")

    lines += [
        "",
        "=" * 65,
        "NOTE: All records are AI-generated and require cataloger review",
        "before submission to OCLC or import to Alma.",
        "Records will carry MARC field 500: 'AI-generated catalog record.'",
        "and field 588 with the generation date.",
        "=" * 65,
    ]

    report_text = "\n".join(lines)

    # Save to deliverables
    deliverables = os.path.join(results_folder, "deliverables")
    report_filename = f"original-cataloging-report-{current_ts}.txt"
    report_path = os.path.join(deliverables, report_filename)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_text)
    print(f"Report saved: {report_path}")

    # Save to AI_Music_Operations if configured
    if ops_dir:
        date_str = datetime.datetime.now().strftime('%Y-%m-%d')
        ops_folder = os.path.join(ops_dir, "original-cataloging", date_str)
        os.makedirs(ops_folder, exist_ok=True)
        ops_report_path = os.path.join(ops_folder, report_filename)
        with open(ops_report_path, 'w', encoding='utf-8') as f:
            f.write(report_text)
        print(f"Report also saved to: {ops_report_path}")

    return report_path


def main():
    print("=" * 65)
    print("STEP 3b: AI ORIGINAL CATALOGING — NO-OCLC-MATCH RECORDS")
    print("=" * 65)

    file_paths = get_file_path_config()
    model_config = get_model_config("step3")  # Reuse step3 model config
    model_name = model_config.get("model", "gpt-4.1")
    current_ts = get_current_timestamp()

    # Find latest results folder
    results_folder = find_latest_results_folder(file_paths["results_prefix"])
    if not results_folder:
        print("No results folder found. Run Steps 1-5 first.")
        return

    print(f"Results folder: {results_folder}")

    # Get workflow JSON from data subfolder (Step 5 moves it there)
    data_folder = os.path.join(results_folder, "data")
    workflow_json_path = get_workflow_json_path(data_folder)
    workflow_data = load_workflow_json(workflow_json_path)

    # Find no-match barcodes
    no_match_barcodes = get_no_match_barcodes(results_folder)
    if not no_match_barcodes:
        print("No records need original cataloging. Exiting.")
        return

    print(f"\nModel: {model_name}")
    print(f"Records to process: {len(no_match_barcodes)}")

    # Decide batch vs individual
    use_batch = bp.should_use_batch(num_requests=len(no_match_barcodes), step_name="step3")
    print(f"Processing mode: {'BATCH' if use_batch else 'INDIVIDUAL'}")

    start_time = datetime.datetime.now()

    # Generate MARC records
    if use_batch:
        marc_results = generate_marc_records_batch(
            no_match_barcodes, workflow_json_path, workflow_data, model_name
        )
    else:
        marc_results = generate_marc_records_individual(
            no_match_barcodes, workflow_data, model_name
        )

    # Save results to workflow JSON
    print("\nSaving MARC data to workflow JSON...")
    for barcode in no_match_barcodes:
        result = marc_results.get(barcode, {"marc_fields": {}, "success": False, "error": "Not processed"})
        extracted = workflow_data.get("records", {}).get(str(barcode), {}).get(
            "step1_metadata_extraction", {}).get("extracted_fields", {})
        save_marc_to_workflow_json(workflow_json_path, str(barcode), result, extracted)
        print(f"  {barcode}: {'OK' if result.get('success') else 'FAILED'}")

    # Reload workflow data with new step3b entries
    workflow_data = load_workflow_json(workflow_json_path)

    # Get images folder path
    images_folder = file_paths["images_folder"]

    # Create HTML review interface
    print("\nCreating HTML review interface...")
    html_result = create_original_cataloging_html(
        results_folder,
        no_match_barcodes,
        workflow_json_path,
        workflow_data,
        images_folder,
        current_ts
    )

    # Create report/receipt
    print("\nGenerating report...")
    report_path, batch_path = create_report(
        results_folder, no_match_barcodes, marc_results, workflow_data, current_ts
    )

    # Log metrics
    duration = (datetime.datetime.now() - start_time).total_seconds()
    successful = sum(1 for b in no_match_barcodes if marc_results.get(b, {}).get("success"))
    failed = len(no_match_barcodes) - successful

    try:
        log_processing_metrics(
            results_folder_path=results_folder,
            step="step3b_original_cataloging",
            batch_metrics=create_batch_summary(
                total_items=len(no_match_barcodes),
                successful_items=successful,
                failed_items=failed,
                total_time=duration,
                total_tokens=0,
                estimated_cost=0,
                processing_mode="BATCH" if use_batch else "INDIVIDUAL"
            )
        )
    except Exception as e:
        print(f"Warning: Could not log metrics: {e}")

    print(f"\n{'='*65}")
    print(f"STEP 3b COMPLETE")
    print(f"{'='*65}")
    print(f"Records processed:  {len(no_match_barcodes)}")
    print(f"MARC generated:     {successful}")
    print(f"Failed:             {failed}")
    print(f"Duration:           {duration:.1f}s")
    print(f"\nHTML review index:  {html_result['index_path']}")
    print(f"Report:             {report_path}")
    print(f"\nNext: Open the HTML index, review records, export CSV,")
    print(f"then run step_3c_oclc_original_record.py")


if __name__ == "__main__":
    main()
