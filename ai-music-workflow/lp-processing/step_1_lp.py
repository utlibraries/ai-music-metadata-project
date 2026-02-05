# Extract metadata from LP images using GPT-4o with batch processing support
import os
import base64
import time
from datetime import datetime
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openai import OpenAI

# Import custom modules
from token_logging import create_token_usage_log, log_individual_response
from batch_processor import BatchProcessor
from model_pricing import calculate_cost, get_model_info
from json_workflow import initialize_workflow_json, update_record_step1, log_error, log_processing_metrics
from shared_utilities import get_workflow_json_path, extract_metadata_fields, group_images_by_barcode, create_batch_summary
from lp_workflow_config import get_current_timestamp, get_file_path_config, get_model_config, get_token_limit_param
from retry_utils import retry_api_call, log_failure

STEP_NAME = "step1"
bp = BatchProcessor(default_step=STEP_NAME)

def should_use_batch_for_this_step(num_requests: int) -> bool:
    return bp.should_use_batch(num_requests=num_requests, step_name=STEP_NAME)

MODEL_CONFIG = get_model_config(STEP_NAME)
DEFAULT_MODEL = MODEL_CONFIG["model"]
DEFAULT_MAX_TOKENS = MODEL_CONFIG["max_tokens"]
DEFAULT_TEMPERATURE = MODEL_CONFIG["temperature"]


client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

def get_llm_prompt():
    return """Analyze these images of a vinyl record and extract the following metadata fields in the specified format. 

### Rules:
1. For Latin alphabet based languages, such as Germanic languages, romance languages, slavic languages: use original language unless otherwise noted in the metadata instructions.
2. For non-Latin alphabet based languages, for example languages written in the cyrillic alphabet, in arabic script, in chinese characters: metadata should be either English transliterations or English translations.  No non latin characters should be included in the metadata.
3. Do not include any notes to explain why something has been included or excluded from a query.
4. Do not include any information that is not visible in the images.  Mark fields where the information is not available or unclear as "Not visible".
5. Take care to spell names and titles correctly. 
6. Metadata must be in this exact format, with no additional fields and no introductory, concluding, or other text.  

### Metadata Format:
Title Information:
  - Main Title: [Title, transliterated if the characters are not Latin]
  - English Title: [Title in English if written on the item]
  - Subtitle: [Subtitle in English or transliterated.  If a subtitle is written in both English and another language, use the English subtitle]
Primary Contributor:
  - Artist/Performer: [Name of primary artist or group (transliterated if the characters are not Latin)]
Additional Contributors:
  - Arrangers: [List of arrangers]
  - Engineers: [List of engineers]
  - Producers: [List of producers]
  - Session Musicians: [List of session musicians with instruments]
Publishers:
  - Name: [Publisher Name]
    Place: [Place of publication]
    Numbers: [Manufacturing numbers, UPC (NOT university library barcode stickers)]
Dates:
  - publicationDate: [Publication Date]
  - recordingDate: [Recording Date]
  - recordingLocation: [Recording location]
Language:
  - sungLanguages: [Languages of sung text]
  - printedLanguages: [All printed languages visible on the images]
  - pressingLanguage: [Primary printed language on the item that is not English - particularly important for items with multiple languages]
Format:
  - generalFormat: [e.g., Sound Recording]
  - specificFormat: [e.g., Vinyl Record, LP]
  - materialTypes: [List of Material Types]
Sound Characteristics:
  - soundConfiguration: [e.g., stereo/mono]
  - recordingType: [analog/digital]
  - specialPlayback: [e.g., Quadraphonic]
Physical Description:
  - size: [Diameter, e.g., 12", 7"]
  - material: [e.g., Vinyl]
  - labelDesign: [Label design and color]
  - physicalCondition: [Condition notes]
  - specialFeatures: [Sleeve details, inserts, etc.]
Contents:
  - tracks: [
      {
        "number": [Track number],
        "title": [Track title (transliterated if the characters are not Latin)],
        "titleOriginalLanguage": [Title in original language if different],
        "composer": [Composer or songwriter name],
        "lyricist": [Lyricist name],
        "duration": [Duration if shown]
      }
    ]
Series:
  - seriesTitle: [Series name if any (transliterated if the characters are not Latin)]
  - seriesNumber: [Number within series]
Subject Information:
  - genre: [Musical genre/style]
  - geographic: [Geographic origins]
  - timePeriod: [Time period]
Notes:
  - generalNotes: [Any additional notes]
  - technicalNotes: [Pressing plant, mastering info]

Include additional information in the notes section. Return only the metadata in the specified format."""

def prepare_batch_requests(image_groups, model_name):
    """Prepare all requests for batch processing."""
    batch_requests = []
    custom_id_mapping = {}
    
    for i, (barcode, image_paths) in enumerate(sorted(image_groups.items())):
        # Take up to first 3 images for each barcode
        image_paths = image_paths[:3]
        prompt_text = get_llm_prompt()
        uploaded_files_info = ""

        for j, img_path in enumerate(image_paths):
            # Determine image type based on filename
            filename = os.path.basename(img_path).lower()
            if filename.endswith('a.png') or filename.endswith('a.jpg') or filename.endswith('a.jpeg'):
                image_type = "FRONT COVER"
            elif filename.endswith('b.png') or filename.endswith('b.jpg') or filename.endswith('b.jpeg'):
                image_type = "BACK COVER"
            elif filename.endswith('c.png') or filename.endswith('c.jpg') or filename.endswith('c.jpeg'):
                image_type = "ADDITIONAL IMAGE"
            else:
                image_type = "IMAGE"
            
            uploaded_files_info += f"[Image {j+1} - {image_type}: {img_path}]\n"

        prompt = prompt_text + "\n" + uploaded_files_info

        # Prepare base64 images
        base64_images = []
        content_types = []
        
        for img_path in image_paths:
            with open(img_path, "rb") as image_file:
                base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                base64_images.append(base64_image)
            
            ext = os.path.splitext(img_path)[1].lower()
            if ext == '.png':
                content_types.append("image/png")
            else:
                content_types.append("image/jpeg")

        # Create messages with appropriate content types
        image_contents = []
        for j, base64_image in enumerate(base64_images):
            image_contents.append({
                "type": "image_url",
                "image_url": {"url": f"data:{content_types[j]};base64,{base64_image}"}
            })

        # Create request data
        request_data = {
            "model": model_name,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    *image_contents
                ]
            }],
            "max_tokens": 2000
        }
        
        batch_requests.append(request_data)
        custom_id_mapping[f"req_{i}"] = {
            "barcode": barcode,
            "image_paths": image_paths,
            "row_number": i + 2  # +2 for header row
        }
    
    return batch_requests, custom_id_mapping

def process_folder_with_batch(folder_path, wb, results_folder_path, workflow_json_path):
    """Process folder using batch processing when appropriate."""
    model_config = get_model_config("step1")
    model_name = model_config.get("model", "gpt-4o")
    
    ws = wb.active
    headers = ['Input Image 1', 'Input Image 2', 'Input Image 3', 'Barcode', 'AI-Generated Metadata']
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        if col == 4:  # Barcode column
            ws.column_dimensions[get_column_letter(col)].width = 17
        else:
            ws.column_dimensions[get_column_letter(col)].width = 30 if col <= 3 else 52

    # Create logs folder
    logs_folder_path = os.path.join(results_folder_path, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)

    # Group images by barcode
    image_groups = group_images_by_barcode(folder_path)
    total_items = len(image_groups)
    
    print(f"\nSTEP 1: METADATA EXTRACTION")
    print(f"Found {total_items} LP image groups to process")
    print(f"Starting metadata extraction using {model_name}...")
    print("-" * 50)

    # Initialize batch processor and check if we should use batch processing
    use_batch = should_use_batch_for_this_step(total_items)
    processor = bp

    
    print(f"Processing mode: {'BATCH' if use_batch else 'INDIVIDUAL'}")
    
    if use_batch:
        print(f"Preparing {total_items} requests for batch processing...")
        
        # Estimate costs
        batch_requests, custom_id_mapping = prepare_batch_requests(image_groups, model_name)
        cost_estimate = processor.estimate_batch_cost(batch_requests, model_name)
        
        print(f"Cost estimate:")
        print(f"   Regular API: ${cost_estimate['regular_cost']:.4f}")
        print(f"   Batch API: ${cost_estimate['batch_cost']:.4f}")
        print(f"   Savings: ${cost_estimate['savings']:.4f} ({cost_estimate['savings_percentage']:.1f}%)")
        
        # Convert to batch format
        formatted_requests = processor.create_batch_requests(
            batch_requests,
            "lp_metadata"
        )


        # Use adaptive batch processing that automatically splits based on file size
        results = processor.submit_adaptive_batch(
            batch_requests=formatted_requests,
            custom_id_mapping=custom_id_mapping,
            description=f"LP Metadata Extraction - {total_items} items - {datetime.now().strftime('%Y-%m-%d')}",
            max_file_size_mb=40  
        )
        
        if results:
            # Process batch results
            processed_results = processor.process_batch_results(results, custom_id_mapping)
            
            print(f"Processing batch results...")
            items_with_issues = 0
            
            # Add results to spreadsheet
            for custom_id, result_data in processed_results["results"].items():
                    # Handle both single batch and chunked batch custom IDs
                    if custom_id.startswith("lp_metadata_"):
                        # Single batch format: lp_metadata_0_hash
                        index = int(custom_id.split("_")[2])
                        mapping_key = f"req_{index}"
                    elif custom_id.startswith("chunk_"):
                        # Chunked batch format: chunk_0_1_hash
                        parts = custom_id.split("_")
                        if len(parts) >= 3:
                            index = int(parts[2])  # Third part is the original request index
                            mapping_key = f"req_{index}"
                        else:
                            continue  # Skip malformed custom_ids
                    else:
                        continue  # Skip unknown custom_id formats
                    
                    if mapping_key in custom_id_mapping:
                        barcode = custom_id_mapping[mapping_key]["barcode"]
                        image_paths = custom_id_mapping[mapping_key]["image_paths"]
                        row_number = custom_id_mapping[mapping_key]["row_number"]
                        
                        if result_data["success"]:
                            metadata_output = result_data["content"]
                            usage = result_data["usage"]
                            
                            # Log individual response
                            log_individual_response(
                                logs_folder_path=logs_folder_path,
                                script_name="step1",
                                row_number=row_number,
                                barcode=barcode,
                                response_text=metadata_output,
                                model_name=model_name,
                                prompt_tokens=usage.get("prompt_tokens", 0),
                                completion_tokens=usage.get("completion_tokens", 0),
                                processing_time=0  # Batch processing doesn't track individual timing
                            )
                            
                            try:
                                extracted_fields = extract_metadata_fields(metadata_output)
                                update_record_step1(
                                    json_path=workflow_json_path,
                                    barcode=barcode,
                                    raw_metadata=metadata_output,
                                    extracted_fields=extracted_fields,
                                    model=model_name,
                                    prompt_tokens=usage.get("prompt_tokens", 0),
                                    completion_tokens=usage.get("completion_tokens", 0),
                                    processing_time=0
                                )
                            except Exception as json_error:
                                log_error(
                                    results_folder_path=results_folder_path,
                                    step="step1",
                                    barcode=barcode,
                                    error_type="json_update_error",
                                    error_message=str(json_error)
                                )
                        else:
                            metadata_output = f"Error: {result_data['error']}"
                            items_with_issues += 1
                            
                            # Log error
                            log_individual_response(
                                logs_folder_path=logs_folder_path,
                                script_name="step1",
                                row_number=row_number,
                                barcode=barcode,
                                response_text=metadata_output,
                                model_name=model_name,
                                prompt_tokens=0,
                                completion_tokens=0,
                                processing_time=0
                            )
                        
                        # Add to spreadsheet
                        row_data = ['', '', '', barcode, metadata_output]
                        ws.append(row_data)
                        
                        # Add thumbnail images in correct columns based on filename endings
                        sorted_images = {'a': None, 'b': None, 'c': None}
                        for img_path in image_paths:
                            filename = os.path.basename(img_path).lower()
                            if filename.endswith('a.png') or filename.endswith('a.jpg') or filename.endswith('a.jpeg'):
                                sorted_images['a'] = img_path
                            elif filename.endswith('b.png') or filename.endswith('b.jpg') or filename.endswith('b.jpeg'):
                                sorted_images['b'] = img_path
                            elif filename.endswith('c.png') or filename.endswith('c.jpg') or filename.endswith('c.jpeg'):
                                sorted_images['c'] = img_path

                        for col_index, (suffix, img_path) in enumerate([('a', sorted_images['a']), ('b', sorted_images['b']), ('c', sorted_images['c'])], start=1):
                            if img_path:
                                img = PILImage.open(img_path)
                                img.thumbnail((200, 200))
                                output = BytesIO()
                                img.save(output, format='PNG')
                                output.seek(0)
                                img_openpyxl = Image(output)
                                img_openpyxl.anchor = ws.cell(row=ws.max_row, column=col_index).coordinate
                                ws.add_image(img_openpyxl)
                        
                        ws.row_dimensions[ws.max_row].height = 215
                        for cell in ws[ws.max_row]:
                            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Return batch processing metrics
            summary = processed_results["summary"]
            return (total_items, items_with_issues, 0,  # 0 for total_time since batch doesn't track individual timing
                   summary["total_prompt_tokens"], summary["total_completion_tokens"], 
                   summary["total_prompt_tokens"] + summary["total_completion_tokens"])
        
        else:
            print("Batch processing failed, falling back to individual processing...")
            use_batch = False
    
    # Fall back to individual processing if batch fails or isn't used
    if not use_batch:
        return process_folder_individual(image_groups, ws, logs_folder_path, model_name, total_items, workflow_json_path, results_folder_path)

def process_folder_individual(image_groups, ws, logs_folder_path, model_name, total_items, workflow_json_path, results_folder_path):
    """Process using individual API calls (original logic)."""
    items_with_issues = 0
    processed_items = 0
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    total_time = 0

    for barcode, image_paths in sorted(image_groups.items()):
        processed_items += 1
        item_start_time = time.time()
        row_number = processed_items + 1

        print(f"\nProcessing LP {processed_items}/{total_items}")
        print(f"   Barcode: {barcode}")
        print(f"   Progress: {(processed_items/total_items)*100:.1f}%")

        try:
            # Take up to first 3 images for each barcode
            image_paths = image_paths[:3]
            prompt_text = get_llm_prompt()
            uploaded_files_info = ""

            for i, img_path in enumerate(image_paths):
                filename = os.path.basename(img_path).lower()
                if filename.endswith('a.png') or filename.endswith('a.jpg') or filename.endswith('a.jpeg'):
                    image_type = "FRONT COVER"
                elif filename.endswith('b.png') or filename.endswith('b.jpg') or filename.endswith('b.jpeg'):
                    image_type = "BACK COVER"
                elif filename.endswith('c.png') or filename.endswith('c.jpg') or filename.endswith('c.jpeg'):
                    image_type = "ADDITIONAL IMAGE"
                else:
                    image_type = "IMAGE"
                
                uploaded_files_info += f"[Image {i+1} - {image_type}: {img_path}]\n"

            prompt = prompt_text + "\n" + uploaded_files_info

            try:
                print(f"Calling OpenAI API...")
                
                base64_images = []
                for img_path in image_paths:
                    with open(img_path, "rb") as image_file:
                        base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                        base64_images.append(base64_image)

                api_start_time = time.time()
                
                content_types = []
                for img_path in image_paths:
                    ext = os.path.splitext(img_path)[1].lower()
                    if ext == '.png':
                        content_types.append("image/png")
                    else:
                        content_types.append("image/jpeg")
                
                image_contents = []
                for i, base64_image in enumerate(base64_images):
                    image_contents.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:{content_types[i]};base64,{base64_image}"}
                    })

                # Use retry wrapper for API call (3 attempts with exponential backoff)
                success, response, error = retry_api_call(
                    client.chat.completions.create,
                    model=model_name,
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            *image_contents
                        ]
                    }],
                    barcode=barcode,
                    **get_token_limit_param(model_name, 2000)
                )

                api_duration = time.time() - api_start_time

                if not success:
                    # All retries failed - log and continue with placeholder
                    print(f"   ⚠ All retries failed for {barcode} - recording failure and continuing")
                    log_failure(barcode, "step1", error, "Will create placeholder record with 0 confidence")

                    # Log error to workflow
                    log_error(
                        results_folder_path=results_folder_path,
                        step="step1",
                        barcode=barcode,
                        error_type="api_error_max_retries",
                        error_message=error
                    )

                    # Create placeholder record in workflow JSON
                    update_record_step1(
                        json_path=workflow_json_path,
                        barcode=barcode,
                        raw_metadata=f"FAILED: All retry attempts exhausted. Error: {error[:500]}",
                        extracted_fields={},
                        model=model_name,
                        prompt_tokens=0,
                        completion_tokens=0,
                        processing_time=0
                    )

                    # Add error to spreadsheet
                    error_message = f"FAILED after retries: {error[:200]}"
                    ws.append(['', '', '', barcode, error_message])
                    ws.row_dimensions[ws.max_row].height = 50
                    for cell in ws[ws.max_row]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

                    continue  # Skip to next item

                prompt_tokens = response.usage.prompt_tokens
                completion_tokens = response.usage.completion_tokens
                total_item_tokens = prompt_tokens + completion_tokens
                
                total_prompt_tokens += prompt_tokens
                total_completion_tokens += completion_tokens
                total_tokens += total_item_tokens

                metadata_output = response.choices[0].message.content.strip()
                
                print(f"API call successful! Tokens: {total_item_tokens:,}")
                
                log_individual_response(
                    logs_folder_path=logs_folder_path,
                    script_name="step1",
                    row_number=row_number,
                    barcode=barcode,
                    response_text=metadata_output,
                    model_name=model_name,
                    prompt_tokens=prompt_tokens,
                    completion_tokens=completion_tokens,
                    processing_time=api_duration
                )
                try:
                    # Extract structured metadata fields
                    extracted_fields = extract_metadata_fields(metadata_output)
                    
                    # Update workflow JSON with Step 1 results
                    update_record_step1(
                        json_path=workflow_json_path,
                        barcode=barcode,
                        raw_metadata=metadata_output,
                        extracted_fields=extracted_fields,
                        model=model_name,
                        prompt_tokens=prompt_tokens,
                        completion_tokens=completion_tokens,
                        processing_time=api_duration
                    )
                except Exception as json_error:
                    log_error(
                        results_folder_path=results_folder_path,
                        step="step1",
                        barcode=barcode,
                        error_type="json_update_error",
                        error_message=str(json_error)
                    )
                row_data = ['', '', '', barcode, metadata_output]
                ws.append(row_data)

                # Add thumbnail images in correct columns based on filename endings
                sorted_images = {'a': None, 'b': None, 'c': None}
                for img_path in image_paths:
                    filename = os.path.basename(img_path).lower()
                    if filename.endswith('a.png') or filename.endswith('a.jpg') or filename.endswith('a.jpeg'):
                        sorted_images['a'] = img_path
                    elif filename.endswith('b.png') or filename.endswith('b.jpg') or filename.endswith('b.jpeg'):
                        sorted_images['b'] = img_path
                    elif filename.endswith('c.png') or filename.endswith('c.jpg') or filename.endswith('c.jpeg'):
                        sorted_images['c'] = img_path

                for col_index, (suffix, img_path) in enumerate([('a', sorted_images['a']), ('b', sorted_images['b']), ('c', sorted_images['c'])], start=1):
                    if img_path:
                        img = PILImage.open(img_path)
                        img.thumbnail((200, 200))
                        output = BytesIO()
                        img.save(output, format='PNG')
                        output.seek(0)
                        img_openpyxl = Image(output)
                        img_openpyxl.anchor = ws.cell(row=ws.max_row, column=col_index).coordinate
                        ws.add_image(img_openpyxl)

                ws.row_dimensions[ws.max_row].height = 215
                for cell in ws[ws.max_row]:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            except Exception as e:
                print(f"API call failed: {str(e)}")
                error_message = f"Error: {str(e)}"
                ws.append(['', '', '', barcode, error_message])
                items_with_issues += 1

        except Exception as e:
            print(f"Processing failed: {str(e)}")
            error_message = f"Error: {str(e)}"
            ws.append(['', '', '', barcode, error_message])
            items_with_issues += 1

        item_duration = time.time() - item_start_time
        total_time += item_duration

    return total_items, items_with_issues, total_time, total_prompt_tokens, total_completion_tokens, total_tokens

def main():
    # Get configuration
    file_paths = get_file_path_config()
    model_config = get_model_config("step1")
    
    model_name = model_config.get("model", "gpt-4o")
    
    # Start timing the entire script execution
    script_start_time = time.time()
    
    images_folder = file_paths["images_folder"]
    
    current_timestamp = get_current_timestamp()
    results_folder_name = f"results-{current_timestamp}"
    results_folder_path = os.path.join(file_paths["output_base"], results_folder_name)

    if not os.path.exists(results_folder_path):
        os.makedirs(results_folder_path)
    workflow_json_path = get_workflow_json_path(results_folder_path)
    if not os.path.exists(workflow_json_path):
        workflow_json_path = initialize_workflow_json(results_folder_path, images_folder)
        print(f"Initialized workflow JSON: {workflow_json_path}")

    
    logs_folder_path = os.path.join(results_folder_path, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)
    
    # Get image groups to show count in model info
    image_groups = group_images_by_barcode(images_folder)
    total_items = len(image_groups)
    
    # Show model pricing info at start
    model_info = get_model_info(model_name)
    if model_info:
        print(f"STEP 1: METADATA EXTRACTION")
        print(f"Using model: {model_name}")
        print(f"Pricing: ${model_info['input_per_1k']:.5f}/1K input, ${model_info['output_per_1k']:.5f}/1K output")
        print(f"Batch discount: {model_info['batch_discount']*100:.0f}%")
        print(f"Total LPs to process: {total_items}")
        print("-" * 50)
    
    wb = Workbook()
    total_items, items_with_issues, total_time, total_prompt_tokens, total_completion_tokens, total_tokens = process_folder_with_batch(images_folder, wb, results_folder_path, workflow_json_path)

    # Apply formatting to all cells
    for row in wb.active.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.active.freeze_panes = 'A2'

    output_file = f"full-workflow-data-lp-{current_timestamp}.xlsx"
    full_output_path = os.path.join(results_folder_path, output_file)
    wb.save(full_output_path)
    
    # Calculate script metrics
    script_duration = time.time() - script_start_time
    
    # Determine if batch processing was used (check if we have many items but zero processing time)
    was_batch_processed = should_use_batch_for_this_step(total_items) and total_time == 0

    
    # Calculate actual cost using the model pricing
    estimated_cost = calculate_cost(
        model_name=model_name,
        prompt_tokens=total_prompt_tokens,
        completion_tokens=total_completion_tokens,
        is_batch=was_batch_processed
    )
    
    # Create standardized token usage log with enhanced metrics
    # Use script_duration for batch processing (total_time is 0 for batch since individual timing isn't tracked)
    log_time = script_duration if was_batch_processed else total_time

    create_token_usage_log(
        logs_folder_path=logs_folder_path,
        script_name="step1",
        model_name=model_name,
        total_items=total_items,
        items_with_issues=items_with_issues,
        total_time=log_time,
        total_prompt_tokens=total_prompt_tokens,
        total_completion_tokens=total_completion_tokens,
        additional_metrics={
            "Total script execution time": f"{script_duration:.2f}s",
            "Processing time percentage": f"{(log_time/script_duration)*100:.1f}%" if script_duration > 0 else "0%",
            "Items successfully processed": total_items - items_with_issues,
            "Processing mode": "BATCH" if was_batch_processed else "INDIVIDUAL",
            "Actual cost": f"${estimated_cost:.4f}",
            "Average tokens per item": f"{total_tokens/total_items:.0f}" if total_items > 0 else "0"
        }
    )
    
    # Log full responses if configured
    print(f"\nSTEP 1 COMPLETED!")
    print(f"Successfully processed: {total_items - items_with_issues}/{total_items} LPs")
    print(f"Items with issues: {items_with_issues}")
    print(f"Total script time: {script_duration:.1f}s ({script_duration/60:.1f} minutes)")
    print(f"Processing time: {total_time:.1f}s")
    print(f"Total tokens: {total_tokens:,} (Input: {total_prompt_tokens:,}, Output: {total_completion_tokens:,})")
    print(f"Processing mode: {'BATCH' if was_batch_processed else 'INDIVIDUAL'}")
    print(f"Actual cost: ${estimated_cost:.4f}")
    
    # Show batch savings if applicable
    if was_batch_processed:
        regular_cost = calculate_cost(model_name, total_prompt_tokens, total_completion_tokens, is_batch=False)
        savings = regular_cost - estimated_cost
        savings_percentage = (savings / regular_cost) * 100 if regular_cost > 0 else 0
        print(f"Regular API cost would have been: ${regular_cost:.4f}")
        print(f"Batch savings: ${savings:.4f} ({savings_percentage:.1f}%)")
    
    print(f"Results saved to: {full_output_path}")
    print(f"Token usage log saved to: {os.path.join(logs_folder_path, 'step1_token_usage_log.txt')}")
    print(f"Full responses log saved to: {os.path.join(logs_folder_path, 'step1_llm_responses_log.txt')}")
    
    try:
        batch_summary = create_batch_summary(
            total_items=total_items,
            successful_items=total_items - items_with_issues,
            failed_items=items_with_issues,
            total_time=total_time,
            total_tokens=total_tokens,
            estimated_cost=estimated_cost,
            processing_mode="BATCH" if was_batch_processed else "INDIVIDUAL"
        )
        
        log_processing_metrics(
            results_folder_path=results_folder_path,
            step="step1_metadata_extraction",
            batch_metrics=batch_summary
        )
    except Exception as metrics_error:
        print(f"Warning: Could not log processing metrics: {metrics_error}")

if __name__ == "__main__":
    main()