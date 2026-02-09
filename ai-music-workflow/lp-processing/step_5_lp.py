# Create output files for final results 
import os
import datetime
import re
import openpyxl
import json
from difflib import SequenceMatcher
from openpyxl import load_workbook

# Custom modules
from json_workflow import update_record_step5, log_error, log_processing_metrics
from shared_utilities import find_latest_results_folder, get_workflow_json_path, create_batch_summary, find_latest_lp_metadata_file, get_bib_info_from_workflow
from lp_workflow_config import get_file_path_config, get_threshold_config, get_current_timestamp, get_step_config, FILE_NAMING

current_timestamp = get_current_timestamp()

def get_holdings_info_from_workflow(oclc_number, workflow_json_path):
    """
    Extract holdings information from formatted OCLC results in workflow JSON.
    """
    try:
        with open(workflow_json_path, 'r', encoding='utf-8') as f:
            workflow_data = json.load(f)
        
        # Search through all records for the target OCLC number
        for barcode, record_data in workflow_data.get("records", {}).items():
            step2_data = record_data.get("step2_detailed_data", {})
            formatted_results = step2_data.get("formatted_oclc_results", "")
            
            # Look for this OCLC number and extract holdings info
            oclc_pattern = rf"OCLC Number: {oclc_number}\n\nHeld by IXA: (Yes|No)\nTotal Institutions Holding: (\d+)"
            match = re.search(oclc_pattern, formatted_results)
            
            if match:
                is_held_by_ixa = match.group(1) == "Yes"
                total_holdings = int(match.group(2))
                
                return {
                    "held_by_ixa": is_held_by_ixa,
                    "total_holdings": total_holdings
                }
        
        return {
            "held_by_ixa": False,
            "total_holdings": 0,
            "error": "Holdings data not found in workflow"
        }
        
    except Exception as e:
        return {
            "held_by_ixa": False,
            "total_holdings": 0,
            "error": str(e)
        }

def clean_title(title):
    """Clean up title by removing strange punctuation but keeping slashes."""
    # Replace any double sword or other special characters, but keep slashes
    title = re.sub(r'[^\w\s\/\-\:\;\,\.\(\)\[\]\&\']', ' ', title)
    # Normalize whitespace
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def extract_title_from_bib_info(data):
    """Extract title from workflow bibliographic information."""
    if isinstance(data, dict):
        if "error" in data:
            return "No title available"
        return data.get("title", "No title available")
    return "No title available"

def extract_author_from_bib_info(data):
    """Extract author from workflow bibliographic information."""
    if isinstance(data, dict) and "contributors" in data:
        contributors = data["contributors"]
        if contributors and len(contributors) > 0:
            return contributors[0]  # Return first contributor
        return "No author available"
    return "No author available"

def extract_publication_date_from_bib_info(data):
    """Extract publication date from workflow bibliographic information."""
    if isinstance(data, dict):
        if "error" in data:
            return "No date available"
        return data.get("publication_date", "No date available")
    return "No date available"

def format_bib_info(data):
    """Format bibliographic information for display."""
    output = []
    
    if not isinstance(data, dict) or "identifier" not in data:
        return "No bibliographic information available."
    
    record = data
    
    title = "N/A"
    if "title" in record and "mainTitles" in record["title"]:
        title = record["title"]["mainTitles"][0].get("text", "N/A")
    
    series_title = "N/A"
    if "series" in record and isinstance(record["series"], list) and len(record["series"]) > 0:
        series_title = record["series"][0].get("title", "N/A")
    
    author = "N/A"
    contributors = []
    if "contributor" in record:
        if "creators" in record["contributor"] and record["contributor"]["creators"]:
            for creator in record["contributor"]["creators"]:
                if "nonPersonName" in creator and "text" in creator["nonPersonName"]:
                    author_name = creator["nonPersonName"]["text"]
                    author = author_name
                    contributors.append(author_name)
                elif "firstName" in creator and "secondName" in creator:
                    first_name = creator.get("firstName", {}).get("text", "")
                    second_name = creator.get("secondName", {}).get("text", "")
                    author_name = f"{first_name} {second_name}".strip()
                    author = author_name
                    contributors.append(author_name)
    
    publisher = "N/A"
    place = "N/A"
    date_pub = "N/A"
    if "publishers" in record and record["publishers"]:
        publisher_info = record["publishers"][0]
        if "publisherName" in publisher_info and "text" in publisher_info["publisherName"]:
            publisher = publisher_info["publisherName"]["text"]
        if "publicationPlace" in publisher_info:
            place = publisher_info["publicationPlace"]
    
    if "date" in record and "publicationDate" in record["date"]:
        date_pub = record["date"]["publicationDate"].replace("\u2117", "c")
    
    content_type = "N/A"
    if "format" in record:
        if "generalFormat" in record["format"]:
            content_type = record["format"]["generalFormat"]
        if "specificFormat" in record["format"]:
            content_type += f" - {record['format']['specificFormat']}"
    
    upc = "N/A"
    if "identifier" in record and "otherStandardIdentifiers" in record["identifier"]:
        for identifier in record["identifier"]["otherStandardIdentifiers"]:
            if identifier.get("type") == "Universal Product Code (UPC)":
                upc = identifier.get("id", "N/A")
                break
    
    # Enhanced content parsing logic
    contents = []
    
    # Method 1: Look for contents in the description field
    if "description" in record:
        # Check for the contentNote field in contents array - this is the format in the example
        if "contents" in record["description"]:
            for content_item in record["description"]["contents"]:
                # Check for contentNote object which contains track listings
                if "contentNote" in content_item and "text" in content_item["contentNote"]:
                    content_text = content_item["contentNote"]["text"]
                    # Common pattern: tracks separated by " -- "
                    if " -- " in content_text:
                        # Split by " -- " and clean up each track
                        tracks = []
                        for track in content_text.split(" -- "):
                            track = track.strip()
                            # Remove trailing period if it exists
                            if track.endswith('.'):
                                track = track[:-1].strip()
                            tracks.append(track)
                        contents.extend(tracks)
                        continue
                
                # Check for titles array format (original method)
                if "titles" in content_item:
                    for title_item in content_item["titles"]:
                        if isinstance(title_item, str):
                            contents.append(title_item)
                        elif isinstance(title_item, dict) and "text" in title_item:
                            contents.append(title_item["text"])
                
                # Check for different content formats
                if "items" in content_item:
                    for item in content_item["items"]:
                        if isinstance(item, str):
                            contents.append(item)
                        elif isinstance(item, dict) and "text" in item:
                            contents.append(item["text"])

        # Method 2: Check for TOC as a note
        if "notes" in record["description"]:
            for note in record["description"]["notes"]:
                # Look for various types of content notes
                is_content_note = False
                if "text" in note and any(marker in note["text"].lower() for marker in ["contents:", "tracks:", "track listing"]):
                    is_content_note = True
                
                if is_content_note and "text" in note:
                    toc_text = note["text"]
                    
                    # Try multiple approaches to parse the track list
                    # Approach 1: Split by common delimiters
                    for delimiter in ["--", ";", "/"]:
                        if delimiter in toc_text:
                            # Extract the content part (after any header like "Contents:" if present)
                            content_part = toc_text
                            for header in ["Contents:", "Tracks:", "Track listing:"]:
                                if header.lower() in toc_text.lower():
                                    content_part = toc_text.split(header, 1)[-1].strip()
                                    break
                                
                            # Split by delimiter and clean up
                            parts = [part.strip() for part in content_part.split(delimiter) if part.strip()]
                            if parts:
                                contents.extend(parts)
                                break

    # Method 3: Check for special MARC21 fields often used for music contents
    if "varFields" in record:
        for field in record.get("varFields", []):
            if field.get("marcTag") in ["505", "500"] and "subfields" in field:
                for subfield in field["subfields"]:
                    if subfield.get("code") == "a" and subfield.get("content"):
                        content = subfield["content"]
                        # Split content by common delimiters in track listings
                        for delimiter in ["--", ";", "/"]:
                            if delimiter in content:
                                parts = [part.strip() for part in content.split(delimiter) if part.strip()]
                                if parts:
                                    contents.extend(parts)
                                    break
    
    # Method 4: Check for $t prefixed content
    if "description" in record and "notes" in record["description"]:
        for note in record["description"]["notes"]:
            if "text" in note and "**$t**" in note["text"]:
                parts = note["text"].split("**$t**")
                # Skip the first part as it's usually empty or a header
                for part in parts[1:]:
                    # Clean up each part and add to contents
                    cleaned_part = part.strip()
                    if cleaned_part:
                        # Remove trailing -- if present
                        if cleaned_part.endswith("--"):
                            cleaned_part = cleaned_part[:-2].strip()
                        contents.append(cleaned_part)
    
    oclc_number = "N/A"
    if "identifier" in record and "oclcNumber" in record["identifier"]:
        oclc_number = record["identifier"]["oclcNumber"]
    
    output.append(f"Title: {title}")
    output.append(f"Series Title: {series_title}")
    output.append(f"Author: {author}")
    output.append(f"Contributors: {', '.join(contributors) if contributors else 'N/A'}")
    output.append(f"Publisher: {publisher}")
    output.append(f"Place of Publication: {place}")
    output.append(f"Date of Publication: {date_pub}")
    output.append(f"Content Type: {content_type}")
    output.append(f"UPC: {upc}")
    
    if contents:
        output.append("Contents:")
        for i, track in enumerate(contents, 1):
            # Clean up the track listing
            cleaned_track = track
            # Remove any trailing punctuation
            if cleaned_track.endswith(('.', ';')):
                cleaned_track = cleaned_track[:-1]
            output.append(f"  {i}. {cleaned_track}")
    
    output.append(f"OCLC Number: {oclc_number}")
    
    return "\n".join(output)

def calculate_title_similarity(title1, title2):
    """Calculate similarity between two titles using SequenceMatcher."""
    return SequenceMatcher(None, title1.lower(), title2.lower()).ratio()

def create_low_confidence_review_text_log(results_folder, step4_file, all_records, workflow_json_path, current_timestamp):
    """
    Create a review Excel spreadsheet for unique low confidence matches with detailed information.
    """
    print("Creating low confidence review spreadsheet...")
    
    low_confidence_records = [record for record in all_records 
                             if record["sort_group"] == "Cataloger Review (Low Confidence)"]
    
    if not low_confidence_records:
        print("No low confidence matches found to review.")
        return None
    
    wb_src = load_workbook(step4_file)
    sheet_src = wb_src.active
    
    barcode_to_source = {}
    for row_idx in range(2, sheet_src.max_row + 1):
        barcode = sheet_src.cell(row=row_idx, column=4).value
        if barcode:
            row_data = {
                "barcode": barcode,
                "metadata": sheet_src.cell(row=row_idx, column=5).value,
                "other_oclc_numbers": sheet_src.cell(row=row_idx, column=11).value
            }
            barcode_to_source[barcode] = row_data
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Low Confidence Review"
    
    headers = [
        "Record #",
        "Barcode",
        "OCLC Number Chosen",
        "Confidence Score",
        "AI-Generated Metadata",
        "Other Potential Matches",
        "OCLC Record Details (JSON)",
        "Total Holdings",
        "Held by IXA"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 70
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 70
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    ws.freeze_panes = 'A2'
    
    processed_count = 0
    for idx, record in enumerate(low_confidence_records, start=2):
        barcode = record["barcode"]
        oclc_number = record["oclc_number"]
        
        print(f"Processing low confidence review record {processed_count + 1}/{len(low_confidence_records)} - Barcode: {barcode}")
        
        source_data = barcode_to_source.get(barcode, {})
        metadata = source_data.get("metadata", "No AI-generated metadata available")
        other_oclc_numbers = source_data.get("other_oclc_numbers", "No other candidates")
        
        ws.cell(row=idx, column=1, value=processed_count + 1)
        ws.cell(row=idx, column=2, value=barcode)
        ws.cell(row=idx, column=3, value=oclc_number if oclc_number else 'No OCLC number')
        ws.cell(row=idx, column=4, value=record.get('confidence_score', 'No confidence score'))
        
        metadata_cell = ws.cell(row=idx, column=5, value=metadata if metadata and metadata.strip() else "No AI-generated metadata available")
        metadata_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        other_cell = ws.cell(row=idx, column=6, value=other_oclc_numbers if other_oclc_numbers and other_oclc_numbers.strip() else "No other candidates")
        other_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        if oclc_number and record["has_valid_oclc"]:
            oclc_data = get_bib_info_from_workflow(oclc_number, workflow_json_path)
            
            import json
            raw_oclc_json = json.dumps(oclc_data, indent=2, ensure_ascii=False)
            oclc_cell = ws.cell(row=idx, column=7, value=raw_oclc_json)
            oclc_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            holdings_info = get_holdings_info_from_workflow(oclc_number, workflow_json_path)
            ws.cell(row=idx, column=8, value=holdings_info.get('total_holdings', 0))
            ws.cell(row=idx, column=9, value='Yes' if holdings_info.get('held_by_ixa', False) else 'No')
        else:
            ws.cell(row=idx, column=7, value="No OCLC record available - no valid OCLC number found")
            ws.cell(row=idx, column=8, value=0)
            ws.cell(row=idx, column=9, value='No')
        
        processed_count += 1
    
    deliverables_folder = os.path.join(results_folder, "deliverables")
    review_file = f"low-confidence-matches-review-{current_timestamp}.xlsx"
    review_path = os.path.join(deliverables_folder, review_file)
    
    wb.save(review_path)
    
    print(f"Low confidence review spreadsheet created with {len(low_confidence_records)} records: {review_path}")
    return review_path

def create_marc_format_text_log(results_folder, all_records, workflow_json_path, current_timestamp):
    """
    Create a MARC-formatted Excel spreadsheet from the original JSON metadata for low confidence records.
    """
    print("Creating MARC-formatted spreadsheet from original metadata...")
    
    low_confidence_records = [record for record in all_records 
                             if record["sort_group"] == "Cataloger Review (Low Confidence)"]
    
    if not low_confidence_records:
        print("No low confidence matches found for MARC formatting.")
        return None
    
    import json
    try:
        with open(workflow_json_path, 'r', encoding='utf-8') as f:
            workflow_data = json.load(f)
    except Exception as e:
        print(f"Error reading workflow JSON: {e}")
        return None
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "MARC Format"
    
    headers = [
        "Barcode",
        "100 - Main Entry",
        "245 - Title Statement",
        "264 - Publication",
        "300 - Physical Description",
        "500 - General Note",
        "505 - Contents Note",
        "650 - Subject"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 20
    
    ws.freeze_panes = 'A2'
    
    def is_valid_field(value):
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() and value.strip().lower() != 'not visible'
        return bool(value)
    
    def safe_get(obj, key, default=""):
        if obj is None:
            return default
        value = obj.get(key, default)
        return value if value is not None else default
    
    processed_count = 0
    for idx, record in enumerate(low_confidence_records, start=2):
        barcode = record["barcode"]
        
        workflow_record = workflow_data.get("records", {}).get(barcode)
        if not workflow_record or "step1_metadata_extraction" not in workflow_record:
            ws.cell(row=idx, column=1, value=barcode)
            ws.cell(row=idx, column=2, value="No original metadata available")
            processed_count += 1
            continue
        
        extracted_fields = workflow_record["step1_metadata_extraction"].get("extracted_fields", {})
        
        title_info = extracted_fields.get("title_information", {}) or {}
        main_title = safe_get(title_info, "main_title")
        subtitle = safe_get(title_info, "subtitle") 
        primary_contributor = safe_get(title_info, "primary_contributor")
        
        publishers = extracted_fields.get("publishers", {}) or {}
        place = safe_get(publishers, "place")
        publisher_name_raw = safe_get(publishers, "name")
        publisher_name = publisher_name_raw.replace("Name: ", "").strip() if publisher_name_raw else ""
        
        dates = extracted_fields.get("dates", {}) or {}
        publication_date = safe_get(dates, "publication_date")
        
        physical = extracted_fields.get("physical_description", {}) or {}
        contents = extracted_fields.get("contents", {}) or {}
        tracks = contents.get("tracks", []) or []
        
        ws.cell(row=idx, column=1, value=barcode)
        
        if is_valid_field(primary_contributor):
            field_100 = f"100 1  {primary_contributor}, $ecomposer, $eperformer."
            ws.cell(row=idx, column=2, value=field_100)
        
        if is_valid_field(main_title):
            title_field = f"245 1 0 {main_title}"
            if is_valid_field(subtitle):
                title_field += f" : $b{subtitle}"
            title_field += f" / $c{primary_contributor}." if is_valid_field(primary_contributor) else "."
            ws.cell(row=idx, column=3, value=title_field)
        elif is_valid_field(primary_contributor):
            ws.cell(row=idx, column=3, value=f"245 1 0 [Title not visible] / $c{primary_contributor}.")
        else:
            ws.cell(row=idx, column=3, value="245 1 0 [Title and contributor not visible]")
        
        if is_valid_field(place) or is_valid_field(publisher_name) or is_valid_field(publication_date):
            pub_field = "264  1 "
            
            if is_valid_field(place):
                if is_valid_field(publisher_name):
                    pub_field += f"{place} : "
                else:
                    pub_field += f"{place} "
            
            if is_valid_field(publisher_name):
                pub_field += f"$b{publisher_name}"
                if is_valid_field(publication_date):
                    pub_field += ", "
                else:
                    pub_field += "."
            
            if is_valid_field(publication_date):
                date_clean = (publication_date or "").replace("©", "").replace("℗", "").strip()
                import re
                year_match = re.search(r'\b(19|20)\d{2}\b', date_clean)
                if year_match:
                    date_clean = f"[{year_match.group()}]"
                pub_field += f"$c{date_clean}"
            
            pub_field = pub_field.rstrip(', ') + "."
            ws.cell(row=idx, column=4, value=pub_field)
        
        ws.cell(row=idx, column=5, value="300    1 audio disc : $banalog ; $c12 in.")

        ws.cell(row=idx, column=6, value="340    vinyl.")
        
        if tracks and isinstance(tracks, list):
            track_list = []
            for track in tracks:
                if isinstance(track, dict):
                    track_title = safe_get(track, "title")
                    if is_valid_field(track_title):
                        if not any(x in track_title.lower() for x in ['standard lp', 'vinyl record', '12', 'vinyl', 'pvc']):
                            track_list.append(track_title)
                elif isinstance(track, str) and is_valid_field(track):
                    if not any(x in track.lower() for x in ['standard lp', 'vinyl record', '12', 'vinyl', 'pvc']):
                        track_list.append(track)
            
            if track_list:
                contents_field = "505 0  " + " -- ".join(track_list) + "."
                contents_cell = ws.cell(row=idx, column=7, value=contents_field)
                contents_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        ws.cell(row=idx, column=8, value="650  0  $aMusic.")
        
        processed_count += 1
    
    deliverables_folder = os.path.join(results_folder, "deliverables")
    marc_file = f"low-confidence-marc-{current_timestamp}.xlsx"
    marc_path = os.path.join(deliverables_folder, marc_file)
    
    wb.save(marc_path)
    
    print(f"MARC format spreadsheet created with {processed_count} records: {marc_path}")
    return marc_path

def create_cataloger_review_spreadsheet(results_folder, all_records, current_timestamp):
    """
    Create a separate Excel workbook for catalogers to review low confidence matches.
    """
    print("Creating cataloger review spreadsheet...")
    
    # Filter for low confidence records only
    low_confidence_records = [record for record in all_records 
                             if record["sort_group"] == "Cataloger Review (Low Confidence)"]
    
    if not low_confidence_records:
        print("No low confidence matches found for cataloger review.")
        return None
    
    # Create new workbook
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cataloger Review"
    
    # Define headers
    headers = [
        "Barcode",
        "Date AI Processed", 
        "AI-Suggested OCLC Number",
        "Title",
        "Date Cataloger Checked",
        "Status",
        "Correct OCLC Number",
        "Notes"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 16  # Barcode
    ws.column_dimensions['B'].width = 18  # Date AI Processed
    ws.column_dimensions['C'].width = 22  # AI Suggested OCLC
    ws.column_dimensions['D'].width = 55  # Title
    ws.column_dimensions['E'].width = 22  # Date Cataloger Checked
    ws.column_dimensions['F'].width = 25  # Status
    ws.column_dimensions['G'].width = 22  # Correct OCLC
    ws.column_dimensions['H'].width = 40  # Notes
    
    # Add data rows
    for row_num, record in enumerate(low_confidence_records, start=2):
        ws.cell(row=row_num, column=1, value=record["barcode"])
        ws.cell(row=row_num, column=2, value=current_timestamp)
        
        # Show the AI-suggested OCLC number (what the workflow chose)
        ai_suggested_oclc = record["oclc_number"] if record["oclc_number"] else "None suggested"
        ws.cell(row=row_num, column=3, value=ai_suggested_oclc)
        
        # Show the title from OCLC record
        ws.cell(row=row_num, column=4, value=record["title"])
        
        # Columns 5, 6 left empty for cataloger input (Date Cataloger Checked, Status)
        
        # Column 7 (G) - Auto-populate formula for Correct OCLC Number
        # If status is "Approved", use the AI-suggested number, otherwise leave blank for manual entry
        ws.cell(row=row_num, column=7, value=f'=IF(F{row_num}="Approved",C{row_num},"")')
        
        # Column 8 (H) left empty for cataloger notes
    
    # Create dropdown validation for Status column - alternative approach
    from openpyxl.worksheet.datavalidation import DataValidation
    
    dv = DataValidation(type="list", formula1='"Approved,Rejected - Different OCLC,Rejected - Needs Original Cataloging"')
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please pick from the list'
    dv.promptTitle = 'List Selection'
    
    # Add validation to the Status column range
    ws.add_data_validation(dv)
    dv.add(f'F2:F{len(low_confidence_records) + 1}')
    
    # Apply conditional formatting for highlighting
    from openpyxl.formatting.rule import FormulaRule
        
    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
    rule = FormulaRule(
        formula=['AND($F2<>"Approved", $G2="")'],
        fill=highlight_fill
    )
        
    # Apply to all data rows
    max_row = len(low_confidence_records) + 1
    data_range = f'A2:H{max_row}'
    ws.conditional_formatting.add(data_range, rule)
    
    # Save the workbook in deliverables subfolder
    deliverables_folder = os.path.join(results_folder, "deliverables")
    review_file = f"tracking-spreadsheet-catalogers-{current_timestamp}.xlsx"
    review_path = os.path.join(deliverables_folder, review_file)
    
    wb.save(review_path)
    
    print(f"Cataloger review spreadsheet created with {len(low_confidence_records)} records: {review_path}")
    return review_path

def move_workflow_data_files(results_folder, data_folder):
    """Move and rename JSON and Excel workflow files to data subfolder"""
    import shutil
    moved_files = 0
    try:
        print(f"Looking for workflow files in: {results_folder}")
        print(f"Target data folder: {data_folder}")
        
        if not os.path.exists(results_folder):
            print(f"Source folder does not exist: {results_folder}")
            return
            
        files_in_results = os.listdir(results_folder)
        print(f"Files in results folder: {len(files_in_results)}")
                
        for filename in files_in_results:
            src = os.path.join(results_folder, filename)
            
            # Move JSON workflow files that start with full-workflow-data-lp-
            if filename.startswith("full-workflow-data-lp-") and filename.endswith(".json"):
                dst = os.path.join(data_folder, filename)
                if os.path.exists(src) and os.path.isfile(src):
                    shutil.move(src, dst)
                    print(f"Moved workflow JSON to: {dst}")
                    moved_files += 1
            # Move JSON workflow files that start with full-workflow-data-lp-
            if filename.startswith("full-workflow-data-lp-") and filename.endswith(".xlsx"):
                dst = os.path.join(data_folder, filename)
                if os.path.exists(src) and os.path.isfile(src):
                    shutil.move(src, dst)
                    print(f"Moved workflow Excel to: {dst}")
                    moved_files += 1
        
        print(f"Successfully moved {moved_files} workflow data files")
                    
    except Exception as e:
        print(f"Warning: Could not move workflow data files: {e}")
        import traceback
        traceback.print_exc()
    
def find_duplicate_groups(all_records, similarity_threshold=0.9, confidence_threshold=80):
    """
    Find groups of duplicate records based on similar OCLC numbers or titles.
    Only includes records with confidence >= confidence_threshold in duplicate detection.
    
    Args:
        all_records: List of all record dictionaries
        similarity_threshold: Threshold for determining similar titles
        confidence_threshold: Minimum confidence to consider for duplicate detection
    
    Returns:
        List of lists, where each inner list contains duplicate records
    """
    # Only consider records with confidence >= threshold for duplicate detection
    high_confidence_records = []
    high_confidence_indices = []
    
    for i, record in enumerate(all_records):
        try:
            confidence = float(record.get("confidence_score", 0))
            if confidence >= confidence_threshold:
                high_confidence_records.append(record)
                high_confidence_indices.append(i)
        except (ValueError, TypeError):
            # If confidence can't be parsed, treat as low confidence
            pass
    
    duplicate_groups = []
    processed_indices = set()
    
    for i, record in enumerate(high_confidence_records):
        original_index = high_confidence_indices[i]
        if original_index in processed_indices:
            continue
            
        oclc_number = str(record.get("oclc_number", "")).strip()
        title = record.get("title", "")
        
        # Find all duplicates for this record (only among high confidence records)
        duplicates = [record]
        duplicate_original_indices = [original_index]
        
        for j, other_record in enumerate(high_confidence_records[i+1:], i+1):
            other_original_index = high_confidence_indices[j]
            if other_original_index in processed_indices:
                continue
                
            other_oclc = str(other_record.get("oclc_number", "")).strip()
            other_title = other_record.get("title", "")
            
            is_duplicate = False
            
            # Check for similar OCLC numbers (might indicate duplicates)
            try:
                if abs(int(oclc_number) - int(other_oclc)) <= 5:
                    is_duplicate = True
            except (ValueError, TypeError):
                pass
            
            # Check for very similar titles
            if not is_duplicate and title and other_title:
                similarity = calculate_title_similarity(title, other_title)
                if similarity >= similarity_threshold:
                    is_duplicate = True
            
            if is_duplicate:
                duplicates.append(other_record)
                duplicate_original_indices.append(other_original_index)
        
        # Mark all indices as processed
        for idx in duplicate_original_indices:
            processed_indices.add(idx)
        
        # Only add to duplicate groups if there are actually duplicates
        if len(duplicates) > 1:
            duplicate_groups.append(duplicates)
    
    return duplicate_groups

def determine_sort_group_for_duplicates(duplicate_group, confidence_threshold=80):
    """
    Determine sort groups for a group of duplicate records.
    Priority: IXA held > highest confidence > first encountered
    
    Args:
        duplicate_group: List of duplicate record dictionaries
        confidence_threshold: Threshold for high confidence matches
    
    Returns:
        None (modifies records in place)
    """
    if len(duplicate_group) <= 1:
        return
    
    # Sort by priority: IXA held first, then by confidence score (highest first), then by original order
    def sort_key(record):
        held_by_ixa = record.get("held_by_ixa", False)
        try:
            confidence = float(record.get("confidence_score", 0))
        except (ValueError, TypeError):
            confidence = 0
        
        # Return tuple for sorting: (not held_by_ixa, -confidence)
        # This puts IXA held items first, then highest confidence
        return (not held_by_ixa, -confidence)
    
    sorted_duplicates = sorted(duplicate_group, key=sort_key)
    
    # The first item (highest priority) gets its normal classification
    primary_record = sorted_duplicates[0]
    
    # Determine classification for primary record
    if primary_record.get("held_by_ixa", False):
        # For IXA held items, check confidence to determine if it's truly the same item
        try:
            conf_score = float(primary_record.get("confidence_score", 0))
            if conf_score >= confidence_threshold:
                primary_record["sort_group"] = "Held by UT Libraries (IXA)"
            else:
                # Low confidence IXA match - treat as uncertain, not as held item
                primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
        except (ValueError, TypeError):
            primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
    else:
        try:
            conf_score = float(primary_record.get("confidence_score", 0))
            if conf_score >= confidence_threshold:
                primary_record["sort_group"] = "Alma Batch Upload (High Confidence)"
            else:
                primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
        except (ValueError, TypeError):
            primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
    
    # All other records in the group are marked as duplicates
    for record in sorted_duplicates[1:]:
        record["sort_group"] = "Duplicate"

def determine_sort_group(record, confidence_threshold=80):
    """
    Determine the sort group for a non-duplicate record.
    
    Args:
        record: Dictionary containing record information
        confidence_threshold: Threshold for high confidence matches
    
    Returns:
        String indicating the sort group
    """
    held_by_ixa = record.get("held_by_ixa", False)
    confidence_score = record.get("confidence_score", 0)
    
    # Check confidence level first - low confidence items are always for review
    try:
        conf_score = float(confidence_score)
        if conf_score < confidence_threshold:
            return "Cataloger Review (Low Confidence)"
    except (ValueError, TypeError):
        return "Cataloger Review (Low Confidence)"
    
    # Only for high confidence items, check if held by IXA
    if held_by_ixa:
        return "Held by UT Libraries (IXA)"
    else:
        return "Alma Batch Upload (High Confidence)"

def create_all_records_spreadsheet():
    # Get configuration
    file_paths = get_file_path_config()
    step5_config = get_step_config("step5")
    threshold_config = get_threshold_config("confidence")
    
    # Find latest results folder using new utility
    results_folder = find_latest_results_folder(file_paths["results_prefix"])
    if not results_folder:
        print("No results folder found! Please run the previous scripts first.")
        return None
        
    print(f"Using results folder: {results_folder}")
    
    # Create subfolders for organized file output
    guides_folder = os.path.join(results_folder, "guides")
    deliverables_folder = os.path.join(results_folder, "deliverables") 
    data_folder = os.path.join(results_folder, "data")

    # Create directories if they don't exist
    for folder in [guides_folder, deliverables_folder, data_folder]:
        os.makedirs(folder, exist_ok=True)
    
    # Initialize workflow JSON path
    workflow_json_path = get_workflow_json_path(results_folder)
    
    step4_file = find_latest_lp_metadata_file(results_folder)
    if not step4_file:
        print("No full-workflow-data-lp file found in the results folder!")

    print(f"Using source file: {step4_file}")

    try: 
        # Open the latest step 4 workbook
        wb_src = load_workbook(step4_file)
        sheet_src = wb_src.active

        # Create a new workbook for all records
        wb_new = openpyxl.Workbook()
        sheet_new = wb_new.active

        # Set header row
        header_row = ["Barcode", "Sort Group", "OCLC Number", "OCLC Title", "OCLC Author", "OCLC Date of Publication", "Confidence Score"]
        sheet_new.append(header_row)

        # Set column widths
        sheet_new.column_dimensions['A'].width = 16  # Barcode
        sheet_new.column_dimensions['B'].width = 28  # Sort Group
        sheet_new.column_dimensions['C'].width = 12  # OCLC Number
        sheet_new.column_dimensions['D'].width = 50  # OCLC Title
        sheet_new.column_dimensions['E'].width = 50  # OCLC Author
        sheet_new.column_dimensions['F'].width = 20  # OCLC Date
        sheet_new.column_dimensions['G'].width = 15  # Confidence Score

        # Get the column indices from the source workbook
        # In source: Column D=Barcode, Column H=OCLC Number, Column I=Confidence Score
        BARCODE_COL_IDX = 4  # Column D
        OCLC_NUM_COL_IDX = 8  # Column H
        CONF_SCORE_COL_IDX = 9  # Column I

        # First pass: collect all records with basic info
        all_records = []
        all_records_dict = {}  # For duplicate detection
        
        print("First pass: Collecting all records...")
        for row_idx in range(2, sheet_src.max_row + 1):  # Skip header row
            barcode = sheet_src.cell(row=row_idx, column=BARCODE_COL_IDX).value
            oclc_number = sheet_src.cell(row=row_idx, column=OCLC_NUM_COL_IDX).value
            confidence_score = sheet_src.cell(row=row_idx, column=CONF_SCORE_COL_IDX).value
            
            # Skip rows with missing barcode (essential identifier)
            if not barcode:
                continue
            
            # Handle records with no OCLC number or invalid OCLC numbers
            oclc_number_str = ""
            has_valid_oclc = False
            
            if oclc_number and str(oclc_number).strip() not in ["", "Not found", "Error processing"]:
                oclc_number_str = str(oclc_number).strip()
                has_valid_oclc = True
            
            record = {
                "barcode": barcode,
                "oclc_number": oclc_number_str,
                "confidence_score": confidence_score if confidence_score else 0,
                "title": "",
                "author": "",
                "publication_date": "",
                "held_by_ixa": False,
                "sort_group": "",
                "has_valid_oclc": has_valid_oclc
            }
            
            all_records.append(record)
            
            # Only add to dict for duplicate detection if it has a valid OCLC number
            if has_valid_oclc:
                all_records_dict[oclc_number_str] = record
            
        print(f"Found {len(all_records)} records to process.")

        # Second pass: Get OCLC data and holdings information from workflow JSON
        print("Second pass: Getting OCLC data from workflow JSON...")
        processed_count = 0
        
        for record in all_records:
            oclc_number = record["oclc_number"]
            has_valid_oclc = record["has_valid_oclc"]
            
            print(f"Processing record {processed_count + 1}/{len(all_records)} - Barcode: {record['barcode']}")
            
            if has_valid_oclc:
                print(f"  OCLC Number: {oclc_number}")
                
                # Get bibliographic information from workflow JSON
                oclc_data = get_bib_info_from_workflow(oclc_number, workflow_json_path)
                print(f"  Retrieved OCLC data: {oclc_data}")  
                record["title"] = extract_title_from_bib_info(oclc_data)
                record["author"] = extract_author_from_bib_info(oclc_data)
                record["publication_date"] = extract_publication_date_from_bib_info(oclc_data)
                
                # Get holdings information from workflow JSON
                holdings_info = get_holdings_info_from_workflow(oclc_number, workflow_json_path)
                record["held_by_ixa"] = holdings_info["held_by_ixa"]
                
                print(f"  Holdings from workflow: IXA={holdings_info['held_by_ixa']}")
                
                # No API delay needed since we're reading from local JSON
            else:
                print(f"  No valid OCLC number - using defaults")
                record["title"] = "No OCLC match found"
                record["author"] = "No author available"
                record["publication_date"] = "No date available"
                record["held_by_ixa"] = False
            
            processed_count += 1
        
        # Third pass: Handle duplicates and determine sort groups
        print("Third pass: Finding duplicate groups...")
        
        # Only look for duplicates among records with valid OCLC numbers AND high confidence
        records_with_oclc = [record for record in all_records if record["has_valid_oclc"]]
        duplicate_groups = find_duplicate_groups(records_with_oclc, confidence_threshold=80)
        
        print(f"Found {len(duplicate_groups)} duplicate groups (high confidence only)")
        for i, group in enumerate(duplicate_groups):
            print(f"  Group {i+1}: {len(group)} duplicates")
        
        # Process duplicate groups
        print("Processing duplicate groups...")
        for group in duplicate_groups:
            determine_sort_group_for_duplicates(group, confidence_threshold=80)
        
        # Process remaining records (including low confidence items that were excluded from duplicate detection)
        print("Processing non-duplicate records...")
        processed_records = set()
        for group in duplicate_groups:
            for record in group:
                processed_records.add(id(record))
        
        for record in all_records:
            if id(record) not in processed_records:
                if record["has_valid_oclc"]:
                    record["sort_group"] = determine_sort_group(record, confidence_threshold=80)
                else:
                    # Records without valid OCLC numbers are always low confidence
                    record["sort_group"] = "Cataloger Review (Low Confidence)"
        
        # Add all records to spreadsheet and log to JSON workflow
        for record in all_records:
            sheet_new.append([
                record["barcode"],
                record["sort_group"],
                record["oclc_number"],
                record["title"],
                record["author"],
                record["publication_date"],
                record["confidence_score"]
            ])
            
            # Update JSON workflow with Step 5 results
            try:
                is_duplicate = record["sort_group"] == "Duplicate"
                update_record_step5(
                    json_path=workflow_json_path,
                    barcode=str(record["barcode"]),
                    sort_group=record["sort_group"],
                    final_oclc_number=record["oclc_number"],
                    is_duplicate=is_duplicate,
                    oclc_title=record["title"],
                    oclc_author=record["author"],
                    oclc_date=record["publication_date"]
                )
            except Exception as json_error:
                print(f"   JSON logging error for {record['barcode']}: {json_error}")
                log_error(
                    results_folder_path=results_folder,
                    step="step5_final_classification",
                    barcode=str(record["barcode"]),
                    error_type="json_update_error",
                    error_message=str(json_error)
                )
        
        # Save the all records spreadsheet
        all_records_file = f"lp-workflow-sorting-{current_timestamp}.xlsx"
        all_records_path = os.path.join(deliverables_folder, all_records_file)
        wb_new.save(all_records_path)
        
        print(f"Sort physical items with {len(all_records)} records: {all_records_path}")
        
        # Create summary statistics
        sort_group_counts = {}
        for record in all_records:
            group = record["sort_group"]
            sort_group_counts[group] = sort_group_counts.get(group, 0) + 1
        
        print("\nSort Group Summary:")
        for group, count in sorted(sort_group_counts.items()):
            print(f"  {group}: {count} records")
        
        # Create high confidence matches for Alma batch upload (text file only)
        unique_matches = [record for record in all_records 
                         if record["sort_group"] == "Alma Batch Upload (High Confidence)"]
        
        text_file = FILE_NAMING["batch_upload_alma"].format(timestamp=current_timestamp)
        text_path = os.path.join(deliverables_folder, text_file)
        
        with open(text_path, 'w', newline='', encoding='utf-8') as f:
            for record in unique_matches:
                line = f"{record['oclc_number']}|{record['barcode']}|{record['title']}\n"
                f.write(line)
        
        print(f"Pipe-delimited text file for Alma processing created: {text_path}")
        
        # Create low confidence review spreadsheet
        review_path = create_low_confidence_review_text_log(
            results_folder, step4_file, all_records, workflow_json_path, current_timestamp
        )
        
        # Create MARC format text log for low confidence records
        marc_path = create_marc_format_text_log(
            results_folder, all_records, workflow_json_path, current_timestamp
        )
        
        # Create cataloger review spreadsheet
        review_spreadsheet_path = create_cataloger_review_spreadsheet(
            results_folder, all_records, current_timestamp
        )

        # Copy both guides to guides subfolder
        try:
            import shutil
            script_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Try multiple potential locations for the guides
            potential_locations = [
                os.path.dirname(os.path.dirname(script_dir)),  # Two levels up
                os.path.dirname(script_dir),  # One level up  
                script_dir,  # Same directory as script
                os.getcwd()  # Current working directory
            ]
            
            guides_found = False
            for project_root in potential_locations:
                cataloger_guide_source = os.path.join(project_root, "CATALOGER_GUIDE.txt")
                technical_guide_source = os.path.join(project_root, "TECHNICAL_GUIDE.txt")
                
                if os.path.exists(cataloger_guide_source):
                    cataloger_guide_dest = os.path.join(guides_folder, "CATALOGER_GUIDE.txt")
                    shutil.copy2(cataloger_guide_source, cataloger_guide_dest)
                    print(f"Cataloger guide copied to: {cataloger_guide_dest}")
                    guides_found = True
                
                if os.path.exists(technical_guide_source):
                    technical_guide_dest = os.path.join(guides_folder, "TECHNICAL_GUIDE.txt")
                    shutil.copy2(technical_guide_source, technical_guide_dest)
                    print(f"Technical guide copied to: {technical_guide_dest}")
                    guides_found = True
                
                if guides_found:
                    break
            
            if not guides_found:
                print("Warning: Guide files not found in any expected location")
                
        except Exception as e:
            print(f"Warning: Could not copy guides: {e}")
            
        # Log final Step 5 processing metrics
        try:
            step5_metrics = create_batch_summary(
                total_items=len(all_records),
                successful_items=len([r for r in all_records if r["sort_group"] != ""]),
                failed_items=len([r for r in all_records if r["sort_group"] == ""]),
                total_time=0,  # Step 5 doesn't track detailed timing
                total_tokens=0,
                estimated_cost=0,
                processing_mode="CLASSIFICATION"
            )
            
            # Add step-specific metrics
            step5_metrics.update({
                "sort_group_counts": sort_group_counts,
                "unique_matches_count": len(unique_matches),
                "duplicate_groups_found": len(duplicate_groups) if 'duplicate_groups' in locals() else 0,
                "records_with_valid_oclc": len([r for r in all_records if r["has_valid_oclc"]]),
                "step": "step5_final_classification"
            })
            
            log_processing_metrics(
                results_folder_path=results_folder,
                step="step5_final_classification", 
                batch_metrics=step5_metrics
            )
            
        except Exception as metrics_error:
            print(f"Warning: Could not log Step 5 processing metrics: {metrics_error}")

        # Close the source workbook to release file handles before moving
        wb_src.close()

        # Move workflow data files to data subfolder after all processing is complete
        print("Moving workflow data files to data subfolder...")
        move_workflow_data_files(results_folder, data_folder)
            
        return {
            "all_records_path": all_records_path,
            "text_file_path": text_path,
            "review_path": review_path,
            "marc_path": marc_path,
            "review_spreadsheet_path": review_spreadsheet_path,
            "guide_path": os.path.join(guides_folder, "CATALOGER_GUIDE.txt"),
            "total_records": len(all_records),
            "sort_group_counts": sort_group_counts,
            "unique_matches_count": len(unique_matches)
        }
        
    except Exception as e:
        print(f"Error creating all records spreadsheet: {str(e)}")
        
        # Log error to JSON workflow
        try:
            log_error(
                results_folder_path=results_folder,
                step="step5_final_classification",
                barcode="unknown",
                error_type="processing_error",
                error_message=str(e)
            )
        except Exception as json_error:
            print(f"JSON error logging failed: {json_error}")
        
        return None

def main():
    result = create_all_records_spreadsheet()
    if result:
        print(f"\n=== All Records Processing Complete ===")
        print(f"Total records processed: {result['total_records']}")
        print(f"All records file: {result['all_records_path']}")
        print(f"Alma processing file: {result['text_file_path']}")
        if result['review_path']:
            print(f"Low confidence review file: {result['review_path']}")
        if result.get('review_spreadsheet_path'):
            print(f"Cataloger review spreadsheet: {result['review_spreadsheet_path']}")
        if result.get('marc_path'):
            print(f"MARC format file: {result['marc_path']}")
        print(f"Unique high-confidence matches: {result['unique_matches_count']}")
        print("\nBreakdown by Sort Group:")
        for group, count in sorted(result['sort_group_counts'].items()):
            print(f"  {group}: {count}")
    else:
        print("Failed to create all records spreadsheet.")

if __name__ == "__main__":
    main()