"""
Step 6: Create Interactive HTML Review Interface for LP Records
Creates paginated HTML files with images for cataloger review of LP metadata matches.
This step is optional and can be skipped for large batches where HTML generation is impractical.
"""

import os
import math
import shutil
from openpyxl import load_workbook

# Custom modules
from shared_utilities import find_latest_results_folder, get_workflow_json_path, get_bib_info_from_workflow, find_latest_lp_metadata_file
from lp_workflow_config import get_file_path_config, get_current_timestamp, get_current_date


def create_paginated_review_html(results_folder, all_records, current_date, workflow_json_path, records_per_page=100):
    """
    Create paginated HTML files with external images and lazy loading for large datasets.
    All HTML files will be in the same folder for maximum compatibility.
    """
    print(f"Creating paginated review with {records_per_page} records per page...")
    
    # Get file path config to find images folder
    file_paths = get_file_path_config()
    images_folder = file_paths["images_folder"]
    
    # Calculate number of pages needed
    total_pages = math.ceil(len(all_records) / records_per_page)
    page_files = []
    
    # Create index page path (in results folder)
    index_file = f"review-index-{current_date}.html"
    index_path = os.path.join(results_folder, index_file)
    
    # Group records by sort group for better organization
    sort_groups = {}
    for record in all_records:
        group = record.get("sort_group", "Unknown")
        if group not in sort_groups:
            sort_groups[group] = []
        sort_groups[group].append(record)
    
    # Create index page
    create_review_index(index_path, sort_groups, current_date, total_pages, records_per_page)
    
    # Create individual pages in the same folder as index
    for page_num in range(1, total_pages + 1):
        start_idx = (page_num - 1) * records_per_page
        end_idx = min(start_idx + records_per_page, len(all_records))
        page_records = all_records[start_idx:end_idx]
        
        # Put page files directly in results folder alongside index
        page_file = f"review-page-{page_num}-{current_date}.html"
        page_path = os.path.join(results_folder, page_file)

        create_single_review_page(
            page_path, page_records, current_date, workflow_json_path, 
            images_folder, results_folder, page_num, total_pages, records_per_page, start_idx
        )
        
        page_files.append(page_path)
        print(f"Created page {page_num}/{total_pages} with {len(page_records)} records")
        
    print(f"Images copied to images subfolder for portability")
    print(f"All HTML files are in the same folder for reliable navigation")
    print(f"To share: Send entire '{os.path.basename(results_folder)}' folder")
    
    return {
        "index_path": index_path,
        "page_files": page_files,
        "total_pages": total_pages
    }

def create_review_index(index_path, sort_groups, current_date, total_pages, records_per_page):
    """Create an index page with links to all review pages and sort group summaries."""
    
    total_records = sum(len(records) for records in sort_groups.values())
    
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>LP Review Index - {current_date}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .header {{ background-color: #2c3e50; color: white; padding: 20px; border-radius: 5px; margin-bottom: 30px; }}
        .summary {{ background-color: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .page-links {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 15px; margin-top: 20px; }}
        .page-link {{ background-color: #3498db; color: white; padding: 15px; text-decoration: none; border-radius: 5px; text-align: center; font-weight: bold; }}
        .page-link:hover {{ background-color: #2980b9; }}
        .sort-group {{ margin: 10px 0; padding: 10px; background-color: #f8f9fa; border-left: 4px solid #3498db; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>LP Cataloger Review Index</h1>
        <p>Generated: {current_date} | Total Records: {total_records} | Pages: {total_pages}</p>
    </div>
    
    <div class="summary">
        <h2>Sort Group Summary</h2>"""
    
    for group, records in sorted(sort_groups.items()):
        html_content += f'<div class="sort-group"><strong>{group}:</strong> {len(records)} records</div>'
    
    html_content += f"""
    </div>
    
    <div class="summary">
        <h2>Review Pages</h2>
        <p>Each page contains up to {records_per_page} records for manageable review.</p>
        <p><strong>All files are in the same folder for reliable navigation.</strong></p>
        <div class="page-links">"""
    
    for page_num in range(1, total_pages + 1):
        start_record = (page_num - 1) * records_per_page + 1
        end_record = min(page_num * records_per_page, total_records)
        page_filename = f"review-page-{page_num}-{current_date}.html"
        
        html_content += f'''
            <a href="{page_filename}" class="page-link">
                Page {page_num}<br>
                Records {start_record}-{end_record}
            </a>'''
    
    html_content += """
        </div>
    </div>
    
    <div class="summary">
        <h2>Export All Decisions</h2>
        <p>Export decisions from all pages at once (requires visiting each page first).</p>
        <button onclick="exportAllDecisions()" style="background: #e74c3c; color: white; border: none; padding: 15px 30px; border-radius: 5px; cursor: pointer; font-weight: bold; font-size: 16px;">
            Export All Decisions to CSV
        </button>
    </div>
    
    <script>
        // Create unique storage namespace for this workflow
        const STORAGE_PREFIX = 'lp-workflow-""" + current_date + """-';
        
        function exportAllDecisions() {
            const catalogerName = prompt('Enter your name for the export file:');
            if (!catalogerName) return;

            const rows = [];
            // Only look at keys for THIS workflow
            for (let i = 0; i < window.localStorage.length; i++) {
                const fullKey = window.localStorage.key(i) || '';
                
                // Skip keys that don't belong to this workflow
                if (!fullKey.startsWith(STORAGE_PREFIX)) continue;
                
                // Now match within our prefix
                const localKey = fullKey.replace(STORAGE_PREFIX, '');
                const m = localKey.match(/^decision-(\d+)$/);
                if (!m) continue;

                const recordId = m[1];
                const decision = window.localStorage.getItem(fullKey);
                const notes = window.localStorage.getItem(STORAGE_PREFIX + 'notes-' + recordId) || '';

                let recordData = null;
                try {
                    const raw = window.localStorage.getItem(STORAGE_PREFIX + 'record-data-' + recordId);
                    if (raw) recordData = JSON.parse(raw);
                } catch (e) {
                    console.log('Error parsing record-data for', recordId, e);
                }

                const barcode    = (recordData && recordData.barcode)    ? recordData.barcode    : ('Record-' + recordId);
                const confidence = (recordData && recordData.confidence) ? recordData.confidence : 'N/A';
                const sortGroup  = (recordData && recordData.sortGroup)  ? recordData.sortGroup  : 'N/A';
                const pageNumber = (recordData && recordData.pageNumber) ? recordData.pageNumber : 'Unknown';

                let correctOclc = '';
                if (decision === 'approved' && recordData && recordData.oclcNumber) {
                    correctOclc = recordData.oclcNumber;
                }

                const decisionLabels = {
                    'approved': 'Approved',
                    'different': 'Different OCLC # Needed',
                    'original': 'Original Cataloging Needed',
                    'review': 'Further Review Needed'
                };

                rows.push({
                    recordId: recordId,
                    barcode: barcode,
                    confidence: confidence,
                    sortGroup: sortGroup,
                    decision: decisionLabels[decision] || decision,
                    aiSuggestedOclc: (recordData && recordData.oclcNumber) ? recordData.oclcNumber : '',
                    correctOclc: correctOclc,
                    notes: notes,
                    cataloger: catalogerName,
                    reviewDate: new Date().toISOString().split('T')[0],
                    pageNumber: pageNumber
                });
            }

            if (rows.length === 0) {
                alert('No decisions found. Open a review page, make a decision, then try again.');
                return;
            }

            rows.sort((a, b) => (parseInt(a.recordId) || 0) - (parseInt(b.recordId) || 0));

            const headers = ['Record', 'Barcode', 'Confidence', 'Initial Sort Group', 'Cataloger Decision', 'AI-Suggested OCLC #', 'Correct OCLC #', 'Notes', 'Cataloger', 'Review Date', 'Page Number'];
            const esc = (s) => {
                const str = String(s == null ? '' : s);
                return /[",\\n\\r]/.test(str) ? '"' + str.replace(/"/g, '""') + '"' : str;
            };

            let csv = headers.join(',') + '\\n';
            for (const r of rows) {
                csv += [
                    r.recordId,
                    esc(r.barcode),
                    esc(r.confidence),
                    esc(r.sortGroup),
                    esc(r.decision),
                    esc(r.aiSuggestedOclc),
                    esc(r.correctOclc),
                    esc(r.notes),
                    esc(r.cataloger),
                    r.reviewDate,
                    r.pageNumber
                ].join(',') + '\\n';
            }

            const blob = new Blob([csv], { type: 'text/csv;charset=utf-8' });
            const a = document.createElement('a');
            a.href = URL.createObjectURL(blob);
            a.download = 'all-cataloger-decisions-' + catalogerName.replace(/[^a-zA-Z0-9]/g, '_') + '-' + new Date().toISOString().split('T')[0] + '.csv';
            document.body.appendChild(a);
            a.click();
            setTimeout(function () {
                URL.revokeObjectURL(a.href);
                a.remove();
            }, 0);

            alert('Exported ' + rows.length + ' decisions to CSV file.');
        }
    </script>
</body>
</html>"""
    
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

def create_single_review_page(page_path, page_records, current_date, workflow_json_path, images_folder, results_folder, page_num, total_pages, records_per_page, start_idx):
    """Create a single review page with direct image loading."""
    
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>LP Review Page {page_num} - {current_date}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .header {{ background-color: #2c3e50; color: white; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
        .navigation {{ background-color: white; padding: 15px; border-radius: 5px; margin-bottom: 20px; text-align: center; }}
        .nav-btn {{ background-color: #3498db; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin: 0 10px; font-weight: bold; }}
        .nav-btn:hover {{ background-color: #2980b9; }}
        .nav-btn.disabled {{ background-color: #95a5a6; pointer-events: none; }}
        .record {{ background-color: white; border: 1px solid #ddd; border-radius: 8px; margin-bottom: 30px; padding: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .record-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #eee; }}
        .barcode {{ font-size: 24px; font-weight: bold; color: #2c3e50; }}
        .confidence {{ font-size: 18px; font-weight: bold; padding: 8px 15px; border-radius: 20px; color: white; }}
        .sort-group {{ font-size: 16px; font-weight: bold; padding: 6px 12px; border-radius: 15px; margin-left: 10px; }}
        .group-alma-batch-upload-high-confidence {{ background-color: #27ae60; color: white; }}
        .group-cataloger-review-low-confidence {{ background-color: #e74c3c; color: white; }}
        .group-held-by-ut-libraries-ixa {{ background-color: #3498db; color: white; }}
        .group-duplicate {{ background-color: #95a5a6; color: white; }}
        .group-unknown {{ background-color: #95a5a6; color: white; }}
        .confidence-low {{ background-color: #e74c3c; }}
        .confidence-medium {{ background-color: #f39c12; }}
        .confidence-high {{ background-color: #27ae60; }}
        .content-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .images-section {{ display: flex; flex-direction: column; gap: 15px; }}
        .image-container {{ text-align: center; }}
        .image-container img {{ max-width: 100%; height: auto; max-height: 500px; border: 2px solid #ddd; border-radius: 5px; cursor: pointer; transition: transform 0.2s; object-fit: contain; }}
        .image-container img:hover {{ transform: scale(1.05); border-color: #3498db; }}
        .image-label {{ font-weight: bold; margin-bottom: 5px; color: #555; }}
        .oclc-section {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; max-height: 80vh; overflow-y: auto; }}
        .oclc-field {{ margin-bottom: 10px; }}
        .oclc-label {{ font-weight: bold; color: #2c3e50; display: inline-block; width: 150px; }}
        .oclc-value {{ color: #333; }}
        .decision-section {{ grid-column: 1 / -1; margin-top: 20px; padding: 15px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px; }}
        .decision-buttons {{ display: flex; gap: 10px; margin-top: 10px; }}
        .decision-btn {{ padding: 8px 15px; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; }}
        .btn-approve {{ background-color: #27ae60; color: white; }}
        .btn-reject {{ background-color: #e74c3c; color: white; }}
        .btn-review {{ background-color: #f39c12; color: white; }}
        .notes-area {{ width: 100%; margin-top: 10px; padding: 8px; border: 1px solid #ddd; border-radius: 5px; resize: vertical; min-height: 60px; }}
        .no-image {{ color: #999; font-style: italic; text-align: center; padding: 20px; border: 2px dashed #ddd; border-radius: 5px; }}
        .sort-btn {{ transition: background-color 0.3s ease; opacity: 1; }}
        .sort-btn:hover {{ opacity: 0.8; }}
        .sort-btn.active {{ background-color: #3498db !important; transform: scale(1.05); }}
        .sort-btn:not(.active) {{ background-color: #95a5a6 !important; }}
        .sorting-controls {{ background-color: white; padding: 15px; margin-bottom: 20px; border: 1px solid #ddd; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
    </style>
</head>
<body>
    <div class="header">
        <h1>LP Review - Page {page_num} of {total_pages}</h1>
        <p>Generated: {current_date} | Records {start_idx + 1}-{start_idx + len(page_records)} of {(total_pages - 1) * records_per_page + len(page_records)}</p>
    </div>
    
    <div class="navigation">
        <a href="review-index-{current_date}.html" class="nav-btn">Back to Index</a>"""
    
    if page_num > 1:
        html_content += f'<a href="review-page-{page_num - 1}-{current_date}.html" class="nav-btn">Previous</a>'
    else:
        html_content += '<span class="nav-btn disabled">Previous</span>'
    
    html_content += f'<span style="margin: 0 20px; font-weight: bold;">Page {page_num} of {total_pages}</span>'
    
    if page_num < total_pages:
        html_content += f'<a href="review-page-{page_num + 1}-{current_date}.html" class="nav-btn">Next</a>'
    else:
        html_content += '<span class="nav-btn disabled">Next</span>'
    
    html_content += f"""
    </div>
    
    <div class="sorting-controls">
        <h3 style="margin: 0 0 10px 0; color: #2c3e50;">Sort Records</h3>
        <div style="display: flex; gap: 10px; align-items: center;">
            <button onclick="sortByOriginalOrder()" id="sortOriginal" class="sort-btn active" style="background: #3498db; color: white; border: none; padding: 8px 15px; border-radius: 5px; cursor: pointer; font-weight: bold;">
                Original Order
            </button>
            <button onclick="sortByConfidence()" id="sortConfidence" class="sort-btn" style="background: #95a5a6; color: white; border: none; padding: 8px 15px; border-radius: 5px; cursor: pointer; font-weight: bold;">
                Sort by Confidence (Low to High)
            </button>
            <span style="margin-left: 20px; color: #666; font-size: 14px;">
                Sorting preserves your decisions and notes
            </span>
        </div>
    </div>
    
    <div class="export-controls" style="background-color: white; padding: 15px; margin-bottom: 20px; border: 1px solid #ddd; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
        <h3 style="margin: 0 0 10px 0; color: #2c3e50;">Export Decisions</h3>
        <div style="display: flex; flex-direction: column; gap: 10px;">
            <div style="display: flex; gap: 15px; align-items: center;">
                <label style="display: flex; align-items: center; gap: 5px;">
                    <input type="radio" name="exportType" value="decisions" checked>
                    <span id="decisionsLabel">Export decisions only (<span id="decisionsCount">0</span> records)</span>
                </label>
                <label style="display: flex; align-items: center; gap: 5px;">
                    <input type="radio" name="exportType" value="all">
                    <span>Export all records on this page ({len(page_records)} records)</span>
                </label>
            </div>
            <div style="display: flex; gap: 10px; align-items: center;">
                <button onclick="exportDecisions()" style="background: #e74c3c; color: white; border: none; padding: 10px 20px; border-radius: 5px; cursor: pointer; font-weight: bold;">
                    Export Page CSV
                </button>
                <span style="color: #666; font-size: 14px;">
                    Choose export type above
                </span>
            </div>
        </div>
    </div>
"""

    # Process records for this page
    for i, record in enumerate(page_records, 1):
        global_record_id = start_idx + i
        barcode = record["barcode"]
        oclc_number = record["oclc_number"]
        confidence_score = record.get("confidence_score", 0)
        
        sort_group = record.get("sort_group", "Unknown")
        sort_group_class = "group-" + sort_group.lower().replace(" ", "-").replace("(", "").replace(")", "")
        
        try:
            conf_value = float(confidence_score)
            if conf_value < 30:
                conf_class = "confidence-low"
            elif conf_value < 60:
                conf_class = "confidence-medium"
            else:
                conf_class = "confidence-high"
        except (ValueError, TypeError):
            conf_class = "confidence-low"
            conf_value = 0
        
        # Find image files and copy them to results folder for portability
        image_files = []
        images_subfolder = os.path.join(results_folder, "images")
        os.makedirs(images_subfolder, exist_ok=True)

        if os.path.exists(images_folder):
            for filename in os.listdir(images_folder):
                if filename.startswith(str(barcode)) and filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                    src_path = os.path.join(images_folder, filename)
                    dest_path = os.path.join(images_subfolder, filename)
                    
                    try:
                        shutil.copy2(src_path, dest_path)
                        print(f"Copied image: {filename}")
                        
                        rel_path = os.path.join("images", filename).replace("\\", "/")
                        image_files.append((rel_path, filename))
                        
                    except Exception as copy_error:
                        print(f"Warning: Could not copy image {filename}: {copy_error}")
                        continue
        else:
            print(f"Warning: Images folder not found at {images_folder}")

        image_files.sort()
        
        html_content += f"""
            <div class="record" id="record-{global_record_id}" data-barcode="{barcode}" data-oclc-number="{oclc_number}" data-sort-group="{sort_group}">
        <div class="record-header">
            <div class="barcode">Record {global_record_id}: Barcode {barcode}</div>
            <div>
                <div class="confidence {conf_class}">{conf_value}% Confidence</div>
            </div>
        </div>
        
        <div class="content-grid">
            <div class="images-section">
                <h3>LP Images</h3>"""
        
        if image_files:
            for j, (img_path, filename) in enumerate(image_files[:3]):
                filename_lower = filename.lower()
                if 'a.' in filename_lower:
                    img_type = "Front Cover"
                elif 'b.' in filename_lower:
                    img_type = "Back Cover"
                elif 'c.' in filename_lower:
                    img_type = "Additional Image"
                else:
                    img_type = f"Image {j+1}"
                
                html_content += f"""
                <div class="image-container">
                    <div class="image-label">{img_type}</div>
                    <img src="{img_path}" 
                         alt="{img_type} for barcode {barcode}"
                         onclick="window.open(this.src, '_blank')"
                         onerror="this.style.display='none'; this.nextElementSibling.style.display='block';">
                    <div class="no-image" style="display: none;">Image not found</div>
                </div>"""
        else:
            html_content += '<div class="no-image">No images found for this barcode</div>'
        
        html_content += """
            </div>
            
            <div class="oclc-section">
                <h3>OCLC Record Information</h3>"""
        
        if oclc_number and record.get("has_valid_oclc", False):
            oclc_data = get_bib_info_from_workflow(oclc_number, workflow_json_path)
            formatted_record = oclc_data.get("full_record_text", "No detailed record available")
            html_content += f"""
                <pre style="background: #f8f9fa; padding: 15px; border: 1px solid #ddd; border-radius: 5px; overflow: auto; max-height: 70vh; font-size: 13px; white-space: pre-wrap; word-wrap: break-word;">{formatted_record}</pre>"""
        else:
            html_content += """
                <div style="background: #f8f9fa; padding: 15px; border: 1px solid #ddd; border-radius: 5px; color: #666; font-style: italic;">
                    No valid OCLC match found for this record.
                </div>"""
        
        html_content += f"""
            </div>
            
            <div class="decision-section">
                <h3>Cataloger Decision</h3>
                <p>Review the images and OCLC record above, then make your decision:</p>
                
                <div class="decision-buttons">
                    <button class="decision-btn btn-approve" onclick="setDecision({global_record_id}, 'approved', event)">
                        Approve OCLC Match
                    </button>
                    <button class="decision-btn btn-reject" onclick="setDecision({global_record_id}, 'different', event)">
                        Different OCLC Needed
                    </button>
                    <button class="decision-btn btn-reject" onclick="setDecision({global_record_id}, 'original', event)">
                        Needs Original Cataloging
                    </button>
                    <button class="decision-btn btn-review" onclick="setDecision({global_record_id}, 'review', event)">
                        Needs More Review
                    </button>
                </div>
                
                <textarea class="notes-area" placeholder="Notes and correct OCLC number (if different)..." 
                         id="notes-{global_record_id}"></textarea>
            </div>
        </div>
    </div>"""

    # Add JavaScript
    html_content += f"""
    <script>
        // Create unique storage namespace for this workflow
        const STORAGE_PREFIX = 'lp-workflow-{current_date}-';

        // Helper functions to use namespaced storage
        function setStorage(key, value) {{
            window.localStorage.setItem(STORAGE_PREFIX + key, value);
        }}

        function getStorage(key) {{
            return window.localStorage.getItem(STORAGE_PREFIX + key);
        }}

        function removeStorage(key) {{
            window.localStorage.removeItem(STORAGE_PREFIX + key);
        }}

        // Get all keys for this workflow
        function getAllWorkflowKeys() {{
            const keys = [];
            for (let i = 0; i < window.localStorage.length; i++) {{
                const key = window.localStorage.key(i);
                if (key && key.startsWith(STORAGE_PREFIX)) {{
                    keys.push(key.replace(STORAGE_PREFIX, ''));
                }}
            }}
            return keys;
        }}
        
        let currentSort = 'original';
        const totalRecordsInDataset = {(total_pages - 1) * records_per_page + len(page_records)};
        const pageStartIndex = {start_idx};
        
        function setDecision(recordId, decision, event) {{
            const record = document.getElementById('record-' + recordId);
            const buttons = record.querySelectorAll('.decision-btn');
            buttons.forEach(btn => btn.style.opacity = '0.5');
        
            event.target.style.opacity = '1';
            event.target.style.transform = 'scale(1.05)';
            
            setStorage('decision-' + recordId, decision);
            document.getElementById('notes-' + recordId).focus();
            
            const barcode = record.getAttribute('data-barcode') || '';
            const confEl = record.querySelector('.confidence');
            const confidenceText = confEl ? confEl.textContent : '';
            const confidence = confidenceText ? confidenceText.replace('% Confidence', '%') : 'N/A';
            const sortGroup = record.getAttribute('data-sort-group') || 'N/A';

            let oclcNumber = '';
            const oclcSection = record.querySelector('.oclc-section pre');
            if (oclcSection) {{
                const oclcText = oclcSection.textContent || '';
                const oclcMatch = oclcText.match(/OCLC Number: (\d+)/);
                if (oclcMatch) {{
                    oclcNumber = oclcMatch[1];
                }}
            }}
            if (!oclcNumber) {{
                const recordData = record.dataset || {{}};
                if (recordData.oclcNumber && recordData.oclcNumber !== 'None suggested') {{
                    oclcNumber = recordData.oclcNumber;
                }}
            }}


            
            const enhancedRecordData = {{
                barcode: barcode,
                confidence: confidence,
                sortGroup: sortGroup,
                oclcNumber: oclcNumber,
                pageNumber: {page_num}
                }};
            setStorage('record-data-' + recordId, JSON.stringify(enhancedRecordData));
            updateDecisionCounts();
        }}
        
        function sortByConfidence() {{
            if (currentSort === 'confidence') return;
            
            saveCurrentState();
            
            const records = Array.from(document.querySelectorAll('.record'));
            const recordsWithConfidence = records.map(record => {{
                const confidenceText = record.querySelector('.confidence').textContent;
                const confidence = parseFloat(confidenceText.replace('% Confidence', '')) || 0;
                return {{ element: record, confidence: confidence }};
            }});
            
            recordsWithConfidence.sort((a, b) => a.confidence - b.confidence);
            
            const container = records[0].parentNode;
            recordsWithConfidence.forEach(item => {{
                container.appendChild(item.element);
            }});
            
            document.getElementById('sortOriginal').classList.remove('active');
            document.getElementById('sortConfidence').classList.add('active');
            currentSort = 'confidence';
            
            restoreUserState();
        }}
        
        function sortByOriginalOrder() {{
            if (currentSort === 'original') return;
            
            saveCurrentState();
            
            const records = Array.from(document.querySelectorAll('.record'));
            const container = records[0].parentNode;
            
            records.sort((a, b) => {{
                const aId = parseInt(a.id.replace('record-', ''));
                const bId = parseInt(b.id.replace('record-', ''));
                return aId - bId;
            }});
            
            records.forEach(record => {{
                container.appendChild(record);
            }});
            
            document.getElementById('sortConfidence').classList.remove('active');
            document.getElementById('sortOriginal').classList.add('active');
            currentSort = 'original';
            
            restoreUserState();
        }}
        
        function saveCurrentState() {{
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                const notesElement = document.getElementById('notes-' + i);
                if (notesElement) {{
                    setStorage('notes-' + i, notesElement.value);
                }}
            }}
        }}
        
        function restoreUserState() {{
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                const decision = getStorage('decision-' + i);
                const record = document.getElementById('record-' + i);
                if (!record) continue;

                // Reset button visuals
                const buttons = record.querySelectorAll('.decision-btn');
                buttons.forEach(btn => {{
                btn.style.opacity = '0.5';
                btn.style.transform = '';
                }});

                // Re-apply selection by matching button text
                if (decision) {{
                buttons.forEach(btn => {{
                    const txt = (btn.textContent || '').toLowerCase();
                    if ((decision === 'approved' && txt.includes('approve')) ||
                        (decision === 'different' && txt.includes('different')) ||
                        (decision === 'original'  && txt.includes('original')) ||
                        (decision === 'review'    && (txt.includes('more review') || txt.includes('review')))) {{
                    btn.style.opacity = '1';
                    btn.style.transform = 'scale(1.05)';
                    }}
                }});
                }}

                // Restore notes text
                const notes = getStorage('notes-' + i);
                const notesElement = document.getElementById('notes-' + i);
                if (notes && notesElement) {{
                notesElement.value = notes;
                }}
            }}

            // Refresh "(N records)" counter
            updateDecisionCounts();
            }}

        
        document.addEventListener('DOMContentLoaded', function() {{
            updateDecisionCounts();
            restoreUserState();
        }});

        document.addEventListener('input', function(e) {{
            if (e.target.classList.contains('notes-area')) {{
                const recordId = e.target.id.split('-')[1];
                setStorage('notes-' + recordId, e.target.value);
            }}
        }});
        
        function updateDecisionCounts() {{
            let decisionsCount = 0;
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                if (getStorage('decision-' + i)) {{
                    decisionsCount++;
                }}
            }}
            const counterEl = document.getElementById('decisionsCount');
            if (counterEl) {{
                counterEl.textContent = decisionsCount;
            }}
        }}
        
        function getDecisionLabel(decision) {{
            const labels = {{
                'approved': 'Approved',
                'different': 'Different OCLC # Needed',
                'original': 'Original Cataloging Needed',
                'review': 'Further Review Needed'
            }};
            return labels[decision] || decision;
        }}
        
        function exportDecisions() {{
            const catalogerName = prompt('Enter your name for the export file:');
            if (!catalogerName) return;
            const exportType = document.querySelector('input[name="exportType"]:checked').value;
            
            const decisions = [];
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                const decision = getStorage('decision-' + i);
                const notes = getStorage('notes-' + i);
                const recordElement = document.getElementById('record-' + i);
                
                if (recordElement) {{
                    if (exportType === 'decisions' && !decision) {{
                        continue;
                    }}

                    const barcode = recordElement.getAttribute('data-barcode') || '';
                    const confEl = recordElement.querySelector('.confidence');
                    const confidenceText = confEl ? confEl.textContent : '';
                    const confidence = confidenceText ? confidenceText.replace('% Confidence', '%') : 'N/A';
                    const sortGroup = recordElement.getAttribute('data-sort-group') || 'N/A';

                    let oclcNumber = '';
                    const oclcSection = recordElement.querySelector('.oclc-section pre');
                    if (oclcSection) {{
                        const oclcText = oclcSection.textContent || '';
                        const oclcMatch = oclcText.match(/OCLC Number: (\d+)/);
                        if (oclcMatch) {{
                            oclcNumber = oclcMatch[1];
                        }}
                    }}
                    

                    
                    if (!oclcNumber) {{
                        const recordData = recordElement.dataset;
                        if (recordData.oclcNumber && recordData.oclcNumber !== 'None suggested' && recordData.oclcNumber !== '') {{
                            oclcNumber = recordData.oclcNumber;
                        }}
                    }}
                    
                    let correctOclc = '';
                    if (decision === 'approved' && oclcNumber) {{
                        correctOclc = oclcNumber;
                    }}
                    
                    var enhancedRecordData = {{
                        barcode: barcode,
                        confidence: confidence,
                        sortGroup: sortGroup,
                        oclcNumber: oclcNumber,
                        pageNumber: {page_num}
                    }};
                    setStorage('record-data-' + i, JSON.stringify(enhancedRecordData));
                    
                    decisions.push({{
                        record: i,
                        barcode: barcode,
                        confidence: confidence,
                        sortGroup: sortGroup,
                        decision: getDecisionLabel(decision) || 'Not reviewed',
                        aiSuggestedOclc: oclcNumber,
                        correctOclc: correctOclc,
                        notes: notes || '',
                        cataloger: catalogerName,
                        reviewDate: new Date().toISOString().split('T')[0],
                        pageNumber: {page_num}
                    }});
                }}
            }}
            
            const headers = ['Record', 'Barcode', 'Confidence', 'Initial Sort Group', 'Cataloger Decision', 'Correct OCLC #', 'Notes', 'Cataloger', 'Review Date', 'Page Number'];
            let csvContent = headers.join(',') + '\\n';

            decisions.forEach(row => {{
            const csvRow = [
                row.record,
                row.barcode,
                '"' + row.confidence + '"',
                '"' + row.sortGroup + '"',
                '"' + row.decision + '"',
                '"' + row.aiSuggestedOclc + '"',
                '"' + row.correctOclc + '"',
                '"' + row.notes.replace(/"/g, '""') + '"',
                '"' + row.cataloger + '"',
                row.reviewDate,
                row.pageNumber
            ].join(',');
            csvContent += csvRow + '\\n';
        }});
            
            const blob = new Blob([csvContent], {{ type: 'text/csv' }});
            const url = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            const exportTypeLabel = exportType === 'decisions' ? 'decisions' : 'all-records';
            a.download = 'cataloger-' + exportTypeLabel + '-page-{page_num}-' + catalogerName.replace(/[^a-zA-Z0-9]/g, '_') + '-' + new Date().toISOString().split('T')[0] + '.csv';
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            window.URL.revokeObjectURL(url);
            
            alert('Exported decisions for ' + decisions.length + ' records to CSV file.');
        }}
    </script>
</body>
</html>"""

    with open(page_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

def load_records_from_step5(step5_file):
    """Load all records from the Step 5 sorting spreadsheet."""
    wb = load_workbook(step5_file)
    sheet = wb.active
    
    all_records = []
    
    for row_idx in range(2, sheet.max_row + 1):
        barcode = sheet.cell(row=row_idx, column=1).value
        sort_group = sheet.cell(row=row_idx, column=2).value
        oclc_number = sheet.cell(row=row_idx, column=3).value
        title = sheet.cell(row=row_idx, column=4).value
        confidence_score = sheet.cell(row=row_idx, column=7).value
        
        if not barcode:
            continue
        
        has_valid_oclc = bool(oclc_number and str(oclc_number).strip() not in ["", "Not found", "Error processing"])
        
        record = {
            "barcode": barcode,
            "sort_group": sort_group or "Unknown",
            "oclc_number": str(oclc_number).strip() if oclc_number else "",
            "title": title or "No title available",
            "confidence_score": confidence_score if confidence_score else 0,
            "has_valid_oclc": has_valid_oclc
        }
        
        all_records.append(record)
    
    return all_records

def main():
    print("Step 6: Creating Interactive HTML Review Interface")
    print("=" * 60)
    
    # Get configuration
    file_paths = get_file_path_config()
    
    # Find latest results folder
    results_folder = find_latest_results_folder(file_paths["results_prefix"])
    if not results_folder:
        print("No results folder found! Please run Steps 1-5 first.")
        return None
    
    print(f"Using results folder: {results_folder}")
    
    # Get workflow JSON path - now in data subfolder after Step 5
    data_folder = os.path.join(results_folder, "data")
    workflow_json_path = get_workflow_json_path(data_folder)
    
    # Find the sorting spreadsheet from Step 5
    deliverables_folder = os.path.join(results_folder, "deliverables")
    sorting_files = [f for f in os.listdir(deliverables_folder) 
                     if f.startswith("sorting-spreadsheet-") and f.endswith(".xlsx")]
    
    if not sorting_files:
        print("No sorting spreadsheet found! Please run Step 5 first.")
        return None
    
    latest_sorting_file = max(sorting_files)
    sorting_file_path = os.path.join(deliverables_folder, latest_sorting_file)
    
    print(f"Loading records from: {latest_sorting_file}")
    
    # Load all records from Step 5 spreadsheet
    all_records = load_records_from_step5(sorting_file_path)
    
    if not all_records:
        print("No records found to create HTML review for.")
        return None
    
    print(f"Loaded {len(all_records)} records")
    
    # Get current date
    from lp_workflow_config import get_current_date
    current_date = get_current_date()
    
    # Create paginated HTML review
    result = create_paginated_review_html(
        results_folder, 
        all_records, 
        current_date, 
        workflow_json_path, 
        records_per_page=100
    )
    
    print(f"\n=== HTML Review Interface Created ===")
    print(f"Index page: {result['index_path']}")
    print(f"Total pages created: {result['total_pages']}")
    print(f"Total records: {len(all_records)}")
    print(f"\nOpen the index page in a web browser to begin review.")
    
    return result

if __name__ == "__main__":
    main()