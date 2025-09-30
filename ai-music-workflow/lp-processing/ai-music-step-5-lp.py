# Create output files for final results 
import os
import datetime
import re
import openpyxl
import json
from difflib import SequenceMatcher
from openpyxl import load_workbook

# Custom modules
from json_workflow import update_record_step5, log_error, log_processing_metrics
from shared_utilities import find_latest_results_folder, get_workflow_json_path, create_batch_summary
from lp_workflow_config import get_file_path_config, get_threshold_config, get_current_timestamp, get_step_config, FILE_NAMING

def find_latest_lp_metadata_file(results_folder):
    # Find files starting with "full-workflow-data-lp" and ending with ".xlsx"
    files = [f for f in os.listdir(results_folder) 
             if f.startswith("full-workflow-data-lp") and f.endswith(".xlsx")]
    if not files:
        return None
    latest_file = max(files)
    return os.path.join(results_folder, latest_file)

def get_bib_info_from_workflow(oclc_number, workflow_json_path):
    """
    Extract bibliographic information from formatted OCLC results in workflow JSON.
    """
    try:
        with open(workflow_json_path, 'r', encoding='utf-8') as f:
            workflow_data = json.load(f)
        
        # Search through all records for the target OCLC number in formatted results
        for barcode, record_data in workflow_data.get("records", {}).items():
            step2_data = record_data.get("step2_detailed_data", {})
            formatted_results = step2_data.get("formatted_oclc_results", "")
            
            # Look for this OCLC number in the formatted results
            oclc_pattern = rf"OCLC Number: {re.escape(oclc_number)}\n\n(.*?)(?=\n-{{40}}\nOCLC Number:|\Z)"
            match = re.search(oclc_pattern, formatted_results, re.DOTALL)
            
            if match:
                # Parse the formatted text to extract key information
                record_text = match.group(1)
                
                # Extract title
                title_match = re.search(r"Title Information:\s*\n\s*- Main Title: (.+?)(?:\n|$)", record_text)
                title = title_match.group(1) if title_match else "No title available"
                
                # Extract contributors/author
                contributors = []
                contributor_matches = re.findall(r"Contributors:\s*\n((?:\s*- .+?\n)*)", record_text)
                if contributor_matches:
                    contributor_lines = contributor_matches[0].strip().split('\n')
                    for line in contributor_lines:
                        if line.strip().startswith('- '):
                            contributor = line.strip()[2:].split(' (')[0]  # Remove role info
                            contributors.append(contributor)
                
                # Extract publication date
                date_match = re.search(r"- publicationDate: (.+?)(?:\n|$)", record_text)
                pub_date = date_match.group(1) if date_match else "No date available"
                
                return {
                    "title": title,
                    "contributors": contributors,
                    "publication_date": pub_date,
                    "full_record_text": record_text
                }
        
        return {"error": "OCLC record not found in workflow data"}
        
    except Exception as e:
        return {"error": str(e)}

def get_holdings_info_from_workflow(oclc_number, workflow_json_path):
    """
    Extract holdings information from formatted OCLC results in workflow JSON.
    """
    try:
        with open(workflow_json_path, 'r', encoding='utf-8') as f:
            workflow_data = json.load(f)
        
        # Search through all records for the target OCLC number
        for barcode, record_data in workflow_data.get("records", {}).items():
            step2_data = record_data.get("step2_detailed_data", {})
            formatted_results = step2_data.get("formatted_oclc_results", "")
            
            # Look for this OCLC number and extract holdings info
            oclc_pattern = rf"OCLC Number: {oclc_number}\n\nHeld by IXA: (Yes|No)\nTotal Institutions Holding: (\d+)"
            match = re.search(oclc_pattern, formatted_results)
            
            if match:
                is_held_by_ixa = match.group(1) == "Yes"
                total_holdings = int(match.group(2))
                
                return {
                    "held_by_ixa": is_held_by_ixa,
                    "total_holdings": total_holdings
                }
        
        return {
            "held_by_ixa": False,
            "total_holdings": 0,
            "error": "Holdings data not found in workflow"
        }
        
    except Exception as e:
        return {
            "held_by_ixa": False,
            "total_holdings": 0,
            "error": str(e)
        }

def clean_title(title):
    """Clean up title by removing strange punctuation but keeping slashes."""
    # Replace any double sword or other special characters, but keep slashes
    title = re.sub(r'[^\w\s\/\-\:\;\,\.\(\)\[\]\&\']', ' ', title)
    # Normalize whitespace
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def extract_title_from_bib_info(data):
    """Extract title from workflow bibliographic information."""
    if isinstance(data, dict):
        if "error" in data:
            return "No title available"
        return data.get("title", "No title available")
    return "No title available"

def extract_author_from_bib_info(data):
    """Extract author from workflow bibliographic information."""
    if isinstance(data, dict) and "contributors" in data:
        contributors = data["contributors"]
        if contributors and len(contributors) > 0:
            return contributors[0]  # Return first contributor
        return "No author available"
    return "No author available"

def extract_publication_date_from_bib_info(data):
    """Extract publication date from workflow bibliographic information."""
    if isinstance(data, dict):
        if "error" in data:
            return "No date available"
        return data.get("publication_date", "No date available")
    return "No date available"

def format_bib_info(data):
    """Format bibliographic information for display."""
    output = []
    
    if not isinstance(data, dict) or "identifier" not in data:
        return "No bibliographic information available."
    
    record = data
    
    title = "N/A"
    if "title" in record and "mainTitles" in record["title"]:
        title = record["title"]["mainTitles"][0].get("text", "N/A")
    
    series_title = "N/A"
    if "series" in record and isinstance(record["series"], list) and len(record["series"]) > 0:
        series_title = record["series"][0].get("title", "N/A")
    
    author = "N/A"
    contributors = []
    if "contributor" in record:
        if "creators" in record["contributor"] and record["contributor"]["creators"]:
            for creator in record["contributor"]["creators"]:
                if "nonPersonName" in creator and "text" in creator["nonPersonName"]:
                    author_name = creator["nonPersonName"]["text"]
                    author = author_name
                    contributors.append(author_name)
                elif "firstName" in creator and "secondName" in creator:
                    first_name = creator.get("firstName", {}).get("text", "")
                    second_name = creator.get("secondName", {}).get("text", "")
                    author_name = f"{first_name} {second_name}".strip()
                    author = author_name
                    contributors.append(author_name)
    
    publisher = "N/A"
    place = "N/A"
    date_pub = "N/A"
    if "publishers" in record and record["publishers"]:
        publisher_info = record["publishers"][0]
        if "publisherName" in publisher_info and "text" in publisher_info["publisherName"]:
            publisher = publisher_info["publisherName"]["text"]
        if "publicationPlace" in publisher_info:
            place = publisher_info["publicationPlace"]
    
    if "date" in record and "publicationDate" in record["date"]:
        date_pub = record["date"]["publicationDate"].replace("\u2117", "c")
    
    content_type = "N/A"
    if "format" in record:
        if "generalFormat" in record["format"]:
            content_type = record["format"]["generalFormat"]
        if "specificFormat" in record["format"]:
            content_type += f" - {record['format']['specificFormat']}"
    
    upc = "N/A"
    if "identifier" in record and "otherStandardIdentifiers" in record["identifier"]:
        for identifier in record["identifier"]["otherStandardIdentifiers"]:
            if identifier.get("type") == "Universal Product Code (UPC)":
                upc = identifier.get("id", "N/A")
                break
    
    # Enhanced content parsing logic
    contents = []
    
    # Method 1: Look for contents in the description field
    if "description" in record:
        # Check for the contentNote field in contents array - this is the format in the example
        if "contents" in record["description"]:
            for content_item in record["description"]["contents"]:
                # Check for contentNote object which contains track listings
                if "contentNote" in content_item and "text" in content_item["contentNote"]:
                    content_text = content_item["contentNote"]["text"]
                    # Common pattern: tracks separated by " -- "
                    if " -- " in content_text:
                        # Split by " -- " and clean up each track
                        tracks = []
                        for track in content_text.split(" -- "):
                            track = track.strip()
                            # Remove trailing period if it exists
                            if track.endswith('.'):
                                track = track[:-1].strip()
                            tracks.append(track)
                        contents.extend(tracks)
                        continue
                
                # Check for titles array format (original method)
                if "titles" in content_item:
                    for title_item in content_item["titles"]:
                        if isinstance(title_item, str):
                            contents.append(title_item)
                        elif isinstance(title_item, dict) and "text" in title_item:
                            contents.append(title_item["text"])
                
                # Check for different content formats
                if "items" in content_item:
                    for item in content_item["items"]:
                        if isinstance(item, str):
                            contents.append(item)
                        elif isinstance(item, dict) and "text" in item:
                            contents.append(item["text"])

        # Method 2: Check for TOC as a note
        if "notes" in record["description"]:
            for note in record["description"]["notes"]:
                # Look for various types of content notes
                is_content_note = False
                if "text" in note and any(marker in note["text"].lower() for marker in ["contents:", "tracks:", "track listing"]):
                    is_content_note = True
                
                if is_content_note and "text" in note:
                    toc_text = note["text"]
                    
                    # Try multiple approaches to parse the track list
                    # Approach 1: Split by common delimiters
                    for delimiter in ["--", ";", "/"]:
                        if delimiter in toc_text:
                            # Extract the content part (after any header like "Contents:" if present)
                            content_part = toc_text
                            for header in ["Contents:", "Tracks:", "Track listing:"]:
                                if header.lower() in toc_text.lower():
                                    content_part = toc_text.split(header, 1)[-1].strip()
                                    break
                                
                            # Split by delimiter and clean up
                            parts = [part.strip() for part in content_part.split(delimiter) if part.strip()]
                            if parts:
                                contents.extend(parts)
                                break

    # Method 3: Check for special MARC21 fields often used for music contents
    if "varFields" in record:
        for field in record.get("varFields", []):
            if field.get("marcTag") in ["505", "500"] and "subfields" in field:
                for subfield in field["subfields"]:
                    if subfield.get("code") == "a" and subfield.get("content"):
                        content = subfield["content"]
                        # Split content by common delimiters in track listings
                        for delimiter in ["--", ";", "/"]:
                            if delimiter in content:
                                parts = [part.strip() for part in content.split(delimiter) if part.strip()]
                                if parts:
                                    contents.extend(parts)
                                    break
    
    # Method 4: Check for $t prefixed content
    if "description" in record and "notes" in record["description"]:
        for note in record["description"]["notes"]:
            if "text" in note and "**$t**" in note["text"]:
                parts = note["text"].split("**$t**")
                # Skip the first part as it's usually empty or a header
                for part in parts[1:]:
                    # Clean up each part and add to contents
                    cleaned_part = part.strip()
                    if cleaned_part:
                        # Remove trailing -- if present
                        if cleaned_part.endswith("--"):
                            cleaned_part = cleaned_part[:-2].strip()
                        contents.append(cleaned_part)
    
    oclc_number = "N/A"
    if "identifier" in record and "oclcNumber" in record["identifier"]:
        oclc_number = record["identifier"]["oclcNumber"]
    
    output.append(f"Title: {title}")
    output.append(f"Series Title: {series_title}")
    output.append(f"Author: {author}")
    output.append(f"Contributors: {', '.join(contributors) if contributors else 'N/A'}")
    output.append(f"Publisher: {publisher}")
    output.append(f"Place of Publication: {place}")
    output.append(f"Date of Publication: {date_pub}")
    output.append(f"Content Type: {content_type}")
    output.append(f"UPC: {upc}")
    
    if contents:
        output.append("Contents:")
        for i, track in enumerate(contents, 1):
            # Clean up the track listing
            cleaned_track = track
            # Remove any trailing punctuation
            if cleaned_track.endswith(('.', ';')):
                cleaned_track = cleaned_track[:-1]
            output.append(f"  {i}. {cleaned_track}")
    
    output.append(f"OCLC Number: {oclc_number}")
    
    return "\n".join(output)

def calculate_title_similarity(title1, title2):
    """Calculate similarity between two titles using SequenceMatcher."""
    return SequenceMatcher(None, title1.lower(), title2.lower()).ratio()

def create_low_confidence_review_text_log(results_folder, step4_file, all_records, workflow_json_path, current_date):
    """
    Create a review text log for unique low confidence matches with detailed information.
    """
    print("Creating low confidence review text log...")
    
    # Filter for unique low confidence matches (including those without OCLC numbers)
    low_confidence_records = [record for record in all_records 
                             if record["sort_group"] == "Cataloger Review (Low Confidence)"]
    
    if not low_confidence_records:
        print("No low confidence matches found to review.")
        return None
    
    # Open the original step 4 file to get additional data
    wb_src = load_workbook(step4_file)
    sheet_src = wb_src.active
    
    # Create mapping from barcode to source row data
    barcode_to_source = {}
    for row_idx in range(2, sheet_src.max_row + 1):  # Skip header row
        barcode = sheet_src.cell(row=row_idx, column=4).value  # Column D
        if barcode:
            # Get relevant columns: D (Barcode), E (AI-Generated Metadata), H (OCLC Number), other potential columns
            row_data = {
                "barcode": barcode,
                "metadata": sheet_src.cell(row=row_idx, column=5).value,  # Column E - AI-Generated Metadata
                "other_oclc_numbers": sheet_src.cell(row=row_idx, column=11).value   # Column K - Other Potential Matches
            }
            barcode_to_source[barcode] = row_data
    
    # Create text log file in deliverables subfolder
    deliverables_folder = os.path.join(results_folder, "deliverables")
    review_file = f"low-confidence-matches-review-{current_date}.txt"
    review_path = os.path.join(deliverables_folder, review_file)
    
    # Process each low confidence record and write to text file
    processed_count = 0
    with open(review_path, 'w', encoding='utf-8') as f:
        # Write header
        f.write("=" * 80 + "\n")
        f.write("LOW CONFIDENCE MATCHES REVIEW LOG\n")
        f.write(f"Generated: {current_date}\n")
        f.write(f"Total Records: {len(low_confidence_records)}\n")
        f.write("=" * 80 + "\n\n")
        
        for record in low_confidence_records:
            barcode = record["barcode"]
            oclc_number = record["oclc_number"]
            
            print(f"Processing low confidence review record {processed_count + 1}/{len(low_confidence_records)} - Barcode: {barcode}")
            
            # Get source data
            source_data = barcode_to_source.get(barcode, {})
            metadata = source_data.get("metadata", "No AI-generated metadata available")
            other_oclc_numbers = source_data.get("other_oclc_numbers", "No other candidates")
            
            # Write record header
            f.write("-" * 60 + "\n")
            f.write(f"RECORD {processed_count + 1}\n")
            f.write("-" * 60 + "\n")
            f.write(f"Barcode: {barcode}\n")
            f.write(f"OCLC Number Chosen: {oclc_number if oclc_number else 'No OCLC number'}\n")
            f.write(f"Confidence Score: {record.get('confidence_score', 'No confidence score')}\n")
            f.write("\n")
            
            # Write AI-generated metadata
            f.write("AI-Generated Metadata:\n")
            if metadata and metadata.strip():
                # Format metadata with proper line breaks
                metadata_lines = metadata.replace('\n', '\n  ')
                f.write(f"  {metadata_lines}\n")
            else:
                f.write("  No AI-generated metadata available\n")
            f.write("\n")
            
            # Write other potential matches
            f.write("Other Potential Matches:\n")
            if other_oclc_numbers and other_oclc_numbers.strip():
                other_lines = other_oclc_numbers.replace('\n', '\n  ')
                f.write(f"  {other_lines}\n")
            else:
                f.write("  No other candidates\n")
            f.write("\n")
            
            # Get detailed OCLC record information from workflow JSON
            if oclc_number and record["has_valid_oclc"]:
                f.write("OCLC Record Details:\n")
                oclc_data = get_bib_info_from_workflow(oclc_number, workflow_json_path)
                
                # Get holdings information
                holdings_info = get_holdings_info_from_workflow(oclc_number, workflow_json_path)
                
                # Write raw OCLC data as JSON
                import json
                raw_oclc_json = json.dumps(oclc_data, indent=2, ensure_ascii=False)
                f.write(f"  Raw OCLC Data:\n  {raw_oclc_json.replace(chr(10), chr(10) + '  ')}\n")
                
                # Add holdings information
                holdings_text = f"\nHoldings Information:\nTotal Institutions Holding: {holdings_info.get('total_holdings', 0)}\nHeld by IXA: {'Yes' if holdings_info.get('held_by_ixa', False) else 'No'}"
                f.write(f"  {holdings_text}\n")
                
            else:
                f.write("OCLC Record Details:\n")
                f.write("  No OCLC record available - no valid OCLC number found\n")
            
            f.write("\n")
            processed_count += 1
        
        # Write summary footer
        f.write("=" * 80 + "\n")
        f.write("END OF REVIEW LOG\n")
        f.write(f"Total Records Processed: {processed_count}\n")
        f.write("=" * 80 + "\n")
    
    print(f"Low confidence review text log created with {len(low_confidence_records)} records: {review_path}")
    return review_path

def create_marc_format_text_log(results_folder, all_records, workflow_json_path, current_date):
    """
    Create a MARC-formatted text log from the original JSON metadata for low confidence records.
    """
    print("Creating MARC-formatted text log from original metadata...")
    
    # Filter for unique low confidence matches (including those without OCLC numbers)
    low_confidence_records = [record for record in all_records 
                             if record["sort_group"] == "Cataloger Review (Low Confidence)"]
    
    if not low_confidence_records:
        print("No low confidence matches found for MARC formatting.")
        return None
    
    # Load workflow JSON to get original metadata
    import json
    try:
        with open(workflow_json_path, 'r', encoding='utf-8') as f:
            workflow_data = json.load(f)
    except Exception as e:
        print(f"Error reading workflow JSON: {e}")
        return None
    
    # Create MARC text log file in deliverables subfolder
    deliverables_folder = os.path.join(results_folder, "deliverables")
    marc_file = f"marc-formatted-low-confidence-matches-{current_date}.txt"
    marc_path = os.path.join(deliverables_folder, marc_file)
    
    def is_valid_field(value):
        """Check if a field value is valid (not None, empty, or 'Not visible')"""
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() and value.strip().lower() != 'not visible'
        return bool(value)
    
    def safe_get(obj, key, default=""):
        """Safely get a value from a dictionary, handling None values."""
        if obj is None:
            return default
        value = obj.get(key, default)
        return value if value is not None else default
    
    processed_count = 0
    with open(marc_path, 'w', encoding='utf-8') as f:
        # Write header
        f.write("=" * 80 + "\n")
        f.write("MARC FORMAT - LOW CONFIDENCE RECORDS\n")
        f.write("This is AI-Generated Metadata from Step 1 formatted in MARC to kickstart original cataloging\n")
        f.write(f"Generated: {current_date}\n")
        f.write(f"Total Records: {len(low_confidence_records)}\n")
        f.write("Note: Only fields with visible/available data are included\n")
        f.write("=" * 80 + "\n\n")
        
        for record in low_confidence_records:
            barcode = record["barcode"]
            
            # Get original metadata from workflow JSON
            workflow_record = workflow_data.get("records", {}).get(barcode)
            if not workflow_record or "step1_metadata_extraction" not in workflow_record:
                f.write(f"Record {processed_count + 1} - Barcode: {barcode}\n")
                f.write("No original metadata available\n")
                f.write("-" * 60 + "\n\n")
                processed_count += 1
                continue
            
            extracted_fields = workflow_record["step1_metadata_extraction"].get("extracted_fields", {})
            
            # Write record header
            f.write(f"Record {processed_count + 1} - Barcode: {barcode}\n")
            f.write("-" * 60 + "\n")
            
            # Get field values with safe handling
            title_info = extracted_fields.get("title_information", {}) or {}
            main_title = safe_get(title_info, "main_title")
            subtitle = safe_get(title_info, "subtitle") 
            primary_contributor = safe_get(title_info, "primary_contributor")
            
            publishers = extracted_fields.get("publishers", {}) or {}
            place = safe_get(publishers, "place")
            publisher_name_raw = safe_get(publishers, "name")
            publisher_name = publisher_name_raw.replace("Name: ", "").strip() if publisher_name_raw else ""
            
            dates = extracted_fields.get("dates", {}) or {}
            publication_date = safe_get(dates, "publication_date")
            
            physical = extracted_fields.get("physical_description", {}) or {}
            contents = extracted_fields.get("contents", {}) or {}
            tracks = contents.get("tracks", []) or []
            
            # 100 - Main Entry (Primary contributor)
            if is_valid_field(primary_contributor):
                f.write(f"100 1  {primary_contributor}, $ecomposer, $eperformer.\n")
            
            # 245 - Title Statement (check this before 264)
            if is_valid_field(main_title):
                title_field = f"245 1 0 {main_title}"
                if is_valid_field(subtitle):
                    title_field += f" : $b{subtitle}"
                title_field += f" / $c{primary_contributor}." if is_valid_field(primary_contributor) else "."
                f.write(f"{title_field}\n")
            elif is_valid_field(primary_contributor):
                # If no title but we have contributor, create minimal title field
                f.write(f"245 1 0 [Title not visible] / $c{primary_contributor}.\n")
            else:
                # If neither title nor contributor, note for cataloger
                f.write("245 1 0 [Title and contributor not visible]\n")
            
            # 264 - Publication Statement
            if is_valid_field(place) or is_valid_field(publisher_name) or is_valid_field(publication_date):
                pub_field = "264  1 "
                
                # Handle place
                if is_valid_field(place):
                    if is_valid_field(publisher_name):
                        pub_field += f"{place} : "  # Colon only if publisher follows
                    else:
                        pub_field += f"{place} "    # No colon if no publisher
                
                # Handle publisher
                if is_valid_field(publisher_name):
                    pub_field += f"$b{publisher_name}"
                    if is_valid_field(publication_date):
                        pub_field += ", "  # Comma only if date follows
                    else:
                        pub_field += "."   # Period if no date follows
                
                # Handle date
                if is_valid_field(publication_date):
                    # Clean up date - remove copyright symbols and extra text
                    date_clean = (publication_date or "").replace("©", "").replace("℗", "").strip()
                    # Extract just the year if possible
                    import re
                    year_match = re.search(r'\b(19|20)\d{2}\b', date_clean)
                    if year_match:
                        date_clean = f"[{year_match.group()}]"
                    pub_field += f"$c{date_clean}"
                
                # Remove any trailing comma and ensure proper punctuation
                pub_field = pub_field.rstrip(', ') + "."
                f.write(f"{pub_field}\n")
            
            # 300 - Physical Description
            f.write("300    1 audio disc : $banalog ; $c12 in.\n")
            
            # 500 - General Note
            f.write("340    vinyl.\n")
            
            # 505 - Contents Note (Track listing)
            if tracks and isinstance(tracks, list):
                track_list = []
                for track in tracks:
                    if isinstance(track, dict):
                        track_title = safe_get(track, "title")
                        if is_valid_field(track_title):
                            # Filter out metadata that got mixed in with track titles
                            if not any(x in track_title.lower() for x in ['standard lp', 'vinyl record', '12', 'vinyl', 'PVC']):
                                track_list.append(track_title)
                    elif isinstance(track, str) and is_valid_field(track):
                        # Handle case where tracks might be stored as strings
                        if not any(x in track.lower() for x in ['standard lp', 'vinyl record', '12', 'vinyl', 'PVC']):
                            track_list.append(track)
                
                if track_list:
                    contents_field = "505 0  " + " -- ".join(track_list) + "."
                    f.write(f"{contents_field}\n")
            
            # 650 - Subject Added Entry (Genre/Form)
            f.write("650  0  $aMusic.\n")
            
            f.write("-" * 60 + "\n\n")
            processed_count += 1
            
        # Write summary footer
        f.write("=" * 80 + "\n")
        f.write("END OF MARC FORMAT LOG\n") 
        f.write(f"Total Records Processed: {processed_count}\n")
        f.write("=" * 80 + "\n")
    
    print(f"MARC format text log created with {processed_count} records: {marc_path}")
    return marc_path

def create_cataloger_review_spreadsheet(results_folder, all_records, current_date):
    """
    Create a separate Excel workbook for catalogers to review low confidence matches.
    """
    print("Creating cataloger review spreadsheet...")
    
    # Filter for low confidence records only
    low_confidence_records = [record for record in all_records 
                             if record["sort_group"] == "Cataloger Review (Low Confidence)"]
    
    if not low_confidence_records:
        print("No low confidence matches found for cataloger review.")
        return None
    
    # Create new workbook
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cataloger Review"
    
    # Define headers
    headers = [
        "Barcode",
        "Date AI Processed", 
        "AI-Suggested OCLC Number",
        "Title",
        "Date Cataloger Checked",
        "Status",
        "Correct OCLC Number",
        "Notes"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 16  # Barcode
    ws.column_dimensions['B'].width = 18  # Date AI Processed
    ws.column_dimensions['C'].width = 22  # AI Suggested OCLC
    ws.column_dimensions['D'].width = 55  # Title
    ws.column_dimensions['E'].width = 22  # Date Cataloger Checked
    ws.column_dimensions['F'].width = 25  # Status
    ws.column_dimensions['G'].width = 22  # Correct OCLC
    ws.column_dimensions['H'].width = 40  # Notes
    
    # Add data rows
    for row_num, record in enumerate(low_confidence_records, start=2):
        ws.cell(row=row_num, column=1, value=record["barcode"])
        ws.cell(row=row_num, column=2, value=current_date)
        
        # Show the AI-suggested OCLC number (what the workflow chose)
        ai_suggested_oclc = record["oclc_number"] if record["oclc_number"] else "None suggested"
        ws.cell(row=row_num, column=3, value=ai_suggested_oclc)
        
        # Show the title from OCLC record
        ws.cell(row=row_num, column=4, value=record["title"])
        
        # Columns 5, 6 left empty for cataloger input (Date Cataloger Checked, Status)
        
        # Column 7 (G) - Auto-populate formula for Correct OCLC Number
        # If status is "Approved", use the AI-suggested number, otherwise leave blank for manual entry
        ws.cell(row=row_num, column=7, value=f'=IF(F{row_num}="Approved",C{row_num},"")')
        
        # Column 8 (H) left empty for cataloger notes
    
    # Create dropdown validation for Status column - alternative approach
    from openpyxl.worksheet.datavalidation import DataValidation
    
    dv = DataValidation(type="list", formula1='"Approved,Rejected - Different OCLC,Rejected - Needs Original Cataloging"')
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please pick from the list'
    dv.promptTitle = 'List Selection'
    
    # Add validation to the Status column range
    ws.add_data_validation(dv)
    dv.add(f'F2:F{len(low_confidence_records) + 1}')
    
    # Apply conditional formatting for highlighting
    from openpyxl.formatting.rule import FormulaRule
    
    # Create conditional formatting rule: highlight if Correct OCLC Number column is empty
    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    formula_rule = FormulaRule(formula=[f'$G2=""'], fill=highlight_fill)
    
    # Apply conditional formatting to all data rows
    data_range = f"A2:H{len(low_confidence_records) + 1}"
    ws.conditional_formatting.add(data_range, formula_rule)
    
    # Save the workbook in deliverables subfolder
    deliverables_folder = os.path.join(results_folder, "deliverables")
    review_file = f"tracking-spreadsheet-catalogers-{current_date}.xlsx"
    review_path = os.path.join(deliverables_folder, review_file)
    
    print(f"Cataloger review spreadsheet created with {len(low_confidence_records)} records: {review_path}")
    return review_path

def create_paginated_review_html(results_folder, all_records, current_date, workflow_json_path, records_per_page=100):
    """
    Create paginated HTML files with external images and lazy loading for large datasets.
    All HTML files will be in the same folder for maximum compatibility.
    """
    print(f"Creating paginated review with {records_per_page} records per page...")
    
    import os
    import math
    from lp_workflow_config import get_file_path_config
    
    # Get file path config to find images folder
    file_paths = get_file_path_config()
    images_folder = file_paths["images_folder"]
    
    # Calculate number of pages needed
    total_pages = math.ceil(len(all_records) / records_per_page)
    page_files = []
    
    # Create index page path (in results folder)
    index_file = f"review-index-{current_date}.html"
    index_path = os.path.join(results_folder, index_file)
    
    # Group records by sort group for better organization
    sort_groups = {}
    for record in all_records:
        group = record.get("sort_group", "Unknown")
        if group not in sort_groups:
            sort_groups[group] = []
        sort_groups[group].append(record)
    
    # Create index page
    create_review_index(index_path, sort_groups, current_date, total_pages, records_per_page)
    
    # Create individual pages in the same folder as index
    for page_num in range(1, total_pages + 1):
        start_idx = (page_num - 1) * records_per_page
        end_idx = min(start_idx + records_per_page, len(all_records))
        page_records = all_records[start_idx:end_idx]
        
        # Put page files directly in results folder alongside index
        page_file = f"review-page-{page_num}-{current_date}.html"
        page_path = os.path.join(results_folder, page_file)

        create_single_review_page(
            page_path, page_records, current_date, workflow_json_path, 
            images_folder, results_folder, page_num, total_pages, records_per_page, start_idx
        )
        
        page_files.append(page_path)
        print(f"Created page {page_num}/{total_pages} with {len(page_records)} records")
        
    print(f"Images copied to images subfolder for portability")
    print(f"All HTML files are in the same folder for reliable navigation")
    print(f"To share: Send entire '{os.path.basename(results_folder)}' folder")
    
    return {
        "index_path": index_path,
        "page_files": page_files,
        "total_pages": total_pages
    }

def create_review_index(index_path, sort_groups, current_date, total_pages, records_per_page):
    """Create an index page with links to all review pages and sort group summaries."""
    
    total_records = sum(len(records) for records in sort_groups.values())
    
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>LP Review Index - {current_date}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .header {{ background-color: #2c3e50; color: white; padding: 20px; border-radius: 5px; margin-bottom: 30px; }}
        .summary {{ background-color: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .page-links {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 15px; margin-top: 20px; }}
        .page-link {{ background-color: #3498db; color: white; padding: 15px; text-decoration: none; border-radius: 5px; text-align: center; font-weight: bold; }}
        .page-link:hover {{ background-color: #2980b9; }}
        .sort-group {{ margin: 10px 0; padding: 10px; background-color: #f8f9fa; border-left: 4px solid #3498db; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>LP Cataloger Review Index</h1>
        <p>Generated: {current_date} | Total Records: {total_records} | Pages: {total_pages}</p>
    </div>
    
    <div class="summary">
        <h2>Sort Group Summary</h2>"""
    
    for group, records in sorted(sort_groups.items()):
        html_content += f'<div class="sort-group"><strong>{group}:</strong> {len(records)} records</div>'
    
    html_content += f"""
    </div>
    
    <div class="summary">
        <h2>Review Pages</h2>
        <p>Each page contains up to {records_per_page} records for manageable review.</p>
        <p><strong>All files are in the same folder for reliable navigation.</strong></p>
        <div class="page-links">"""
    
    for page_num in range(1, total_pages + 1):
        start_record = (page_num - 1) * records_per_page + 1
        end_record = min(page_num * records_per_page, total_records)
        # Simple filename - no subfolder path needed
        page_filename = f"review-page-{page_num}-{current_date}.html"
        
        html_content += f'''
            <a href="{page_filename}" class="page-link">
                Page {page_num}<br>
                Records {start_record}-{end_record}
            </a>'''
    
    html_content += """
        </div>
    </div>
    </div>
    
    <div class="summary">
        <h2>Export All Decisions</h2>
        <p>Export decisions from all pages at once (requires visiting each page first).</p>
        <button onclick="exportAllDecisions()" style="background: #e74c3c; color: white; border: none; padding: 15px 30px; border-radius: 5px; cursor: pointer; font-weight: bold; font-size: 16px;">
            Export All Decisions to CSV
        </button>
    </div>
    
    <script>
        function exportAllDecisions() {
            const catalogerName = prompt('Enter your name for the export file:');
            if (!catalogerName) return;
            
            const allDecisions = [];
            
            // Scan localStorage for all decisions
            for (let i = 0; i < localStorage.length; i++) {
                const key = localStorage.key(i);
                if (key.startsWith('decision-')) {
                    const recordId = key.replace('decision-', '');
                    const decision = localStorage.getItem(key);
                    const notes = localStorage.getItem('notes-' + recordId);
                    
                    // Get enhanced record data if available
                    const recordDataKey = 'record-data-' + recordId;
                    let recordData = null;
                    try {
                        const storedData = localStorage.getItem(recordDataKey);
                        if (storedData) {
                            recordData = JSON.parse(storedData);
                        }
                    } catch (e) {
                        // If parsing fails, recordData stays null
                        console.log('Error parsing record data for record ' + recordId + ':', e);
                    }
                    
                    // Determine correct OCLC based on decision
                    let correctOclc = '';
                    if (decision === 'approved' && recordData && recordData.oclcNumber) {
                        correctOclc = recordData.oclcNumber;
                    }
                    
                    allDecisions.push({
                        recordId: recordId,
                        barcode: recordData ? recordData.barcode : ('Record-' + recordId),
                        confidence: recordData ? recordData.confidence : 'N/A',
                        sortGroup: recordData ? recordData.sortGroup : 'N/A',
                        decision: decision,
                        correctOclc: correctOclc,
                        notes: notes || '',
                        cataloger: catalogerName,
                        reviewDate: new Date().toISOString().split('T')[0],
                        pageNumber: recordData ? recordData.pageNumber : 'Unknown'
                    });
                }
            }
            
            if (allDecisions.length === 0) {
                alert('No decisions found. Please review some records first.');
                return;
            }
            
            // Sort decisions by record ID for consistent ordering
            allDecisions.sort((a, b) => parseInt(a.recordId) - parseInt(b.recordId));
            
            // Create CSV content matching page export format
            const headers = ['Record', 'Barcode', 'Confidence', 'Sort Group', 'Decision', 'Correct OCLC #', 'Notes', 'Cataloger', 'Review Date', 'Page Number'];
            let csvContent = headers.join(',') + '\\n';

            allDecisions.forEach(row => {
                const csvRow = [
                    row.recordId,
                    row.barcode,
                    '"' + row.confidence + '"',
                    '"' + row.sortGroup + '"',
                    '"' + row.decision + '"',
                    '"' + row.correctOclc + '"',
                    '"' + row.notes.replace(/"/g, '""') + '"',
                    '"' + row.cataloger + '"',
                    row.reviewDate,
                    row.pageNumber
                ].join(',');
                csvContent += csvRow + '\\n';
            });
            
            // Download CSV
            const blob = new Blob([csvContent], { type: 'text/csv' });
            const url = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = `all-cataloger-decisions-${catalogerName.replace(/[^a-zA-Z0-9]/g, '_')}-${new Date().toISOString().split('T')[0]}.csv`;
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            window.URL.revokeObjectURL(url);
            
            // Debug information
            console.log(`Found ${allDecisions.length} decisions across all pages`);
            const pageBreakdown = {};
            allDecisions.forEach(decision => {
                const page = decision.pageNumber;
                pageBreakdown[page] = (pageBreakdown[page] || 0) + 1;
            });
            console.log('Decisions per page:', pageBreakdown);
            
            alert(`Exported ${allDecisions.length} decisions to CSV file.`);
        }
    </script>
</body>
</html>"""
    
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

def create_single_review_page(page_path, page_records, current_date, workflow_json_path, images_folder, results_folder, page_num, total_pages, records_per_page, start_idx):
    """Create a single review page with direct image loading."""
    
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>LP Review Page {page_num} - {current_date}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .header {{ background-color: #2c3e50; color: white; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
        .navigation {{ background-color: white; padding: 15px; border-radius: 5px; margin-bottom: 20px; text-align: center; }}
        .nav-btn {{ background-color: #3498db; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin: 0 10px; font-weight: bold; }}
        .nav-btn:hover {{ background-color: #2980b9; }}
        .nav-btn.disabled {{ background-color: #95a5a6; pointer-events: none; }}
        .record {{ background-color: white; border: 1px solid #ddd; border-radius: 8px; margin-bottom: 30px; padding: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .record-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #eee; }}
        .barcode {{ font-size: 24px; font-weight: bold; color: #2c3e50; }}
        .confidence {{ font-size: 18px; font-weight: bold; padding: 8px 15px; border-radius: 20px; color: white; }}
        .sort-group {{ font-size: 16px; font-weight: bold; padding: 6px 12px; border-radius: 15px; margin-left: 10px; }}
        .group-alma-batch-upload-high-confidence {{ background-color: #27ae60; color: white; }}
        .group-cataloger-review-low-confidence {{ background-color: #e74c3c; color: white; }}
        .group-held-by-ut-libraries-ixa {{ background-color: #3498db; color: white; }}
        .group-duplicate {{ background-color: #95a5a6; color: white; }}
        .group-unknown {{ background-color: #95a5a6; color: white; }}
        .confidence-low {{ background-color: #e74c3c; }}
        .confidence-medium {{ background-color: #f39c12; }}
        .confidence-high {{ background-color: #27ae60; }}
        .content-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .images-section {{ display: flex; flex-direction: column; gap: 15px; }}
        .image-container {{ text-align: center; }}
        .image-container img {{ max-width: 100%; height: auto; max-height: 500px; border: 2px solid #ddd; border-radius: 5px; cursor: pointer; transition: transform 0.2s; object-fit: contain; }}
        .image-container img:hover {{ transform: scale(1.05); border-color: #3498db; }}
        .image-label {{ font-weight: bold; margin-bottom: 5px; color: #555; }}
        .oclc-section {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; max-height: 80vh; overflow-y: auto; }}
        .oclc-field {{ margin-bottom: 10px; }}
        .oclc-label {{ font-weight: bold; color: #2c3e50; display: inline-block; width: 150px; }}
        .oclc-value {{ color: #333; }}
        .decision-section {{ grid-column: 1 / -1; margin-top: 20px; padding: 15px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px; }}
        .decision-buttons {{ display: flex; gap: 10px; margin-top: 10px; }}
        .decision-btn {{ padding: 8px 15px; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; }}
        .btn-approve {{ background-color: #27ae60; color: white; }}
        .btn-reject {{ background-color: #e74c3c; color: white; }}
        .btn-review {{ background-color: #f39c12; color: white; }}
        .notes-area {{ width: 100%; margin-top: 10px; padding: 8px; border: 1px solid #ddd; border-radius: 5px; resize: vertical; min-height: 60px; }}
        .no-image {{ color: #999; font-style: italic; text-align: center; padding: 20px; border: 2px dashed #ddd; border-radius: 5px; }}
        .sort-btn {{ transition: background-color 0.3s ease; opacity: 1; }}
        .sort-btn:hover {{ opacity: 0.8; }}
        .sort-btn.active {{ background-color: #3498db !important; transform: scale(1.05); }}
        .sort-btn:not(.active) {{ background-color: #95a5a6 !important; }}
        .sorting-controls {{ background-color: white; padding: 15px; margin-bottom: 20px; border: 1px solid #ddd; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
    </style>
</head>
<body>
    <div class="header">
        <h1>LP Review - Page {page_num} of {total_pages}</h1>
        <p>Generated: {current_date} | Records {start_idx + 1}-{start_idx + len(page_records)} of {(total_pages - 1) * records_per_page + len(page_records)}</p>
    </div>
    
    <div class="navigation">
        <a href="review-index-{current_date}.html" class="nav-btn">Back to Index</a>"""
    
    if page_num > 1:
        html_content += f'<a href="review-page-{page_num - 1}-{current_date}.html" class="nav-btn">← Previous</a>'
    else:
        html_content += '<span class="nav-btn disabled">← Previous</span>'
    
    html_content += f'<span style="margin: 0 20px; font-weight: bold;">Page {page_num} of {total_pages}</span>'
    
    if page_num < total_pages:
        html_content += f'<a href="review-page-{page_num + 1}-{current_date}.html" class="nav-btn">Next →</a>'
    else:
        html_content += '<span class="nav-btn disabled">Next →</span>'
    
    html_content += f"""
    </div>
    
    <div class="sorting-controls">
        <h3 style="margin: 0 0 10px 0; color: #2c3e50;">Sort Records</h3>
        <div style="display: flex; gap: 10px; align-items: center;">
            <button onclick="sortByOriginalOrder()" id="sortOriginal" class="sort-btn active" style="background: #3498db; color: white; border: none; padding: 8px 15px; border-radius: 5px; cursor: pointer; font-weight: bold;">
                Original Order
            </button>
            <button onclick="sortByConfidence()" id="sortConfidence" class="sort-btn" style="background: #95a5a6; color: white; border: none; padding: 8px 15px; border-radius: 5px; cursor: pointer; font-weight: bold;">
                Sort by Confidence (Low to High)
            </button>
            <span style="margin-left: 20px; color: #666; font-size: 14px;">
                Sorting preserves your decisions and notes
            </span>
        </div>
    </div>
    
    <div class="export-controls" style="background-color: white; padding: 15px; margin-bottom: 20px; border: 1px solid #ddd; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
        <h3 style="margin: 0 0 10px 0; color: #2c3e50;">Export Decisions</h3>
        <div style="display: flex; flex-direction: column; gap: 10px;">
            <div style="display: flex; gap: 15px; align-items: center;">
                <label style="display: flex; align-items: center; gap: 5px;">
                    <input type="radio" name="exportType" value="decisions" checked>
                    <span id="decisionsLabel">Export decisions only (<span id="decisionsCount">0</span> records)</span>
                </label>
                <label style="display: flex; align-items: center; gap: 5px;">
                    <input type="radio" name="exportType" value="all">
                    <span>Export all records on this page ({len(page_records)} records)</span>
                </label>
            </div>
            <div style="display: flex; gap: 10px; align-items: center;">
                <button onclick="exportDecisions()" style="background: #e74c3c; color: white; border: none; padding: 10px 20px; border-radius: 5px; cursor: pointer; font-weight: bold;">
                    Export Page CSV
                </button>
                <span style="color: #666; font-size: 14px;">
                    Choose export type above
                </span>
            </div>
        </div>
    </div>
"""

    # Process records for this page
    for i, record in enumerate(page_records, 1):
        global_record_id = start_idx + i
        barcode = record["barcode"]
        oclc_number = record["oclc_number"]
        confidence_score = record.get("confidence_score", 0)
        
        # Get sort group for styling
        sort_group = record.get("sort_group", "Unknown")
        sort_group_class = "group-" + sort_group.lower().replace(" ", "-").replace("(", "").replace(")", "")
        
        # Determine confidence class
        try:
            conf_value = float(confidence_score)
            if conf_value < 30:
                conf_class = "confidence-low"
            elif conf_value < 60:
                conf_class = "confidence-medium"
            else:
                conf_class = "confidence-high"
        except (ValueError, TypeError):
            conf_class = "confidence-low"
            conf_value = 0
        
        # Find image files and copy them to results folder for portability
        image_files = []
        images_subfolder = os.path.join(results_folder, "images")
        os.makedirs(images_subfolder, exist_ok=True)

        if os.path.exists(images_folder):
            for filename in os.listdir(images_folder):
                if filename.startswith(str(barcode)) and filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                    # Copy image to results subfolder for portability
                    src_path = os.path.join(images_folder, filename)
                    dest_path = os.path.join(images_subfolder, filename)
                    
                    try:
                        import shutil
                        # Always try to copy the image
                        shutil.copy2(src_path, dest_path)
                        print(f"Copied image: {filename}")
                        
                        rel_path = os.path.join("images", filename).replace("\\", "/")
                        image_files.append((rel_path, filename))
                        
                    except Exception as copy_error:
                        print(f"Warning: Could not copy image {filename}: {copy_error}")
                        # Don't add to image_files if copy failed
                        continue
        else:
            print(f"Warning: Images folder not found at {images_folder}")

        image_files.sort()
        
        html_content += f"""
    <div class="record" id="record-{global_record_id}" data-barcode="{barcode}" data-oclc-number="{oclc_number}">
        <div class="record-header">
            <div class="barcode">Record {global_record_id}: Barcode {barcode}</div>
            <div>
                <div class="confidence {conf_class}">{conf_value}% Confidence</div>
                <div class="sort-group {sort_group_class}">{sort_group}</div>
            </div>
        </div>
        
        <div class="content-grid">
            <div class="images-section">
                <h3>LP Images</h3>"""
        
        # Add images with direct loading
        if image_files:
            for j, (img_path, filename) in enumerate(image_files[:3]):
                filename_lower = filename.lower()
                if 'a.' in filename_lower:
                    img_type = "Front Cover"
                elif 'b.' in filename_lower:
                    img_type = "Back Cover"
                elif 'c.' in filename_lower:
                    img_type = "Additional Image"
                else:
                    img_type = f"Image {j+1}"
                
                html_content += f"""
                <div class="image-container">
                    <div class="image-label">{img_type}</div>
                    <img src="{img_path}" 
                         alt="{img_type} for barcode {barcode}"
                         onclick="window.open(this.src, '_blank')"
                         onerror="this.style.display='none'; this.nextElementSibling.style.display='block';">
                    <div class="no-image" style="display: none;">Image not found</div>
                </div>"""
        else:
            html_content += '<div class="no-image">No images found for this barcode</div>'
        
        html_content += """
            </div>
            
            <div class="oclc-section">
                <h3>OCLC Record Information</h3>"""
        
        # Add OCLC information from workflow JSON
        if oclc_number and record.get("has_valid_oclc", False):
            oclc_data = get_bib_info_from_workflow(oclc_number, workflow_json_path)
            formatted_record = oclc_data.get("full_record_text", "No detailed record available")
            html_content += f"""
                <pre style="background: #f8f9fa; padding: 15px; border: 1px solid #ddd; border-radius: 5px; overflow: auto; max-height: 70vh; font-size: 13px; white-space: pre-wrap; word-wrap: break-word;">{formatted_record}</pre>"""
        else:
            html_content += """
                <div style="background: #f8f9fa; padding: 15px; border: 1px solid #ddd; border-radius: 5px; color: #666; font-style: italic;">
                    No valid OCLC match found for this record.
                </div>"""
        
        html_content += f"""
            </div>
            
            <div class="decision-section">
                <h3>Cataloger Decision</h3>
                <p>Review the images and OCLC record above, then make your decision:</p>
                
                <div class="decision-buttons">
                    <button class="decision-btn btn-approve" onclick="setDecision({global_record_id}, 'approved', event)">
                        Approve OCLC Match
                    </button>
                    <button class="decision-btn btn-reject" onclick="setDecision({global_record_id}, 'different', event)">
                        Different OCLC Needed
                    </button>
                    <button class="decision-btn btn-reject" onclick="setDecision({global_record_id}, 'original', event)">
                        Needs Original Cataloging
                    </button>
                    <button class="decision-btn btn-review" onclick="setDecision({global_record_id}, 'review', event)">
                        Needs More Review
                    </button>
                </div>
                
                <textarea class="notes-area" placeholder="Notes and correct OCLC number (if different)..." 
                         id="notes-{global_record_id}"></textarea>
            </div>
        </div>
    </div>"""

    # Add JavaScript
    html_content += f"""
    <script>
        let currentSort = 'original';
        const totalRecordsInDataset = {(total_pages - 1) * records_per_page + len(page_records)};
        const pageStartIndex = {start_idx};
        
        // Decision and sorting functions
        function setDecision(recordId, decision, event) {{
            const record = document.getElementById('record-' + recordId);
            const buttons = record.querySelectorAll('.decision-btn');
            buttons.forEach(btn => btn.style.opacity = '0.5');
        
            event.target.style.opacity = '1';
            event.target.style.transform = 'scale(1.05)';
            
            localStorage.setItem('decision-' + recordId, decision);
            document.getElementById('notes-' + recordId).focus();
            
            // Store enhanced record data immediately when decision is made
            const barcode = record.getAttribute('data-barcode');
            const confidenceText = record.querySelector('.confidence').textContent;
            const confidence = confidenceText.replace('% Confidence', '%');
            const sortGroup = record.querySelector('.sort-group').textContent;
            
            // Get OCLC number
            let oclcNumber = '';
            const oclcSection = record.querySelector('.oclc-section pre');
            if (oclcSection) {{
                const oclcText = oclcSection.textContent;
                const oclcMatch = oclcText.match(/OCLC Number: (\\d+)/);
                if (oclcMatch) {{
                    oclcNumber = oclcMatch[1];
                }}
            }}
            if (!oclcNumber) {{
                const recordData = record.dataset;
                if (recordData.oclcNumber && recordData.oclcNumber !== 'None suggested' && recordData.oclcNumber !== '') {{
                    oclcNumber = recordData.oclcNumber;
                }}
            }}
            
            // Store enhanced record data for index page access
            const enhancedRecordData = {{
                barcode: barcode,
                confidence: confidence,
                sortGroup: sortGroup,
                oclcNumber: oclcNumber,
                pageNumber: {page_num}
            }};
            localStorage.setItem('record-data-' + recordId, JSON.stringify(enhancedRecordData));
            updateDecisionCounts();
        }}
        
        function sortByConfidence() {{
            if (currentSort === 'confidence') return;
            
            saveCurrentState();
            
            const records = Array.from(document.querySelectorAll('.record'));
            const recordsWithConfidence = records.map(record => {{
                const confidenceText = record.querySelector('.confidence').textContent;
                const confidence = parseFloat(confidenceText.replace('% Confidence', '')) || 0;
                return {{ element: record, confidence: confidence }};
            }});
            
            recordsWithConfidence.sort((a, b) => a.confidence - b.confidence);
            
            const container = records[0].parentNode;
            recordsWithConfidence.forEach(item => {{
                container.appendChild(item.element);
            }});
            
            document.getElementById('sortOriginal').classList.remove('active');
            document.getElementById('sortConfidence').classList.add('active');
            currentSort = 'confidence';
            
            restoreUserState();
        }}
        
        function sortByOriginalOrder() {{
            if (currentSort === 'original') return;
            
            saveCurrentState();
            
            const records = Array.from(document.querySelectorAll('.record'));
            const container = records[0].parentNode;
            
            records.sort((a, b) => {{
                const aId = parseInt(a.id.replace('record-', ''));
                const bId = parseInt(b.id.replace('record-', ''));
                return aId - bId;
            }});
            
            records.forEach(record => {{
                container.appendChild(record);
            }});
            
            document.getElementById('sortConfidence').classList.remove('active');
            document.getElementById('sortOriginal').classList.add('active');
            currentSort = 'original';
            
            restoreUserState();
        }}
        
        function saveCurrentState() {{
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                const notesElement = document.getElementById('notes-' + i);
                if (notesElement) {{
                    localStorage.setItem('notes-' + i, notesElement.value);
                }}
            }}
        }}
        
        function restoreUserState() {{
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                const decision = localStorage.getItem('decision-' + i);
                if (decision) {{
                    const record = document.getElementById('record-' + i);
                    if (record) {{
                        const buttons = record.querySelectorAll('.decision-btn');
                        buttons.forEach(btn => {{
                            if ((decision === 'approved' && btn.textContent.includes('Approve')) ||
                                (decision === 'different' && btn.textContent.includes('Different')) ||
                                (decision === 'original' && btn.textContent.includes('Original')) ||
                                (decision === 'review' && btn.textContent.includes('More Review'))) {{
                                btn.style.opacity = '1';
                                btn.style.transform = 'scale(1.05)';
                            }} else {{
                                btn.style.opacity = '0.5';
                            }}
                        }});
                    }}
                }}
                
                const notes = localStorage.getItem('notes-' + i);
                const notesElement = document.getElementById('notes-' + i);
                if (notes && notesElement) {{
                    notesElement.value = notes;
                }}
            }}
        }}
        
        // Initialize on page load
        document.addEventListener('DOMContentLoaded', function() {{
            updateDecisionCounts();
            restoreUserState();
        }});

        // Save notes on input
        document.addEventListener('input', function(e) {{
            if (e.target.classList.contains('notes-area')) {{
                const recordId = e.target.id.split('-')[1];
                localStorage.setItem('notes-' + recordId, e.target.value);
            }}
        }});
        // Update decision count when decisions are made
        function updateDecisionCounts() {{
            let decisionsCount = 0;
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                if (localStorage.getItem('decision-' + i)) {{
                    decisionsCount++;
                }}
            }}
            document.getElementById('decisionsCount').textContent = decisionsCount;
        }}
        function exportDecisions() {{
            const catalogerName = prompt('Enter your name for the export file:');
            if (!catalogerName) return;
            const exportType = document.querySelector('input[name="exportType"]:checked').value;
            
            const decisions = [];
            for (let i = pageStartIndex + 1; i <= pageStartIndex + {len(page_records)}; i++) {{
                const decision = localStorage.getItem('decision-' + i);
                const notes = localStorage.getItem('notes-' + i);
                const recordElement = document.getElementById('record-' + i);
                
                if (recordElement) {{
                    // Skip records without decisions if exporting decisions only
                    if (exportType === 'decisions' && !decision) {{
                        continue;
                    }}

                    const barcode = recordElement.getAttribute('data-barcode');
                    const confidenceText = recordElement.querySelector('.confidence').textContent;
                    const confidence = confidenceText.replace('% Confidence', '%');
                    const sortGroup = recordElement.querySelector('.sort-group').textContent;

                    // Get OCLC number from the record data we passed in
                    let oclcNumber = '';
                    
                    // Extract OCLC number from the OCLC section text content
                    const oclcSection = recordElement.querySelector('.oclc-section pre');
                    if (oclcSection) {{
                        const oclcText = oclcSection.textContent;
                        const oclcMatch = oclcText.match(/OCLC Number: (\d+)/);
                        if (oclcMatch) {{
                            oclcNumber = oclcMatch[1];
                        }}
                    }}
                    
                    // If not found in pre section, try to get it from the record data
                    if (!oclcNumber) {{
                        const recordData = recordElement.dataset;
                        if (recordData.oclcNumber && recordData.oclcNumber !== 'None suggested' && recordData.oclcNumber !== '') {{
                            oclcNumber = recordData.oclcNumber;
                        }}
                    }}
                    // Determine correct OCLC based on decision
                    let correctOclc = '';
                    if (decision === 'approved' && oclcNumber) {{
                        correctOclc = oclcNumber;
                    }}
                    
                    // Store enhanced record data in localStorage for index page access
                    var enhancedRecordData = {{
                        barcode: barcode,
                        confidence: confidence,
                        sortGroup: sortGroup,
                        oclcNumber: oclcNumber,
                        pageNumber: {page_num}
                    }};
                    localStorage.setItem('record-data-' + i, JSON.stringify(enhancedRecordData));
                    
                    decisions.push({{
                        record: i,
                        barcode: barcode,
                        confidence: confidence,
                        sortGroup: sortGroup,
                        decision: decision || 'Not reviewed',
                        correctOclc: correctOclc,
                        notes: notes || '',
                        cataloger: catalogerName,
                        reviewDate: new Date().toISOString().split('T')[0],
                        pageNumber: {page_num}
                    }});
                }}
            }}
            
            // Create CSV content
            const headers = ['Record', 'Barcode', 'Confidence', 'Sort Group', 'Decision', 'Correct OCLC #', 'Notes', 'Cataloger', 'Review Date', 'Page Number'];
            let csvContent = headers.join(',') + '\\n';

            decisions.forEach(row => {{
            const csvRow = [
                row.record,
                row.barcode,
                '"' + row.confidence + '"',
                '"' + row.sortGroup + '"',
                '"' + row.decision + '"',
                '"' + row.correctOclc + '"',
                '"' + row.notes.replace(/"/g, '""') + '"',
                '"' + row.cataloger + '"',
                row.reviewDate,
                row.pageNumber
            ].join(',');
            csvContent += csvRow + '\\n';
        }});
            
            const blob = new Blob([csvContent], {{ type: 'text/csv' }});
            const url = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            const exportTypeLabel = exportType === 'decisions' ? 'decisions' : 'all-records';
            a.download = 'cataloger-' + exportTypeLabel + '-page-{page_num}-' + catalogerName.replace(/[^a-zA-Z0-9]/g, '_') + '-' + new Date().toISOString().split('T')[0] + '.csv';
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            window.URL.revokeObjectURL(url);
            
            alert('Exported decisions for ' + decisions.length + ' records to CSV file.');
        }}
    </script>
</body>
</html>"""

    with open(page_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

def move_workflow_data_files(results_folder, data_folder):
    """Move JSON and Excel workflow files to data subfolder"""
    import shutil
    moved_files = 0
    try:
        print(f"Looking for workflow files in: {results_folder}")
        print(f"Target data folder: {data_folder}")
        
        if not os.path.exists(results_folder):
            print(f"Source folder does not exist: {results_folder}")
            return
            
        files_in_results = os.listdir(results_folder)
        print(f"Files in results folder: {len(files_in_results)}")
        
        for filename in files_in_results:
            src = os.path.join(results_folder, filename)
            
            # Move JSON workflow files that start with full-workflow-data-lp-
            if filename.startswith("full-workflow-data-lp-") and filename.endswith(".json"):
                dst = os.path.join(data_folder, filename)
                if os.path.exists(src) and os.path.isfile(src):
                    shutil.move(src, dst)
                    print(f"Moved workflow JSON to: {dst}")
                    moved_files += 1
            
            # Move full workflow Excel files
            elif filename.startswith("full-workflow-data-lp") and filename.endswith(".xlsx"):
                dst = os.path.join(data_folder, filename)
                if os.path.exists(src) and os.path.isfile(src):
                    shutil.move(src, dst)
                    print(f"Moved workflow Excel to: {dst}")
                    moved_files += 1
        
        print(f"Successfully moved {moved_files} workflow data files")
                    
    except Exception as e:
        print(f"Warning: Could not move workflow data files: {e}")
        import traceback
        traceback.print_exc()
    
def find_duplicate_groups(all_records, similarity_threshold=0.9, confidence_threshold=80):
    """
    Find groups of duplicate records based on similar OCLC numbers or titles.
    Only includes records with confidence >= confidence_threshold in duplicate detection.
    
    Args:
        all_records: List of all record dictionaries
        similarity_threshold: Threshold for determining similar titles
        confidence_threshold: Minimum confidence to consider for duplicate detection
    
    Returns:
        List of lists, where each inner list contains duplicate records
    """
    # Only consider records with confidence >= threshold for duplicate detection
    high_confidence_records = []
    high_confidence_indices = []
    
    for i, record in enumerate(all_records):
        try:
            confidence = float(record.get("confidence_score", 0))
            if confidence >= confidence_threshold:
                high_confidence_records.append(record)
                high_confidence_indices.append(i)
        except (ValueError, TypeError):
            # If confidence can't be parsed, treat as low confidence
            pass
    
    duplicate_groups = []
    processed_indices = set()
    
    for i, record in enumerate(high_confidence_records):
        original_index = high_confidence_indices[i]
        if original_index in processed_indices:
            continue
            
        oclc_number = str(record.get("oclc_number", "")).strip()
        title = record.get("title", "")
        
        # Find all duplicates for this record (only among high confidence records)
        duplicates = [record]
        duplicate_original_indices = [original_index]
        
        for j, other_record in enumerate(high_confidence_records[i+1:], i+1):
            other_original_index = high_confidence_indices[j]
            if other_original_index in processed_indices:
                continue
                
            other_oclc = str(other_record.get("oclc_number", "")).strip()
            other_title = other_record.get("title", "")
            
            is_duplicate = False
            
            # Check for similar OCLC numbers (might indicate duplicates)
            try:
                if abs(int(oclc_number) - int(other_oclc)) <= 5:
                    is_duplicate = True
            except (ValueError, TypeError):
                pass
            
            # Check for very similar titles
            if not is_duplicate and title and other_title:
                similarity = calculate_title_similarity(title, other_title)
                if similarity >= similarity_threshold:
                    is_duplicate = True
            
            if is_duplicate:
                duplicates.append(other_record)
                duplicate_original_indices.append(other_original_index)
        
        # Mark all indices as processed
        for idx in duplicate_original_indices:
            processed_indices.add(idx)
        
        # Only add to duplicate groups if there are actually duplicates
        if len(duplicates) > 1:
            duplicate_groups.append(duplicates)
    
    return duplicate_groups

def determine_sort_group_for_duplicates(duplicate_group, confidence_threshold=80):
    """
    Determine sort groups for a group of duplicate records.
    Priority: IXA held > highest confidence > first encountered
    
    Args:
        duplicate_group: List of duplicate record dictionaries
        confidence_threshold: Threshold for high confidence matches
    
    Returns:
        None (modifies records in place)
    """
    if len(duplicate_group) <= 1:
        return
    
    # Sort by priority: IXA held first, then by confidence score (highest first), then by original order
    def sort_key(record):
        held_by_ixa = record.get("held_by_ixa", False)
        try:
            confidence = float(record.get("confidence_score", 0))
        except (ValueError, TypeError):
            confidence = 0
        
        # Return tuple for sorting: (not held_by_ixa, -confidence)
        # This puts IXA held items first, then highest confidence
        return (not held_by_ixa, -confidence)
    
    sorted_duplicates = sorted(duplicate_group, key=sort_key)
    
    # The first item (highest priority) gets its normal classification
    primary_record = sorted_duplicates[0]
    
    # Determine classification for primary record
    if primary_record.get("held_by_ixa", False):
        # For IXA held items, check confidence to determine if it's truly the same item
        try:
            conf_score = float(primary_record.get("confidence_score", 0))
            if conf_score >= confidence_threshold:
                primary_record["sort_group"] = "Held by UT Libraries (IXA)"
            else:
                # Low confidence IXA match - treat as uncertain, not as held item
                primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
        except (ValueError, TypeError):
            primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
    else:
        try:
            conf_score = float(primary_record.get("confidence_score", 0))
            if conf_score >= confidence_threshold:
                primary_record["sort_group"] = "Alma Batch Upload (High Confidence)"
            else:
                primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
        except (ValueError, TypeError):
            primary_record["sort_group"] = "Cataloger Review (Low Confidence)"
    
    # All other records in the group are marked as duplicates
    for record in sorted_duplicates[1:]:
        record["sort_group"] = "Duplicate"

def determine_sort_group(record, confidence_threshold=80):
    """
    Determine the sort group for a non-duplicate record.
    
    Args:
        record: Dictionary containing record information
        confidence_threshold: Threshold for high confidence matches
    
    Returns:
        String indicating the sort group
    """
    held_by_ixa = record.get("held_by_ixa", False)
    confidence_score = record.get("confidence_score", 0)
    
    # Check confidence level first - low confidence items are always for review
    try:
        conf_score = float(confidence_score)
        if conf_score < confidence_threshold:
            return "Cataloger Review (Low Confidence)"
    except (ValueError, TypeError):
        return "Cataloger Review (Low Confidence)"
    
    # Only for high confidence items, check if held by IXA
    if held_by_ixa:
        return "Held by UT Libraries (IXA)"
    else:
        return "Alma Batch Upload (High Confidence)"

def create_all_records_spreadsheet():
    # Get configuration
    file_paths = get_file_path_config()
    step5_config = get_step_config("step5")
    threshold_config = get_threshold_config("confidence")
    
    # Find latest results folder using new utility
    results_folder = find_latest_results_folder(file_paths["results_prefix"])
    if not results_folder:
        print("No results folder found! Please run the previous scripts first.")
        return None
        
    print(f"Using results folder: {results_folder}")
    
    # Create subfolders for organized file output
    guides_folder = os.path.join(results_folder, "guides")
    deliverables_folder = os.path.join(results_folder, "deliverables") 
    data_folder = os.path.join(results_folder, "data")

    # Create directories if they don't exist
    for folder in [guides_folder, deliverables_folder, data_folder]:
        os.makedirs(folder, exist_ok=True)
    
    # Initialize workflow JSON path
    workflow_json_path = get_workflow_json_path(results_folder)
    
    step4_file = find_latest_lp_metadata_file(results_folder)
    if not step4_file:
        print("No full-workflow-data-lp file found in the results folder!")

    print(f"Using source file: {step4_file}")

    try: 
        # Open the latest step 4 workbook
        wb_src = load_workbook(step4_file)
        sheet_src = wb_src.active

        # Create a new workbook for all records
        wb_new = openpyxl.Workbook()
        sheet_new = wb_new.active

        # Set header row
        header_row = ["Barcode", "Sort Group", "OCLC Number", "OCLC Title", "OCLC Author", "OCLC Date of Publication", "Confidence Score"]
        sheet_new.append(header_row)

        # Set column widths
        sheet_new.column_dimensions['A'].width = 16  # Barcode
        sheet_new.column_dimensions['B'].width = 28  # Sort Group
        sheet_new.column_dimensions['C'].width = 12  # OCLC Number
        sheet_new.column_dimensions['D'].width = 50  # OCLC Title
        sheet_new.column_dimensions['E'].width = 50  # OCLC Author
        sheet_new.column_dimensions['F'].width = 20  # OCLC Date
        sheet_new.column_dimensions['G'].width = 15  # Confidence Score

        # Get the column indices from the source workbook
        # In source: Column D=Barcode, Column H=OCLC Number, Column I=Confidence Score
        BARCODE_COL_IDX = 4  # Column D
        OCLC_NUM_COL_IDX = 8  # Column H
        CONF_SCORE_COL_IDX = 9  # Column I

        # First pass: collect all records with basic info
        all_records = []
        all_records_dict = {}  # For duplicate detection
        
        print("First pass: Collecting all records...")
        for row_idx in range(2, sheet_src.max_row + 1):  # Skip header row
            barcode = sheet_src.cell(row=row_idx, column=BARCODE_COL_IDX).value
            oclc_number = sheet_src.cell(row=row_idx, column=OCLC_NUM_COL_IDX).value
            confidence_score = sheet_src.cell(row=row_idx, column=CONF_SCORE_COL_IDX).value
            
            # Skip rows with missing barcode (essential identifier)
            if not barcode:
                continue
            
            # Handle records with no OCLC number or invalid OCLC numbers
            oclc_number_str = ""
            has_valid_oclc = False
            
            if oclc_number and str(oclc_number).strip() not in ["", "Not found", "Error processing"]:
                oclc_number_str = str(oclc_number).strip()
                has_valid_oclc = True
            
            record = {
                "barcode": barcode,
                "oclc_number": oclc_number_str,
                "confidence_score": confidence_score if confidence_score else 0,
                "title": "",
                "author": "",
                "publication_date": "",
                "held_by_ixa": False,
                "sort_group": "",
                "has_valid_oclc": has_valid_oclc
            }
            
            all_records.append(record)
            
            # Only add to dict for duplicate detection if it has a valid OCLC number
            if has_valid_oclc:
                all_records_dict[oclc_number_str] = record
            
        print(f"Found {len(all_records)} records to process.")

        # Second pass: Get OCLC data and holdings information from workflow JSON
        print("Second pass: Getting OCLC data from workflow JSON...")
        processed_count = 0
        
        for record in all_records:
            oclc_number = record["oclc_number"]
            has_valid_oclc = record["has_valid_oclc"]
            
            print(f"Processing record {processed_count + 1}/{len(all_records)} - Barcode: {record['barcode']}")
            
            if has_valid_oclc:
                print(f"  OCLC Number: {oclc_number}")
                
                # Get bibliographic information from workflow JSON
                oclc_data = get_bib_info_from_workflow(oclc_number, workflow_json_path)
                record["title"] = extract_title_from_bib_info(oclc_data)
                record["author"] = extract_author_from_bib_info(oclc_data)
                record["publication_date"] = extract_publication_date_from_bib_info(oclc_data)
                
                # Get holdings information from workflow JSON
                holdings_info = get_holdings_info_from_workflow(oclc_number, workflow_json_path)
                record["held_by_ixa"] = holdings_info["held_by_ixa"]
                
                print(f"  Holdings from workflow: IXA={holdings_info['held_by_ixa']}")
                
                # No API delay needed since we're reading from local JSON
            else:
                print(f"  No valid OCLC number - using defaults")
                record["title"] = "No OCLC match found"
                record["author"] = "No author available"
                record["publication_date"] = "No date available"
                record["held_by_ixa"] = False
            
            processed_count += 1
        
        # Third pass: Handle duplicates and determine sort groups
        print("Third pass: Finding duplicate groups...")
        
        # Only look for duplicates among records with valid OCLC numbers AND high confidence
        records_with_oclc = [record for record in all_records if record["has_valid_oclc"]]
        duplicate_groups = find_duplicate_groups(records_with_oclc, confidence_threshold=80)
        
        print(f"Found {len(duplicate_groups)} duplicate groups (high confidence only)")
        for i, group in enumerate(duplicate_groups):
            print(f"  Group {i+1}: {len(group)} duplicates")
        
        # Process duplicate groups
        print("Processing duplicate groups...")
        for group in duplicate_groups:
            determine_sort_group_for_duplicates(group, confidence_threshold=80)
        
        # Process remaining records (including low confidence items that were excluded from duplicate detection)
        print("Processing non-duplicate records...")
        processed_records = set()
        for group in duplicate_groups:
            for record in group:
                processed_records.add(id(record))
        
        for record in all_records:
            if id(record) not in processed_records:
                if record["has_valid_oclc"]:
                    record["sort_group"] = determine_sort_group(record, confidence_threshold=80)
                else:
                    # Records without valid OCLC numbers are always low confidence
                    record["sort_group"] = "Cataloger Review (Low Confidence)"
        
        # Add all records to spreadsheet and log to JSON workflow
        for record in all_records:
            sheet_new.append([
                record["barcode"],
                record["sort_group"],
                record["oclc_number"],
                record["title"],
                record["author"],
                record["publication_date"],
                record["confidence_score"]
            ])
            
            # Update JSON workflow with Step 5 results
            try:
                is_duplicate = record["sort_group"] == "Duplicate"
                update_record_step5(
                    json_path=workflow_json_path,
                    barcode=str(record["barcode"]),
                    sort_group=record["sort_group"],
                    final_oclc_number=record["oclc_number"],
                    is_duplicate=is_duplicate,
                    oclc_title=record["title"],
                    oclc_author=record["author"],
                    oclc_date=record["publication_date"]
                )
            except Exception as json_error:
                print(f"   JSON logging error for {record['barcode']}: {json_error}")
                log_error(
                    results_folder_path=results_folder,
                    step="step5_final_classification",
                    barcode=str(record["barcode"]),
                    error_type="json_update_error",
                    error_message=str(json_error)
                )
        
        # Save the all records spreadsheet
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        all_records_file = f"sorting-spreadsheet-{current_date}.xlsx"
        all_records_path = os.path.join(deliverables_folder, all_records_file)
        wb_new.save(all_records_path)
        
        print(f"Sort physical items with {len(all_records)} records: {all_records_path}")
        
        # Create summary statistics
        sort_group_counts = {}
        for record in all_records:
            group = record["sort_group"]
            sort_group_counts[group] = sort_group_counts.get(group, 0) + 1
        
        print("\nSort Group Summary:")
        for group, count in sorted(sort_group_counts.items()):
            print(f"  {group}: {count} records")
        
        # Create high confidence matches for Alma batch upload (text file only)
        unique_matches = [record for record in all_records 
                         if record["sort_group"] == "Alma Batch Upload (High Confidence)"]
        
        current_timestamp = get_current_timestamp()
        text_file = FILE_NAMING["batch_upload_alma"].format(timestamp=current_timestamp)
        text_path = os.path.join(deliverables_folder, text_file)
        
        with open(text_path, 'w', newline='', encoding='utf-8') as f:
            for record in unique_matches:
                line = f"{record['oclc_number']}|{record['barcode']}|{record['title']}\n"
                f.write(line)
        
        print(f"Pipe-delimited text file for Alma processing created: {text_path}")
        
        # Create low confidence review spreadsheet
        review_path = create_low_confidence_review_text_log(
            results_folder, step4_file, all_records, workflow_json_path, current_date
        )
        
        # Create MARC format text log for low confidence records
        marc_path = create_marc_format_text_log(
            results_folder, all_records, workflow_json_path, current_date
        )
        
        # Create cataloger review spreadsheet
        review_spreadsheet_path = create_cataloger_review_spreadsheet(
            results_folder, all_records, current_date
        )
        
        # Create paginated easy review HTML files with lazy loading
        paginated_review_result = create_paginated_review_html(
            results_folder, all_records, current_date, workflow_json_path, records_per_page=100
        )

        # Copy both guides to guides subfolder
        try:
            import shutil
            script_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Try multiple potential locations for the guides
            potential_locations = [
                os.path.dirname(os.path.dirname(script_dir)),  # Two levels up
                os.path.dirname(script_dir),  # One level up  
                script_dir,  # Same directory as script
                os.getcwd()  # Current working directory
            ]
            
            guides_found = False
            for project_root in potential_locations:
                cataloger_guide_source = os.path.join(project_root, "CATALOGER_GUIDE.txt")
                technical_guide_source = os.path.join(project_root, "TECHNICAL_GUIDE.txt")
                
                if os.path.exists(cataloger_guide_source):
                    cataloger_guide_dest = os.path.join(guides_folder, "CATALOGER_GUIDE.txt")
                    shutil.copy2(cataloger_guide_source, cataloger_guide_dest)
                    print(f"Cataloger guide copied to: {cataloger_guide_dest}")
                    guides_found = True
                
                if os.path.exists(technical_guide_source):
                    technical_guide_dest = os.path.join(guides_folder, "TECHNICAL_GUIDE.txt")
                    shutil.copy2(technical_guide_source, technical_guide_dest)
                    print(f"Technical guide copied to: {technical_guide_dest}")
                    guides_found = True
                
                if guides_found:
                    break
            
            if not guides_found:
                print("Warning: Guide files not found in any expected location")
                
        except Exception as e:
            print(f"Warning: Could not copy guides: {e}")
            
        # Log final Step 5 processing metrics
        try:
            step5_metrics = create_batch_summary(
                total_items=len(all_records),
                successful_items=len([r for r in all_records if r["sort_group"] != ""]),
                failed_items=len([r for r in all_records if r["sort_group"] == ""]),
                total_time=0,  # Step 5 doesn't track detailed timing
                total_tokens=0,
                estimated_cost=0,
                processing_mode="CLASSIFICATION"
            )
            
            # Add step-specific metrics
            step5_metrics.update({
                "sort_group_counts": sort_group_counts,
                "unique_matches_count": len(unique_matches),
                "duplicate_groups_found": len(duplicate_groups) if 'duplicate_groups' in locals() else 0,
                "records_with_valid_oclc": len([r for r in all_records if r["has_valid_oclc"]]),
                "step": "step5_final_classification"
            })
            
            log_processing_metrics(
                results_folder_path=results_folder,
                step="step5_final_classification", 
                batch_metrics=step5_metrics
            )
            
        except Exception as metrics_error:
            print(f"Warning: Could not log Step 5 processing metrics: {metrics_error}")
            
        # Move workflow data files to data subfolder after all processing is complete
        print("Moving workflow data files to data subfolder...")
        move_workflow_data_files(results_folder, data_folder)
            
        return {
            "all_records_path": all_records_path,
            "text_file_path": text_path,
            "review_path": review_path,
            "marc_path": marc_path,
            "review_spreadsheet_path": review_spreadsheet_path,
            "paginated_review": paginated_review_result,
            "guide_path": os.path.join(guides_folder, "CATALOGER_GUIDE.txt"),
            "total_records": len(all_records),
            "sort_group_counts": sort_group_counts,
            "unique_matches_count": len(unique_matches)
        }
        
    except Exception as e:
        print(f"Error creating all records spreadsheet: {str(e)}")
        
        # Log error to JSON workflow
        try:
            log_error(
                results_folder_path=results_folder,
                step="step5_final_classification",
                barcode="unknown",
                error_type="processing_error",
                error_message=str(e)
            )
        except Exception as json_error:
            print(f"JSON error logging failed: {json_error}")
        
        return None

def main():
    result = create_all_records_spreadsheet()
    if result:
        print(f"\n=== All Records Processing Complete ===")
        print(f"Total records processed: {result['total_records']}")
        print(f"All records file: {result['all_records_path']}")
        print(f"Alma processing file: {result['text_file_path']}")
        if result['review_path']:
            print(f"Low confidence review file: {result['review_path']}")
        if result.get('review_spreadsheet_path'):
            print(f"Cataloger review spreadsheet: {result['review_spreadsheet_path']}")
        if result.get('marc_path'):
            print(f"MARC format file: {result['marc_path']}")
        if result.get('paginated_review'):
            paginated_info = result['paginated_review']
            print(f"Paginated review index: {paginated_info['index_path']}")
            print(f"Created {paginated_info['total_pages']} review pages")
        print(f"Unique high-confidence matches: {result['unique_matches_count']}")
        print("\nBreakdown by Sort Group:")
        for group, count in sorted(result['sort_group_counts'].items()):
            print(f"  {group}: {count}")
    else:
        print("Failed to create all records spreadsheet.")

if __name__ == "__main__":
    main()