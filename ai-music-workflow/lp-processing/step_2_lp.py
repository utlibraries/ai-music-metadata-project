# Query OCLC API with the extracted metadata
import os
import json
import requests
import time
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import re

# Custom modules
from json_workflow import update_record_step2, log_oclc_api_search, log_error, log_processing_metrics
from shared_utilities import find_latest_results_folder, get_workflow_json_path, extract_metadata_fields
from lp_workflow_config import get_file_path_config
    
api_calls = {'count': 0, 'reset_time': time.time()}

def get_access_token(client_id, client_secret):
    token_url = "https://oauth.oclc.org/token"
    data = {
        "grant_type": "client_credentials",
        "scope": "wcapi"
    }
    response = requests.post(token_url, data=data, auth=(client_id, client_secret))
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        raise Exception(f"Failed to get access token: {response.text}")

def construct_queries_from_metadata(metadata, workflow_data=None, barcode=None):
    """Generate all possible query combinations from the JSON structure without limiting to just 5."""
    def safe_get(value):
        if not value or not isinstance(value, str):
            return None
        if any(x in value.lower() for x in ["not visible", "not available", "n/a", "unavailable", "unknown", " [none]", "none", "not present", "not listed", "not applicable"]):
            return None
        cleaned = re.sub(r'^-\s', '', value.strip())
        cleaned = re.sub(r'^(Primary Contributor:|Artist/Performer:|Name:)\s', '', cleaned)
        return cleaned if cleaned else None

    # Extract from JSON structure
    title_info = metadata.get('title_information', {})
    title = safe_get(title_info.get('main_title'))
    subtitle = safe_get(title_info.get('subtitle'))
    artist = safe_get(title_info.get('primary_contributor'))
    
    publishers = metadata.get('publishers', {})
    publisher = safe_get(publishers.get('name'))
    pub_numbers = safe_get(publishers.get('numbers'))
    
    dates = metadata.get('dates', {})
    pub_year = safe_get(dates.get('publication_date'))
    
    # Extract tracks from JSON structure
    contents = metadata.get('contents', {})
    tracks = contents.get('tracks', [])
    
    first_track = None
    second_track = None
    third_track = None
    
    if len(tracks) > 0:
        first_track = safe_get(tracks[0].get('title'))
    if len(tracks) > 1:
        second_track = safe_get(tracks[1].get('title'))
    if len(tracks) > 2:
        third_track = safe_get(tracks[2].get('title'))

    queries = []
    
    # PRIORITY 1: Get cleaned UPC/catalog numbers from Step 1.5 if available
    cleaned_numbers = []
    if workflow_data and barcode:
        barcode_str = str(barcode)
        if (isinstance(workflow_data, dict) and 
            "records" in workflow_data and 
            barcode_str in workflow_data["records"]):
            
            step1_5_data = workflow_data["records"][barcode_str].get("step1_5_metadata_cleaning", {})
            valid_numbers = step1_5_data.get("valid_numbers_extracted")
            
            if valid_numbers and isinstance(valid_numbers, str):
                # Split multiple numbers if comma-separated
                for num in valid_numbers.split(','):
                    num = num.strip()
                    if num:
                        cleaned_numbers.append(num)
    
    # PRIORITY 2: Fallback to extracting from metadata if no Step 1.5 data
    if not cleaned_numbers and isinstance(pub_numbers, str):
        # Look for UPC/EAN labeled codes first
        labeled_matches = re.finditer(r'(UPC|EAN):\s*([^,\]]+)', pub_numbers, re.IGNORECASE)
        for match in labeled_matches:
            potential_code = match.group(2).strip()
            digits_only = re.sub(r'\D', '', potential_code)
            if len(digits_only) in [12, 13]:
                cleaned_numbers.append(digits_only)
        
        # If no labeled codes found, look for digit sequences
        if not cleaned_numbers:
            code_candidates = re.findall(r'\d[\d\s-]{10,}\d', pub_numbers)
            for candidate in code_candidates:
                digits_only = re.sub(r'\D', '', candidate)
                if len(digits_only) in [12, 13]:
                    cleaned_numbers.append(digits_only)
        
        # If still no UPCs found, treat the whole thing as a catalog number
        if not cleaned_numbers and pub_numbers:
            cleaned_numbers.append(pub_numbers)

    # PRIORITY QUERIES: Add cleaned numbers first (UPCs or catalog numbers)
    for number in cleaned_numbers:
        # Check if it's a UPC (12-13 digits only)
        digits_only = re.sub(r'\D', '', number)
        if len(digits_only) in [12, 13] and digits_only == number.replace(' ', ''):
            # It's a UPC - add as-is
            queries.append(number.replace(' ', ''))
        else:
            # It's a catalog number - add as-is (no quotes needed for single terms)
            queries.append(number)

    # Continue with existing query patterns using cleaned_numbers instead of product_codes
    if artist and first_track and second_track:
        queries.append(f'"{artist}" "{first_track}" "{second_track}"')
        
    if first_track and second_track:
        queries.append(f'"{first_track}" "{second_track}"')
        
    if title:
        if all([title, subtitle, artist]):
            queries.append(f'"{title}" "{subtitle}" "{artist}"')
            
        if all([title, first_track, second_track]):
            queries.append(f'"{title}" "{first_track}" "{second_track}"')
            
        if all([title, artist, publisher]):
            queries.append(f'"{title}" "{artist}" {publisher}')

        if title and artist:
            queries.append(f'"{title}" "{artist}"')

        if all([title, first_track]):
            queries.append(f'"{title}" "{first_track}"')
            
        if all([title, subtitle]):
            queries.append(f'"{title}" "{subtitle}"')

        if title and publisher:
            queries.append(f'"{title}" {publisher}')

        # Use the first cleaned number if available
        if title and cleaned_numbers:
            queries.append(f'"{title}" {cleaned_numbers[0]}')
        
    # Use the first cleaned number in combination queries
    if artist and publisher and cleaned_numbers:
        queries.append(f'"{artist}" {publisher} {cleaned_numbers[0]}')

    if artist and publisher and pub_year:
        queries.append(f'{artist} {publisher} {pub_year}')

    if artist and publisher and first_track:
        queries.append(f'{artist} {publisher} {first_track}')
        
    if artist and second_track:
        queries.append(f'{artist} {second_track}')
    
    if first_track:
        queries.append(f'"{first_track}"')

    if second_track:
        queries.append(f'"{second_track}"')

    if artist:
        queries.append(f'{artist}')

    # Remove duplicates while preserving order
    seen = set()
    unique_queries = []
    for q in queries:
        if q not in seen:
            seen.add(q)
            unique_queries.append(q)

    return unique_queries

def format_oclc_results(json_response, access_token):
    try:
        data = json.loads(json_response)
        if not isinstance(data, dict):
            return "Error: Invalid JSON response"
            
        total_records = data.get('numberOfRecords', 0)
        if total_records == 0:
            return "No records found"
            
        formatted_results = []
        valid_records = []
        
        for record in data.get('bibRecords', []):
            include_record = False
            if 'format' in record and 'specificFormat' in record['format']:
                specific_format = record['format']['specificFormat']
                if isinstance(specific_format, str) and any(lp_term in specific_format.lower() for lp_term in ["lp", "vinyl", "long play", "audiobook", "spoken", "spoken word", "talking book", "sound recording"]):
                    include_record = True
            
            if include_record:
                valid_records.append(record)
        
        filtered_total = len(valid_records)
        if filtered_total == 0:
            return "No matching records with LP format found"

        formatted_results.append(f"Total Records Found (LP format only): {filtered_total}\n")

        for idx, record in enumerate(valid_records[:5], 1):
            formatted_results.append(f"\nRecord {idx}:")
            formatted_results.append("-" * 40)
            
            oclc_number = None
            if 'identifier' in record and 'oclcNumber' in record['identifier']:
                oclc_number = record['identifier']['oclcNumber']
                formatted_results.append(f"OCLC Number: {oclc_number}")
            
            if oclc_number:
                # FIXED: Now correctly unpacking all 3 return values
                is_held_by_IXA, total_holding_count, holding_institutions = get_holdings_info(oclc_number, access_token)
                formatted_results.append(f"\nHeld by IXA: {'Yes' if is_held_by_IXA else 'No'}")
                formatted_results.append(f"Total Institutions Holding: {total_holding_count}")
            
            if 'identifier' in record:
                formatted_results.append("\nIdentifier:")
                # Add OCLC number
                if 'oclcNumber' in record['identifier']:
                    formatted_results.append(f"  - oclcNumber: {record['identifier']['oclcNumber']}")
                
                # Add catalog numbers from standardNumbers (028 field)
                if 'standardNumbers' in record['identifier']:
                    for std_num in record['identifier']['standardNumbers']:
                        if isinstance(std_num, dict):
                            number = std_num.get('standardNumber', '')
                            source = std_num.get('source', '')
                            if number:
                                if source:
                                    formatted_results.append(f"  - Catalog Number: {number} ({source})")
                                else:
                                    formatted_results.append(f"  - Catalog Number: {number}")
                
                # Add UPC if it exists
                if 'otherStandardIdentifiers' in record['identifier']:
                    for id_item in record['identifier']['otherStandardIdentifiers']:
                        if isinstance(id_item, dict) and id_item.get('type') == 'Universal Product Code (UPC)':
                            formatted_results.append(f"  - UPC: {id_item.get('id', 'N/A')}")
            
            if 'title' in record:
                formatted_results.append("Title Information:")
                if 'mainTitles' in record['title']:
                    for title in record['title']['mainTitles']:
                        formatted_results.append(f"  - Main Title: {title.get('text', 'N/A')}")
                if 'subtitles' in record['title']:
                    for subtitle in record['title']['subtitles']:
                        formatted_results.append(f"  - Subtitle: {subtitle.get('text', 'N/A')}")
            
            if 'contributor' in record:
                formatted_results.append("Contributors:")
                for creator_type in ['creators', 'contributors']:
                    if creator_type in record['contributor']:
                        for person in record['contributor'][creator_type]:
                            if 'firstName' in person and 'secondName' in person:
                                name = f"{person.get('firstName', {}).get('text', '')} {person.get('secondName', {}).get('text', '')}"
                            elif 'nonPersonName' in person:
                                name = person['nonPersonName'].get('text', '')
                            else:
                                name = 'N/A'
                            role = person.get('type', 'N/A')
                            formatted_results.append(f"  - {name.strip()} ({role})")
                        
            if 'publishers' in record:
                formatted_results.append("Publishers:")
                for pub in record['publishers']:
                    pub_name = pub.get('publisherName', {}).get('text', 'N/A')
                    pub_place = pub.get('publicationPlace', 'N/A')
                    formatted_results.append(f"  - Name: {pub_name}")
                    formatted_results.append(f"    Place: {pub_place}")
            
            if 'date' in record:
                formatted_results.append("Dates:")
                # Only include publicationDate
                if 'publicationDate' in record['date']:
                    formatted_results.append(f"  - publicationDate: {record['date']['publicationDate']}")
            
            if 'language' in record:
                formatted_results.append("Language:")
                for key, value in record['language'].items():
                    formatted_results.append(f"  - {key}: {value}")
                        
            if 'description' in record:
                formatted_results.append("Description:")
                if 'physicalDescription' in record['description']:
                    formatted_results.append(f"  - Physical: {record['description']['physicalDescription']}")
                if 'contents' in record['description']:
                    for content in record['description']['contents']:
                        if 'contentNote' in content:
                            formatted_results.append(f"  - Content: {content['contentNote'].get('text', '')}")
                        
            formatted_results.append("-" * 40)
            
        return "\n".join(formatted_results)
        
    except json.JSONDecodeError:
        return "Error: Invalid JSON response"
    except Exception as e:
        return f"Error formatting results: {str(e)}"
    
def format_oclc_api_response_for_accumulation(data, access_token, seen_oclc_numbers=None):
    if seen_oclc_numbers is None:
        seen_oclc_numbers = set()
        
    try:
        if not isinstance(data, dict):
            return "Error: Invalid JSON response"
            
        total_records = data.get('numberOfRecords', 0)
        if total_records == 0:
            return "No records found"
            
        formatted_results = []
        valid_records = []
        
        for record in data.get('bibRecords', []):
            # Skip records we've already seen
            oclc_number = None
            if 'identifier' in record and 'oclcNumber' in record['identifier']:
                oclc_number = record['identifier']['oclcNumber']
                if oclc_number in seen_oclc_numbers:
                    continue
            
            include_record = False
            if 'format' in record and 'specificFormat' in record['format']:
                specific_format = record['format']['specificFormat']
                if isinstance(specific_format, str) and any(lp_term in specific_format.lower() for lp_term in ["lp", "vinyl", "long play", "audiobook", "spoken", "spoken word", "talking book", "sound recording"]):
                    include_record = True
            
            if include_record:
                valid_records.append(record)
        
        filtered_total = len(valid_records)
        if filtered_total == 0:
            return "No matching records with LP format found"

        for record in valid_records[:5]:
            # Add a divider line between records
            formatted_results.append("\n" + "-" * 40)
            
            oclc_number = None
            if 'identifier' in record and 'oclcNumber' in record['identifier']:
                oclc_number = record['identifier']['oclcNumber']
                formatted_results.append(f"OCLC Number: {oclc_number}")
            
            if oclc_number:
                is_held_by_IXA, total_holding_count, holding_institutions = get_holdings_info(oclc_number, access_token)
                formatted_results.append(f"\nHeld by IXA: {'Yes' if is_held_by_IXA else 'No'}")
                formatted_results.append(f"Total Institutions Holding: {total_holding_count}")
            
            if 'identifier' in record:
                formatted_results.append("\nIdentifier:")
                # Add OCLC number
                if 'oclcNumber' in record['identifier']:
                    formatted_results.append(f"  - oclcNumber: {record['identifier']['oclcNumber']}")
                
                # Add catalog numbers from standardNumbers (028 field)
                if 'standardNumbers' in record['identifier']:
                    for std_num in record['identifier']['standardNumbers']:
                        if isinstance(std_num, dict):
                            number = std_num.get('standardNumber', '')
                            source = std_num.get('source', '')
                            if number:
                                if source:
                                    formatted_results.append(f"  - Catalog Number: {number} ({source})")
                                else:
                                    formatted_results.append(f"  - Catalog Number: {number}")
                
                # Add UPC if it exists
                if 'otherStandardIdentifiers' in record['identifier']:
                    for id_item in record['identifier']['otherStandardIdentifiers']:
                        if isinstance(id_item, dict) and id_item.get('type') == 'Universal Product Code (UPC)':
                            formatted_results.append(f"  - UPC: {id_item.get('id', 'N/A')}")
            
            if 'title' in record:
                formatted_results.append("Title Information:")
                if 'mainTitles' in record['title']:
                    for title in record['title']['mainTitles']:
                        formatted_results.append(f"  - Main Title: {title.get('text', 'N/A')}")
                if 'subtitles' in record['title']:
                    for subtitle in record['title']['subtitles']:
                        formatted_results.append(f"  - Subtitle: {subtitle.get('text', 'N/A')}")
                if 'seriesTitles' in record['title']:
                    for series in record['title']['seriesTitles']:
                        formatted_results.append(f"  - Series Title: {series.get('seriesTitle', 'N/A')}")
            
            if 'contributor' in record:
                formatted_results.append("Contributors:")
                for creator_type in ['creators', 'contributors']:
                    if creator_type in record['contributor']:
                        for person in record['contributor'][creator_type]:
                            if 'firstName' in person and 'secondName' in person:
                                name = f"{person.get('firstName', {}).get('text', '')} {person.get('secondName', {}).get('text', '')}"
                            elif 'nonPersonName' in person:
                                name = person['nonPersonName'].get('text', '')
                            else:
                                name = 'N/A'
                            role = person.get('type', 'N/A')
                            formatted_results.append(f"  - {name.strip()} ({role})")
            
            if 'publishers' in record:
                formatted_results.append("Publishers:")
                for pub in record['publishers']:
                    pub_name = pub.get('publisherName', {}).get('text', 'N/A')
                    pub_place = pub.get('publicationPlace', 'N/A')
                    formatted_results.append(f"  - Name: {pub_name}")
                    formatted_results.append(f"    Place: {pub_place}")
            
            if 'date' in record:
                formatted_results.append("Dates:")
                # Only include publicationDate
                if 'publicationDate' in record['date']:
                    formatted_results.append(f"  - publicationDate: {record['date']['publicationDate']}")
            
            if 'language' in record:
                formatted_results.append("Language:")
                for key, value in record['language'].items():
                    formatted_results.append(f"  - {key}: {value}")
                        
            if 'musicInfo' in record:
                formatted_results.append("Music Information:")
                for key, value in record['musicInfo'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            if 'description' in record:
                formatted_results.append("Description:")
                if 'physicalDescription' in record['description']:
                    formatted_results.append(f"  - Physical: {record['description']['physicalDescription']}")
                if 'contents' in record['description']:
                    for content in record['description']['contents']:
                        if 'contentNote' in content:
                            content_text = content['contentNote'].get('text', '')
                            # Smart content handling for large multi-disc sets
                            if len(content_text) > 1500:
                                # Count structural elements
                                disc_count = content_text.count('Disc ')
                                chapter_count = content_text.count('Chapter ')
                                track_patterns = len(re.findall(r'(?:--|\d+\.|\(\d+:\d+\))', content_text))
                                
                                # Determine if this is a large compilation
                                is_large_compilation = (
                                    disc_count > 4 or 
                                    chapter_count > 20 or 
                                    track_patterns > 100 or
                                    len(content_text) > 5000
                                )
                                
                                if is_large_compilation:
                                    # Extract sample tracks and artists
                                    track_pattern = r'([^-\n]+?)\s*(?:\(\d+:\d+\)|--)'
                                    sample_tracks = re.findall(track_pattern, content_text[:2000])
                                    sample_tracks = [t.strip() for t in sample_tracks[:10] if len(t.strip()) > 3]
                                    
                                    summary = f"LARGE MULTI-DISC COMPILATION: {disc_count} discs"
                                    if chapter_count > 0:
                                        summary += f", {chapter_count} chapters"
                                    summary += ". "
                                    if sample_tracks:
                                        summary += f"Sample tracks: {', '.join(sample_tracks)}. "
                                    summary += f"[Original: {len(content_text):,} characters]"
                                    content_text = summary
                                else:
                                    content_text = content_text[:1500]
                                    last_break = content_text.rfind(' -- ')
                                    if last_break > 1000:
                                        content_text = content_text[:last_break]
                                    content_text += "... [Content truncated for analysis]"
                            
                            formatted_results.append(f"  - Content: {content_text}")
                            
            if 'note' in record:
                formatted_results.append("Notes:")
                if isinstance(record['note'], dict):
                    for key, value in record['note'].items():
                        formatted_results.append(f"  - {key}: {value}")
                elif isinstance(record['note'], list):
                    for note in record['note']:
                        formatted_results.append(f"  - {note}")
                                    
            formatted_results.append("-" * 40)
            
        return "\n".join(formatted_results), filtered_total, None
        
    except Exception as e:
        return f"Error formatting results: {str(e)}", 0, None
    
def get_holdings_info(oclc_number, access_token):
    global api_calls
    api_calls['count'] += 1
    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs-holdings"
    
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    
    params = {
        "oclcNumber": oclc_number,
        "limit": 50
    }
    
    try:
        response = requests.get(endpoint, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        is_held_by_IXA = False
        total_holding_count = 0
        holding_institutions = []
        
        if "briefRecords" in data and len(data["briefRecords"]) > 0:
            record = data["briefRecords"][0]
            
            if "institutionHolding" in record:
                holdings = record["institutionHolding"]
                total_holding_count = holdings.get("totalHoldingCount", 0)
                
                if "briefHoldings" in holdings:
                    for holding in holdings["briefHoldings"]:
                        symbol = holding.get("oclcSymbol", "")
                        inst_name = holding.get("institutionName", "")
                        if inst_name:
                            import html
                            inst_name = html.unescape(inst_name)
                        formatted_holding = f"{symbol} ({inst_name})" if inst_name else symbol
                        holding_institutions.append(formatted_holding)
                        if symbol == "IXA":
                            is_held_by_IXA = True
        
        return is_held_by_IXA, total_holding_count, holding_institutions
    
    except requests.RequestException as e:
        print(f"Error getting holdings for OCLC number {oclc_number}: {str(e)}")
        return False, 0, []
    
def truncate_contributors(performers, max_performers=3):
    return performers[:max_performers]

def remove_non_latin(text):
    cleaned = re.sub(r'[^\w\s\-\/\(\)áéíóúãõñâêîôûÁÉÍÓÚÃÕÑÂÊÎÔÛ]', '', text)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = re.sub(r'\(\s*\)', '', cleaned)
    return cleaned.strip()

def query_oclc_api(queries, barcode, limit=10):
    global api_calls
    current_time = time.time()
    if current_time - api_calls['reset_time'] >= 86400:
        api_calls = {'count': 0, 'reset_time': current_time}

    if api_calls['count'] >= 50000:
        return "Rate limit exceeded. Please try again later.", {}

    client_id = os.environ.get("OCLC_CLIENT_ID")
    client_secret = os.environ.get("OCLC_SECRET")
    
    if not client_id or not client_secret:
        return (
            "Error: OCLC_CLIENT_ID and OCLC_SECRET must be set in environment variables",
            "Early exit before API call (missing credentials).",
            []
        )

    try:
        access_token = get_access_token(client_id, client_secret)
    except Exception as e:
        return (
            f"Error getting access token: {str(e)}",
            "Early exit before API call (token retrieval failed).",
            []
        )

    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs"

    if not isinstance(queries, list):
        return (
            "Error: Invalid query format",
            "Queries must be a list of strings",
            []
        )


    cleaned_queries = []
    for q in queries:
        if isinstance(q, str) and q.strip():
            cleaned = q.replace(str(barcode), "x").strip()
            if len(cleaned) >= 3:
                cleaned_queries.append(cleaned)

    if not cleaned_queries:
        return (
            "No valid queries could be constructed",
            "Please check the metadata format",
            []
        )


    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }

    query_log = ["Attempted Queries:"]
    attempted_queries = []
    raw_api_responses = []
    
    # Track unique OCLC numbers to avoid duplicates
    seen_oclc_numbers = set()
    accumulated_results = []
    total_records_found = 0
    max_results_to_show = 10  # Our target - 10 LP results

    # Try queries until we have 10 unique results or exhausted all queries
    for idx, query in enumerate(cleaned_queries, 1):
        # Check if we already have enough results - NEW EXIT CONDITION
        if len(seen_oclc_numbers) >= max_results_to_show:
            query_log.append(f"\nReached target of {max_results_to_show} unique LP records. Stopping further queries.")
            break  # Stop the loop once we have 10 unique results
            
        query_log.append(f"\nQuery {idx}: {query}")
        attempted_queries.append(idx)
        
        params = {
            "q": query,
            "limit": limit,
            "offset": 1,
            "itemType": "music",
            "inCatalogLanguage": "eng"
            # Removed itemSubType to allow both music-lp and audiobook-lp
        }

        try:
            response = requests.get(endpoint, params=params, headers=headers)
            api_calls['count'] += 1
            response.raise_for_status()
            data = response.json()
            
            raw_api_responses.append({
                "query_number": idx,
                "query_text": query,
                "api_response": data,  # Raw JSON from OCLC
                "total_records": data.get("numberOfRecords", 0),
                "timestamp": datetime.datetime.now().isoformat()
            })
            
            total_records = data.get("numberOfRecords", 0)
            if total_records > 1000:
                query_log.append(f"Too many results ({total_records}), skipping")
                continue
                
            if total_records > 0:
                # Extract OCLC numbers from the current response to check for new records
                current_oclc_numbers = set()
                for record in data.get('bibRecords', []):
                    # Only consider LP format records
                    include_record = False
                    if 'format' in record and 'specificFormat' in record['format']:
                        specific_format = record['format']['specificFormat']
                        if isinstance(specific_format, str) and any(lp_term in specific_format.lower() for lp_term in [
    "lp", "vinyl", "long play", "audiobook", "spoken", "spoken word", "talking book", "sound recording"
]):
                            include_record = True
                    
                    if not include_record:
                        continue
                        
                    # Check OCLC number
                    if 'identifier' in record and 'oclcNumber' in record['identifier']:
                        oclc_number = record['identifier']['oclcNumber']
                        if oclc_number and oclc_number not in seen_oclc_numbers:
                            current_oclc_numbers.add(oclc_number)
                
                if current_oclc_numbers:
                    # We found new unique LP format records
                    results, record_count, _ = format_oclc_api_response_for_accumulation(data, access_token, seen_oclc_numbers)
                    
                    if results and "No matching records with LP format found" not in results and "No records found" not in results:
                        accumulated_results.append(results)
                        total_records_found += record_count
                        # Add the new OCLC numbers to our seen set
                        seen_oclc_numbers.update(current_oclc_numbers)
                        query_log.append(f"Added new LP format matches (now have {len(seen_oclc_numbers)} unique records)")
                        
                        # Check if we've reached our target - NEW CHECK
                        if len(seen_oclc_numbers) >= max_results_to_show:
                            query_log.append(f"Reached target of {max_results_to_show} unique LP records.")
                    else:
                        query_log.append(f"No new LP format matches found")
                else:
                    query_log.append(f"No new unique LP format matches found")
            else:
                query_log.append(f"No matches found")
            
        except requests.RequestException as e:
            api_calls['count'] += 1
            query_log.append(f"Query failed: {str(e)}")
            raw_api_responses.append({
                "query_number": idx,
                "query_text": query,
                "api_response": None,
                "error": str(e),
                "timestamp": datetime.datetime.now().isoformat()
            })

    # Combine all accumulated results with a single count at the top
    if accumulated_results:
        # Add the total count at the top, but limit to max_results_to_show
        displayed_record_count = min(total_records_found, max_results_to_show)
        total_header = f"Total LP Format Records Found: {total_records_found} (Displaying: {displayed_record_count})"

        # Limit the actual content to show only max_results_to_show records
        limited_results = []
        record_count = 0
        
        # Parse through accumulated results to extract and limit individual records
        for result_set in accumulated_results:
            if record_count >= max_results_to_show:
                break
                
            # Split by record divider
            record_sections = result_set.split("-" * 40)
            
            for section in record_sections:
                if record_count >= max_results_to_show:
                    break
                    
                if section.strip() and "OCLC Number:" in section:
                    limited_results.append("-" * 40)
                    limited_results.append(section.strip())
                    record_count += 1
        
        combined_results = total_header + "\n\n" + "\n".join(limited_results)
        return combined_results, "\n".join(query_log), raw_api_responses
    else:
        return "No matching records with LP format found after trying all queries", "\n".join(query_log), raw_api_responses

def process_metadata_file(input_file, results_folder_path, workflow_json_path):
    items_with_issues = 0
    total_rows = 0
    processed_rows = 0
    total_queries_sent = 0
    total_records_found_across_all = 0
    start_time = time.time()
    wb = load_workbook(input_file)
    ws = wb.active

    # Create a temporary workbook for periodic saving
    temp_wb = Workbook()
    temp_ws = temp_wb.active
    
    # Copy headers from main workbook
    for col_idx, cell in enumerate(ws[1], 1):
        temp_ws.cell(row=1, column=col_idx, value=cell.value)
        # Copy column widths
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(col_idx)
        if column_letter in ws.column_dimensions:
            temp_ws.column_dimensions[column_letter].width = ws.column_dimensions[column_letter].width

    if ws['F1'].value != 'OCLC Query':
        ws.insert_cols(6)
        ws['F1'] = 'OCLC Query'
        temp_ws['F1'] = 'OCLC Query'
    if ws['G1'].value != 'OCLC API Results':
        ws.insert_cols(7)
        ws['G1'] = 'OCLC API Results'
        temp_ws['G1'] = 'OCLC API Results'
        
    ws.column_dimensions['F'].width = 52
    ws.column_dimensions['G'].width = 52
    temp_ws.column_dimensions['F'].width = 52
    temp_ws.column_dimensions['G'].width = 52
    
    # Temporary file path
    temp_output_file = "temp_lp_metadata_progress.xlsx"
    temp_output_path = os.path.join(results_folder_path, temp_output_file)
    
    total_rows = ws.max_row
    processed_rows = 0
    
    for row in range(2, total_rows + 1):
        metadata_str = ws.cell(row=row, column=5).value  # Column E
        barcode = ws.cell(row=row, column=4).value       # Column D
        if not metadata_str or metadata_str.startswith('Error'):
            # Still copy this row to temp workbook
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                temp_cell = temp_ws.cell(row=row, column=col_idx, value=cell_value)
                if ws.cell(row=row, column=col_idx).alignment:
                    temp_cell.alignment = Alignment(vertical='top', wrap_text=True)
            continue

        try:
            metadata_fields = extract_metadata_fields(metadata_str)
            # Prefer JSON 'extracted_fields' from the workflow; fallback to legacy text parser if not present
            from json_workflow import load_workflow_json  # local import to avoid top-level changes

            workflow_data = load_workflow_json(workflow_json_path)
            barcode_str = str(barcode) if barcode is not None else ""

            metadata_fields = {}
            if isinstance(workflow_data, dict) and "records" in workflow_data and barcode_str in workflow_data["records"]:
                metadata_fields = (
                    workflow_data["records"][barcode_str]
                    .get("step1_metadata_extraction", {})
                    .get("extracted_fields", {}) or {}
                )

            # Fallback to legacy text parser when JSON fields are unavailable
            if not isinstance(metadata_fields, dict) or not metadata_fields:
                metadata_fields = extract_metadata_fields(metadata_str)

            if not isinstance(metadata_fields, dict) or not metadata_fields:
                raise ValueError("Invalid metadata format for query construction")

            queries = construct_queries_from_metadata(metadata_fields, workflow_data, barcode)
            results, query_log, raw_api_responses = query_oclc_api(queries, barcode)
            
            # Update main workbook with results
            ws.cell(row=row, column=6, value=query_log)
            ws.cell(row=row, column=7, value=results)
            ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            # Update temp workbook with results
            temp_ws.cell(row=row, column=6, value=query_log)
            temp_ws.cell(row=row, column=7, value=results)
            temp_ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            temp_ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            # Copy image cells and other data from main to temp
            for col_idx in range(1, 6):  # Columns A-E
                cell_value = ws.cell(row=row, column=col_idx).value
                temp_cell = temp_ws.cell(row=row, column=col_idx, value=cell_value)
                if ws.cell(row=row, column=col_idx).alignment:
                    temp_cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Now do JSON logging
            try:
                # Count queries attempted and records found
                queries_attempted = len(queries)

                # Parse total records from results
                total_records_found = 0
                if "Total LP Format Records Found:" in results:
                    match = re.search(r'Total LP Format Records Found:\s*(\d+)', results)
                    if match:
                        total_records_found = int(match.group(1))

                # Update workflow JSON with comprehensive Step 2 results
                update_record_step2(
                    json_path=workflow_json_path,
                    barcode=barcode,
                    queries_attempted=queries_attempted,
                    total_records_found=total_records_found
                )

                # Also log the detailed OCLC data to the main workflow JSON
                from json_workflow import load_workflow_json, save_workflow_json
                workflow_data = load_workflow_json(workflow_json_path)
                
                if barcode in workflow_data["records"]:
                    # Add detailed query and result information to the main workflow
                    workflow_data["records"][barcode]["step2_detailed_data"] = {
                        "constructed_queries": queries,
                        "query_execution_log": query_log,
                        "formatted_oclc_results": results,
                        "raw_api_responses_count": len(raw_api_responses),
                        "processing_summary": {
                            "unique_queries_generated": len(queries),
                            "api_calls_made": len([r for r in raw_api_responses if r.get("api_response") is not None]),
                            "api_errors": len([r for r in raw_api_responses if r.get("error") is not None]),
                            "total_oclc_records_found": total_records_found
                        }
                    }
                    
                    # Update the timestamp
                    workflow_data["records"][barcode]["updated_at"] = datetime.datetime.now().isoformat()
                    
                    # Save the updated workflow data
                    save_workflow_json(workflow_json_path, workflow_data)

                # Log comprehensive OCLC API search data
                log_oclc_api_search(
                    results_folder_path=results_folder_path,
                    barcode=barcode,
                    queries=queries,
                    raw_api_responses=raw_api_responses,
                    formatted_results=results,  # What goes in Excel
                    query_log=query_log,
                    queries_attempted=queries_attempted,
                    total_records_found=total_records_found
                )
                # Log metrics
                total_queries_sent += queries_attempted
                total_records_found_across_all += total_records_found
                
            except Exception as json_error:
                log_error(
                    results_folder_path=results_folder_path,
                    step="step2",
                    barcode=barcode,
                    error_type="json_update_error",
                    error_message=str(json_error)
                )

        except Exception as e:
            print(f"   Error processing row {row}: {str(e)}")
            error_message = f"Error: {str(e)}"
            items_with_issues += 1
            log_error(
                results_folder_path=results_folder_path,
                step="step2",
                barcode=barcode,
                error_type="oclc_api_error",
                error_message=str(e),
                additional_context={"queries_attempted": len(queries) if 'queries' in locals() else 0}
            )
            
            # Update both workbooks with error
            ws.cell(row=row, column=6, value="Error processing")
            ws.cell(row=row, column=7, value=error_message)
            ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            temp_ws.cell(row=row, column=6, value="Error processing")
            temp_ws.cell(row=row, column=7, value=error_message)
            temp_ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            temp_ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            
            # Copy other columns from main to temp
            for col_idx in range(1, 6):  # Columns A-E
                cell_value = ws.cell(row=row, column=col_idx).value
                temp_cell = temp_ws.cell(row=row, column=col_idx, value=cell_value)
                if ws.cell(row=row, column=col_idx).alignment:
                    temp_cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Single increment at end of each iteration
        processed_rows += 1
        print(f"Processed row {row}/{total_rows}")
        
        # Save temporary workbook every 10 rows
        if processed_rows % 10 == 0:
            try:
                temp_wb.save(temp_output_path)
                print(f"Progress saved ({processed_rows}/{total_rows-1} data rows)")
            except Exception as save_error:
                print(f"Warning: Could not save temporary progress: {save_error}")
                
    time.sleep(0.1)
            
    # Clean up temporary file
    try:
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
            print(f"Temporary progress file removed: {temp_output_path}")
    except Exception as remove_error:
        print(f"Warning: Could not remove temporary progress file: {remove_error}")

    # Calculate final metrics
    end_time = time.time()
    total_processing_time = end_time - start_time
    
    # Log Step 2 processing metrics
    try:
        step2_metrics = {
            "total_items": total_rows - 1,  # Subtract 1 for header row
            "successful_items": processed_rows - items_with_issues,
            "failed_items": items_with_issues,
            "success_rate": ((processed_rows - items_with_issues) / processed_rows * 100) if processed_rows > 0 else 0,
            "total_time_seconds": total_processing_time,
            "total_time_minutes": total_processing_time / 60,
            "average_time_per_item": total_processing_time / processed_rows if processed_rows > 0 else 0,
            "total_queries_sent": total_queries_sent,
            "average_queries_per_item": total_queries_sent / processed_rows if processed_rows > 0 else 0,
            "total_oclc_records_found": total_records_found_across_all,
            "average_records_found_per_item": total_records_found_across_all / processed_rows if processed_rows > 0 else 0,
            "processing_mode": "INDIVIDUAL",
            "api_calls_made": api_calls['count'],  # Track API usage
            "timestamp": datetime.datetime.now().isoformat()
        }
        
        log_processing_metrics(
            results_folder_path=results_folder_path,
            step="step2_oclc_search",
            batch_metrics=step2_metrics
        )
        
    except Exception as metrics_error:
        print(f"Warning: Could not log Step 2 processing metrics: {metrics_error}")

    return wb

def main():
    file_paths = get_file_path_config()
    results_folder = find_latest_results_folder(file_paths["results_prefix"])
    workflow_json_path = get_workflow_json_path(results_folder)
    
    if not results_folder:
        print("No results folder found! Run the first script first.")
        return
        
    print(f"Using results folder: {results_folder}")
    
    # Look for previous step files in the results folder.
    input_files = [f for f in os.listdir(results_folder) 
               if f.startswith('full-workflow-data-lp') and f.endswith('.xlsx')]
    
    if not input_files:
        print("No step 1 files found in the results folder!")
        return
        
    latest_file = max(input_files)
    input_file = os.path.join(results_folder, latest_file)
    
    print(f"Processing file: {input_file}")
    wb = process_metadata_file(input_file, results_folder, workflow_json_path)
    
    # Save back to the same file (in-place modification)
    wb.save(input_file)
    print(f"Results saved to {input_file}")

if __name__ == "__main__":
    main()