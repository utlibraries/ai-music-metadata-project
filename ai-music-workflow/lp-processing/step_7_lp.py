"""
Step 7: Apply Cataloger Decisions to Workflow Outputs
Processes cataloger review decisions from exported CSV and updates all deliverable files.
"""

import os
import shutil
import csv
import re
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Import OCLC functions from Step 2
from step_2_lp import get_access_token, get_holdings_info

# Import Alma verification for accurate holdings check
from alma_api_utils import verify_holdings_in_alma

# Import Step 5 functions for generating low confidence content
from step_5_lp import (
    get_bib_info_from_workflow,
    get_holdings_info_from_workflow,
)

# Import shared utilities
from json_workflow import update_record_step7, log_processing_metrics, load_workflow_json
from shared_utilities import get_workflow_json_path
from lp_workflow_config import get_current_timestamp

# Import Step 6 functions for HTML regeneration
from step_6_lp import create_paginated_review_html, load_records_from_step5

def clean_date_format(date_string):
    """
    Ensure date is in YYYY-MM-DD format, cleaning any timestamps.
    Handles formats like: 2025-10-20-10-41-54 or 2025-10-20
    """
    if not date_string:
        return date_string
    
    date_str = str(date_string)
    # Extract first 3 parts (YYYY-MM-DD)
    parts = date_str.split('-')
    if len(parts) >= 3:
        return f"{parts[0]}-{parts[1]}-{parts[2]}"
    return date_str

def normalize_barcode(barcode):
    """
    Normalize barcode to consistent format for comparison.
    Handles numeric formatting issues between CSV and Excel.
    """
    if barcode is None:
        return None
    barcode_str = str(barcode).strip()
    
    if 'E+' in barcode_str.upper() or 'e+' in barcode_str:
        try:
            return str(int(float(barcode_str))).zfill(15)
        except (ValueError, OverflowError):
            return barcode_str
    
    return barcode_str.zfill(15) if barcode_str.isdigit() else barcode_str

def load_cataloger_decisions(csv_path):
    """
    Load cataloger decisions from exported CSV.
    Returns dict mapping barcode -> decision data.
    """
    decisions = {}
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            barcode = normalize_barcode(row['Barcode'])
            decisions[barcode] = {
                'record_id': row['Record'],
                'initial_sort_group': row['Initial Sort Group'],
                'decision': row['Cataloger Decision'],
                'ai_suggested_oclc': row['AI-Suggested OCLC #'].strip(),
                'correct_oclc': row['Correct OCLC #'].strip(),
                'notes': row['Notes'],
                'cataloger': row['Cataloger'],
                'review_date': row['Review Date'],
                'confidence': row['Confidence']
            }
    
    return decisions


def backup_original_files(results_folder):
    """
    Create 'original-outputs' folder and move original deliverable files there.
    Only backs up files if they don't already exist in backup folder to preserve
    the truly original outputs across multiple runs.
    Returns dict of {file_type: (original_path, backup_path)}
    """
    deliverables_folder = os.path.join(results_folder, "deliverables")
    backup_folder = os.path.join(results_folder, "original-outputs")
    
    os.makedirs(backup_folder, exist_ok=True)
    
    backed_up_files = {}
    
    file_patterns = {
        'sorting': 'lp-workflow-sorting-',
        'batch_upload': 'batch-upload-alma-lp-',
        'tracking': 'tracking-spreadsheet-catalogers-',
        'review': 'low-confidence-matches-review-',
        'marc': 'low-confidence-marc-'
    }
    
    for file_type, prefix in file_patterns.items():
        files = [f for f in os.listdir(deliverables_folder) if f.startswith(prefix)]
        if files:
            latest_file = max(files)
            original_path = os.path.join(deliverables_folder, latest_file)
            backup_path = os.path.join(backup_folder, latest_file)
            
            if not os.path.exists(backup_path):
                shutil.copy2(original_path, backup_path)
                backed_up_files[file_type] = (original_path, backup_path)
                print(f"   Backed up {file_type}: {latest_file}")
            else:
                print(f"   Backup already exists for {file_type}: {latest_file} (preserving original)")
                backed_up_files[file_type] = (original_path, backup_path)
    
    return backed_up_files


def fetch_oclc_data(oclc_number):
    """
    Make OCLC API call to get title, author, and holdings for new OCLC number.
    Uses the holdings API which returns all needed data in one call.
    Returns: (title, author, is_held_by_ixa) or (None, None, False) on error
    """
    try:
        client_id = os.environ.get("OCLC_CLIENT_ID")
        client_secret = os.environ.get("OCLC_SECRET")

        if not client_id or not client_secret:
            print("   Error: OCLC credentials not found in environment")
            return None, None, False

        access_token = get_access_token(client_id, client_secret)

        # Use the holdings API which returns title, creator, and holdings
        is_held_by_ixa, total_holdings, holding_institutions = get_holdings_info(oclc_number, access_token)

        # Get title and author from holdings API
        import requests
        base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
        endpoint = f"{base_url}/bibs-holdings"

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Accept": "application/json"
        }

        params = {
            "oclcNumber": oclc_number,
            "limit": 1
        }

        response = requests.get(endpoint, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()

        title = "No title available"
        author = "No author available"

        if "briefRecords" in data and len(data["briefRecords"]) > 0:
            record = data["briefRecords"][0]
            title = record.get("title", "No title available")
            author = record.get("creator", "No author available")

        return title, author, is_held_by_ixa

    except Exception as e:
        print(f"   Error fetching OCLC data for {oclc_number}: {e}")
        return None, None, False


def fetch_full_oclc_bib_data(oclc_number):
    """
    Fetch full bibliographic data from OCLC for display in HTML.
    Returns a formatted text representation matching step 2's detailed format.

    Uses the bibs endpoint to get full bibRecords (not briefRecords) which include:
    - Full title information (main, subtitle, series)
    - All contributors with roles
    - Publisher details with place
    - Physical description
    - Contents/track listings
    - Music information
    - Notes
    """
    try:
        client_id = os.environ.get("OCLC_CLIENT_ID")
        client_secret = os.environ.get("OCLC_SECRET")

        if not client_id or not client_secret:
            return None

        access_token = get_access_token(client_id, client_secret)

        import requests
        base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"

        # Use bibs endpoint with OCLC number query to get full record
        endpoint = f"{base_url}/bibs"

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Accept": "application/json"
        }

        params = {
            "q": f"no:{oclc_number}",
            "limit": 1
        }

        response = requests.get(endpoint, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()

        # Check for full bibRecords
        if "bibRecords" not in data or len(data["bibRecords"]) == 0:
            # Fallback to abbreviated format if full record not found
            return _fetch_abbreviated_oclc_data(oclc_number, access_token)

        record = data["bibRecords"][0]

        # Get holdings info
        is_held_by_ixa, total_holdings, _ = get_holdings_info(oclc_number, access_token)

        # Get Alma verification
        alma_result = verify_holdings_in_alma(str(oclc_number))
        alma_verified = "Yes" if alma_result.get("alma_verified", False) else "No"

        # Format the record text matching step 2's detailed format
        formatted_results = []
        formatted_results.append(f"OCLC Number: {oclc_number}")
        formatted_results.append(f"\nHeld by IXA: {'Yes' if is_held_by_ixa else 'No'}  <<<  ALMA VERIFIED: {alma_verified}")
        formatted_results.append(f"Total Institutions Holding: {total_holdings}")

        # Identifier section
        if 'identifier' in record:
            formatted_results.append("\nIdentifier:")
            if 'oclcNumber' in record['identifier']:
                formatted_results.append(f"  - oclcNumber: {record['identifier']['oclcNumber']}")
            # Add UPC if it exists
            if 'otherStandardIdentifiers' in record['identifier']:
                for id_item in record['identifier']['otherStandardIdentifiers']:
                    if isinstance(id_item, dict) and id_item.get('type') == 'Universal Product Code (UPC)':
                        formatted_results.append(f"  - UPC: {id_item.get('id', 'N/A')}")

        # Title Information
        title_text = "No title available"
        if 'title' in record:
            formatted_results.append("Title Information:")
            if 'mainTitles' in record['title']:
                for title in record['title']['mainTitles']:
                    title_text = title.get('text', 'N/A')
                    formatted_results.append(f"  - Main Title: {title_text}")
            if 'subtitles' in record['title']:
                for subtitle in record['title']['subtitles']:
                    formatted_results.append(f"  - Subtitle: {subtitle.get('text', 'N/A')}")
            if 'seriesTitles' in record['title']:
                for series in record['title']['seriesTitles']:
                    formatted_results.append(f"  - Series Title: {series.get('seriesTitle', 'N/A')}")

        # Contributors
        contributors = []
        if 'contributor' in record:
            formatted_results.append("Contributors:")
            for creator_type in ['creators', 'contributors']:
                if creator_type in record['contributor']:
                    for person in record['contributor'][creator_type]:
                        if 'firstName' in person and 'secondName' in person:
                            name = f"{person.get('firstName', {}).get('text', '')} {person.get('secondName', {}).get('text', '')}"
                        elif 'nonPersonName' in person:
                            name = person['nonPersonName'].get('text', '')
                        else:
                            name = 'N/A'
                        role = person.get('type', 'N/A')
                        formatted_results.append(f"  - {name.strip()} ({role})")
                        contributors.append(name.strip())

        # Publishers
        if 'publishers' in record:
            formatted_results.append("Publishers:")
            for pub in record['publishers']:
                pub_name = pub.get('publisherName', {}).get('text', 'N/A')
                pub_place = pub.get('publicationPlace', 'N/A')
                formatted_results.append(f"  - Name: {pub_name}")
                formatted_results.append(f"    Place: {pub_place}")

        # Dates
        pub_date = ""
        if 'date' in record:
            formatted_results.append("Dates:")
            if 'publicationDate' in record['date']:
                pub_date = record['date']['publicationDate']
                formatted_results.append(f"  - publicationDate: {pub_date}")

        # Language
        if 'language' in record:
            formatted_results.append("Language:")
            for key, value in record['language'].items():
                formatted_results.append(f"  - {key}: {value}")

        # Music Information
        if 'musicInfo' in record:
            formatted_results.append("Music Information:")
            for key, value in record['musicInfo'].items():
                formatted_results.append(f"  - {key}: {value}")

        # Description (Physical and Contents)
        if 'description' in record:
            formatted_results.append("Description:")
            if 'physicalDescription' in record['description']:
                formatted_results.append(f"  - Physical: {record['description']['physicalDescription']}")

            if 'contents' in record['description']:
                for content in record['description']['contents']:
                    # Handle the titles array format from OCLC
                    if 'titles' in content and isinstance(content['titles'], list):
                        formatted_results.append("  - Content:")
                        for i, title in enumerate(content['titles'], 1):
                            formatted_results.append(f"    {i}. {title}")
                    # Also handle contentNote format as fallback
                    elif 'contentNote' in content and 'text' in content['contentNote']:
                        content_text = content['contentNote']['text']
                        # Smart content handling for large multi-disc sets
                        if len(content_text) > 1500:
                            disc_count = content_text.count('Disc ')
                            chapter_count = content_text.count('Chapter ')
                            track_patterns = len(re.findall(r'(?:--|\d+\.|\(\d+:\d+\))', content_text))

                            is_large_compilation = (
                                disc_count > 4 or
                                chapter_count > 20 or
                                track_patterns > 100 or
                                len(content_text) > 5000
                            )

                            if is_large_compilation:
                                track_pattern = r'([^-\n]+?)\s*(?:\(\d+:\d+\)|--)'
                                sample_tracks = re.findall(track_pattern, content_text[:2000])
                                sample_tracks = [t.strip() for t in sample_tracks[:10] if len(t.strip()) > 3]

                                summary = f"LARGE MULTI-DISC COMPILATION: {disc_count} discs"
                                if chapter_count > 0:
                                    summary += f", {chapter_count} chapters"
                                summary += ". "

                                if sample_tracks:
                                    summary += f"Sample tracks: {', '.join(sample_tracks)}. "

                                summary += f"[Original: {len(content_text):,} characters - This is a large compilation/box set, not a single album]"
                                content_text = summary
                            else:
                                content_text = content_text[:1500]
                                last_break = content_text.rfind(' -- ')
                                if last_break > 1000:
                                    content_text = content_text[:last_break]
                                content_text += "... [Content truncated for analysis]"

                        formatted_results.append(f"  - Content: {content_text}")

        # Notes
        if 'note' in record:
            formatted_results.append("Notes:")
            if isinstance(record['note'], dict):
                for key, value in record['note'].items():
                    formatted_results.append(f"  - {key}: {value}")
            elif isinstance(record['note'], list):
                for note in record['note']:
                    formatted_results.append(f"  - {note}")

        formatted_results.append("\n[Record provided by cataloger - Different OCLC # selected]")

        formatted_text = "\n".join(formatted_results)

        return {
            "oclc_number": oclc_number,
            "title": title_text,
            "contributors": contributors if contributors else [''],
            "publication_date": pub_date,
            "full_record_text": formatted_text
        }

    except Exception as e:
        print(f"   Warning: Could not fetch full bib data for OCLC {oclc_number}: {e}")
        return None


def _fetch_abbreviated_oclc_data(oclc_number, access_token):
    """
    Fallback function to fetch abbreviated OCLC data when full record is not available.
    Uses bibs-holdings endpoint which returns briefRecords.
    """
    try:
        import requests
        base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
        endpoint = f"{base_url}/bibs-holdings"

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Accept": "application/json"
        }

        params = {
            "oclcNumber": oclc_number,
            "limit": 1
        }

        response = requests.get(endpoint, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()

        if "briefRecords" not in data or len(data["briefRecords"]) == 0:
            return None

        record = data["briefRecords"][0]

        # Get holdings info
        is_held_by_ixa, total_holdings, _ = get_holdings_info(oclc_number, access_token)

        # Get Alma verification
        alma_result = verify_holdings_in_alma(str(oclc_number))
        alma_verified = "Yes" if alma_result.get("alma_verified", False) else "No"

        formatted_text = f"""OCLC Number: {oclc_number}

Title Information:
  - Main Title: {record.get('title', 'No title available')}

Contributors:
  - {record.get('creator', 'No contributor information')}

Publication Information:
  - Date: {record.get('date', 'No date available')}
  - Publisher: {record.get('publisher', 'No publisher information')}

Material Type: {record.get('generalFormat', 'Unknown')} / {record.get('specificFormat', 'Unknown')}

Language: {record.get('language', 'Unknown')}

Holdings Information:
  - Total Libraries: {total_holdings}
  - Held by IXA: {"Yes" if is_held_by_ixa else "No"}  <<<  ALMA VERIFIED: {alma_verified}

[Record provided by cataloger - Different OCLC # selected]
[Note: Full bibliographic details not available - showing abbreviated record]
"""

        return {
            "oclc_number": oclc_number,
            "title": record.get('title', 'No title available'),
            "contributors": [record.get('creator', '')],
            "publication_date": record.get('date', ''),
            "full_record_text": formatted_text
        }

    except Exception as e:
        print(f"   Warning: Could not fetch abbreviated bib data for OCLC {oclc_number}: {e}")
        return None


def determine_changes(barcode, decision_data, current_state, decisions_history_current):
    """
    Determine what changes need to be made based on decisions history.
    Now uses decisions_history_current as the source of truth.
    
    Args:
        barcode: The barcode being processed
        decision_data: Data from CSV (for metadata like cataloger name, notes)
        current_state: Current state from sorting spreadsheet
        decisions_history_current: Current decision from decisions-history.xlsx (SOURCE OF TRUTH)
    
    Returns: dict of changes to apply
    """
    if barcode in decisions_history_current:
        decision_info = decisions_history_current[barcode]
        decision = decision_info['cataloger_decision']
        new_oclc = decision_info['chosen_oclc']
        final_confidence = decision_info['final_confidence']
        held_by_ixa = decision_info['held_by_ixa']
    else:
        decision = decision_data['decision']
        new_oclc = decision_data['correct_oclc']
        ai_conf = decision_data.get('confidence', 0)
        final_confidence = calculate_final_confidence(decision, new_oclc, ai_conf)
        held_by_ixa = "No"
    
    initial_status = current_state['sort_group']
    
    changes = {
        'barcode': barcode,
        'new_status': None,
        'new_oclc': None,
        'new_title': None,
        'new_oclc_bib_data': None,  # Full bib data for new OCLC
        'original_oclc': current_state.get('oclc_number', ''),
        'original_title': current_state.get('title', ''),
        'original_status': initial_status,
        'clear_oclc': False,
        'clear_title': False,
        'remove_from_batch_upload': False,
        'add_to_batch_upload': False,
        'remove_from_low_confidence': False,
        'add_to_low_confidence': False,
        'update_notes': decision_data.get('notes', ''),
        'cataloger': decision_data.get('cataloger', ''),
        'review_date': decision_data.get('review_date', ''),
        'final_confidence': final_confidence
    }
    
    if final_confidence >= 80:
        if held_by_ixa == "Yes":
            changes['new_status'] = 'Held by UT Libraries (IXA)'
            changes['remove_from_batch_upload'] = True
            changes['remove_from_low_confidence'] = True
        else:
            changes['new_status'] = 'Alma Batch Upload (High Confidence)'
            changes['add_to_batch_upload'] = True
            changes['remove_from_low_confidence'] = True
        
        if new_oclc and new_oclc != changes['original_oclc']:
            changes['new_oclc'] = new_oclc
            title, author, _ = fetch_oclc_data(new_oclc)  # Ignore OCLC holdings (unreliable)
            if title:
                changes['new_title'] = title

            # Fetch full bib data for HTML display
            full_bib_data = fetch_full_oclc_bib_data(new_oclc)
            if full_bib_data:
                changes['new_oclc_bib_data'] = full_bib_data

            # Re-verify holdings against Alma for accurate status
            alma_result = verify_holdings_in_alma(str(new_oclc))
            new_held_by_ixa = "Yes" if alma_result.get("alma_verified", False) else "No"

            # Update sort group if holdings status changed
            if new_held_by_ixa != held_by_ixa:
                if new_held_by_ixa == "Yes":
                    changes['new_status'] = 'Held by UT Libraries (IXA)'
                    changes['remove_from_batch_upload'] = True
                    changes['add_to_batch_upload'] = False
                else:
                    changes['new_status'] = 'Alma Batch Upload (High Confidence)'
                    changes['add_to_batch_upload'] = True
                    changes['remove_from_batch_upload'] = False
    
    elif final_confidence == 0:
        changes['new_status'] = 'Cataloger Review (Low Confidence)'
        changes['remove_from_batch_upload'] = True
        changes['add_to_low_confidence'] = True
        
        if decision == "Different OCLC # Needed" and not new_oclc:
            changes['clear_oclc'] = True
            changes['clear_title'] = True
    
    return changes

def load_current_state_from_sorting(sorting_file):
    """
    Load current state of all records from sorting spreadsheet.
    Returns dict mapping barcode -> {sort_group, oclc_number, title, author, confidence}
    """
    wb = load_workbook(sorting_file)
    ws = wb.active
    
    # Add new column header if it doesn't exist
    if ws.cell(row=1, column=8).value != "Date Cataloger Checked":
        ws.cell(row=1, column=8, value="Date Cataloger Checked")
    
    # Set column H width to 18
    ws.column_dimensions['H'].width = 20
    
    current_state = {}
    
    for row_idx in range(2, ws.max_row + 1):
        barcode = ws.cell(row=row_idx, column=1).value
        if barcode:
            current_state[str(barcode)] = {
                'row': row_idx,
                'sort_group': ws.cell(row=row_idx, column=2).value,
                'oclc_number': ws.cell(row=row_idx, column=3).value,
                'title': ws.cell(row=row_idx, column=4).value,
                'author': ws.cell(row=row_idx, column=5).value,
                'publication_date': ws.cell(row=row_idx, column=6).value,
                'confidence': ws.cell(row=row_idx, column=7).value
            }
    
    return current_state, wb, ws


def update_sorting_spreadsheet(ws, barcode, changes, current_state):
    """Update a single record in the sorting spreadsheet."""
    row_idx = current_state[barcode]['row']
    
    if changes['new_status']:
        ws.cell(row=row_idx, column=2, value=changes['new_status'])
    
    # Handle OCLC number - either clear or update
    if changes['clear_oclc']:
        ws.cell(row=row_idx, column=3, value=None)
    elif changes['new_oclc']:
        ws.cell(row=row_idx, column=3, value=changes['new_oclc'])
    
    # Handle title - either clear or update
    if changes['clear_title']:
        ws.cell(row=row_idx, column=4, value=None)
    elif changes['new_title']:
        ws.cell(row=row_idx, column=4, value=changes['new_title'])
    
    # Update confidence - use final_confidence
    if changes.get('final_confidence') is not None:
        ws.cell(row=row_idx, column=7, value=changes['final_confidence'])
    
    # Update Date Cataloger Checked (column 8)
    if changes.get('review_date'):
        ws.cell(row=row_idx, column=8, value=changes['review_date'])

def update_batch_upload_file(batch_file, all_changes, current_state):
    """
    Rebuild the batch upload file based on changes.
    """
    # Read existing batch upload records
    existing_records = {}
    if os.path.exists(batch_file):
        with open(batch_file, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split('|')
                if len(parts) >= 2:
                    barcode = parts[1]
                    existing_records[barcode] = line.strip()
    
    # Apply changes
    for barcode, changes in all_changes.items():
        if changes['remove_from_batch_upload']:
            existing_records.pop(barcode, None)
        
        if changes['add_to_batch_upload']:
            oclc = changes['new_oclc'] if changes['new_oclc'] else current_state[barcode]['oclc_number']
            title = changes['new_title'] if changes['new_title'] else current_state[barcode]['title']
            existing_records[barcode] = f"{oclc}|{barcode}|{title}"
    
    # Write updated file
    with open(batch_file, 'w', encoding='utf-8') as f:
        for record in existing_records.values():
            f.write(record + '\n')


def add_record_to_low_confidence_review(review_file, barcode, oclc_number, confidence, workflow_json_path):
    """
    Add a single record to the low confidence review Excel spreadsheet.
    """
    workflow_data = load_workflow_json(workflow_json_path)
    
    if barcode not in workflow_data.get("records", {}):
        print(f"      Warning: Barcode {barcode} not found in workflow JSON")
        return
    
    workflow_record = workflow_data["records"][barcode]
    
    step1_data = workflow_record.get("step1_metadata_extraction", {})
    metadata_text = step1_data.get("raw_ai_metadata", "No AI-generated metadata available")
    
    step2_data = workflow_record.get("step2_detailed_data", {})
    formatted_results = step2_data.get("formatted_oclc_results", "")
    
    other_oclc_numbers = []
    if formatted_results:
        oclc_pattern = r"OCLC Number: (\d+)"
        matches = re.findall(oclc_pattern, formatted_results)
        if matches:
            other_oclc_numbers = [num for num in matches if num != oclc_number]
    
    if other_oclc_numbers:
        other_matches = "\n".join([f"OCLC Number: {num}" for num in other_oclc_numbers])
    else:
        other_matches = "No other candidates"
    
    wb = load_workbook(review_file)
    ws = wb.active
    
    next_row = ws.max_row + 1
    record_num = next_row - 1
    
    ws.cell(row=next_row, column=1, value=record_num)
    ws.cell(row=next_row, column=2, value=barcode)
    ws.cell(row=next_row, column=3, value=oclc_number if oclc_number else 'No OCLC number')
    ws.cell(row=next_row, column=4, value=confidence)  
    
    metadata_cell = ws.cell(row=next_row, column=5, value=metadata_text if metadata_text and metadata_text.strip() else "No AI-generated metadata available")
    metadata_cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    other_cell = ws.cell(row=next_row, column=6, value=other_matches)
    other_cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    if oclc_number:
        oclc_data = get_bib_info_from_workflow(oclc_number, workflow_json_path)
        
        raw_oclc_json = json.dumps(oclc_data, indent=2, ensure_ascii=False)
        oclc_cell = ws.cell(row=next_row, column=7, value=raw_oclc_json)
        oclc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        holdings_info = get_holdings_info_from_workflow(oclc_number, workflow_json_path)
        ws.cell(row=next_row, column=8, value=holdings_info.get('total_holdings', 0))
        ws.cell(row=next_row, column=9, value='Yes' if holdings_info.get('held_by_ixa', False) else 'No')
    else:
        ws.cell(row=next_row, column=7, value="No OCLC record available - no valid OCLC number found")
        ws.cell(row=next_row, column=8, value=0)
        ws.cell(row=next_row, column=9, value='No')
    
    wb.save(review_file)

def add_record_to_marc_file(marc_file, barcode, workflow_json_path):
    """
    Add a single record to the MARC Excel spreadsheet.
    """
    workflow_data = load_workflow_json(workflow_json_path)
    
    if barcode not in workflow_data.get("records", {}):
        print(f"      Warning: Barcode {barcode} not found in workflow JSON")
        return
    
    workflow_record = workflow_data["records"][barcode]
    step1_data = workflow_record.get("step1_metadata_extraction", {})
    extracted_fields = step1_data.get("extracted_fields", {})
    
    if not extracted_fields:
        print(f"      Warning: No extracted fields for barcode {barcode}")
        return
    
    def is_valid_field(value):
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() and value.strip().lower() != 'not visible'
        return bool(value)
    
    def safe_get(obj, key, default=""):
        if obj is None:
            return default
        value = obj.get(key, default)
        return value if value is not None else default
    
    title_info = extracted_fields.get("title_information", {}) or {}
    main_title = safe_get(title_info, "main_title")
    subtitle = safe_get(title_info, "subtitle")
    primary_contributor = safe_get(title_info, "primary_contributor")
    
    publishers = extracted_fields.get("publishers", {}) or {}
    place = safe_get(publishers, "place")
    publisher_name_raw = safe_get(publishers, "name")
    publisher_name = publisher_name_raw.replace("Name: ", "").strip() if publisher_name_raw else ""
    
    dates = extracted_fields.get("dates", {}) or {}
    publication_date = safe_get(dates, "publication_date")
    
    contents = extracted_fields.get("contents", {}) or {}
    tracks = contents.get("tracks", []) or []
    
    wb = load_workbook(marc_file)
    ws = wb.active
    
    next_row = ws.max_row + 1
    
    ws.cell(row=next_row, column=1, value=barcode)
    
    if is_valid_field(primary_contributor):
        field_100 = f"100 1  {primary_contributor}, $ecomposer, $eperformer."
        ws.cell(row=next_row, column=2, value=field_100)
    
    if is_valid_field(main_title):
        title_field = f"245 1 0 {main_title}"
        if is_valid_field(subtitle):
            title_field += f" : $b{subtitle}"
        title_field += f" / $c{primary_contributor}." if is_valid_field(primary_contributor) else "."
        ws.cell(row=next_row, column=3, value=title_field)
    elif is_valid_field(primary_contributor):
        ws.cell(row=next_row, column=3, value=f"245 1 0 [Title not visible] / $c{primary_contributor}.")
    else:
        ws.cell(row=next_row, column=3, value="245 1 0 [Title and contributor not visible]")
    
    if is_valid_field(place) or is_valid_field(publisher_name) or is_valid_field(publication_date):
        pub_field = "264  1 "
        
        if is_valid_field(place):
            if is_valid_field(publisher_name):
                pub_field += f"{place} : "
            else:
                pub_field += f"{place} "
        
        if is_valid_field(publisher_name):
            pub_field += f"$b{publisher_name}"
            if is_valid_field(publication_date):
                pub_field += ", "
            else:
                pub_field += "."
        
        if is_valid_field(publication_date):
            date_clean = (publication_date or "").replace("©", "").replace("℗", "").strip()
            year_match = re.search(r'\b(19|20)\d{2}\b', date_clean)
            if year_match:
                date_clean = f"[{year_match.group()}]"
            pub_field += f"$c{date_clean}"
        
        pub_field = pub_field.rstrip(', ') + "."
        ws.cell(row=next_row, column=4, value=pub_field)
    
    ws.cell(row=next_row, column=5, value="300    1 audio disc : $banalog ; $c12 in.")
    ws.cell(row=next_row, column=6, value="340    vinyl.")
    
    if tracks and isinstance(tracks, list):
        track_list = []
        for track in tracks:
            if isinstance(track, dict):
                track_title = safe_get(track, "title")
                if is_valid_field(track_title):
                    if not any(x in track_title.lower() for x in ['standard lp', 'vinyl record', '12', 'vinyl', 'pvc']):
                        track_list.append(track_title)
            elif isinstance(track, str) and is_valid_field(track):
                if not any(x in track.lower() for x in ['standard lp', 'vinyl record', '12', 'vinyl', 'pvc']):
                    track_list.append(track)
        
        if track_list:
            contents_field = "505 0  " + " -- ".join(track_list) + "."
            contents_cell = ws.cell(row=next_row, column=7, value=contents_field)
            contents_cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws.cell(row=next_row, column=8, value="650  0  $aMusic.")
    
    wb.save(marc_file)


def get_ai_processing_date_from_tracking_filename(tracking_file):
    """
    Extract the AI processing date from the tracking spreadsheet filename.
    Format: tracking-spreadsheet-catalogers-YYYY-MM-DD-HH-MM-SS.xlsx
    Returns: YYYY-MM-DD
    """
    filename = os.path.basename(tracking_file)
    # Extract date pattern: YYYY-MM-DD
    match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    if match:
        return match.group(1)
    return None


def get_original_ai_processing_date(tracking_file, barcode):
    """
    Get the original AI processing date for a barcode from the tracking spreadsheet.
    Returns None if not found.
    """
    if not os.path.exists(tracking_file):
        return None
    
    wb = load_workbook(tracking_file)
    ws = wb.active
    
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=1).value) == str(barcode):
            date_value = ws.cell(row=row_idx, column=2).value
            return clean_date_format(date_value)  # Clean the date format
    
    return None


def update_tracking_spreadsheet_record(tracking_file, barcode, changes):
    """
    Update an existing record in the tracking spreadsheet.
    Used when a record already exists in low confidence tracking and needs updates.
    """
    wb = load_workbook(tracking_file)
    ws = wb.active
    
    # Find the row with this barcode
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=1).value) == str(barcode):
            target_row = row_idx
            break
    
    if target_row:
        # Update Date Cataloger Checked (column 5)
        ws.cell(row=target_row, column=5, value=changes['review_date'])
        
        # Update Status (column 6)
        ws.cell(row=target_row, column=6, value="Not Approved")
        
        # Clear Correct OCLC Number (column 7)
        ws.cell(row=target_row, column=7, value="")
        
        # Update Notes (column 8)
        ws.cell(row=target_row, column=8, value="Different OCLC # Needed")
        
        wb.save(tracking_file)
        return True
    
    return False

def setup_tracking_spreadsheet_conditional_formatting(tracking_file):
    """
    Apply conditional formatting to tracking spreadsheet.
    Highlights rows yellow ONLY if Status != "Approved" AND Correct OCLC is empty.
    This makes highlighting dynamic - it updates as catalogers work.
    """
    from openpyxl.formatting.rule import FormulaRule
    
    wb = load_workbook(tracking_file)
    ws = wb.active
    
    # Clear existing conditional formatting
    ws.conditional_formatting._cf_rules.clear()
    
    # Yellow highlight
    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Rule: Highlight if Status (column F) is NOT "Approved" AND Correct OCLC (column G) is empty
    # Using FormulaRule with AND condition
    # For row 2: =AND($F2<>"Approved", $G2="")
    # The $ before column letter makes it absolute, no $ before row number makes it relative
    rule = FormulaRule(
        formula=['AND($F2<>"Approved", $G2="")'],
        fill=highlight_fill
    )
    
    # Apply to entire data range (columns A-H, starting from row 2)
    max_row = ws.max_row if ws.max_row > 2 else 1000  # Allow for future rows
    ws.conditional_formatting.add(f'A2:H{max_row}', rule)
    
    wb.save(tracking_file)
    print(f"   Applied conditional formatting to tracking spreadsheet")

def add_record_to_tracking_spreadsheet(tracking_file, barcode, oclc_number, title, ai_processing_date, changes, status=None):
    """
    Add a single record to the tracking spreadsheet.
    Rows are highlighted ONLY if Status != "Approved" AND Correct OCLC is empty.
    """
    from openpyxl.formatting.rule import FormulaRule
    
    wb = load_workbook(tracking_file)
    ws = wb.active
    
    next_row = ws.max_row + 1
    
    ws.cell(row=next_row, column=1, value=barcode)
    ws.cell(row=next_row, column=2, value=ai_processing_date)
    ws.cell(row=next_row, column=3, value=oclc_number if oclc_number else "None suggested")
    ws.cell(row=next_row, column=4, value=title)
    ws.cell(row=next_row, column=5, value=changes.get('review_date', ''))
    
    if status:
        ws.cell(row=next_row, column=6, value=status)
    
    ws.cell(row=next_row, column=7, value=f'=IF(F{next_row}="Approved",C{next_row},"")')
    
    if changes.get('update_notes'):
        ws.cell(row=next_row, column=8, value=changes['update_notes'])
    
    # NO static highlighting - let conditional formatting handle it
    
    dv = DataValidation(type="list", formula1='"Approved,Different OCLC # Needed,Original Cataloging Needed,Further Review Needed"')
    ws.add_data_validation(dv)
    dv.add(f'F{next_row}')
    
    wb.save(tracking_file)

def map_decision_to_status(decision):
    """Map cataloger decision to tracking spreadsheet status."""
    status_map = {
        "Approved": "Approved",
        "Different OCLC # Needed": "Rejected - Different OCLC",
        "Original Cataloging Needed": "Rejected - Needs Original Cataloging",
        "Further Review Needed": ""  # Leave empty for further review
    }
    return status_map.get(decision, "")

def update_all_tracking_dates(tracking_file, decisions):
    """
    Update Date Cataloger Checked (column 5) and Status (column 6) for all records 
    in the tracking spreadsheet that appear in the cataloger decisions CSV.
    """
    if not os.path.exists(tracking_file):
        return
    
    wb = load_workbook(tracking_file)
    ws = wb.active
    
    updated_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        barcode = str(ws.cell(row=row_idx, column=1).value)
        if barcode in decisions:
            # Update Date Cataloger Checked (column 5)
            ws.cell(row=row_idx, column=5, value=decisions[barcode]['review_date'])
            
            # Update Status (column 6) if it's empty and not "Approved"
            current_status = ws.cell(row=row_idx, column=6).value
            if not current_status or (current_status and current_status.strip() != "Approved"):
                decision_text = decisions[barcode]['decision']
                if decision_text:  # Only update if we have a decision
                    ws.cell(row=row_idx, column=6, value=decision_text)
            
            updated_count += 1
    
    wb.save(tracking_file)
    print(f"      Updated Date Cataloger Checked and Status for {updated_count} records in tracking spreadsheet")

def remove_record_from_low_confidence_review(review_file, barcode):
    """
    Remove a record from the low confidence review Excel spreadsheet.
    """
    if not os.path.exists(review_file):
        return
    
    normalized_barcode = str(barcode).strip()
    
    wb = load_workbook(review_file)
    ws = wb.active
    
    rows_to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        cell_value = str(ws.cell(row=row_idx, column=2).value).strip()
        if cell_value == normalized_barcode:
            rows_to_delete.append(row_idx)
    
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    
    renumber_records(ws)
    
    wb.save(review_file)


def remove_record_from_low_confidence_marc(marc_file, barcode):
    """
    Remove a record from the MARC Excel spreadsheet.
    """
    if not os.path.exists(marc_file):
        return
    
    normalized_barcode = str(barcode).strip()
    
    wb = load_workbook(marc_file)
    ws = wb.active
    
    rows_to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        cell_value = str(ws.cell(row=row_idx, column=1).value).strip()
        if cell_value == normalized_barcode:
            rows_to_delete.append(row_idx)
    
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    
    wb.save(marc_file)


def remove_record_from_tracking_spreadsheet(tracking_file, barcode):
    """
    Remove a record from the tracking spreadsheet.
    """
    if not os.path.exists(tracking_file):
        return
    
    normalized_barcode = str(barcode).strip()
    
    wb = load_workbook(tracking_file)
    ws = wb.active
    
    rows_to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        cell_value = str(ws.cell(row=row_idx, column=1).value).strip()
        if cell_value == normalized_barcode:
            rows_to_delete.append(row_idx)
    
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    
    wb.save(tracking_file)


def remove_record_from_low_confidence_files(review_file, marc_file, tracking_file, barcode):
    """
    Remove a record from all low confidence Excel files.
    """
    remove_record_from_low_confidence_review(review_file, barcode)
    remove_record_from_low_confidence_marc(marc_file, barcode)
    remove_record_from_tracking_spreadsheet(tracking_file, barcode)

def renumber_records(worksheet):
    """
    Renumber the Record # column (column 1) in the review spreadsheet after deletions.
    """
    for idx, row_idx in enumerate(range(2, worksheet.max_row + 1), start=1):
        worksheet.cell(row=row_idx, column=1, value=idx)

def remove_duplicate_records_from_low_confidence_files(review_file, marc_file):
    """
    Remove duplicate records from low confidence Excel files based on barcode.
    """
    print("   Removing duplicate records from low confidence files...")
    
    if os.path.exists(review_file):
        wb = load_workbook(review_file)
        ws = wb.active
        
        seen_barcodes = set()
        rows_to_delete = []
        
        for row_idx in range(2, ws.max_row + 1):
            barcode = str(ws.cell(row=row_idx, column=2).value).strip()
            
            if barcode in seen_barcodes:
                rows_to_delete.append(row_idx)
            else:
                seen_barcodes.add(barcode)
        
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)
        
        if rows_to_delete:
            renumber_records(ws)
            print(f"      Removed {len(rows_to_delete)} duplicate(s) from review file")
        
        wb.save(review_file)
    
    if os.path.exists(marc_file):
        wb = load_workbook(marc_file)
        ws = wb.active
        
        seen_barcodes = set()
        rows_to_delete = []
        
        for row_idx in range(2, ws.max_row + 1):
            barcode = str(ws.cell(row=row_idx, column=1).value).strip()
            
            if barcode in seen_barcodes:
                rows_to_delete.append(row_idx)
            else:
                seen_barcodes.add(barcode)
        
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)
        
        if rows_to_delete:
            print(f"      Removed {len(rows_to_delete)} duplicate(s) from MARC file")
        
        wb.save(marc_file)

def load_decisions_history(results_folder):
    """Load existing decisions history spreadsheet."""
    deliverables_folder = os.path.join(results_folder, "deliverables")
    decisions_file = os.path.join(deliverables_folder, "decisions-history.xlsx")
    
    if not os.path.exists(decisions_file):
        print("   Error: decisions-history.xlsx not found. Please run Step 6 first.")
        return None
    
    wb = load_workbook(decisions_file)
    current_sheet = wb["Current Decisions"]
    current_decisions = {}
    
    for row_idx in range(2, current_sheet.max_row + 1):
        barcode = normalize_barcode(current_sheet.cell(row=row_idx, column=1).value)
        if not barcode:
            continue
        
        # UPDATED - No Record ID column
        current_decisions[barcode] = {
            'ai_suggested_oclc': str(current_sheet.cell(row=row_idx, column=2).value or ''),
            'ai_confidence': current_sheet.cell(row=row_idx, column=3).value or 0,
            'cataloger_decision': current_sheet.cell(row=row_idx, column=4).value,
            'chosen_oclc': str(current_sheet.cell(row=row_idx, column=5).value or ''),
            'final_confidence': current_sheet.cell(row=row_idx, column=6).value or 0,
            'held_by_ixa': current_sheet.cell(row=row_idx, column=7).value,
            'decided_by': current_sheet.cell(row=row_idx, column=8).value,
            'decided_at': current_sheet.cell(row=row_idx, column=9).value,
            'notes': current_sheet.cell(row=row_idx, column=10).value or '',
            'version': current_sheet.cell(row=row_idx, column=11).value or 'v1'
        }
    
    history_sheet = wb["Decision History"]
    version_counts = {}
    
    for row_idx in range(2, history_sheet.max_row + 1):
        barcode = normalize_barcode(history_sheet.cell(row=row_idx, column=1).value)
        if not barcode:
            continue
        version_counts[barcode] = version_counts.get(barcode, 0) + 1
    
    wb.close()
    
    return {
        'current_decisions': current_decisions,
        'version_counts': version_counts,
        'file_path': decisions_file
    }

def calculate_final_confidence(decision, correct_oclc, ai_confidence):
    """
    Calculate final confidence based on cataloger decision.
    
    Rules:
    - Not Reviewed: Keep original AI confidence (could be 0-100%)
    - Approved: 100%
    - Different OCLC Needed + valid OCLC provided: 100%
    - Different OCLC Needed (no OCLC): 0%
    - Original Cataloging Needed: 0%
    - Further Review Needed: 0%
    """
    decision = str(decision).strip()
    correct_oclc = str(correct_oclc).strip()
    
    # If not yet reviewed by cataloger, keep AI confidence
    if decision == "Not Reviewed":
        return ai_confidence
    
    # Cataloger decisions override with 0% or 100%
    if decision == "Approved":
        return 100
    
    if decision == "Different OCLC # Needed":
        if correct_oclc and correct_oclc not in ["", "0", "None"]:
            return 100
        else:
            return 0
    
    if decision in ["Original Cataloging Needed", "Further Review Needed"]:
        return 0
    
    # Default for any unknown decision
    return 0


def validate_decision(decision_data):
    """
    Validate that decision follows business rules.
    
    Rule: If decision is "Approved", chosen OCLC must match AI suggested OCLC
    
    Returns: (is_valid, error_message)
    """
    decision = str(decision_data.get('decision', '')).strip()
    ai_oclc = str(decision_data.get('ai_suggested_oclc', '')).strip()
    correct_oclc = str(decision_data.get('correct_oclc', '')).strip()
    
    if decision == "Approved":
        if correct_oclc != ai_oclc:
            return False, f"Decision is 'Approved' but OCLC changed from {ai_oclc} to {correct_oclc}"
    
    return True, None


def update_decisions_history(decisions_history_data, new_decisions):
    """Update decisions history spreadsheet with new cataloger decisions."""
    from datetime import datetime
    
    decisions_file = decisions_history_data['file_path']
    current_decisions = decisions_history_data['current_decisions']
    version_counts = decisions_history_data['version_counts']
    
    wb = load_workbook(decisions_file)
    current_sheet = wb["Current Decisions"]
    history_sheet = wb["Decision History"]
    
    api_calls_made = 0
    validation_errors = []
    
    for barcode, decision_data in new_decisions.items():
        is_valid, error_msg = validate_decision(decision_data)
        if not is_valid:
            validation_errors.append(f"Barcode {barcode}: {error_msg}")
            print(f"   Warning: {error_msg}")
            continue
        
        current_version_num = version_counts.get(barcode, 1)
        new_version_num = current_version_num + 1
        new_version = f"v{new_version_num}"
        
        # Determine chosen OCLC based on decision type
        decision = decision_data['decision']
        if decision == "Approved":
            # Approved means cataloger agrees with AI suggestion
            chosen_oclc = decision_data['ai_suggested_oclc']
        elif decision_data['correct_oclc']:
            # Cataloger provided a new OCLC number
            chosen_oclc = decision_data['correct_oclc']
        else:
            # No OCLC provided (rejected AI but no replacement, or needs original cataloging)
            chosen_oclc = ""
        
        # Calculate confidence with the CHOSEN OCLC
        final_confidence = calculate_final_confidence(
            decision_data['decision'],
            chosen_oclc,
            current_decisions.get(barcode, {}).get('ai_confidence', 0)
        )
        
        held_by_ixa = current_decisions.get(barcode, {}).get('held_by_ixa', 'No')

        # Only query APIs if there's a NEW, DIFFERENT OCLC number
        if chosen_oclc and chosen_oclc != decision_data['ai_suggested_oclc']:
            print(f"   Querying OCLC for new number: {chosen_oclc}")
            title, author, _ = fetch_oclc_data(chosen_oclc)  # Ignore OCLC holdings (unreliable)
            if title and author:
                api_calls_made += 1
            # Re-verify holdings against Alma for accurate status
            alma_result = verify_holdings_in_alma(str(chosen_oclc))
            held_by_ixa = "Yes" if alma_result.get("alma_verified", False) else "No"
            print(f"   Alma verification: held_by_ixa = {held_by_ixa}")
        
        current_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        new_row_data = [
            barcode,
            decision_data['ai_suggested_oclc'],
            current_decisions.get(barcode, {}).get('ai_confidence', 0),
            decision_data['decision'],
            chosen_oclc,  # Now correctly blank when appropriate
            final_confidence,
            held_by_ixa,
            decision_data['cataloger'],
            current_timestamp,
            decision_data['notes'],
            new_version
        ]
        
        history_row = history_sheet.max_row + 1
        for col_idx, value in enumerate(new_row_data, start=1):
            cell = history_sheet.cell(row=history_row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        found_row = None
        for row_idx in range(2, current_sheet.max_row + 1):
            if normalize_barcode(current_sheet.cell(row=row_idx, column=1).value) == barcode:
                found_row = row_idx
                break
        
        if found_row:
            for col_idx, value in enumerate(new_row_data, start=1):
                cell = current_sheet.cell(row=found_row, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            current_row = current_sheet.max_row + 1
            for col_idx, value in enumerate(new_row_data, start=1):
                cell = current_sheet.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        version_counts[barcode] = new_version_num
        
        current_decisions[barcode] = {
            'ai_suggested_oclc': decision_data['ai_suggested_oclc'],
            'ai_confidence': current_decisions.get(barcode, {}).get('ai_confidence', 0),
            'cataloger_decision': decision_data['decision'],
            'chosen_oclc': chosen_oclc,
            'final_confidence': final_confidence,
            'held_by_ixa': held_by_ixa,
            'decided_by': decision_data['cataloger'],
            'decided_at': current_timestamp,
            'notes': decision_data['notes'],
            'version': new_version
        }
    
    wb.save(decisions_file)
    
    return {
        'api_calls_made': api_calls_made,
        'validation_errors': validation_errors,
        'updated_decisions': current_decisions
    }

def apply_cataloger_decisions(csv_path, results_folder):
    """
    Apply cataloger decisions from CSV to all workflow outputs.
    Now uses decisions-history.xlsx as the single source of truth.
    """
    print("\n" + "=" * 60)
    print("APPLYING CATALOGER DECISIONS")
    print("=" * 60)
    
    print("\n1. Loading decisions history...")
    decisions_history_data = load_decisions_history(results_folder)
    
    if not decisions_history_data:
        return None
    
    print(f"   Loaded {len(decisions_history_data['current_decisions'])} existing decisions")
    
    print("\n2. Loading cataloger decisions from CSV...")
    decisions = load_cataloger_decisions(csv_path)
    print(f"   Loaded {len(decisions)} decisions from CSV")
    
    print("\n3. Updating decisions history spreadsheet...")
    workflow_json_path = get_workflow_json_path(os.path.join(results_folder, "data"))
    update_result = update_decisions_history(
        decisions_history_data,
        decisions
    )
    
    print(f"   Updated {len(decisions)} decisions")
    print(f"   API calls made: {update_result['api_calls_made']}")
    
    if update_result['validation_errors']:
        print(f"   Validation warnings: {len(update_result['validation_errors'])}")
        for error in update_result['validation_errors'][:5]:
            print(f"      {error}")
    
    current_decisions = update_result['updated_decisions']
    
    print("\n4. Backing up original files...")
    backed_up_files = backup_original_files(results_folder)
    
    deliverables_folder = os.path.join(results_folder, "deliverables")
    
    sorting_files = [f for f in os.listdir(deliverables_folder) 
                     if f.startswith("lp-workflow-sorting-")]
    batch_upload_files = [f for f in os.listdir(deliverables_folder) 
                          if f.startswith("batch-upload-alma-lp-")]
    tracking_files = [f for f in os.listdir(deliverables_folder) 
                      if f.startswith("tracking-spreadsheet-catalogers-")]
    review_files = [f for f in os.listdir(deliverables_folder) 
                    if f.startswith("low-confidence-matches-review-")]
    marc_files = [f for f in os.listdir(deliverables_folder) 
                  if f.startswith("low-confidence-marc-")]
    
    if not sorting_files:
        print("   Error: No sorting spreadsheet found")
        return None
    
    sorting_file = os.path.join(deliverables_folder, max(sorting_files))
    batch_upload_file = os.path.join(deliverables_folder, max(batch_upload_files)) if batch_upload_files else None
    tracking_file = os.path.join(deliverables_folder, max(tracking_files)) if tracking_files else None
    review_file = os.path.join(deliverables_folder, max(review_files)) if review_files else None
    marc_file = os.path.join(deliverables_folder, max(marc_files)) if marc_files else None
    
    print("\n5. Determining changes from decisions history...")
    current_state_map = {}
    wb = load_workbook(sorting_file)
    sheet = wb.active
    
    for row_idx in range(2, sheet.max_row + 1):
        barcode = normalize_barcode(sheet.cell(row=row_idx, column=1).value)
        if barcode:
            current_state_map[barcode] = {
                'sort_group': sheet.cell(row=row_idx, column=2).value,
                'oclc_number': str(sheet.cell(row=row_idx, column=3).value or ''),
                'title': sheet.cell(row=row_idx, column=4).value
            }
    wb.close()
    
    all_changes = {}
    api_calls_made = update_result['api_calls_made']
    
    for barcode in decisions.keys():
        if barcode in current_state_map:
            changes = determine_changes(
                barcode,
                decisions[barcode],
                current_state_map[barcode],
                current_decisions
            )
            all_changes[barcode] = changes
    
    print(f"   Determined changes for {len(all_changes)} records")
    
    # Just need to update the summary at the end to include decisions info:
    summary = {
        'total': len(decisions),
        'api_calls': api_calls_made,
        'validation_errors': len(update_result['validation_errors']),
        'files_updated': len(backed_up_files),
        'backed_up_files': list(backed_up_files.keys()),
        'decisions_history_updated': True,
        'status_changes': {
            'promoted_to_high_confidence': sum(1 for c in all_changes.values() 
                if c['new_status'] == 'Alma Batch Upload (High Confidence)' 
                and c['original_status'] == 'Cataloger Review (Low Confidence)'),
            'demoted_to_low_confidence': sum(1 for c in all_changes.values() 
                if c['new_status'] == 'Cataloger Review (Low Confidence)' 
                and c['original_status'] != 'Cataloger Review (Low Confidence)'),
            'changed_to_held': sum(1 for c in all_changes.values() 
                if c['new_status'] == 'Held by UT Libraries (IXA)' 
                and c['original_status'] != 'Held by UT Libraries (IXA)'),
            'oclc_numbers_changed': sum(1 for c in all_changes.values() if c['new_oclc']),
        }
    }
    
    # Step 6: Update sorting spreadsheet
    print("\n6. Updating sorting spreadsheet...")
    current_state, wb, ws = load_current_state_from_sorting(sorting_file)
    
    for barcode, changes in all_changes.items():
        if barcode in current_state:
            update_sorting_spreadsheet(ws, barcode, changes, current_state)
    
    wb.save(sorting_file)
    print(f"   Updated {len(all_changes)} records in sorting spreadsheet")
    
    # Step 7: Update batch upload file
    if batch_upload_file:
        print("\n7. Updating batch upload file...")
        update_batch_upload_file(batch_upload_file, all_changes, current_state)
        print("   Batch upload file updated")
    
    # Step 8: Update low confidence files
    if review_file and marc_file and tracking_file:
        print("\n8. Updating low confidence files...")
        
        # Remove duplicates first
        print("      Removing duplicates...")
        remove_duplicate_records_from_low_confidence_files(review_file, marc_file)
        
        # Remove ALL records being processed (will re-add if needed)
        print("      Removing all records being reviewed/updated...")
        removed_count = 0
        for barcode, changes in all_changes.items():
            # Remove from low confidence files regardless of new status
            # (We'll add back if needed in the next step)
            remove_record_from_low_confidence_files(review_file, marc_file, tracking_file, barcode)
            removed_count += 1
        print(f"      Removed {removed_count} records")
        
        # Now add back ONLY those that should be in low confidence
        print("      Adding items that should be in low confidence...")
        added_count = 0
        for barcode, changes in all_changes.items():
            if changes['add_to_low_confidence']:
                if changes['clear_oclc']:
                    oclc_num = None
                else:
                    oclc_num = changes['new_oclc'] if changes['new_oclc'] else changes['original_oclc']
                
                if changes['clear_title']:
                    title = "Title cleared - awaiting new OCLC"
                else:
                    title = changes['new_title'] if changes['new_title'] else changes['original_title']
                
                review_oclc = changes['original_oclc'] if changes['original_oclc'] else None
                
                ai_processing_date = get_ai_processing_date_from_tracking_filename(tracking_file)
                if not ai_processing_date:
                    ai_processing_date = get_original_ai_processing_date(tracking_file, barcode)
                    if not ai_processing_date:
                        ai_processing_date = clean_date_format(get_current_timestamp())

                add_record_to_low_confidence_review(review_file, barcode, review_oclc, changes['final_confidence'], workflow_json_path)
                add_record_to_marc_file(marc_file, barcode, workflow_json_path)
                
                add_record_to_tracking_spreadsheet(
                    tracking_file, barcode, oclc_num, title, ai_processing_date, changes,
                    status=decisions[barcode]['decision']
                )
                added_count += 1
        
        print(f"      Added {added_count} items to low confidence files")
        print("   Low confidence files updated")
        
        print("   Updating cataloger review dates in tracking spreadsheet...")
        update_all_tracking_dates(tracking_file, decisions)
        setup_tracking_spreadsheet_conditional_formatting(tracking_file)
    
    # Step 9: Update workflow JSON
    print("\n9. Updating workflow JSON...")
    try:
        workflow_data = load_workflow_json(workflow_json_path)
        
        for barcode, changes in all_changes.items():
            if barcode in workflow_data.get("records", {}):
                try:
                    review_date_str = str(changes['review_date']) if changes['review_date'] else ""
                    cataloger_name_str = str(changes['cataloger']) if changes['cataloger'] else ""
                    notes_str = str(changes['update_notes']) if changes['update_notes'] else ""
                    original_status_str = str(changes['original_status']) if changes['original_status'] else ""
                    new_status_str = str(changes['new_status']) if changes['new_status'] else ""
                    original_oclc_str = str(changes['original_oclc']) if changes['original_oclc'] else ""
                    new_oclc_str = str(changes['new_oclc']) if changes['new_oclc'] else ""
                    decision_str = str(decisions[barcode]['decision']) if decisions[barcode]['decision'] else ""

                    update_record_step7(
                        json_path=workflow_json_path,
                        barcode=barcode,
                        cataloger_decision=decision_str,
                        original_status=original_status_str,
                        new_status=new_status_str,
                        original_oclc=original_oclc_str,
                        new_oclc=new_oclc_str,
                        cataloger_name=cataloger_name_str,
                        review_date=review_date_str,
                        notes=notes_str,
                        new_oclc_bib_data=changes.get('new_oclc_bib_data')
                    )
                except Exception as record_error:
                    print(f"   Warning: Could not update record {barcode} in workflow JSON: {record_error}")
        
        print("   ✓ Workflow JSON updated")
    except Exception as json_error:
        print(f"   Warning: Could not update workflow JSON: {json_error}")

    # Step 10: Regenerate HTML review interface with updated data
    print("\n10. Regenerating HTML review interface...")
    try:
        # First, delete old HTML files to avoid duplicates
        import glob
        old_html_files = glob.glob(os.path.join(results_folder, "review-index-*.html"))
        old_html_files.extend(glob.glob(os.path.join(results_folder, "review-page-*.html")))
        for old_file in old_html_files:
            try:
                os.remove(old_file)
            except Exception:
                pass  # Ignore errors deleting old files
        if old_html_files:
            print(f"   Removed {len(old_html_files)} old HTML files")

        # Load the updated records from the sorting spreadsheet
        updated_records = load_records_from_step5(sorting_file)

        # Generate new timestamp for the regenerated HTML
        html_timestamp = get_current_timestamp()

        # Regenerate the HTML pages
        html_result = create_paginated_review_html(
            results_folder,
            updated_records,
            html_timestamp,
            workflow_json_path,
            records_per_page=100
        )

        print(f"   ✓ HTML regenerated: {html_result['total_pages']} pages created")
        print(f"   Index: {html_result['index_path']}")
        summary['html_regenerated'] = True
        summary['html_pages'] = html_result['total_pages']
    except Exception as html_error:
        print(f"   Warning: Could not regenerate HTML: {html_error}")
        summary['html_regenerated'] = False

    # Step 11: Log metrics
    try:
        step7_metrics = {
            "total_decisions_processed": summary['total'],
            "api_calls_made": summary['api_calls'],
            "files_modified": summary['files_updated'],
            "status_changes": summary['status_changes'],
            "csv_source": os.path.basename(csv_path),
            "timestamp": datetime.now().isoformat()
        }
        
        log_processing_metrics(
            results_folder_path=results_folder,
            step="step7_cataloger_decisions",
            batch_metrics=step7_metrics
        )
    except Exception as metrics_error:
        print(f"   Warning: Could not log Step 7 metrics: {metrics_error}")
    
    return summary

def validate_results_folder(folder_path):
    if not os.path.exists(folder_path):
        print(f"\nError: Folder path does not exist: {folder_path}")
        return False
    
    if not os.path.isdir(folder_path):
        print(f"\nError: Path is not a directory: {folder_path}")
        return False
    
    deliverables_folder = os.path.join(folder_path, "deliverables")
    if not os.path.exists(deliverables_folder):
        print(f"\nError: No 'deliverables' subfolder found in: {folder_path}")
        print("   Please ensure you are pointing to a valid results folder from previous workflow steps.")
        return False
    
    return True


def main():
    print("=" * 60)
    print("Step 7: Applying Cataloger Decisions to Workflow Outputs")
    print("=" * 60)
    
    # Prompt for cataloger decisions CSV path
    csv_path = input("\nEnter path to cataloger decisions CSV: ").strip()
    csv_path = csv_path.strip('"').strip("'")
    
    if not os.path.exists(csv_path):
        print(f"\nError: CSV file not found at {csv_path}")
        return
    
    # Prompt for results folder path instead of automatically finding it
    print("\nEnter path to the results folder containing the deliverables to update.")
    print("This should be the folder created by previous workflow steps (e.g., 'lp-workflow-results-2025-01-15-143022')")
    results_folder = input("Results folder path: ").strip()
    results_folder = results_folder.strip('"').strip("'")
    
    # Validate the results folder
    if not validate_results_folder(results_folder):
        return
    
    print(f"\nUsing results folder: {results_folder}")
    
    # Confirm before proceeding
    response = input("\nThis will modify deliverable files. Original files will be backed up. Continue? (yes/no): ").strip().lower()
    if response not in ['yes', 'y']:
        print("Operation cancelled.")
        return
    
    try:
        summary = apply_cataloger_decisions(csv_path, results_folder)
        
        print("\n" + "=" * 60)
        print("STEP 7 COMPLETE")
        print("=" * 60)
        print(f"Processed {summary['total']} cataloger decisions")
        print(f"\nUpdated deliverables: {os.path.join(results_folder, 'deliverables')}/")
        print(f"Copied original to: {os.path.join(results_folder, 'original-outputs')}/")

        if summary.get('html_regenerated'):
            print(f"\nHTML review interface regenerated with {summary.get('html_pages', 0)} pages")
            print("Catalogers can now review updated records using the new HTML files.")

    except Exception as e:
        print(f"\nError during processing: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()