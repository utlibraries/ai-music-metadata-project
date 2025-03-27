import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def split_by_confidence_score(input_dir):
    # Find the most recent step3 file
    step3_files = [f for f in os.listdir(input_dir) if f.startswith('ai-music-step-3-lp')]
    if not step3_files:
        print("No step3 files found!")
        return

    latest_file = max(step3_files)
    workbook_path = os.path.join(input_dir, latest_file)
    print(f"Processing file: {workbook_path}")

    # Load the source workbook
    wb_source = load_workbook(workbook_path)
    sheet_source = wb_source.active

    # Create two new workbooks
    wb_high = Workbook()
    sheet_high = wb_high.active
    wb_low = Workbook()
    sheet_low = wb_low.active

    # Copy headers and set column widths
    headers = [cell.value for cell in sheet_source[1]]
    sheet_high.append(headers)
    sheet_low.append(headers)

    # Set column widths
    for col in range(1, 4):  # Image columns
        sheet_high.column_dimensions[get_column_letter(col)].width = 30
        sheet_low.column_dimensions[get_column_letter(col)].width = 30
    sheet_high.column_dimensions[get_column_letter(4)].width = 15  # Internal Identifier
    sheet_low.column_dimensions[get_column_letter(4)].width = 15
    for col in range(5, sheet_source.max_column + 1):  # Other columns
        if col in [8, 9]:  # Columns H and I
            sheet_high.column_dimensions[get_column_letter(col)].width = 25
            sheet_low.column_dimensions[get_column_letter(col)].width = 25
        else:
            sheet_high.column_dimensions[get_column_letter(col)].width = 55
            sheet_low.column_dimensions[get_column_letter(col)].width = 55

    # Base path for image folders
    base_image_path = "/Users/hannahmoutran/Library/CloudStorage/Box-Box/AI Music Metadata Project"

    # Keep track of current row in each new sheet
    high_row = 2  # Start after header
    low_row = 2   # Start after header

    # Process each row starting from row 2 (after header)
    for row in range(2, sheet_source.max_row + 1):
        confidence_score = sheet_source[f'I{row}'].value
        internal_identifier = sheet_source[f'D{row}'].value
        
        try:
            # Convert confidence score to float/int if it's a string
            if isinstance(confidence_score, str):
                confidence_score = float(confidence_score.strip('%'))
            
            # Determine target sheet and row
            if confidence_score > 85:
                target_sheet = sheet_high
                target_row = high_row
                high_row += 1
            else:
                target_sheet = sheet_low
                target_row = low_row
                low_row += 1

            # Copy row data
            for col in range(1, sheet_source.max_column + 1):
                target_sheet.cell(row=target_row, column=col).value = sheet_source.cell(row=row, column=col).value

            # Add images
            if internal_identifier:
                record_folder = os.path.join(base_image_path, internal_identifier)
                if os.path.exists(record_folder):
                    images = [f for f in os.listdir(record_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
                    for i, img_name in enumerate(images[:3], start=1):
                        img_path = os.path.join(record_folder, img_name)
                        img = PILImage.open(img_path)
                        img.thumbnail((200, 200))

                        output = BytesIO()
                        img.save(output, format='PNG')
                        output.seek(0)

                        img_excel = Image(output)
                        img_excel.anchor = target_sheet.cell(row=target_row, column=i).coordinate
                        target_sheet.add_image(img_excel)

            # Set row height and cell alignment
            target_sheet.row_dimensions[target_row].height = 215
            for cell in target_sheet[target_row]:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        except (ValueError, TypeError) as e:
            print(f"Error processing row {row}: {e}")
            continue

    # Freeze header row
    sheet_high.freeze_panes = 'A2'
    sheet_low.freeze_panes = 'A2'

    # Save the new workbooks
    current_date = datetime.now().strftime("%Y-%m-%d")
    high_confidence_file = f"ai-music-step-4-high-confidence-lp-{current_date}.xlsx"
    low_confidence_file = f"ai-music-step-4-low-confidence-lp-{current_date}.xlsx"
    
    full_high_path = os.path.join(input_dir, high_confidence_file)
    full_low_path = os.path.join(input_dir, low_confidence_file)
    
    wb_high.save(full_high_path)
    wb_low.save(full_low_path)
    
    print(f"High confidence results saved to: {high_confidence_file}")
    print(f"Low confidence results saved to: {low_confidence_file}")

if __name__ == "__main__":
    input_dir = "/Users/hannahmoutran/Library/CloudStorage/Box-Box/AI Music Metadata Project"
    split_by_confidence_score(input_dir)