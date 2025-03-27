import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

def get_relator_code(role):
    """Convert role text to MARC relator code."""
    role_map = {
        'conductor': 'cnd',
        'performer': 'prf',
        'composer': 'cmp',
        'arranger': 'arr',
        'instrumentalist': 'itr',
        'vocalist': 'voc',
        'orchestra': 'prf'
    }
    role = role.lower().strip()
    return role_map.get(role, 'prf')

def parse_metadata(metadata_str):
    """Convert the metadata string into a structured dictionary."""
    lines = metadata_str.strip().split('\n')
    parsed = {}
    current_section = None
    current_items = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if not line.startswith('- '):  # Section header
            if current_section and current_items:
                parsed[current_section] = current_items
            current_section = line.rstrip(':')
            current_items = []
        else:  # List item
            current_items.append(line[2:])  # Remove "- " prefix
            
    if current_section and current_items:
        parsed[current_section] = current_items
        
    return parsed

def format_marc_record(metadata_dict):
    """Convert parsed metadata into MARC format."""
    marc_lines = []
    
    # Fixed Fields
    marc_lines.append("007  s|bd|df|eu|fn|gg|hn|in|me|nu")
    marc_lines.append("040  \\\\$aIXA$beng$cIXA")
    marc_lines.append("049  \\\\$aIXAA")
    
    # Language
    if 'Language' in metadata_dict:
        for lang in metadata_dict['Language']:
            if lang.startswith('sungLanguage:'):
                sung = lang.split('sungLanguage:', 1)[1].strip()
                if sung not in ['[Not specified]', '[Not Applicable]']:
                    marc_lines.append(f"041  0\\$a{sung}")
            elif lang.startswith('printedLanguage:'):
                printed = lang.split('printedLanguage:', 1)[1].strip()
                marc_lines.append(f"041  0\\$a{'$a'.join(printed.split(', '))}")
    
    # Title
    if 'Title Information' in metadata_dict:
        for title in metadata_dict['Title Information']:
            if title.startswith('Main Title:'):
                title_text = title.split('Main Title:', 1)[1].strip()
                if '(' in title_text and ')' in title_text:
                    main_title, rest = title_text.split('(', 1)
                    alt_title, creator = rest.split(')', 1)
                    marc_lines.append(f"245  10$a{main_title.strip()}$h[sound recording] :$b({alt_title.strip()}){creator}")
                else:
                    marc_lines.append(f"245  10$a{title_text}$h[sound recording]")
    
    # Publication Info and Catalog Numbers
    if 'Publishers' in metadata_dict:
        pub_elements = []
        for pub in metadata_dict['Publishers']:
            if pub.startswith('Name:'):
                pub_name = pub.split('Name:', 1)[1].strip()
                pub_elements.append(f"$b{pub_name}")
            elif pub.startswith('Numbers:'):
                numbers = pub.split('Numbers:', 1)[1].strip()
                if numbers.startswith('[') and numbers.endswith(']'):
                    number = numbers[1:-1]  # Remove brackets
                    if '(' in number:  # Handle "(catalog)" suffix
                        number = number.split('(')[0].strip()
                    if number not in ['Not Available', 'Music', 'Not specified']:
                        marc_lines.append(f"028  01$a{number}")
            elif pub.startswith('Place:'):
                place = pub.split('Place:', 1)[1].strip()
                if place in ['[Not specified]', '[place of publication not identified]']:
                    pub_elements.append("$a[Place of publication not identified]")
                else:
                    pub_elements.append(f"$a{place}")
        
        # Add publication date
        if 'Dates' in metadata_dict:
            for date in metadata_dict['Dates']:
                if date.startswith('publicationDate:'):
                    pub_date = date.split('publicationDate:', 1)[1].strip()
                    pub_elements.append(f"$c{pub_date}")
                elif date.startswith('recordingLocation:'):
                    location = date.split('recordingLocation:', 1)[1].strip()
                    if location not in ['[Not specified]', '[Not Available]']:
                        marc_lines.append(f"518  \\\\$aRecorded in {location}.")
        
        if pub_elements:
            marc_lines.append("260  \\\\" + "".join(pub_elements))
    
    # Physical Description
    if 'Physical Description' in metadata_dict:
        size = None
        material = None
        for desc in metadata_dict['Physical Description']:
            if desc.startswith('size:'):
                size = desc.split('size:', 1)[1].strip()
            elif desc.startswith('material:'):
                material = desc.split('material:', 1)[1].strip()
        
        if size:
            physical_desc = f"300  \\\\$a1 audio disc :$bstereo ;$c{size}"
            if material and material not in ['[Not specified]', '[Not Available]']:
                physical_desc += f"$e{material}"
            marc_lines.append(physical_desc)
    
    # Content/Media/Carrier Types
    marc_lines.append("336  \\\\$aperformed music$bprm$2rdacontent")
    marc_lines.append("337  \\\\$aaudio$bs$2rdamedia")
    marc_lines.append("338  \\\\$aaudio disc$bsd$2rdacarrier")
    
    # Notes
    if 'Notes' in metadata_dict:
        for note in metadata_dict['Notes']:
            if note.startswith('generalNotes:'):
                note_text = note.split('generalNotes:', 1)[1].strip()
                if note_text.startswith('[') and note_text.endswith(']'):
                    note_text = note_text[1:-1]  # Remove outer brackets
                    notes = eval(note_text)  # Safe here as we control the input
                    for n in notes:
                        if isinstance(n, dict) and 'text' in n and n['text'] != '[Not specified]':
                            marc_lines.append(f"500  \\\\$a{n['text']}")
            elif note.startswith('performerNotes:'):
                perf_note = note.split('performerNotes:', 1)[1].strip()
                if perf_note.startswith('[') and perf_note.endswith(']'):
                    perf_note = perf_note[1:-1]  # Remove []
                    if perf_note != '[Not specified]':
                        marc_lines.append(f"511  0\\$a{perf_note}")
            elif note.startswith('participantNote:'):
                part_note = note.split('participantNote:', 1)[1].strip()
                if part_note.startswith('[') and part_note.endswith(']'):
                    part_note = part_note[1:-1]  # Remove []
                    if part_note != '[Not specified]':
                        marc_lines.append(f"500  \\\\$a{part_note}")
    
    # Sound Characteristics
    if 'Sound Characteristics' in metadata_dict:
        for char in metadata_dict['Sound Characteristics']:
            if char.startswith('playbackSpeed:'):
                speed = char.split('playbackSpeed:', 1)[1].strip()
                if speed not in ['[Not Available]', '[Not specified]']:
                    marc_lines.append(f"344  \\\\$c{speed}")
            if char.startswith('soundConfiguration:'):
                config = char.split('soundConfiguration:', 1)[1].strip()
                if config not in ['[Not Available]', '[Not specified]']:
                    marc_lines.append(f"344  \\\\$g{config}")
    
    # Subject Information
    if 'Subject Information' in metadata_dict:
        for subject in metadata_dict['Subject Information']:
            if subject.startswith('genre:'):
                genres = subject.split('genre:', 1)[1].strip()
                for genre in genres.split(', '):
                    if genre not in ['[Not Available]', '[Not specified]']:
                        marc_lines.append(f"650  \\0$a{genre}")
            elif subject.startswith('geographic:'):
                geo = subject.split('geographic:', 1)[1].strip()
                if geo not in ['[Not Available]', '[Not specified]']:
                    marc_lines.append(f"651  \\0$a{geo}")
            elif subject.startswith('timePeriod:'):
                period = subject.split('timePeriod:', 1)[1].strip()
                if period not in ['[Not Available]', '[Not specified]']:
                    marc_lines.append(f"648  \\0$a{period}")
    
    # Contents
    if 'Contents' in metadata_dict:
        for content in metadata_dict['Contents']:
            if content.startswith('tracks:'):
                tracks_text = content.split('tracks:', 1)[1].strip()
                if tracks_text not in ['[Not Available]', '[Not specified]']:
                    try:
                        tracks = eval(tracks_text)
                        contents = []
                        for track in tracks:
                            track_info = []
                            if track.get('number'):
                                track_info.append(track['number'] + ".")
                            if track.get('title'):
                                track_info.append(track['title'])
                            if track.get('composer'):
                                track_info.append(f" / {track['composer']}")
                            if track.get('duration'):
                                track_info.append(f" ({track['duration']})")
                            contents.append(" ".join(track_info))
                        if contents:
                            marc_lines.append(f"505  0\\$a{' -- '.join(contents)}")
                    except:
                        pass
    
    # Contributors
    if 'Contributors' in metadata_dict:
        for i, contributor in enumerate(metadata_dict['Contributors']):
            if contributor.startswith('Additional Contributors:'):
                continue
            elif contributor.startswith('Conductor:'):
                name = contributor.split(':', 1)[1].strip()
                marc_lines.append(f"700  1\\$a{name}$4cnd")
            elif contributor.startswith('Orchestra:'):
                name = contributor.split(':', 1)[1].strip()
                marc_lines.append(f"700  1\\$a{name}$4prf")
            elif '(' in contributor:
                name, role = contributor.split('(', 1)
                role = role.rstrip(')')
                if i == 0:
                    marc_lines.append(f"100  1\\$a{name.strip()}$4{get_relator_code(role)}")
                else:
                    marc_lines.append(f"700  1\\$a{name.strip()}$4{get_relator_code(role)}")
    
    return chr(10).join(marc_lines)

def process_spreadsheet(input_dir):
    # Find the most recent low confidence file
    low_conf_files = [f for f in os.listdir(input_dir) if f.startswith('ai-music-step-4-low-confidence-')]
    if not low_conf_files:
        print("No low confidence files found!")
        return

    latest_file = max(low_conf_files)
    workbook_path = os.path.join(input_dir, latest_file)
    print(f"Processing file: {workbook_path}")

    # Load the workbook
    wb = load_workbook(workbook_path)
    ws = wb.active

    # Insert new column for MARC records if it doesn't exist
    if ws['F1'].value != 'MARC Record':
        ws.insert_cols(6)
        ws['F1'] = 'MARC Record'
    ws.column_dimensions['F'].width = 100

    # Process each row
    for row in range(2, ws.max_row + 1):
        metadata = ws.cell(row=row, column=5).value
        if metadata and not isinstance(metadata, (int, float)):  # Skip non-string cells
            try:
                parsed_metadata = parse_metadata(metadata)
                marc_record = format_marc_record(parsed_metadata)
                
                # Write MARC record to new column
                marc_cell = ws.cell(row=row, column=6)
                marc_cell.value = marc_record
                marc_cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                print(f"Processed row {row}/{ws.max_row}")
            except Exception as e:
                print(f"Error processing row {row}: {str(e)}")
                ws.cell(row=row, column=6).value = f"Error: {str(e)}"

    # Save the updated workbook
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"ai-music-step-5-low-confidence-lp-{current_date}.xlsx"
    full_output_path = os.path.join(input_dir, output_file)
    
    wb.save(full_output_path)
    print(f"Results saved to {full_output_path}")

if __name__ == "__main__":
    input_dir = "/Users/hannahmoutran/Library/CloudStorage/Box-Box/AI Music Metadata Project"
    process_spreadsheet(input_dir)