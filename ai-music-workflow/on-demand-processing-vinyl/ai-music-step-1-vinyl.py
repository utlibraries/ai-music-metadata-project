import os
import json
import base64
import re
import time
from datetime import datetime
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openai import OpenAI
from collections import defaultdict

client = OpenAI(api_key=os.getenv('OPENAI_HMRC_API_KEY'))
total_tokens = 0
prompt_tokens = 0
completion_tokens = 0
max_tokens_per_item = 0
min_tokens_per_item = float('inf')
items_processed = 0

def get_llm_prompt():
    return """Analyze these images of a vinyl record and extract the following metadata fields in the specified format. 

### Rules:
1. For Latin alphabet based languages, such as Germanic languages, romance languages, slavic languages: use original language unless otherwise noted in the metadata instructions.
2. For non-Latin alphabet based languages, for example languages written in the cyrillic alphabet, in arabic script, in chinese characters: metadata should be either English transliterations or English translations.  No non latin characters should be included in the metadata.
3. Do not include any notes to explain why something has been included or excluded from a query.
4. Do not include any information that is not visible in the images.  Mark fields where the information is not available or unclear as "Not visible".
5. Take care to spell names and titles correctly. 
6. Metadata must be in this exact format, with no additional fields and no introductory, concluding, or other text.  

### Metadata Format:
Title Information:
  - Main Title: [Title, transliterated if the characters are not Latin]
  - English Title: [Title in English if written on the item]
  - Subtitle: [Subtitle in English or transliterated.  If a subtitle is written in both English and another language, use the English subtitle]
Primary Contributor:
  - Artist/Performer: [Name of primary artist or group (transliterated if the characters are not Latin)]
Additional Contributors:
  - Arrangers: [List of arrangers]
  - Engineers: [List of engineers]
  - Producers: [List of producers]
  - Session Musicians: [List of session musicians with instruments]
Publishers:
  - Name: [Publisher Name]
    Place: [Place of publication]
    Numbers: [Manufacturing numbers, UPC (NOT university library barcode stickers)]
Dates:
  - publicationDate: [Publication Date]
  - recordingDate: [Recording Date]
  - recordingLocation: [Recording location]
Language:
  - sungLanguages: [Languages of sung text]
  - printedLanguages: [All printed languages visible on the images]
  - pressingLanguage: [Primary printed language on the item that is not English - particularly important for items with multiple languages]
Format:
  - generalFormat: [e.g., Sound Recording]
  - specificFormat: [e.g., Vinyl Record, LP]
  - materialTypes: [List of Material Types]
Sound Characteristics:
  - soundConfiguration: [e.g., stereo/mono]
  - recordingType: [analog/digital]
  - specialPlayback: [e.g., Quadraphonic]
Physical Description:
  - size: [Diameter, e.g., 12", 7"]
  - material: [e.g., Vinyl]
  - labelDesign: [Label design and color]
  - physicalCondition: [Condition notes]
  - specialFeatures: [Sleeve details, inserts, etc.]
Contents:
  - tracks: [
      {
        "number": [Track number],
        "title": [Track title (transliterated if the characters are not Latin)],
        "titleOriginalLanguage": [Title in original language if different],
        "composer": [Composer or songwriter name],
        "lyricist": [Lyricist name],
        "duration": [Duration if shown]
      }
    ]
Series:
  - seriesTitle: [Series name if any (transliterated if the characters are not Latin)]
  - seriesNumber: [Number within series]
Subject Information:
  - genre: [Musical genre/style]
  - geographic: [Geographic origins]
  - timePeriod: [Time period]
Notes:
  - generalNotes: [Any additional notes]
  - technicalNotes: [Pressing plant, mastering info]

Include additional information in the notes section. Return only the metadata in the specified format."""

def group_images_by_barcode(folder_path):
    """Group image files by their barcode number."""
    image_groups = defaultdict(list)

    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.jpg', '.jpeg')):
            # Match filenames with numeric barcode followed by a, b, or c
            match = re.match(r'(\d+)[a-c]\.jpe?g', filename)
            if match:
                barcode = match.group(1)  # Extract the numeric barcode
                image_groups[barcode].append(os.path.join(folder_path, filename))
    
    # Sort files within each group by the suffix (a, b, c)
    for barcode in image_groups:
        image_groups[barcode].sort(key=lambda x: os.path.basename(x)[-5])  # Sort by letter

    return image_groups

def process_folder(folder_path, wb):
    global total_tokens, prompt_tokens, completion_tokens, max_tokens_per_item, min_tokens_per_item, items_processed
    ws = wb.active
    headers = ['Input Image 1', 'Input Image 2', 'Input Image 3', 'Barcode', 'AI-Generated Metadata']
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        if col == 4:  # Barcode column
            ws.column_dimensions[get_column_letter(col)].width = 15
        else:
            ws.column_dimensions[get_column_letter(col)].width = 30 if col <= 3 else 52

    image_groups = group_images_by_barcode(folder_path)
    total_items = len(image_groups)
    items_with_issues = 0
    processed_items = 0

    for barcode, image_paths in sorted(image_groups.items()):
        processed_items += 1
        items_processed += 1

        try:
            # Take up to first 3 images for each barcode
            image_paths = image_paths[:3]
            prompt_text = get_llm_prompt()
            uploaded_files_info = ""

            for img_path in image_paths:
                uploaded_files_info += f"[Image file path: {img_path}]\n"

            prompt = prompt_text + "\n" + uploaded_files_info

            try:
                base64_images = []
                for img_path in image_paths:
                    with open(img_path, "rb") as image_file:
                        base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                        base64_images.append(base64_image)

                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            *[{
                                "type": "image_url",
                                "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                            } for base64_image in base64_images]
                        ]
                    }],
                    max_tokens=2000
                )

                # Update token counts
                item_total_tokens = response.usage.total_tokens
                total_tokens += item_total_tokens
                prompt_tokens += response.usage.prompt_tokens
                completion_tokens += response.usage.completion_tokens
                
                # Update min/max tokens
                max_tokens_per_item = max(max_tokens_per_item, item_total_tokens)
                min_tokens_per_item = min(min_tokens_per_item, item_total_tokens)

                # Print detailed token information for each item
                print(f"\nToken details for barcode {barcode}:")
                print(f"  Prompt tokens: {response.usage.prompt_tokens}")
                print(f"  Completion tokens: {response.usage.completion_tokens}")
                print(f"  Total tokens: {item_total_tokens}")
                print(f"  Average tokens per item so far: {total_tokens / items_processed:.2f}")

                metadata_output = response.choices[0].message.content.strip()
                print(f"Extracted Metadata for barcode {barcode}: {metadata_output}")

                row_data = ['', '', '', barcode, metadata_output]
                ws.append(row_data)

                # Add thumbnail images to Excel
                for i, img_path in enumerate(image_paths, start=1):
                    img = PILImage.open(img_path)
                    img.thumbnail((200, 200))

                    output = BytesIO()
                    img.save(output, format='PNG')
                    output.seek(0)

                    img_openpyxl = Image(output)
                    img_openpyxl.anchor = ws.cell(row=ws.max_row, column=i).coordinate
                    ws.add_image(img_openpyxl)

                ws.row_dimensions[ws.max_row].height = 215

                for cell in ws[ws.max_row]:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            except Exception as e:
                print(f"Error generating content for barcode {barcode}: {str(e)}")
                ws.append(['', '', '', barcode, f"Error: {str(e)}"])
                items_with_issues += 1

        except Exception as e:
            print(f"Error processing barcode {barcode}: {str(e)}")
            ws.append(['', '', '', barcode, f"Error: {str(e)}"])
            items_with_issues += 1

        print(f"Processed {processed_items}/{total_items} items. Current barcode: {barcode}")

    print("\nFinal Token Statistics:")
    print(f"Total tokens used: {total_tokens:,}")
    print(f"Total prompt tokens: {prompt_tokens:,}")
    print(f"Total completion tokens: {completion_tokens:,}")
    print(f"Average tokens per item: {total_tokens / items_processed:.2f}")
    print(f"Maximum tokens for a single item: {max_tokens_per_item:,}")
    print(f"Minimum tokens for a single item: {min_tokens_per_item:,}")
    print(f"Processed {total_items} items. {items_with_issues} items had issues.")
    
    return total_items, items_with_issues

def main():
    start_time = time.time()
    
    base_dir = "/Users/hannahmoutran/Library/CloudStorage/Box-Box/ai-music-metadata-project"
    images_folder = os.path.join(base_dir, "vinyl-scans-9")
    
    # Create results folder with today's date
    current_date = datetime.now().strftime("%Y-%m-%d")
    results_folder_name = f"results-lp-{current_date}"
    results_folder_path = os.path.join(base_dir, results_folder_name)
    
    # Create the results folder if it doesn't exist
    if not os.path.exists(results_folder_path):
        os.makedirs(results_folder_path)
        print(f"Created results folder: {results_folder_path}")
    
    wb = Workbook()
    total_items, items_with_issues = process_folder(images_folder, wb)

    for row in wb.active.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.active.freeze_panes = 'A2'

    output_file = f"ai-music-step-1-9-scans-lp-4o-{current_date}.xlsx"
    full_output_path = os.path.join(results_folder_path, output_file)

    wb.save(full_output_path)
    
    end_time = time.time()
    execution_time = end_time - start_time
    
    print(f"\nExecution Summary:")
    print(f"Results saved to {full_output_path}")
    print(f"Total execution time: {execution_time:.2f} seconds")
    print(f"Average time per item: {execution_time / total_items:.2f} seconds")
    
if __name__ == "__main__":
    main()