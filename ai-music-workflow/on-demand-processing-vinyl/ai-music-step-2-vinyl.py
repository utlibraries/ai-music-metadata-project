import os
import json
import requests
import time
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import re

api_calls = {'count': 0, 'reset_time': time.time()}

def get_access_token(client_id, client_secret):
    token_url = "https://oauth.oclc.org/token"
    data = {
        "grant_type": "client_credentials",
        "scope": "wcapi"
    }
    response = requests.post(token_url, data=data, auth=(client_id, client_secret))
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        raise Exception(f"Failed to get access token: {response.text}")
    
def extract_metadata_fields(metadata_str):
    """
    Extracts specific metadata fields from the structured metadata string.
    """
    fields = {
        "Main Title": None,
        "English Title": None,
        "Subtitle": None,
        "Primary Contributor": {"Artist/Performer": None},
        "Publishers": [{"Name": None}],
        "Language": {"pressingLanguage": None},
        "Contents": {"tracks": []},
    }

    # Patterns that match based on the hierarchical structure
    patterns = {
        "Main Title": r"-\s*Main Title:\s*([^\n]+?)(?=\n|$)",
        "English Title": r"-\s*English Title:\s*([^\n]*?)(?=\n|$)",
        "Subtitle": r"-\s*Subtitle:\s*([^\n]*?)(?=\n|$)",
        "Artist": r"-\s*Artist/Performer:\s*([^\n]+?)(?=\n|$)",
        "Publisher": r"-\s*Name:\s*([^\n]+?)(?=\n|\s*Place:)",
        "pressingLanguage": r"-\s*pressingLanguage:\s*([^\n]+?)(?=\n|$)"
    }

    def clean_value(value):
        if not value or value.lower() == "not visible":
            return None
        # Remove any field labels or hierarchy markers
        value = re.sub(r'^-\s*', '', value)
        value = re.sub(r'^(Main Title:|English Title:|Subtitle:|Primary Contributor:|Artist/Performer:|Name:)\s*', '', value)
        return value.strip()

    # Extract fields
    for field, pattern in patterns.items():
        match = re.search(pattern, metadata_str)
        if match:
            value = clean_value(match.group(1))
            if value:
                if field == "Artist":
                    fields["Primary Contributor"]["Artist/Performer"] = value
                elif field == "Publisher":
                    fields["Publishers"][0]["Name"] = value
                elif field == "pressingLanguage":
                    fields["Language"]["pressingLanguage"] = value
                else:
                    fields[field] = value
            elif field == "English Title":  # Store empty English title to distinguish from missing
                fields["English Title"] = value

    # Extract tracks using JSON-style parsing
    track_matches = re.finditer(r'"title":\s*"([^"]+)"', metadata_str)
    for match in track_matches:
        track_title = clean_value(match.group(1))
        if track_title:
            fields["Contents"]["tracks"].append({"title": track_title})

    return fields

def construct_queries_from_metadata(metadata):
    def safe_get(value):
        if not value or not isinstance(value, str) or value.lower() in ["not visible", "not applicable"]:
            return None
        cleaned = re.sub(r'^-\s*', '', value.strip())
        cleaned = re.sub(r'^(Primary Contributor:|Artist/Performer:|Name:)\s*', '', cleaned)
        return cleaned if cleaned else None

    # Get fields - prioritize English Title
    english_title = safe_get(metadata.get('English Title'))
    main_title = safe_get(metadata.get('Main Title'))
    title = english_title if english_title else main_title
    
    if not title:
        return []

    subtitle = safe_get(metadata.get('Subtitle'))
    artist = safe_get(metadata.get('Primary Contributor', {}).get('Artist/Performer'))
    publisher = safe_get(metadata.get('Publishers', [{}])[0].get('Name'))
    pressing_language = safe_get(metadata.get('Language', {}).get('pressingLanguage'))
    tracks = metadata.get('Contents', {}).get('tracks', [])
    first_track_title = safe_get(tracks[0].get('title')) if tracks else None
    lp = "LP"

    queries = []

    # 1. [Title, Subtitle, Artist/Performer, Pressing Language, LP]
    if title:  # Only proceed if we have a title
        components = [title]
        if subtitle:
            components.append(subtitle)
        if artist:
            components.append(artist)
        if pressing_language:
            components.append(pressing_language)
        components.append(lp)
        queries.append(" ".join(components))

    # 2. [Title, Subtitle, Pressing Language, LP]
    if title:  # Only proceed if we have a title
        components = [title]
        if subtitle:
            components.append(subtitle)
        if pressing_language:
            components.append(pressing_language)
        components.append(lp)
        second_query = " ".join(components)
        if second_query not in queries:
            queries.append(second_query)
        
    # 3. [Artist, Publisher, LP]
    if artist and publisher:  # Only proceed if we have both artist and publisher
        components = [artist, publisher, lp]
        third_query = " ".join(components)
        if third_query not in queries:
            queries.append(third_query)

    # 4. [Track Title, Artist/Performer, Publisher, LP]
    if first_track_title and artist and publisher:  # Only proceed if we have all components
        components = [first_track_title, artist, publisher, lp]
        fourth_query = " ".join(components)
        if fourth_query not in queries:
            queries.append(fourth_query)

    # 5. [Title, Publisher, LP]
    if title and publisher:  # Only proceed if we have both title and publisher
        components = [title, publisher, lp]
        fifth_query = " ".join(components)
        if fifth_query not in queries:
            queries.append(fifth_query)

    # Filter out any queries that are too short
    return [q for q in queries if len(q.strip()) > 5]

def format_oclc_results(json_response):
    try:
        data = json.loads(json_response)
        if not isinstance(data, dict):
            return "Error: Invalid JSON response"
            
        total_records = data.get('numberOfRecords', 0)
        if total_records == 0:
            return "No records found"
            
        formatted_results = []
        formatted_results.append(f"Total Records Found: {total_records}\n")
        
        for idx, record in enumerate(data.get('bibRecords', [])[:5], 1):
            formatted_results.append(f"\nRecord {idx}:")
            formatted_results.append("-" * 40)
            
            # Identifier
            if 'identifier' in record:
                formatted_results.append("Identifier:")
                for key, value in record['identifier'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            # Title
            if 'title' in record:
                formatted_results.append("Title Information:")
                if 'mainTitles' in record['title']:
                    for title in record['title']['mainTitles']:
                        formatted_results.append(f"  - Main Title: {title.get('text', 'N/A')}")
                if 'subtitles' in record['title']:
                    for subtitle in record['title']['subtitles']:
                        formatted_results.append(f"  - Subtitle: {subtitle.get('text', 'N/A')}")
            
            # Contributors
            if 'contributor' in record:
                formatted_results.append("Contributors:")
                for creator_type in ['creators', 'contributors']:
                    if creator_type in record['contributor']:
                        for person in record['contributor'][creator_type]:
                            if 'firstName' in person and 'secondName' in person:
                                name = f"{person.get('firstName', {}).get('text', '')} {person.get('secondName', {}).get('text', '')}"
                            elif 'nonPersonName' in person:
                                name = person['nonPersonName'].get('text', '')
                            else:
                                name = 'N/A'
                            role = person.get('type', 'N/A')
                            formatted_results.append(f"  - {name.strip()} ({role})")
            
            # Subjects
            if 'subjects' in record:
                formatted_results.append("Subjects:")
                for subject in record['subjects']:
                    if 'text' in subject:
                        formatted_results.append(f"  - {subject['text']}")
            
            # Classification
            if 'classification' in record:
                formatted_results.append("Classification:")
                for key, value in record['classification'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            # Publishers
            if 'publishers' in record:
                formatted_results.append("Publishers:")
                for pub in record['publishers']:
                    pub_name = pub.get('publisherName', {}).get('text', 'N/A')
                    pub_place = pub.get('publicationPlace', 'N/A')
                    formatted_results.append(f"  - Name: {pub_name}")
                    formatted_results.append(f"    Place: {pub_place}")
            
            # Date
            if 'date' in record:
                formatted_results.append("Dates:")
                for key, value in record['date'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            # Language
            if 'language' in record:
                formatted_results.append("Language:")
                for key, value in record['language'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            # Format
            if 'format' in record:
                formatted_results.append("Format:")
                for key, value in record['format'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            # Music Info
            if 'musicInfo' in record:
                formatted_results.append("Music Information:")
                for key, value in record['musicInfo'].items():
                    formatted_results.append(f"  - {key}: {value}")
            
            # Description
            if 'description' in record:
                formatted_results.append("Description:")
                if 'physicalDescription' in record['description']:
                    formatted_results.append(f"  - Physical: {record['description']['physicalDescription']}")
                if 'contents' in record['description']:
                    for content in record['description']['contents']:
                        if 'contentNote' in content:
                            formatted_results.append(f"  - Content: {content['contentNote'].get('text', '')}")
            
            # Digital Access
            if 'digitalAccessAndLocations' in record:
                formatted_results.append("Digital Access:")
                for access in record['digitalAccessAndLocations']:
                    formatted_results.append(f"  - {access}")
            
            # Notes
            if 'note' in record:
                formatted_results.append("Notes:")
                if isinstance(record['note'], dict):
                    for key, value in record['note'].items():
                        formatted_results.append(f"  - {key}: {value}")
                elif isinstance(record['note'], list):
                    for note in record['note']:
                        formatted_results.append(f"  - {note}")
            
            formatted_results.append("-" * 40)
            
        return "\n".join(formatted_results)
        
    except json.JSONDecodeError:
        return "Error: Invalid JSON response"
    except Exception as e:
        return f"Error formatting results: {str(e)}"

def clean_text(text):
    # Map accented characters to ASCII equivalents
    accent_map = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'ã': 'a', 'õ': 'o', 'ñ': 'n',
        'â': 'a', 'ê': 'e', 'î': 'i', 'ô': 'o', 'û': 'u'
    }
    # Replace accented characters first
    for accented, ascii in accent_map.items():
        text = text.replace(accented, ascii)
    # Then remove remaining non-ASCII
    return ''.join(c for c in text if ord(c) < 128).strip()

def clean_title(title):
    # Extract romanized version if in parentheses
    if '(' in title and ')' in title:
        romanized = title[title.find('(')+1:title.find(')')]
        return clean_text(romanized)
    # Use part after slash if present
    elif '/' in title:
        return clean_text(title.split('/')[1])
    # Apply ASCII cleaning to original title
    return clean_text(title)

def truncate_contributors(performers, max_performers=3):
    return performers[:max_performers]

def remove_non_latin(text):
    # Keep spaces, basic punctuation, and alphanumeric characters
    cleaned = re.sub(r'[^\x00-\x7F\s\-\/\(\)]', '', text)
    # Remove multiple spaces
    cleaned = re.sub(r'\s+', ' ', cleaned)
    # Remove empty parentheses
    cleaned = re.sub(r'\(\s*\)', '', cleaned)
    return cleaned.strip()

def query_oclc_api(metadata, barcode, limit=5):
    global api_calls

    current_time = time.time()
    if current_time - api_calls['reset_time'] >= 86400:
        api_calls = {'count': 0, 'reset_time': current_time}

    if api_calls['count'] >= 50000:
        return "Rate limit exceeded. Please try again later.", {}

    client_id = os.environ.get("OCLC_CLIENT_ID")
    client_secret = os.environ.get("OCLC_SECRET")
    
    if not client_id or not client_secret:
        return "Error: OCLC_CLIENT_ID and OCLC_SECRET must be set in environment variables", {}

    try:
        access_token = get_access_token(client_id, client_secret)
        print(f"Access token obtained: {access_token[:5]}...")
    except Exception as e:
        return f"Error getting access token: {str(e)}", {}

    base_url = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
    endpoint = f"{base_url}/bibs"

    queries = metadata.get("Queries", [])
    if not isinstance(queries, list):
        return "Error: Invalid query format", "Queries must be a list of strings"

    # Clean queries
    cleaned_queries = []
    for q in queries:
        if isinstance(q, str) and q.strip():
            cleaned = remove_non_latin(q.replace(str(barcode), ""))
            if len(cleaned.strip()) >= 3:  # Ensure query has meaningful content
                cleaned_queries.append(cleaned.strip())

    if not cleaned_queries:
        return "No valid queries could be constructed", "Please check the metadata format"

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }

    query_log = []
    for idx, query in enumerate(cleaned_queries):
        params = {
            "q": query,
            "limit": limit,
            "offset": 1,
            "itemType": "music",
            "inCatalogLanguage": "eng",
            "itemSubType": "music-lp"
        }

        query_log.append(f"Query {idx + 1}: {query}")

        try:
            response = requests.get(endpoint, params=params, headers=headers)
            api_calls['count'] += 1
            response.raise_for_status()
            data = response.json()

            num_records = data.get("numberOfRecords", 0)
            
            # Skip if more than 1000 records found (too broad)
            if num_records > 1000:
                query_log.append(f"Query {idx + 1} skipped: Too many results ({num_records} records)")
                continue
                
            if num_records > 0:
                formatted_results = format_oclc_results(response.text)
                success_log = f"Query {idx + 1} succeeded"
                query_log.append(success_log)
                return formatted_results, "\n".join(query_log)
            
        except requests.RequestException as e:
            api_calls['count'] += 1
            query_log.append(f"Query {idx + 1} failed: {str(e)}")

    return "No matching records found", "\n".join(query_log)
    
def process_metadata_file(input_file):
    wb = load_workbook(input_file)
    ws = wb.active

    if ws['F1'].value != 'OCLC Query':
        ws.insert_cols(6)
        ws['F1'] = 'OCLC Query'
    if ws['G1'].value != 'OCLC API Results':
        ws.insert_cols(7)
        ws['G1'] = 'OCLC API Results'
        
    ws.column_dimensions['F'].width = 52
    ws.column_dimensions['G'].width = 52
    
    total_rows = ws.max_row
    for row in range(2, total_rows + 1):
        metadata_str = ws.cell(row=row, column=5).value  # Column E
        barcode = ws.cell(row=row, column=4).value       # Column D
        if not metadata_str or metadata_str.startswith('Error'):
            continue

        try:
            metadata_fields = extract_metadata_fields(metadata_str)
            if not isinstance(metadata_fields, dict) or not metadata_fields.get("Main Title"):
                raise ValueError("Missing 'Main Title' in metadata")
            
            queries = construct_queries_from_metadata(metadata_fields)
            results, query_log = query_oclc_api({"Queries": queries}, barcode)
            
            ws.cell(row=row, column=6, value=query_log)
            ws.cell(row=row, column=7, value=results)
            ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)
            print(f"Processed row {row}/{total_rows}")
            time.sleep(0.1)

        except Exception as e:
            error_message = f"Error processing row {row}: {str(e)}"
            print(error_message)
            ws.cell(row=row, column=6, value="Error processing")
            ws.cell(row=row, column=7, value=error_message)
            ws.cell(row=row, column=6).alignment = Alignment(vertical='top', wrap_text=True)
            ws.cell(row=row, column=7).alignment = Alignment(vertical='top', wrap_text=True)

    return wb  # Return the workbook after processing

def find_latest_results_folder(base_dir):
    """Find the latest results-lp-YYYY-MM-DD folder in the base directory."""
    results_folders = [folder for folder in os.listdir(base_dir) 
                      if os.path.isdir(os.path.join(base_dir, folder)) 
                      and folder.startswith('results-lp-')]
    
    if not results_folders:
        return None
    
    # Extract dates from folder names and find the latest
    latest_folder = max(results_folders, 
                        key=lambda folder: datetime.strptime(folder.replace('results-lp-', ''), '%Y-%m-%d'))
    
    return os.path.join(base_dir, latest_folder)

def main():
    start_time = time.time()  # Added time tracking
    
    base_dir = "/Users/hannahmoutran/Library/CloudStorage/Box-Box/ai-music-metadata-project"
    
    # Find the latest results folder
    results_folder = find_latest_results_folder(base_dir)
    if not results_folder:
        print("No results folder found! Running the first script is required before this one.")
        return
    
    print(f"Using results folder: {results_folder}")
    
    # Look for step 1 file in the results folder
    step1_files = [f for f in os.listdir(results_folder) 
                  if f.startswith('ai-music-step-1-9-scans-lp-4o-') and f.endswith('.xlsx')]
    
    if not step1_files:
        print(f"No initial metadata files found in {results_folder}!")
        return
        
    # Select the latest file based on the date in the filename
    latest_file = max(step1_files, 
                      key=lambda x: datetime.strptime(x.replace('ai-music-step-1-9-scans-lp-4o-', '').replace('.xlsx', ''), "%Y-%m-%d"))  

    input_file = os.path.join(results_folder, latest_file)
    
    print(f"Processing file: {input_file}")
    wb = process_metadata_file(input_file)
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"ai-music-step-2-9-scans-lp-4o-{current_date}.xlsx"  
    full_output_path = os.path.join(results_folder, output_file)
    
    wb.save(full_output_path)
    
    end_time = time.time()  # Added time tracking
    execution_time = end_time - start_time  # Calculate total execution time
    
    print(f"Results saved to {full_output_path}")
    print("Summary: Process completed.")
    print(f"Total execution time: {execution_time:.2f} seconds")  # Added execution time print

if __name__ == "__main__":
    main()